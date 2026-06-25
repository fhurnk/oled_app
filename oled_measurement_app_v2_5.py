#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
OLED Measurement App
====================
GUI-приложение для ведения серии OLED-измерений:
- создание / открытие серии напыления;
- назначение названий 4 четвертей подложкодержателя;
- автоматическое создание списка 4 x 3 x 4 пикселей;
- журнал серии в Excel;
- ВАЯХ одного пикселя или всей серии;
- спектры только после известной ВАЯХ;
- стабильность по току только после известной ВАЯХ;
- дневные папки измерений.

Важно:
- xtralien и seabreeze импортируются только в момент измерения, поэтому приложение можно
  открыть без подключенного оборудования и проверить журнал / структуру серии.
- Перед первым запуском проверь COM_PORT и параметры по умолчанию в окнах измерений.
"""

from __future__ import annotations

import gc
import ctypes
import json
import math
import os
import random
import re
import sys
import threading
import time
import types
import traceback
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from copy import deepcopy
from typing import Any, Callable, Dict, Iterable, List, Optional, Tuple

import numpy as np
import tkinter as tk
from tkinter import filedialog, messagebox, simpledialog, ttk
from tkinter import font as tkfont
from tkinter.scrolledtext import ScrolledText

from openpyxl import Workbook, load_workbook
from openpyxl.chart import ScatterChart, Series, Reference
from openpyxl.chart.axis import ChartLines
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

APP_VERSION = "1.6.0"
SCRIPT_DIR = Path(__file__).resolve().parent
CONFIG_FILE = "series_config.json"
JOURNAL_FILE = "series_journal.xlsx"
DEFAULT_ROOT = "OLED_series"
APP_SETTINGS_FILE = "oled_app_settings.json"
SIM_CONFIG_FILE = "oled_simulator_config.json"
HARDWARE_MODE_REAL = "real"
HARDWARE_MODE_SIM = "simulator"
MEASUREMENT_FOLDER_NAMES = {
    "IVL": "01_IVL_VAH",
    "SPECTRUM": "02_SPECTRA",
    "STABILITY": "03_STABILITY",
}


def enable_windows_dpi_awareness() -> None:
    if sys.platform != "win32":
        return
    try:
        ctypes.windll.user32.SetProcessDpiAwarenessContext(ctypes.c_void_p(-4))
        return
    except Exception:
        pass
    try:
        ctypes.windll.shcore.SetProcessDpiAwareness(2)
        return
    except Exception:
        pass
    try:
        ctypes.windll.user32.SetProcessDPIAware()
    except Exception:
        pass


class MeasurementStopped(Exception):
    pass

PIXELS_SHEET = "Pixels"
MEASUREMENTS_SHEET = "Measurements"
SERIES_SHEET = "Series"
QUARTERS_SHEET = "Quarters"

PIXEL_HEADERS = [
    "Pixel ID",
    "Quarter code",
    "Quarter number",
    "Substrate number",
    "Pixel number",
    "Last status",
    "Opening voltage (V)",
    "Last IVL date",
    "Last IVL file",
    "Last IVL max current (mA)",
    "Last IVL max photodiode (uA)",
    "Last spectrum date",
    "Last spectrum file",
    "Last spectrum peak count",
    "Last spectrum peaks nm",
    "Last spectrum max intensity (counts/s)",
    "Last stability date",
    "Last stability file",
    "Last updated",
]

MEASUREMENT_HEADERS = [
    "Date time",
    "Measurement day",
    "Type",
    "Pixel ID",
    "Status",
    "File",
    "Params JSON",
    "Notes",
]


# -----------------------------------------------------------------------------
# Общие утилиты
# -----------------------------------------------------------------------------

def now_str() -> str:
    return datetime.now().strftime("%d-%m-%Y %H:%M:%S")


def timestamp_for_file() -> str:
    return datetime.now().strftime("%d-%m-%Y_%Hh%Mm%Ss")


def today_iso() -> str:
    return date.today().isoformat()


def safe_filename(text: str, fallback: str = "item") -> str:
    text = (text or "").strip()
    text = re.sub(r"[^0-9A-Za-zА-Яа-яЁё._\-]+", "_", text)
    text = re.sub(r"_+", "_", text).strip("_.-")
    return text or fallback


def as_float_or_none(value) -> Optional[float]:
    if value in (None, "", "—"):
        return None
    try:
        return float(value)
    except Exception:
        return None


def current_density_mA_cm2(current_mA: Any, pixel_area_mm2: Any) -> Optional[float]:
    current = as_float_or_none(current_mA)
    area = as_float_or_none(pixel_area_mm2)
    if current is None or area is None or area <= 0:
        return None
    return float(current) / (float(area) / 100.0)


def luminance_cd_m2(photo_uA: Any, conversion_cd_m2_per_uA: Any) -> Optional[float]:
    photo = as_float_or_none(photo_uA)
    coeff = as_float_or_none(conversion_cd_m2_per_uA)
    if photo is None or coeff is None:
        return None
    return float(photo) * float(coeff)


def relative_to_or_abs(path: Path, base: Path) -> str:
    try:
        return str(path.resolve().relative_to(base.resolve()))
    except Exception:
        return str(path.resolve())


def resolve_series_file(series_folder: Path, file_value: Any) -> Optional[Path]:
    if not file_value:
        return None
    path = Path(str(file_value))
    if not path.is_absolute():
        path = Path(series_folder) / path
    return path if path.exists() else None


def read_spectrum_metrics_from_workbook(path: Optional[Path]) -> Dict[str, Any]:
    if not path or not path.exists():
        return {}
    try:
        wb = load_workbook(path, data_only=True, read_only=True)
        ws = wb.active
        header_row = 14
        headers = {}
        for candidate in range(1, min(ws.max_row, 40) + 1):
            row_headers = {str(ws.cell(candidate, c).value or ""): c for c in range(1, ws.max_column + 1)}
            if "Peaks detected" in row_headers or "Max intensity processed (counts)" in row_headers or "Max intensity processed (counts/s)" in row_headers or "Max intensity (counts)" in row_headers:
                header_row = candidate
                headers = row_headers
                break
        peaks_col = headers.get("Peaks detected")
        peaks_nm_col = headers.get("Peaks nm")
        max_int_col = headers.get("Max intensity processed (counts)") or headers.get("Max intensity processed (counts/s)") or headers.get("Max intensity (counts)")
        best = {"peak_count": "", "peaks_nm": "", "max_intensity": ""}
        best_intensity = -1.0
        for row in range(header_row + 1, ws.max_row + 1):
            intensity = as_float_or_none(ws.cell(row, max_int_col).value) if max_int_col else None
            if intensity is None:
                continue
            if intensity >= best_intensity:
                best_intensity = float(intensity)
                best = {
                    "peak_count": ws.cell(row, peaks_col).value if peaks_col else "",
                    "peaks_nm": ws.cell(row, peaks_nm_col).value if peaks_nm_col else "",
                    "max_intensity": round(float(intensity), 1),
                }
        wb.close()
        return best
    except Exception:
        return {}


def ensure_day_folder(series_folder: Path) -> Path:
    day_folder = series_folder / "measurements" / today_iso()
    day_folder.mkdir(parents=True, exist_ok=True)
    return day_folder


def ensure_measurement_folder(
    series_folder: Path,
    measurement_type: str,
    pixel_id: str,
    pixel_row: Optional[Dict[str, Any]] = None,
) -> Path:
    measurement_folder = MEASUREMENT_FOLDER_NAMES.get(
        str(measurement_type).upper(),
        safe_filename(str(measurement_type), fallback="measurement"),
    )
    pixel_row = pixel_row or {}

    quarter_number = pixel_row.get("Quarter number") or "unknown"
    quarter_code = pixel_row.get("Quarter code") or "Q"
    substrate_number = pixel_row.get("Substrate number") or "unknown"

    quarter_name = safe_filename(f"{quarter_code}{quarter_number}", fallback=f"Q{quarter_number}")
    substrate_folder = safe_filename(f"{quarter_name}_{substrate_number}", fallback=f"{quarter_name}_unknown")
    pixel_folder = safe_filename(pixel_id, fallback="pixel")

    output_dir = (
        series_folder
        / "measurements"
        / measurement_folder
        / today_iso()
        / quarter_name
        / substrate_folder
        / pixel_folder
    )
    output_dir.mkdir(parents=True, exist_ok=True)
    return output_dir


def pixel_status_color(status: str) -> str:
    status = str(status or "").upper()
    if status == "WORKING":
        return "#8FD694"
    if status == "NO_CONTACT":
        return "#F2D96B"
    if status in {"NONWORKING", "BURNED", "FAILED", "CURRENT_LIMIT_STOP", "CURRENT_LIMIT"}:
        return "#F28B82"
    return "#D9D9D9"


def ivl_status_marker(status: str) -> str:
    status = str(status or "").upper()
    if status == "WORKING":
        return "↑ WORKING"
    if status == "NO_CONTACT":
        return "→ NO_CONTACT"
    if status in {"NONWORKING", "FAILED"}:
        return "↓ " + status
    if status in {"BURNED", "CURRENT_LIMIT_STOP", "CURRENT_LIMIT"}:
        return "↯ " + status
    return "· " + (status or "")


def build_holder_layout(width: int = 930, height: int = 620) -> Dict[int, Dict[str, Any]]:
    """Геометрия подложкодержателя для карты и окна создания серии.

    Тексты сведены к минимуму: остаются большие номера четвертей и короткие
    подписи самих подложек. Даты и легенда разведены по вертикали, чтобы не
    наезжали на подложки при увеличенном масштабе Windows.
    """
    box_w = 86
    box_h = 52

    # Подложки расположены внутри овала с большим запасом между соседями.
    left_x1, left_x2, left_x3 = 170, 305, 238
    right_x1, right_x2, right_x3 = width - 390, width - 255, width - 322
    top_y1, top_y3 = 170, 260
    bottom_y1, bottom_y3 = 320, 405

    quarter_layout = {
        2: {
            "number_xy": (48, 92),
            "name_xy": (92, 36),
            "entry_xy": (78, 62),
            "substrates": [(left_x1, top_y1), (left_x2, top_y1), (left_x3, top_y3)],
        },
        1: {
            "number_xy": (width - 48, 92),
            "name_xy": (width - 248, 36),
            "entry_xy": (width - 230, 62),
            "substrates": [(right_x1, top_y1), (right_x2, top_y1), (right_x3, top_y3)],
        },
        3: {
            "number_xy": (48, height - 118),
            "name_xy": (92, height - 155),
            "entry_xy": (78, height - 130),
            "substrates": [(left_x1, bottom_y1), (left_x2, bottom_y1), (left_x3, bottom_y3)],
        },
        4: {
            "number_xy": (width - 48, height - 118),
            "name_xy": (width - 248, height - 155),
            "entry_xy": (width - 230, height - 130),
            "substrates": [(right_x1, bottom_y1), (right_x2, bottom_y1), (right_x3, bottom_y3)],
        },
    }

    for q, info in quarter_layout.items():
        detailed = []
        for substrate_number, (x, y) in enumerate(info["substrates"], start=1):
            detailed.append({"substrate_number": substrate_number, "x": x, "y": y, "w": box_w, "h": box_h})
        info["substrates"] = detailed
    return quarter_layout


def short_date_for_map(value: str) -> str:
    """Короткая дата для карты подложкодержателя, чтобы подписи не слипались."""
    text = str(value or "").strip()
    try:
        return datetime.strptime(text, "%Y-%m-%d").strftime("%d.%m.%y")
    except Exception:
        return text

# -----------------------------------------------------------------------------
# Настройки приложения, режим реального оборудования / эмулятора
# -----------------------------------------------------------------------------

DEFAULT_SIMULATOR_CONFIG: Dict[str, Any] = {
    "active": True,
    "global": {
        "random_seed": 42,
        "current_noise_relative": 0.01,
        "photodiode_noise_uA": 0.01,
        "voltage_noise_V": 0.002,
        "spectrum_noise_relative": 0.008,
    },
    "default_pixel": {
        "mode": "working",
        "turn_on_voltage_V": 2.65,
        "current_at_5V_mA": 6.0,
        "iv_exponent": 2.0,
        "leakage_mA": 0.0005,
        "max_current_mA": 30.0,
        "short_resistance_ohm": 120.0,
        "burnout_voltage_V": None,
        "burnout_after_s": None,
        "photodiode_gain_uA_per_mA": 0.55,
        "degradation_tau_s": None,
        "voltage_drift_V_per_s": 0.0,
        "spectrum": {
            "dark_offset_counts": 120.0,
            "dark_counts_per_s": 250.0,
            "background_counts_per_s": 50.0,
            "counts_per_mA_per_s": 1_200_000.0,
            "saturation_counts": 65535.0,
            "peaks": [
                {"center_nm": 465.0, "fwhm_nm": 42.0, "amplitude": 0.75},
                {"center_nm": 535.0, "fwhm_nm": 70.0, "amplitude": 1.00},
                {"center_nm": 625.0, "fwhm_nm": 85.0, "amplitude": 0.62},
            ],
        },
    },
    "pixels": {
        "Q1_1_1": {"mode": "working", "turn_on_voltage_V": 2.60, "current_at_5V_mA": 6.5, "photodiode_gain_uA_per_mA": 0.60},
        "Q1_1_2": {"mode": "weak", "photodiode_gain_uA_per_mA": 0.05, "spectrum": {"counts_per_mA_per_s": 130_000.0}},
        "Q1_1_3": {"mode": "nonworking", "photodiode_gain_uA_per_mA": 0.0},
        "Q1_1_4": {"mode": "no_contact", "leakage_mA": 0.001},
        "Q1_2_1": {"mode": "working", "burnout_voltage_V": 3.45, "short_resistance_ohm": 80.0},
        "Q1_2_2": {"mode": "working", "degradation_tau_s": 35.0, "voltage_drift_V_per_s": 0.001},
        "Q1_2_3": {
            "mode": "working",
            "turn_on_voltage_V": 2.75,
            "current_at_5V_mA": 5.5,
            "spectrum": {
                "peaks": [
                    {"center_nm": 455.0, "fwhm_nm": 38.0, "amplitude": 0.95},
                    {"center_nm": 540.0, "fwhm_nm": 72.0, "amplitude": 0.90},
                    {"center_nm": 610.0, "fwhm_nm": 92.0, "amplitude": 0.70},
                ]
            },
        },
    },
}

DEFAULT_APP_SETTINGS: Dict[str, Any] = {
    "default_root": str(SCRIPT_DIR / DEFAULT_ROOT),
    "hardware_mode": HARDWARE_MODE_SIM,
    "com_port": "SIM",
    "simulator_config_path": str(SCRIPT_DIR / SIM_CONFIG_FILE),
    "ivl_advanced": {
        "photodiode_bias_V": -5.0,
        "photodiode_range": 4,
        "photodiode_threshold_uA": 0.5,
        "burnout_current_threshold_mA": 10.0,
        "mark_current_limit_as_burnout": False,
        "no_contact_max_led_current_mA": 0.05,
        "burned_confirmation_cycles": 1,
    },
    "spectrum_advanced": {
        "photodiode_bias_V": -5.0,
        "photodiode_range": 4,
        "target_intensity": 40000.0,
        "intensity_min": 20000.0,
        "intensity_max": 55000.0,
        "saturation_level": 60000.0,
        "min_peak_width_nm": 15.0,
        "max_peak_width_nm": 150.0,
        "t_int_initial_s": 0.01,
        "t_int_min_s": 0.001,
        "t_int_max_s": 10.0,
        "discard_first_scan_after_tint_change": True,
        "kp": 0.3,
        "ki": 0.05,
        "max_iterations": 20,
        "tolerance": 0.05,
        "peak_search_mode_for_tint": "auto",
        "settle_time_voltage_s": 0.1,
        "settle_time_spectrum_s": 0.05,
        "dark_spectrum_enabled": False,
        "dark_spectrum_scans": 3,
        "baseline_correction_enabled": True,
        "peak_detection_enabled": False,
    },
    "measurement_units": {
        "pixel_area_mm2": 1.0,
        "luminance_cd_m2_per_uA": 1.0,
    },
    "stability_advanced": {
        "voltage_step_max": 0.02,
        "current_control_kp": 0.01,
        "photodiode_bias_V": -5.0,
        "photodiode_threshold_uA": 0.1,
        "photodiode_range": 4,
    },
    "ui": {
        "last_ivl_graph_mode": "raw",
    },
    "measurement_defaults": {
        "ivl": {
            "sweep_start_V": "0",
            "sweep_end_V": "5",
            "step_V": "0.02",
            "time_per_point_s": "0.01",
            "cycles": "1",
            "delay_between_cycles_s": "1",
            "current_limit_mA": "10",
        },
        "spectrum": {
            "voltage_end_V": "5",
            "voltage_step_V": "0.1",
            "current_limit_mA": "6",
            "led_type": "auto",
            "use_opening_voltage": True,
        },
        "stability": {
            "current_setpoint_mA": "3.5",
            "voltage_limit_V": "5",
            "current_limit_mA": "10",
            "measurement_time_s": "86400",
            "sample_interval_s": "1",
            "autosave_interval_s": "600",
        },
    },
}


def deep_update(base: Dict[str, Any], update: Dict[str, Any]) -> Dict[str, Any]:
    out = deepcopy(base)
    for key, value in (update or {}).items():
        if isinstance(value, dict) and isinstance(out.get(key), dict):
            out[key] = deep_update(out[key], value)
        else:
            out[key] = value
    return out


def app_settings_path() -> Path:
    return SCRIPT_DIR / APP_SETTINGS_FILE


def load_app_settings() -> Dict[str, Any]:
    path = app_settings_path()
    settings = deepcopy(DEFAULT_APP_SETTINGS)
    if path.exists():
        try:
            loaded = json.loads(path.read_text(encoding="utf-8"))
            settings = deep_update(settings, loaded)
        except Exception:
            pass
    return settings


def save_app_settings(settings: Dict[str, Any]) -> None:
    app_settings_path().write_text(json.dumps(settings, ensure_ascii=False, indent=2), encoding="utf-8")


def fit_toplevel_to_content(win: tk.Toplevel, min_width: int, min_height: int, padding: int = 36) -> None:
    try:
        win.update_idletasks()
        screen_w = int(win.winfo_screenwidth())
        screen_h = int(win.winfo_screenheight())
        req_w = int(win.winfo_reqwidth()) + padding
        req_h = int(win.winfo_reqheight()) + padding
        width = min(max(min_width, req_w), max(320, screen_w - 80))
        height = min(max(min_height, req_h), max(260, screen_h - 100))
        x = max(0, (screen_w - width) // 2)
        y = max(0, (screen_h - height) // 3)
        win.geometry(f"{width}x{height}+{x}+{y}")
        win.minsize(min(width, min_width), min(height, min_height))
    except Exception:
        win.geometry(f"{min_width}x{min_height}")


def ensure_default_sim_config(config_path: Optional[Path] = None) -> Path:
    path = Path(config_path) if config_path else SCRIPT_DIR / SIM_CONFIG_FILE
    if not path.exists():
        path.write_text(json.dumps(DEFAULT_SIMULATOR_CONFIG, ensure_ascii=False, indent=2), encoding="utf-8")
    return path


def hardware_mode_label(settings: Dict[str, Any]) -> str:
    return "Эмулятор" if settings.get("hardware_mode") == HARDWARE_MODE_SIM else "Реальное оборудование"


# -----------------------------------------------------------------------------
# Встроенный эмулятор xtralien + seabreeze, включается только флажком в настройках
# -----------------------------------------------------------------------------

def probe_hardware(settings: Dict[str, Any]) -> Dict[str, Any]:
    mode = settings.get("hardware_mode", HARDWARE_MODE_SIM)
    if mode == HARDWARE_MODE_SIM:
        cfg_path = Path(settings.get("simulator_config_path") or SCRIPT_DIR / SIM_CONFIG_FILE)
        ensure_default_sim_config(cfg_path)
        return {
            "level": "ok",
            "title": "Эмулятор готов",
            "details": f"SIM, конфиг: {cfg_path.name}",
            "smu": "SIM OK",
            "spectrometer": "SIM OK",
        }

    result = {
        "level": "ok",
        "title": "Оборудование готово",
        "details": "",
        "smu": "",
        "spectrometer": "",
    }
    messages: List[str] = []
    com_port = str(settings.get("com_port") or "COM3")
    uninstall_simulator_modules()

    try:
        import xtralien
        with xtralien.Device(com_port) as smu:
            try:
                smu.smu1.set.voltage(0, response=0)
                smu.smu2.set.voltage(0, response=0)
            except Exception:
                pass
        result["smu"] = f"xtralien OK ({com_port})"
    except Exception as exc:
        result["level"] = "error"
        result["title"] = "SMU не отвечает"
        result["smu"] = f"xtralien ERROR ({com_port})"
        messages.append(f"SMU: {exc}")

    try:
        import seabreeze.spectrometers as sb
        devices = list(sb.list_devices())
        if devices:
            result["spectrometer"] = f"Спектрометр OK ({len(devices)})"
        else:
            if result["level"] != "error":
                result["level"] = "warning"
                result["title"] = "SMU готов, спектрометр не найден"
            result["spectrometer"] = "Спектрометр не найден"
            messages.append("seabreeze: устройств не найдено")
    except Exception as exc:
        if result["level"] != "error":
            result["level"] = "warning"
            result["title"] = "SMU готов, ошибка спектрометра"
        result["spectrometer"] = "seabreeze ERROR"
        messages.append(f"seabreeze: {exc}")

    result["details"] = "; ".join(messages) if messages else f"{result['smu']}; {result['spectrometer']}"
    return result


_SIM_STATE: Dict[str, Any] = {
    "pixel_id": "SIM_PIXEL",
    "session_start_monotonic": time.monotonic(),
    "smu1_voltage_set_V": 0.0,
    "smu2_voltage_set_V": 0.0,
    "smu1_enabled": False,
    "smu2_enabled": False,
    "led_current_mA": 0.0,
    "photodiode_uA": 0.0,
    "latched_burned": False,
    "rng": random.Random(42),
}


def sim_load_config(config_path: Optional[Path]) -> Dict[str, Any]:
    cfg = deepcopy(DEFAULT_SIMULATOR_CONFIG)
    path = Path(config_path) if config_path else SCRIPT_DIR / SIM_CONFIG_FILE
    if path.exists():
        try:
            cfg = deep_update(cfg, json.loads(path.read_text(encoding="utf-8")))
        except Exception:
            pass
    return cfg


def sim_global_settings() -> Dict[str, Any]:
    return sim_load_config(Path(os.environ.get("OLED_SIM_CONFIG", SCRIPT_DIR / SIM_CONFIG_FILE))).get("global", {})


def sim_current_pixel_id() -> str:
    return os.environ.get("OLED_SIM_PIXEL_ID", "SIM_PIXEL")


def sim_profile_for_pixel(pixel_id: Optional[str] = None) -> Dict[str, Any]:
    cfg = sim_load_config(Path(os.environ.get("OLED_SIM_CONFIG", SCRIPT_DIR / SIM_CONFIG_FILE)))
    pixel_id = pixel_id or sim_current_pixel_id()
    profile = deepcopy(cfg.get("default_pixel", {}))
    pixels_cfg = cfg.get("pixels", {})
    explicit = pixels_cfg.get(pixel_id)
    if explicit is None:
        match = re.search(r"(\d+)_(\d+)_(\d+)$", str(pixel_id))
        if match:
            explicit = pixels_cfg.get(f"Q{match.group(1)}_{match.group(2)}_{match.group(3)}")
    if explicit is None:
        explicit = sim_generated_profile_for_pixel(pixel_id)
    profile = deep_update(profile, explicit)
    mode = str(profile.get("mode", "working")).lower()
    if mode == "weak" and "photodiode_gain_uA_per_mA" not in profile:
        profile["photodiode_gain_uA_per_mA"] = 0.05
    return profile


def sim_generated_profile_for_pixel(pixel_id: str) -> Dict[str, Any]:
    match = re.search(r"(\d+)_(\d+)_(\d+)$", str(pixel_id))
    if not match:
        return {}
    q, substrate, pix = (int(match.group(1)), int(match.group(2)), int(match.group(3)))
    selector = (q * 17 + substrate * 5 + pix * 3) % 10
    if selector == 0:
        return {"mode": "no_contact", "leakage_mA": 0.001}
    if selector in {1, 2}:
        return {
            "mode": "nonworking",
            "turn_on_voltage_V": 2.8 + 0.05 * q,
            "current_at_5V_mA": 2.0 + 0.3 * substrate,
            "photodiode_gain_uA_per_mA": 0.0,
        }
    if selector == 3:
        return {
            "mode": "weak",
            "turn_on_voltage_V": 2.55 + 0.04 * q,
            "current_at_5V_mA": 4.0 + 0.2 * pix,
            "photodiode_gain_uA_per_mA": 0.04,
            "spectrum": {"counts_per_mA_per_s": 120_000.0},
        }
    if selector == 4:
        return {
            "mode": "working",
            "turn_on_voltage_V": 2.45 + 0.06 * q,
            "current_at_5V_mA": 5.0 + 0.4 * substrate,
            "burnout_voltage_V": 3.4 + 0.1 * pix,
            "short_resistance_ohm": 90.0,
        }
    return {
        "mode": "working",
        "turn_on_voltage_V": 2.45 + 0.08 * q + 0.02 * pix,
        "current_at_5V_mA": 4.5 + 0.6 * substrate + 0.25 * pix,
        "photodiode_gain_uA_per_mA": 0.35 + 0.05 * ((q + substrate + pix) % 5),
    }


def sim_start_session(pixel_id: Optional[str] = None) -> None:
    pixel_id = pixel_id or sim_current_pixel_id()
    seed_base = int(sim_global_settings().get("random_seed", 42))
    seed = seed_base + sum(ord(ch) for ch in pixel_id)
    _SIM_STATE.update({
        "pixel_id": pixel_id,
        "session_start_monotonic": time.monotonic(),
        "smu1_voltage_set_V": 0.0,
        "smu2_voltage_set_V": 0.0,
        "smu1_enabled": False,
        "smu2_enabled": False,
        "led_current_mA": 0.0,
        "photodiode_uA": 0.0,
        "latched_burned": False,
        "rng": random.Random(seed),
    })


def sim_elapsed_s() -> float:
    return max(0.0, time.monotonic() - float(_SIM_STATE.get("session_start_monotonic", time.monotonic())))


def sim_noise(value: float, relative: float = 0.0, absolute: float = 0.0) -> float:
    rng: random.Random = _SIM_STATE.get("rng") or random.Random(42)
    sigma = abs(value) * float(relative) + float(absolute)
    if sigma <= 0:
        return value
    return value + rng.gauss(0.0, sigma)


def sim_is_no_light_mode(mode: str) -> bool:
    return mode in {"nonworking", "no_light", "dead", "burned", "short", "no_contact", "open"}


def sim_set_channel_voltage(channel: int, voltage: float) -> None:
    _SIM_STATE[f"smu{channel}_voltage_set_V"] = float(voltage)


def sim_set_channel_enabled(channel: int, enabled: bool) -> None:
    _SIM_STATE[f"smu{channel}_enabled"] = bool(enabled)


def sim_set_channel_range(channel: int, range_index: int) -> None:
    _SIM_STATE[f"smu{channel}_range"] = int(range_index)


def sim_led_current_mA(voltage_V: Optional[float] = None) -> float:
    p = sim_profile_for_pixel(_SIM_STATE.get("pixel_id"))
    mode = str(p.get("mode", "working")).lower()
    g = sim_global_settings()
    t = sim_elapsed_s()
    voltage = float(_SIM_STATE.get("smu1_voltage_set_V", 0.0) if voltage_V is None else voltage_V)

    if not _SIM_STATE.get("smu1_enabled", False):
        current = 0.0
    elif mode in {"no_contact", "open"}:
        current = float(p.get("leakage_mA", 0.0005))
    else:
        burnout_after_s = p.get("burnout_after_s")
        if burnout_after_s is not None and t >= float(burnout_after_s):
            _SIM_STATE["latched_burned"] = True
        burnout_voltage = p.get("burnout_voltage_V")
        if burnout_voltage is not None and voltage >= float(burnout_voltage):
            _SIM_STATE["latched_burned"] = True
        if mode in {"burned", "short"} or _SIM_STATE.get("latched_burned", False):
            r_short = max(float(p.get("short_resistance_ohm", 120.0)), 1e-9)
            current = max(0.0, voltage) * 1000.0 / r_short
        else:
            v_turn = float(p.get("turn_on_voltage_V", 2.65)) + float(p.get("voltage_drift_V_per_s", 0.0)) * t
            leakage = float(p.get("leakage_mA", 0.0005))
            if voltage <= v_turn:
                current = leakage * max(voltage / max(v_turn, 1e-9), 0.0)
            else:
                denom = max(5.0 - v_turn, 0.1)
                ratio = max((voltage - v_turn) / denom, 0.0)
                exponent = max(float(p.get("iv_exponent", 2.0)), 0.1)
                current = float(p.get("current_at_5V_mA", 6.0)) * (ratio ** exponent) + leakage
                current = min(current, float(p.get("max_current_mA", 30.0)))
    current = max(0.0, sim_noise(current, relative=float(g.get("current_noise_relative", 0.0))))
    _SIM_STATE["led_current_mA"] = current
    return current


def sim_photodiode_current_uA() -> float:
    p = sim_profile_for_pixel(_SIM_STATE.get("pixel_id"))
    mode = str(p.get("mode", "working")).lower()
    g = sim_global_settings()
    current_mA = sim_led_current_mA()
    if not _SIM_STATE.get("smu2_enabled", False):
        photo = 0.0
    elif sim_is_no_light_mode(mode) or _SIM_STATE.get("latched_burned", False):
        photo = 0.0
    else:
        gain = float(p.get("photodiode_gain_uA_per_mA", 0.55))
        tau = p.get("degradation_tau_s")
        degradation = 1.0 if tau in (None, 0, "") else math.exp(-sim_elapsed_s() / max(float(tau), 1e-9))
        photo = current_mA * gain * degradation
    photo = max(0.0, sim_noise(photo, absolute=float(g.get("photodiode_noise_uA", 0.0))))
    _SIM_STATE["photodiode_uA"] = photo
    return photo


def sim_measured_voltage(channel: int) -> float:
    g = sim_global_settings()
    set_v = float(_SIM_STATE.get(f"smu{channel}_voltage_set_V", 0.0))
    return sim_noise(set_v, absolute=float(g.get("voltage_noise_V", 0.0)))


def sim_spectrum_counts(wavelengths_nm: np.ndarray, integration_time_s: float) -> np.ndarray:
    p = sim_profile_for_pixel(_SIM_STATE.get("pixel_id"))
    g = sim_global_settings()
    sp = p.get("spectrum", {})
    mode = str(p.get("mode", "working")).lower()
    t_int = max(float(integration_time_s), 1e-9)
    dark = float(sp.get("dark_offset_counts", 120.0)) + float(sp.get("dark_counts_per_s", 250.0)) * t_int
    background = float(sp.get("background_counts_per_s", 50.0)) * t_int
    signal = np.zeros_like(wavelengths_nm, dtype=np.float64)
    current_mA = float(_SIM_STATE.get("led_current_mA", 0.0)) or sim_led_current_mA()
    if not sim_is_no_light_mode(mode) and not _SIM_STATE.get("latched_burned", False):
        tau = p.get("degradation_tau_s")
        degradation = 1.0 if tau in (None, 0, "") else math.exp(-sim_elapsed_s() / max(float(tau), 1e-9))
        scale = current_mA * float(sp.get("counts_per_mA_per_s", 1_200_000.0)) * t_int * degradation
        for peak in sp.get("peaks", []):
            center = float(peak.get("center_nm", 530.0))
            fwhm = max(float(peak.get("fwhm_nm", 70.0)), 1e-9)
            amplitude = float(peak.get("amplitude", 1.0))
            sigma = fwhm / 2.354820045
            signal += amplitude * np.exp(-0.5 * ((wavelengths_nm - center) / sigma) ** 2)
        signal *= scale
    counts = dark + background + signal
    rel_noise = float(g.get("spectrum_noise_relative", 0.0))
    if rel_noise > 0:
        rng: random.Random = _SIM_STATE.get("rng") or random.Random(42)
        noise = np.array([rng.gauss(0.0, max(math.sqrt(max(v, 1.0)), rel_noise * max(v, 1.0))) for v in counts])
        counts = counts + noise
    saturation = float(sp.get("saturation_counts", 65535.0))
    return np.clip(counts, 0.0, saturation)


class _SimSetter:
    def __init__(self, channel: "_SimSMUChannel"):
        self.channel = channel

    def enabled(self, value, response=0):
        sim_set_channel_enabled(self.channel.index, bool(value))
        return None

    def voltage(self, value, response=0):
        sim_set_channel_voltage(self.channel.index, float(value))
        return None

    def range(self, value, response=0):
        sim_set_channel_range(self.channel.index, int(value))
        return None


class _SimSMUChannel:
    def __init__(self, index: int):
        self.index = index
        self.set = _SimSetter(self)

    def measure(self):
        if self.index == 1:
            return [(sim_measured_voltage(1), sim_led_current_mA() / 1000.0)]
        return [(sim_measured_voltage(2), -sim_photodiode_current_uA() / 1_000_000.0)]


class _SimDevice:
    def __init__(self, com_port="SIM", *args, **kwargs):
        self.com_port = com_port
        self.smu1 = _SimSMUChannel(1)
        self.smu2 = _SimSMUChannel(2)

    def __enter__(self):
        sim_start_session(os.environ.get("OLED_SIM_PIXEL_ID", "SIM_PIXEL"))
        return self

    def __exit__(self, exc_type, exc, tb):
        sim_set_channel_voltage(1, 0.0)
        sim_set_channel_voltage(2, 0.0)
        sim_set_channel_enabled(1, False)
        sim_set_channel_enabled(2, False)
        return False


class _SimulatedSpectrometerDevice:
    def __repr__(self):
        return "OLED simulated spectrometer"


class _SimSpectrometer:
    model = "OLED-SIM-SPECTROMETER"

    def __init__(self, device):
        self.device = device
        self._integration_time_s = 0.01
        self._wavelengths = np.arange(350.0, 851.0, 1.0, dtype=np.float64)

    def wavelengths(self):
        return self._wavelengths.copy()

    def integration_time_micros(self, value):
        self._integration_time_s = max(float(value) / 1_000_000.0, 1e-6)
        return None

    def intensities(self):
        return sim_spectrum_counts(self._wavelengths, self._integration_time_s).astype(np.float64)


def install_simulator_modules(pixel_id: str, config_path: Optional[Path]) -> None:
    os.environ["OLED_SIM_PIXEL_ID"] = str(pixel_id)
    if config_path:
        os.environ["OLED_SIM_CONFIG"] = str(config_path)

    xtralien_mod = types.ModuleType("xtralien")
    xtralien_mod.__oled_app_simulator__ = True
    xtralien_mod.Device = _SimDevice

    seabreeze_mod = types.ModuleType("seabreeze")
    seabreeze_mod.__oled_app_simulator__ = True
    seabreeze_mod.use = lambda backend=None: None

    spectrometers_mod = types.ModuleType("seabreeze.spectrometers")
    spectrometers_mod.__oled_app_simulator__ = True
    spectrometers_mod.list_devices = lambda: [_SimulatedSpectrometerDevice()]
    spectrometers_mod.Spectrometer = _SimSpectrometer
    seabreeze_mod.spectrometers = spectrometers_mod

    sys.modules["xtralien"] = xtralien_mod
    sys.modules["seabreeze"] = seabreeze_mod
    sys.modules["seabreeze.spectrometers"] = spectrometers_mod


def uninstall_simulator_modules() -> None:
    for name in ["xtralien", "seabreeze", "seabreeze.spectrometers"]:
        mod = sys.modules.get(name)
        if getattr(mod, "__oled_app_simulator__", False):
            del sys.modules[name]


def prepare_hardware_environment(pixel_id: str, app_settings: Optional[Dict[str, Any]], log: Callable[[str], None]) -> None:
    settings = app_settings or load_app_settings()
    if settings.get("hardware_mode") == HARDWARE_MODE_SIM:
        cfg_path = Path(settings.get("simulator_config_path") or SCRIPT_DIR / SIM_CONFIG_FILE)
        ensure_default_sim_config(cfg_path)
        install_simulator_modules(pixel_id, cfg_path)
        log(f"Режим оборудования: эмулятор. Пиксель {pixel_id}; конфиг {cfg_path}")
    else:
        uninstall_simulator_modules()
        log("Режим оборудования: реальное оборудование. Будут использованы установленные xtralien/seabreeze.")

def light_border() -> Border:
    side = Side(style="thin", color="D9D9D9")
    return Border(left=side, right=side, top=side, bottom=side)


def style_header_row(ws, row: int, min_col: int, max_col: int):
    fill = PatternFill("solid", fgColor="D9E1F2")
    font = Font(bold=True)
    border = light_border()
    for col in range(min_col, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border


def autosize_columns(ws, min_width: int = 10, max_width: int = 42):
    for col_cells in ws.columns:
        max_len = 0
        letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            value = cell.value
            if value is not None:
                max_len = max(max_len, len(str(value)))
        ws.column_dimensions[letter].width = max(min_width, min(max_width, max_len + 2))


def parse_float(text: str, field_name: str) -> float:
    try:
        return float(str(text).replace(",", "."))
    except Exception:
        raise ValueError(f"Поле '{field_name}' должно быть числом")


def parse_int(text: str, field_name: str) -> int:
    try:
        return int(float(str(text).replace(",", ".")))
    except Exception:
        raise ValueError(f"Поле '{field_name}' должно быть целым числом")


# -----------------------------------------------------------------------------
# Серия и журнал
# -----------------------------------------------------------------------------

@dataclass
class PixelInfo:
    pixel_id: str
    quarter_code: str
    quarter_number: int
    substrate_number: int
    pixel_number: int


def generate_pixels(quarter_names: Dict[str, str]) -> List[PixelInfo]:
    pixels: List[PixelInfo] = []
    for q in range(1, 5):
        code = safe_filename(quarter_names.get(str(q), f"Q{q}"), fallback=f"Q{q}")
        for substrate in range(1, 4):
            for pix in range(1, 5):
                # Формат: CR1_2_3, где CR — название четверти,
                # 1 — номер четверти, 2 — подложка, 3 — пиксель.
                pixel_id = f"{code}{q}_{substrate}_{pix}"
                pixels.append(PixelInfo(pixel_id, code, q, substrate, pix))
    return pixels


class SeriesJournal:
    def __init__(self, series_folder: Path, config: Dict):
        self.series_folder = Path(series_folder)
        self.config = config
        self.path = self.series_folder / JOURNAL_FILE

    def initialize_or_update(self):
        if self.path.exists():
            wb = load_workbook(self.path)
        else:
            wb = Workbook()
            wb.remove(wb.active)

        self._ensure_series_sheet(wb)
        self._ensure_quarters_sheet(wb)
        self._ensure_pixels_sheet(wb)
        self._ensure_measurements_sheet(wb)

        wb.save(self.path)
        wb.close()

    def _ensure_series_sheet(self, wb: Workbook):
        if SERIES_SHEET in wb.sheetnames:
            ws = wb[SERIES_SHEET]
            ws.delete_rows(1, ws.max_row)
        else:
            ws = wb.create_sheet(SERIES_SHEET, 0)

        ws["A1"] = "OLED series journal"
        ws["A1"].font = Font(bold=True, size=14)
        rows = [
            ("App version", APP_VERSION),
            ("Created at", self.config.get("created_at", "")),
            ("Deposition date", self.config.get("deposition_date", "")),
            ("Keyword", self.config.get("keyword", "")),
            ("Series folder", str(self.series_folder.resolve())),
            ("Naming rule", "{quarter_code}{quarter_number}_{substrate_number}_{pixel_number}"),
            ("Example", "CR1_2_3"),
        ]
        for idx, (k, v) in enumerate(rows, start=3):
            ws.cell(row=idx, column=1, value=k).font = Font(bold=True)
            ws.cell(row=idx, column=2, value=v)
        autosize_columns(ws)

    def _ensure_quarters_sheet(self, wb: Workbook):
        if QUARTERS_SHEET in wb.sheetnames:
            ws = wb[QUARTERS_SHEET]
            ws.delete_rows(1, ws.max_row)
        else:
            ws = wb.create_sheet(QUARTERS_SHEET)

        headers = ["Quarter number", "Quarter code/name", "Generated pixel prefix example"]
        ws.append(headers)
        style_header_row(ws, 1, 1, len(headers))
        quarter_names = self.config.get("quarter_names", {})
        for q in range(1, 5):
            code = safe_filename(quarter_names.get(str(q), f"Q{q}"), fallback=f"Q{q}")
            ws.append([q, code, f"{code}{q}_1_1"])
        autosize_columns(ws)

    def _ensure_pixels_sheet(self, wb: Workbook):
        if PIXELS_SHEET in wb.sheetnames:
            ws = wb[PIXELS_SHEET]
            existing = self._read_sheet_as_dicts(ws)
            existing_by_id = {row.get("Pixel ID"): row for row in existing if row.get("Pixel ID")}
            ws.delete_rows(1, ws.max_row)
        else:
            ws = wb.create_sheet(PIXELS_SHEET)
            existing_by_id = {}

        ws.append(PIXEL_HEADERS)
        style_header_row(ws, 1, 1, len(PIXEL_HEADERS))

        pixels = generate_pixels(self.config.get("quarter_names", {}))
        for p in pixels:
            old = existing_by_id.get(p.pixel_id, {})
            ws.append([
                p.pixel_id,
                p.quarter_code,
                p.quarter_number,
                p.substrate_number,
                p.pixel_number,
                old.get("Last status", "UNKNOWN"),
                old.get("Opening voltage (V)", ""),
                old.get("Last IVL date", ""),
                old.get("Last IVL file", ""),
                old.get("Last IVL max current (mA)", ""),
                old.get("Last IVL max photodiode (uA)", ""),
                old.get("Last spectrum date", ""),
                old.get("Last spectrum file", ""),
                old.get("Last spectrum peak count", ""),
                old.get("Last spectrum peaks nm", ""),
                old.get("Last spectrum max intensity (counts/s)", ""),
                old.get("Last stability date", ""),
                old.get("Last stability file", ""),
                old.get("Last updated", ""),
            ])
        ws.freeze_panes = "A2"
        autosize_columns(ws, max_width=38)

    def _ensure_measurements_sheet(self, wb: Workbook):
        if MEASUREMENTS_SHEET not in wb.sheetnames:
            ws = wb.create_sheet(MEASUREMENTS_SHEET)
            ws.append(MEASUREMENT_HEADERS)
            style_header_row(ws, 1, 1, len(MEASUREMENT_HEADERS))
            ws.freeze_panes = "A2"
            autosize_columns(ws, max_width=55)
        else:
            ws = wb[MEASUREMENTS_SHEET]
            if ws.max_row == 0 or ws.cell(row=1, column=1).value != MEASUREMENT_HEADERS[0]:
                ws.delete_rows(1, ws.max_row)
                ws.append(MEASUREMENT_HEADERS)
                style_header_row(ws, 1, 1, len(MEASUREMENT_HEADERS))

    @staticmethod
    def _read_sheet_as_dicts(ws) -> List[Dict]:
        headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
        rows = []
        for r in range(2, ws.max_row + 1):
            row = {}
            has_value = False
            for c, header in enumerate(headers, start=1):
                val = ws.cell(row=r, column=c).value
                if val not in (None, ""):
                    has_value = True
                row[header] = val
            if has_value:
                rows.append(row)
        return rows

    def list_pixels(self) -> List[Dict]:
        wb = load_workbook(self.path, data_only=True)
        ws = wb[PIXELS_SHEET]
        rows = self._read_sheet_as_dicts(ws)
        wb.close()
        return rows

    def list_measurements(self) -> List[Dict]:
        if not self.path.exists():
            return []
        wb = load_workbook(self.path, data_only=True)
        if MEASUREMENTS_SHEET not in wb.sheetnames:
            wb.close()
            return []
        ws = wb[MEASUREMENTS_SHEET]
        rows = self._read_sheet_as_dicts(ws)
        wb.close()
        return rows

    def get_pixel(self, pixel_id: str) -> Optional[Dict]:
        for row in self.list_pixels():
            if row.get("Pixel ID") == pixel_id:
                return row
        return None

    def has_any_ivl(self) -> bool:
        if not self.path.exists():
            return False
        wb = load_workbook(self.path, data_only=True)
        if MEASUREMENTS_SHEET not in wb.sheetnames:
            wb.close()
            return False
        ws = wb[MEASUREMENTS_SHEET]
        headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
        try:
            type_col = headers.index("Type") + 1
        except ValueError:
            wb.close()
            return False
        for r in range(2, ws.max_row + 1):
            if ws.cell(row=r, column=type_col).value == "IVL":
                wb.close()
                return True
        wb.close()
        return False

    def update_after_measurement(
        self,
        measurement_type: str,
        pixel_id: str,
        status: str,
        file_path: Optional[Path],
        params: Dict,
        notes: str = "",
        opening_voltage: Optional[float] = None,
        max_current_mA: Optional[float] = None,
        max_photo_uA: Optional[float] = None,
        spectrum_peak_count: Optional[int] = None,
        spectrum_peaks_nm: str = "",
        spectrum_max_intensity: Optional[float] = None,
    ):
        wb = load_workbook(self.path)
        ws_pixels = wb[PIXELS_SHEET]
        ws_meas = wb[MEASUREMENTS_SHEET]

        rel_file = relative_to_or_abs(file_path, self.series_folder) if file_path else ""
        date_text = now_str()
        day_text = today_iso()

        # Measurements log
        ws_meas.append([
            date_text,
            day_text,
            measurement_type,
            pixel_id,
            status,
            rel_file,
            json.dumps(params, ensure_ascii=False),
            notes,
        ])

        # Pixel table update
        headers = [ws_pixels.cell(row=1, column=c).value for c in range(1, ws_pixels.max_column + 1)]
        col = {h: i + 1 for i, h in enumerate(headers)}
        row_idx = None
        for r in range(2, ws_pixels.max_row + 1):
            if ws_pixels.cell(row=r, column=col["Pixel ID"]).value == pixel_id:
                row_idx = r
                break

        if row_idx:
            if measurement_type in {"IVL", "STABILITY"}:
                ws_pixels.cell(row=row_idx, column=col["Last status"], value=status)
            ws_pixels.cell(row=row_idx, column=col["Last updated"], value=date_text)
            if opening_voltage is not None:
                ws_pixels.cell(row=row_idx, column=col["Opening voltage (V)"], value=float(opening_voltage))

            if measurement_type == "IVL":
                ws_pixels.cell(row=row_idx, column=col["Last IVL date"], value=date_text)
                ws_pixels.cell(row=row_idx, column=col["Last IVL file"], value=rel_file)
                if max_current_mA is not None:
                    ws_pixels.cell(row=row_idx, column=col["Last IVL max current (mA)"], value=float(max_current_mA))
                if max_photo_uA is not None:
                    ws_pixels.cell(row=row_idx, column=col["Last IVL max photodiode (uA)"], value=float(max_photo_uA))
            elif measurement_type == "SPECTRUM":
                ws_pixels.cell(row=row_idx, column=col["Last spectrum date"], value=date_text)
                ws_pixels.cell(row=row_idx, column=col["Last spectrum file"], value=rel_file)
                if spectrum_peak_count is not None and "Last spectrum peak count" in col:
                    ws_pixels.cell(row=row_idx, column=col["Last spectrum peak count"], value=int(spectrum_peak_count))
                if spectrum_peaks_nm and "Last spectrum peaks nm" in col:
                    ws_pixels.cell(row=row_idx, column=col["Last spectrum peaks nm"], value=str(spectrum_peaks_nm))
                if spectrum_max_intensity is not None and "Last spectrum max intensity (counts/s)" in col:
                    ws_pixels.cell(row=row_idx, column=col["Last spectrum max intensity (counts/s)"], value=float(spectrum_max_intensity))
            elif measurement_type == "STABILITY":
                ws_pixels.cell(row=row_idx, column=col["Last stability date"], value=date_text)
                ws_pixels.cell(row=row_idx, column=col["Last stability file"], value=rel_file)

        for ws in [ws_pixels, ws_meas]:
            autosize_columns(ws, max_width=55)
        wb.save(self.path)
        wb.close()


class SeriesManager:
    def __init__(self, series_folder: Path):
        self.series_folder = Path(series_folder)
        self.config_path = self.series_folder / CONFIG_FILE
        if not self.config_path.exists():
            raise FileNotFoundError(f"В папке нет {CONFIG_FILE}: {self.config_path}")
        self.config = json.loads(self.config_path.read_text(encoding="utf-8"))
        self.journal = SeriesJournal(self.series_folder, self.config)
        self.journal.initialize_or_update()

    @classmethod
    def create_new(
        cls,
        root_folder: Path,
        deposition_date: str,
        keyword: str,
        quarter_names: Dict[str, str],
    ) -> "SeriesManager":
        keyword_safe = safe_filename(keyword, fallback="")
        folder_name = f"{deposition_date}"
        if keyword_safe:
            folder_name += f"_{keyword_safe}"
        folder_name = safe_filename(folder_name, fallback="series")

        series_folder = Path(root_folder) / folder_name
        base_folder = series_folder
        suffix = 2
        while series_folder.exists():
            series_folder = Path(f"{base_folder}_{suffix}")
            suffix += 1

        series_folder.mkdir(parents=True, exist_ok=False)
        (series_folder / "measurements").mkdir(exist_ok=True)

        config = {
            "app_version": APP_VERSION,
            "created_at": now_str(),
            "deposition_date": deposition_date,
            "keyword": keyword,
            "quarter_names": {
                str(q): safe_filename(quarter_names.get(str(q), f"Q{q}"), fallback=f"Q{q}")
                for q in range(1, 5)
            },
        }
        (series_folder / CONFIG_FILE).write_text(json.dumps(config, ensure_ascii=False, indent=2), encoding="utf-8")
        manager = cls(series_folder)
        return manager


# -----------------------------------------------------------------------------
# Измерение ВАЯХ
# -----------------------------------------------------------------------------

@dataclass
class IVLParams:
    com_port: str = "COM3"
    sweep_start: float = 0.0
    sweep_end: float = 5.0
    sweep_increment: float = 0.02
    sweep_time_per_point: float = 0.01
    num_cycles: int = 1
    delay_between_cycles: float = 1.0
    current_limit_mA: float = 10.0
    photodiode_bias_V: float = -5.0
    photodiode_range: int = 4
    photodiode_threshold_uA: float = 0.5
    burnout_current_threshold_mA: float = 10.0
    mark_current_limit_as_burnout: bool = False
    no_contact_max_led_current_mA: float = 0.05
    burned_confirmation_cycles: int = 1
    pixel_area_mm2: float = 1.0
    luminance_cd_m2_per_uA: float = 1.0

    def as_dict(self) -> Dict:
        return self.__dict__.copy()


def safe_shutdown_smu(smu):
    try:
        smu.smu1.set.voltage(0, response=0)
    except Exception:
        pass
    try:
        smu.smu2.set.voltage(0, response=0)
    except Exception:
        pass
    time.sleep(0.2)
    try:
        smu.smu1.set.enabled(False, response=0)
    except Exception:
        pass
    try:
        smu.smu2.set.enabled(False, response=0)
    except Exception:
        pass


def define_ivl_pixel_status(
    max_photo_uA: float,
    max_led_current_mA: float,
    current_limit_reached: bool,
    params: IVLParams,
) -> Tuple[str, str]:
    light_detected = max_photo_uA >= params.photodiode_threshold_uA
    burnout_by_high_current = max_led_current_mA >= params.burnout_current_threshold_mA

    if burnout_by_high_current:
        return "BURNED", "Пробой / сгорание по току"
    if light_detected:
        return "WORKING", "Рабочий: фототок выше порога"
    if max_led_current_mA <= params.no_contact_max_led_current_mA:
        return "NO_CONTACT", "Нет контакта: ток почти нулевой"
    return "NONWORKING", "Нерабочий: ток есть, фототока нет"


def detect_opening_voltage(cycle_data: List[Dict], threshold_uA: float) -> Optional[float]:
    for row in cycle_data:
        if row.get("Photodiode current (uA)", 0) >= threshold_uA:
            return float(row.get("Voltage OLED / LED measured (V)", row.get("Voltage set (V)", 0)))
    return None


def describe_ivl_first_measurement(cycles: List[Dict]) -> str:
    """Человеческое описание первого промера ВАЯХ для журнала серии."""
    if not cycles:
        return "ВАЯХ не выполнена"
    burned = next(
        (c for c in cycles if c.get("status") == "BURNED"),
        None,
    )
    if burned is not None:
        cycle_number = int(burned.get("cycle", 1) or 1)
        if cycle_number == 1:
            return "Светодиод сгорел/пробился на первом цикле ВАЯХ"
        return f"Светодиод сгорел/пробился на цикле {cycle_number} ВАЯХ"

    first = str(cycles[0].get("status", "") or "").upper()
    if first == "WORKING":
        return "На первом промере светодиод рабочий"
    if first == "NONWORKING":
        return "Светодиод сразу нерабочий: ток есть, фототока нет"
    if first == "NO_CONTACT":
        return "На первом промере нет контакта с подложкой"
    if first == "BURNED":
        return "Светодиод сгорел/пробился на первом цикле ВАЯХ"
    return f"Первый промер завершился статусом {first or 'UNKNOWN'}"


def run_ivl_cycle(
    smu,
    pixel_id: str,
    cycle_number: int,
    params: IVLParams,
    log: Callable[[str], None],
    progress_callback: Optional[Callable[[int, Dict[str, Any]], None]] = None,
) -> Dict:
    log(f"\nВАЯХ {pixel_id}, цикл {cycle_number}: до {params.sweep_end:.3f} В, лимит {params.current_limit_mA:.3f} мА")

    smu.smu1.set.enabled(True, response=0)
    smu.smu2.set.enabled(True, response=0)
    try:
        smu.smu2.set.range(params.photodiode_range, response=0)
    except Exception:
        pass

    time.sleep(0.1)
    smu.smu1.set.voltage(0, response=0)
    smu.smu2.set.voltage(params.photodiode_bias_V, response=0)
    time.sleep(0.2)

    data: List[Dict] = []
    current_limit_reached = False

    voltage_values = np.arange(
        params.sweep_start,
        params.sweep_end + params.sweep_increment / 2,
        params.sweep_increment,
    )
    voltage_values = np.round(voltage_values, 6)

    for idx, set_v in enumerate(voltage_values, start=1):
        smu.smu1.set.voltage(float(set_v), response=0)
        smu.smu2.set.voltage(params.photodiode_bias_V, response=0)
        time.sleep(params.sweep_time_per_point)

        voltage_led, current_led = smu.smu1.measure()[0]
        voltage_pd, current_pd = smu.smu2.measure()[0]
        current_led_mA = current_led * 1000.0
        current_pd_uA = -current_pd * 1_000_000.0

        point_row = {
            "Point": idx,
            "Voltage set (V)": float(set_v),
            "Voltage OLED / LED measured (V)": float(voltage_led),
            "Current OLED / LED (mA)": float(current_led_mA),
            "Current density (mA/cm^2)": current_density_mA_cm2(current_led_mA, params.pixel_area_mm2),
            "Voltage photodiode measured (V)": float(voltage_pd),
            "Photodiode current (uA)": float(current_pd_uA),
            "Luminance (cd/m^2)": luminance_cd_m2(current_pd_uA, params.luminance_cd_m2_per_uA),
        }
        data.append(point_row)
        if progress_callback is not None:
            try:
                progress_callback(cycle_number, point_row)
            except MeasurementStopped:
                safe_shutdown_smu(smu)
                raise

        if idx == 1 or idx % 25 == 0:
            log(f"  {idx}/{len(voltage_values)}: V={voltage_led:.3f} В, I={current_led_mA:.3f} мА, PD={current_pd_uA:.3f} мкА")

        if current_led_mA >= params.current_limit_mA:
            current_limit_reached = True
            log(f"  Аварийный стоп: ток {current_led_mA:.3f} мА >= {params.current_limit_mA:.3f} мА")
            try:
                smu.smu1.set.voltage(0, response=0)
            except Exception:
                pass
            break

    safe_shutdown_smu(smu)

    max_photo = max([r["Photodiode current (uA)"] for r in data], default=0.0)
    max_current = max([r["Current OLED / LED (mA)"] for r in data], default=0.0)
    status, status_desc = define_ivl_pixel_status(max_photo, max_current, current_limit_reached, params)
    opening = detect_opening_voltage(data, params.photodiode_threshold_uA)

    log(f"  Цикл {cycle_number}: статус {status}, max I={max_current:.3f} мА, max PD={max_photo:.3f} мкА")
    return {
        "cycle": cycle_number,
        "status": status,
        "status_desc": status_desc,
        "current_limit_reached": current_limit_reached,
        "max_photo_uA": max_photo,
        "max_current_mA": max_current,
        "opening_voltage": opening,
        "data": data,
    }


def save_ivl_workbook(pixel_id: str, output_dir: Path, params: IVLParams, cycles: List[Dict]) -> Path:
    filename = output_dir / f"IVL_{safe_filename(pixel_id)}_{timestamp_for_file()}.xlsx"
    ivl_diagnosis = describe_ivl_first_measurement(cycles)
    wb = Workbook()
    ws_sum = wb.active
    ws_sum.title = "Summary"

    ws_sum["A1"] = "IVL / ВАЯХ"
    ws_sum["A1"].font = Font(bold=True, size=14)
    meta = [
        ("Pixel", pixel_id),
        ("Created", now_str()),
        ("COM port", params.com_port),
        ("Sweep", f"{params.sweep_start}–{params.sweep_end} В, step {params.sweep_increment} В"),
        ("Cycles requested", params.num_cycles),
        ("Current limit (mA)", params.current_limit_mA),
        ("Photodiode threshold (uA)", params.photodiode_threshold_uA),
        ("Pixel area (mm^2)", params.pixel_area_mm2),
        ("Luminance conversion (cd/m^2 per uA)", params.luminance_cd_m2_per_uA),
        ("Burned confirmation cycles", params.burned_confirmation_cycles),
        ("Первый промер / диагноз", ivl_diagnosis),
        ("Naming rule", "{quarter_code}{quarter_number}_{substrate_number}_{pixel_number}"),
    ]
    for idx, (k, v) in enumerate(meta, start=3):
        ws_sum.cell(row=idx, column=1, value=k).font = Font(bold=True)
        ws_sum.cell(row=idx, column=2, value=v)

    summary_headers = [
        "Cycle",
        "Status",
        "Status description",
        "Max current (mA)",
        "Max photodiode (uA)",
        "Opening voltage detected (V)",
        "Current limit reached",
        "First measurement diagnosis",
    ]
    start = 13
    ws_sum.append([])
    for c, header in enumerate(summary_headers, start=1):
        ws_sum.cell(row=start, column=c, value=header)
    style_header_row(ws_sum, start, 1, len(summary_headers))

    for r_idx, cyc in enumerate(cycles, start=start + 1):
        ws_sum.cell(row=r_idx, column=1, value=cyc["cycle"])
        ws_sum.cell(row=r_idx, column=2, value=cyc["status"])
        ws_sum.cell(row=r_idx, column=3, value=cyc["status_desc"])
        ws_sum.cell(row=r_idx, column=4, value=cyc["max_current_mA"])
        ws_sum.cell(row=r_idx, column=5, value=cyc["max_photo_uA"])
        ws_sum.cell(row=r_idx, column=6, value=cyc["opening_voltage"])
        ws_sum.cell(row=r_idx, column=7, value="YES" if cyc["current_limit_reached"] else "NO")
        ws_sum.cell(row=r_idx, column=8, value=ivl_diagnosis if cyc["cycle"] == 1 else "")

    for cyc in cycles:
        ws = wb.create_sheet(f"Cycle_{cyc['cycle']}")
        ws["A1"] = f"Pixel {pixel_id} | Cycle {cyc['cycle']} | {cyc['status']}"
        ws["A1"].font = Font(bold=True, size=13)
        headers = [
            "Point",
            "Voltage set (V)",
        "Voltage OLED / LED measured (V)",
        "Current OLED / LED (mA)",
        "Current density (mA/cm^2)",
        "Voltage photodiode measured (V)",
        "Photodiode current (uA)",
        "Luminance (cd/m^2)",
    ]
        header_row = 4
        for c, h in enumerate(headers, start=1):
            ws.cell(row=header_row, column=c, value=h)
        style_header_row(ws, header_row, 1, len(headers))
        for row_idx, row in enumerate(cyc["data"], start=header_row + 1):
            for col_idx, h in enumerate(headers, start=1):
                ws.cell(row=row_idx, column=col_idx, value=row.get(h))
        ws.freeze_panes = f"A{header_row + 1}"

        if len(cyc["data"]) >= 2:
            chart = ScatterChart()
            chart.title = f"{pixel_id} | Cycle {cyc['cycle']} | {cyc['status']}"
            chart.x_axis.title = "Voltage OLED / LED (V)"
            chart.y_axis.title = "Current OLED / LED (mA) / Photodiode (uA)"
            chart.x_axis.majorGridlines = ChartLines()
            min_row = header_row + 1
            max_row = header_row + len(cyc["data"])
            xvalues = Reference(ws, min_col=2, min_row=min_row, max_row=max_row)
            y_current = Reference(ws, min_col=4, min_row=header_row, max_row=max_row)
            y_photo = Reference(ws, min_col=7, min_row=header_row, max_row=max_row)
            chart.series.append(Series(y_current, xvalues, title_from_data=True))
            chart.series.append(Series(y_photo, xvalues, title_from_data=True))
            ws.add_chart(chart, "H4")
        autosize_columns(ws, max_width=34)

    for ws in wb.worksheets:
        autosize_columns(ws, max_width=42)
    wb.save(filename)
    wb.close()
    return filename


def run_ivl_measurement(
    pixel_id: str,
    output_dir: Path,
    params: IVLParams,
    log: Callable[[str], None],
    app_settings: Optional[Dict[str, Any]] = None,
    progress_callback: Optional[Callable[[int, Dict[str, Any]], None]] = None,
) -> Dict:
    prepare_hardware_environment(pixel_id, app_settings, log)
    import xtralien

    cycles: List[Dict] = []
    cycles_to_run = max(1, int(params.num_cycles))
    burned_confirmations_left = max(0, int(params.burned_confirmation_cycles))
    with xtralien.Device(params.com_port) as smu:
        cycle = 1
        while cycle <= cycles_to_run:
            cyc = run_ivl_cycle(smu, pixel_id, cycle, params, log, progress_callback=progress_callback)
            cycles.append(cyc)
            if cyc["status"] == "BURNED" and burned_confirmations_left > 0:
                burned_confirmations_left -= 1
                cycles_to_run = max(cycles_to_run, cycle + 1)
                log("  BURNED: запускается дополнительный подтверждающий цикл.")
            elif cyc["status"] in {"BURNED", "NO_CONTACT", "NONWORKING"}:
                log(f"  Дальнейшие циклы остановлены: {cyc['status']}")
                break
            if cycle < cycles_to_run:
                time.sleep(params.delay_between_cycles)
            cycle += 1

    filename = save_ivl_workbook(pixel_id, output_dir, params, cycles)
    best_opening = next((c.get("opening_voltage") for c in cycles if c.get("opening_voltage") is not None), None)
    max_current = max([c["max_current_mA"] for c in cycles], default=0.0)
    max_photo = max([c["max_photo_uA"] for c in cycles], default=0.0)
    ivl_diagnosis = describe_ivl_first_measurement(cycles)
    burned_cycle = next(
        (int(c.get("cycle", 1) or 1) for c in cycles if c.get("status") == "BURNED"),
        None,
    )
    final_status = "BURNED" if burned_cycle is not None else cycles[-1]["status"] if cycles else "FAILED"
    return {
        "file": filename,
        "status": final_status,
        "opening_voltage": best_opening,
        "max_current_mA": max_current,
        "max_photo_uA": max_photo,
        "ivl_diagnosis": ivl_diagnosis,
        "burned_cycle": burned_cycle,
        "first_cycle_status": cycles[0]["status"] if cycles else "FAILED",
    }


# -----------------------------------------------------------------------------
# Измерение спектров
# -----------------------------------------------------------------------------

@dataclass
class SpectrumParams:
    com_port: str = "COM3"
    voltage_start: float = 2.0
    voltage_end: float = 5.0
    voltage_step: float = 0.1
    opening_voltage: Optional[float] = None
    voltage_start_source: str = "opening"
    current_limit_mA: float = 6.0
    photodiode_bias_V: float = -5.0
    photodiode_range: int = 4
    target_intensity: float = 40000.0
    intensity_min: float = 20000.0
    intensity_max: float = 55000.0
    saturation_level: float = 60000.0
    min_peak_width_nm: float = 15.0
    max_peak_width_nm: float = 150.0
    t_int_initial_s: float = 0.01
    t_int_min_s: float = 0.001
    t_int_max_s: float = 10.0
    discard_first_scan_after_tint_change: bool = True
    kp: float = 0.3
    ki: float = 0.05
    max_iterations: int = 20
    tolerance: float = 0.05
    led_type: str = "auto"
    peak_search_mode_for_tint: str = "auto"
    settle_time_voltage_s: float = 0.1
    settle_time_spectrum_s: float = 0.05
    dark_spectrum_enabled: bool = False
    dark_spectrum_scans: int = 3
    baseline_correction_enabled: bool = True
    peak_detection_enabled: bool = False
    pixel_area_mm2: float = 1.0
    luminance_cd_m2_per_uA: float = 1.0

    def as_dict(self) -> Dict:
        return self.__dict__.copy()


class SpectrumHelper:
    def __init__(self, params: SpectrumParams, log: Callable[[str], None]):
        self.params = params
        self.log = log
        self._last_integration_time_us: Optional[int] = None
        self.last_optimization_started_saturated = False
        self.last_optimization_started_saturated_at_10ms = False
        self.adaptive_initial_time_enabled = False

    def init_spectrometer(self):
        import seabreeze
        seabreeze.use("cseabreeze")
        from seabreeze.spectrometers import Spectrometer, list_devices

        devices = list_devices()
        if not devices:
            raise RuntimeError("Спектрометр не найден. Проверь USB и драйвер SeaBreeze.")
        spec = Spectrometer(devices[0])
        wavelengths = spec.wavelengths()
        self.log(f"Спектрометр: {spec.model}; диапазон {wavelengths[0]:.1f}–{wavelengths[-1]:.1f} нм")
        return spec

    def set_integration_time(self, spec, integration_time_s: float) -> Tuple[float, bool]:
        p = self.params
        integration_time_s = float(np.clip(integration_time_s, p.t_int_min_s, p.t_int_max_s))
        integration_time_us = max(1, int(round(integration_time_s * 1e6)))
        changed = self._last_integration_time_us != integration_time_us
        spec.integration_time_micros(integration_time_us)
        self._last_integration_time_us = integration_time_us
        return integration_time_us / 1e6, changed

    def get_spectrum(self, spec, integration_time_s: float, discard_stale_after_change: bool = True):
        actual_t, changed = self.set_integration_time(spec, integration_time_s)
        wavelengths = spec.wavelengths()
        if discard_stale_after_change and self.params.discard_first_scan_after_tint_change and changed:
            time.sleep(self.params.settle_time_spectrum_s)
            try:
                _ = spec.intensities()
            except Exception as exc:
                self.log(f"  ⚠ Не удалось сбросить первый спектр после смены T_int: {exc}")
        time.sleep(self.params.settle_time_spectrum_s)
        intensities = spec.intensities().astype(np.float64)
        return wavelengths, intensities, actual_t

    def get_dark_spectrum(self, spec, integration_time_s: float):
        spectra = []
        wavelengths = None
        actual_t = integration_time_s
        for i in range(self.params.dark_spectrum_scans):
            wavelengths, intensities, actual_t = self.get_spectrum(spec, integration_time_s, discard_stale_after_change=(i == 0))
            spectra.append(intensities)
            time.sleep(0.05)
        return wavelengths, np.mean(spectra, axis=0), actual_t

    @staticmethod
    def smooth_array(values: np.ndarray, window: int = 9) -> np.ndarray:
        values = np.asarray(values, dtype=np.float64)
        if values.size < 3:
            return values.copy()
        window = int(max(3, min(window, values.size if values.size % 2 == 1 else values.size - 1)))
        if window < 3:
            return values.copy()
        kernel = np.ones(window, dtype=np.float64) / window
        padded = np.pad(values, (window // 2, window // 2), mode="edge")
        return np.convolve(padded, kernel, mode="valid")

    def estimate_baseline(self, wavelengths: np.ndarray, intensities: np.ndarray) -> Tuple[np.ndarray, Tuple[float, float], float]:
        wavelengths = np.asarray(wavelengths, dtype=np.float64)
        intensities = np.asarray(intensities, dtype=np.float64)
        if intensities.size < 3:
            value = float(np.nanmean(intensities)) if intensities.size else 0.0
            return np.full_like(intensities, value), (float("nan"), float("nan")), value

        finite = np.where(np.isfinite(wavelengths) & np.isfinite(intensities))[0]
        if finite.size < 3:
            value = float(np.nanmean(intensities[finite])) if finite.size else 0.0
            return np.full_like(intensities, value), (float("nan"), float("nan")), value

        n = int(finite.size)
        window = int(np.clip(round(n * 0.06), 8, 40))
        window = min(window, n)
        best_start = 0
        best_score = float("inf")
        for start in range(0, n - window + 1):
            idx = finite[start:start + window]
            x = wavelengths[idx]
            y = intensities[idx]
            if np.any(~np.isfinite(x)) or np.any(~np.isfinite(y)):
                continue
            span = float(max(x[-1] - x[0], 1e-9))
            slope = float(abs(y[-1] - y[0]) / span)
            scatter = float(np.nanstd(y))
            level_penalty = float(max(np.nanmean(y), 0.0)) * 1e-6
            score = scatter + slope * span * 0.25 + level_penalty
            if score < best_score:
                best_score = score
                best_start = start

        baseline_idx = finite[best_start:best_start + window]
        baseline_value = float(np.nanmean(intensities[baseline_idx]))
        baseline = np.full_like(intensities, baseline_value)
        wl_range = (float(wavelengths[baseline_idx[0]]), float(wavelengths[baseline_idx[-1]]))
        return baseline, wl_range, baseline_value

    def find_peaks_by_derivatives(self, wavelengths: np.ndarray, intensities: np.ndarray) -> List[Dict[str, float]]:
        wavelengths = np.asarray(wavelengths, dtype=np.float64)
        intensities = np.asarray(intensities, dtype=np.float64)
        if intensities.size < 5:
            return []

        smooth = self.smooth_array(intensities, window=9)
        d1 = np.gradient(smooth, wavelengths)
        d2 = np.gradient(d1, wavelengths)
        noise = float(np.nanmedian(np.abs(smooth - self.smooth_array(smooth, window=21)))) if smooth.size >= 21 else 0.0
        threshold = max(float(np.nanmax(smooth)) * 0.05, noise * 5.0, 1e-9)
        peaks: List[Dict[str, float]] = []

        for i in range(1, smooth.size - 1):
            sign_change = d1[i - 1] > 0 and d1[i + 1] < 0
            local_max = smooth[i] >= smooth[i - 1] and smooth[i] >= smooth[i + 1]
            concave_down = d2[i] < 0
            if sign_change and local_max and concave_down and smooth[i] >= threshold:
                half = smooth[i] / 2.0
                left = i
                while left > 0 and smooth[left] >= half:
                    left -= 1
                right = i
                while right < smooth.size - 1 and smooth[right] >= half:
                    right += 1
                peaks.append({
                    "wavelength_nm": float(wavelengths[i]),
                    "intensity": float(smooth[i]),
                    "fwhm_nm": float(max(wavelengths[right] - wavelengths[left], 0.0)),
                })

        peaks.sort(key=lambda item: item["intensity"], reverse=True)
        return peaks[:8]

    def process_spectrum(self, wavelengths: np.ndarray, spectrum: np.ndarray, dark: Optional[np.ndarray], integration_time_s: float) -> Dict[str, Any]:
        raw = np.asarray(spectrum, dtype=np.float64)
        dark_corrected = raw.copy()
        if dark is not None and self.params.dark_spectrum_enabled:
            dark_corrected = dark_corrected - np.asarray(dark, dtype=np.float64)
        if self.params.baseline_correction_enabled:
            baseline, baseline_region, baseline_value = self.estimate_baseline(wavelengths, raw)
        else:
            baseline = np.zeros_like(raw)
            baseline_region = (float("nan"), float("nan"))
            baseline_value = 0.0
        baseline_corrected = raw - baseline
        normalized = baseline_corrected / max(float(integration_time_s), 1e-9)
        peaks = self.find_peaks_by_derivatives(wavelengths, normalized) if self.params.peak_detection_enabled else []
        return {
            "raw": raw,
            "dark_corrected": dark_corrected,
            "baseline": baseline,
            "baseline_region": baseline_region,
            "baseline_value": baseline_value,
            "baseline_corrected": baseline_corrected,
            "normalized": normalized,
            "peaks": peaks,
        }

    @staticmethod
    def peak_range(wavelengths, mode: str) -> Tuple[float, float]:
        ranges = {
            "red": (580, 700),
            "green": (480, 580),
            "blue": (400, 500),
            "other": (300, 1000),
            "auto": (380, 780),
            "visible": (380, 780),
            "all": (float(wavelengths[0]), float(wavelengths[-1])),
        }
        return ranges.get(mode, ranges["auto"])

    def find_peak_region(self, wavelengths, intensities, mode: str):
        wl_min, wl_max = self.peak_range(wavelengths, mode)
        mask = (wavelengths >= wl_min) & (wavelengths <= wl_max)
        if not np.any(mask):
            return None, None, None
        roi_wl = wavelengths[mask]
        roi_int = intensities[mask]
        if len(roi_int) == 0 or np.all(~np.isfinite(roi_int)):
            return None, None, None
        idx = int(np.nanargmax(roi_int))
        peak_wl = float(roi_wl[idx])
        peak_int = float(roi_int[idx])
        half = peak_int / 2.0
        above = roi_int >= half
        if np.any(above):
            inds = np.where(above)[0]
            fwhm = float(roi_wl[inds[-1]] - roi_wl[inds[0]])
        else:
            fwhm = 0.0
        return peak_wl, peak_int, fwhm

    def analyze_quality(self, wavelengths, intensities, mode: str):
        p = self.params
        peak_wl, peak_int, fwhm = self.find_peak_region(wavelengths, intensities, mode)
        if peak_int is None:
            return 0.0, 0.0, 0.0, False, True, False, False, "NO_PEAK"
        is_sat = bool(np.any(intensities >= p.saturation_level))
        is_weak = bool(peak_int < p.intensity_min)
        is_wide = bool(fwhm > p.max_peak_width_nm)
        is_narrow = bool(fwhm < p.min_peak_width_nm)
        if is_sat:
            status = "SATURATED"
        elif is_weak:
            status = "TOO_WEAK"
        elif is_wide:
            status = "TOO_WIDE"
        elif is_narrow:
            status = "TOO_NARROW"
        elif p.intensity_min <= peak_int <= p.intensity_max:
            status = "GOOD"
        else:
            status = "OK"
        return peak_int, peak_wl, fwhm, is_sat, is_weak, is_wide, is_narrow, status

    def optimize_integration_time(self, spec):
        p = self.params
        t_int = p.t_int_initial_s
        integral = 0.0
        best = None
        best_score = float("inf")
        self.last_optimization_started_saturated = False
        self.last_optimization_started_saturated_at_10ms = False

        self.log(f"  Подбор T_int: цель {p.target_intensity:.0f} counts, область {p.peak_search_mode_for_tint}")
        for iteration in range(1, p.max_iterations + 1):
            wl, inten, actual_t = self.get_spectrum(spec, t_int)
            peak_int, peak_wl, fwhm, is_sat, is_weak, _, _, status = self.analyze_quality(wl, inten, p.peak_search_mode_for_tint)
            if iteration == 1 and is_sat:
                self.last_optimization_started_saturated = True
                self.last_optimization_started_saturated_at_10ms = actual_t >= 0.0095
            if is_sat:
                score = float("inf")
            elif is_weak:
                score = abs(peak_int - p.target_intensity) * 10.0
            else:
                score = abs(peak_int - p.target_intensity)
            if score < best_score and not is_sat:
                best_score = score
                best = (actual_t, wl.copy(), inten.copy(), peak_int, peak_wl, fwhm, status)

            self.log(f"    {iteration}: T={actual_t*1000:.2f} мс, peak={peak_int:.0f} @ {peak_wl:.1f} нм, {status}")

            if status == "GOOD" or (status == "OK" and abs(peak_int - p.target_intensity) / p.target_intensity < p.tolerance):
                best = (actual_t, wl.copy(), inten.copy(), peak_int, peak_wl, fwhm, status)
                break
            if is_sat and actual_t <= p.t_int_min_s + 1e-4:
                best = (actual_t, wl.copy(), inten.copy(), peak_int, peak_wl, fwhm, status)
                break
            if is_weak and actual_t >= p.t_int_max_s - 1e-3:
                best = (actual_t, wl.copy(), inten.copy(), peak_int, peak_wl, fwhm, status)
                break

            current_int = max(peak_int, 1.0)
            error = (p.target_intensity - current_int) / p.target_intensity
            if abs(error) > 0.9:
                kp = p.kp * 5
                ki = p.ki * 3
            elif abs(error) > 0.7:
                kp = p.kp * 3
                ki = p.ki * 2
            elif abs(error) > 0.5:
                kp = p.kp * 2
                ki = p.ki * 1.5
            else:
                kp = p.kp
                ki = p.ki
            integral += ki * error
            integral = float(np.clip(integral, -2.0, 2.0))
            adjustment = float(np.clip(kp * error + integral, -0.9, 5.0))
            t_int = float(np.clip(actual_t * (1.0 + adjustment), p.t_int_min_s, p.t_int_max_s))

        if best is None:
            return None
        t, wl, inten, _, _, _, _ = best
        dark = None
        if p.dark_spectrum_enabled:
            _, dark, _ = self.get_dark_spectrum(spec, t)
        return t, wl, inten, dark


def create_spectrum_workbook(filename: Path, pixel_id: str, params: SpectrumParams, voltage_array: Iterable[float]) -> Workbook:
    wb = Workbook()
    ws_sum = wb.active
    ws_sum.title = "Сводка"
    ws_spec = wb.create_sheet("Спектры")
    ws_norm = wb.create_sheet("Processed counts per s")
    ws_raw = wb.create_sheet("Raw spectra")
    ws_dark = wb.create_sheet("Dark corrected")
    ws_baseline = wb.create_sheet("Baseline")
    ws_desc = wb.create_sheet("Описание полей")

    ws_sum["A1"] = "Спектро-электронное сканирование OLED"
    ws_sum["A1"].font = Font(bold=True, size=14)
    meta = [
        ("Pixel", pixel_id),
        ("Created", now_str()),
        ("Voltage range", f"{params.voltage_start}–{params.voltage_end} В, step {params.voltage_step} В"),
        ("Opening voltage stored (V)", params.opening_voltage if params.opening_voltage is not None else ""),
        ("Voltage start source", params.voltage_start_source),
        ("Current limit", f"{params.current_limit_mA} мА"),
        ("Pixel area (mm^2)", params.pixel_area_mm2),
        ("Luminance conversion (cd/m^2 per uA)", params.luminance_cd_m2_per_uA),
        ("LED_TYPE final", params.led_type),
        ("Peak search for T_int", params.peak_search_mode_for_tint),
        ("Derivative peak detection", "YES" if params.peak_detection_enabled else "NO"),
        ("T_int range", f"{params.t_int_min_s*1000:.2f}–{params.t_int_max_s*1000:.2f} мс"),
        ("Discard first scan after T_int change", "YES" if params.discard_first_scan_after_tint_change else "NO"),
        ("Baseline correction", "YES" if params.baseline_correction_enabled else "NO"),
        ("Saved intensity units", "Спектры: raw counts minus mean background from a flat raw-spectrum segment; Processed counts per s: same divided by T_int"),
    ]
    for idx, (k, v) in enumerate(meta, start=3):
        ws_sum.cell(row=idx, column=1, value=k).font = Font(bold=True)
        ws_sum.cell(row=idx, column=2, value=v)

    headers = [
        "Point",
        "V set (V)",
        "V LED measured (V)",
        "I LED (mA)",
        "J LED (mA/cm^2)",
        "V photodiode measured (V)",
        "I photodiode (uA)",
        "Luminance (cd/m^2)",
        "T_int saved spectrum (ms)",
        "Peak (nm)",
        "Max intensity processed (counts)",
        "FWHM (nm)",
        "Status",
        "Spectra column",
        "Peaks detected",
        "Peaks nm",
        "Background mean raw counts",
        "Background region nm",
    ]
    header_row = 21
    for c, h in enumerate(headers, start=1):
        ws_sum.cell(row=header_row, column=c, value=h)
    style_header_row(ws_sum, header_row, 1, len(headers))
    ws_sum.freeze_panes = f"A{header_row + 1}"

    meta_labels = [
        "Название столбца",
        "Point",
        "V set (V)",
        "V LED measured (V)",
        "I LED (mA)",
        "J LED (mA/cm^2)",
        "V photodiode measured (V)",
        "I photodiode (uA)",
        "Luminance (cd/m^2)",
        "T_int saved spectrum (ms)",
        "Peak (nm)",
        "Max intensity processed (counts/s)",
        "FWHM (nm)",
        "Status",
        "Comment",
        "Peaks detected",
        "Peaks nm",
        "Background mean raw counts",
        "Background region nm",
    ]
    for r, label in enumerate(meta_labels, start=1):
        ws_spec.cell(row=r, column=1, value=label).font = Font(bold=True)
        ws_spec.cell(row=r, column=1).fill = PatternFill("solid", fgColor="E2F0D9")

    data_header_row = 20
    ws_spec.cell(row=data_header_row, column=1, value="Wavelength (nm)").font = Font(bold=True)
    ws_spec.cell(row=data_header_row, column=1).fill = PatternFill("solid", fgColor="D9E1F2")
    for idx, voltage in enumerate(voltage_array, start=1):
        col = idx + 1
        ws_spec.cell(row=1, column=col, value=f"Point {idx}: {float(voltage):.3f} V")
        ws_spec.cell(row=2, column=col, value=idx)
        ws_spec.cell(row=3, column=col, value=float(voltage))
        ws_spec.cell(row=data_header_row, column=col, value=f"Intensity point {idx}, counts corrected")
        ws_spec.cell(row=data_header_row, column=col).font = Font(bold=True)
        ws_spec.cell(row=data_header_row, column=col).fill = PatternFill("solid", fgColor="D9E1F2")

    ws_spec.freeze_panes = f"B{data_header_row + 1}"
    extra_sheets = [
        (ws_norm, "Processed counts per s"),
        (ws_raw, "Raw counts"),
        (ws_dark, "Dark-corrected counts"),
        (ws_baseline, "Background mean counts"),
    ]
    for extra_ws, data_label in extra_sheets:
        for r, label in enumerate(meta_labels, start=1):
            extra_ws.cell(row=r, column=1, value=label).font = Font(bold=True)
            extra_ws.cell(row=r, column=1).fill = PatternFill("solid", fgColor="E2F0D9")
        extra_ws.cell(row=data_header_row, column=1, value="Wavelength (nm)").font = Font(bold=True)
        extra_ws.cell(row=data_header_row, column=1).fill = PatternFill("solid", fgColor="D9E1F2")
        for idx, voltage in enumerate(voltage_array, start=1):
            col = idx + 1
            extra_ws.cell(row=1, column=col, value=f"Point {idx}: {float(voltage):.3f} V")
            extra_ws.cell(row=2, column=col, value=idx)
            extra_ws.cell(row=3, column=col, value=float(voltage))
            extra_ws.cell(row=data_header_row, column=col, value=f"{data_label} point {idx}")
            extra_ws.cell(row=data_header_row, column=col).font = Font(bold=True)
            extra_ws.cell(row=data_header_row, column=col).fill = PatternFill("solid", fgColor="D9E1F2")
        extra_ws.freeze_panes = f"B{data_header_row + 1}"
    ws_desc.append(["Лист", "Поле", "Что означает"])
    style_header_row(ws_desc, 1, 1, 3)
    desc_rows = [
        ("Сводка", "T_int saved spectrum", "Время интегрирования именно того спектра, который записан в столбец."),
        ("Сводка", "Peaks detected / Peaks nm", "Заполняется только если включен флажок поиска пиков производными."),
        ("Спектры", "Intensity", "Raw counts минус среднее значение фона из найденного плоского участка raw-спектра, counts."),
        ("Processed counts per s", "Processed counts/s", "Копия обработанного спектра, деленная на время интегрирования."),
        ("Raw spectra", "Raw counts", "Сырые counts напрямую со спектрометра."),
        ("Dark corrected", "Dark-corrected counts", "Диагностический лист: Raw минус dark spectrum, если dark включен; основная обработка его не использует."),
        ("Baseline", "Background mean counts", "Константный уровень фона: среднее значение найденного плоского участка raw-спектра."),
        ("Спектры", "Строки 1–15", "Метаданные каждого спектра."),
        ("Спектры", "Строка 21+", "Длины волн и интенсивности."),
        ("Важно", "Первый спектр после смены T_int", "Сбрасывается, чтобы не записать старый буфер спектрометра."),
    ]
    for row in desc_rows:
        ws_desc.append(row)

    for ws in wb.worksheets:
        autosize_columns(ws, max_width=45)
    wb.save(filename)
    return wb


def run_spectrum_measurement(
    pixel_id: str,
    output_dir: Path,
    params: SpectrumParams,
    log: Callable[[str], None],
    app_settings: Optional[Dict[str, Any]] = None,
    progress_callback: Optional[Callable[[int, float, float, np.ndarray, np.ndarray, np.ndarray, List[Dict[str, float]], str], None]] = None,
) -> Dict:
    prepare_hardware_environment(pixel_id, app_settings, log)
    import xtralien

    helper = SpectrumHelper(params, log)
    spec = helper.init_spectrometer()

    voltage_array = np.arange(params.voltage_start, params.voltage_end + params.voltage_step / 2, params.voltage_step)
    voltage_array = np.round(voltage_array, 6)
    filename = output_dir / f"SPECTRUM_{safe_filename(pixel_id)}_{timestamp_for_file()}.xlsx"
    wb = create_spectrum_workbook(filename, pixel_id, params, voltage_array)
    ws_sum = wb["Сводка"]
    ws_spec = wb["Спектры"]
    ws_norm = wb["Processed counts per s"]
    ws_raw = wb["Raw spectra"]
    ws_dark = wb["Dark corrected"]
    ws_baseline = wb["Baseline"]
    summary_header_row = 21
    spectra_data_start = 21

    final_status = "FAILED"
    best_spectrum_metrics = {
        "spectrum_peak_count": None,
        "spectrum_peaks_nm": "",
        "spectrum_max_intensity": None,
    }

    with xtralien.Device(params.com_port) as smu:
        try:
            smu.smu1.set.enabled(True, response=0)
            smu.smu2.set.enabled(True, response=0)
            try:
                smu.smu2.set.range(params.photodiode_range, response=0)
            except Exception:
                pass
            smu.smu1.set.voltage(0, response=0)
            smu.smu2.set.voltage(params.photodiode_bias_V, response=0)
            time.sleep(0.3)

            for idx, voltage in enumerate(voltage_array, start=1):
                log(f"\nСпектр {pixel_id}: точка {idx}/{len(voltage_array)}, V={voltage:.3f} В")
                smu.smu1.set.voltage(float(voltage), response=0)
                smu.smu2.set.voltage(params.photodiode_bias_V, response=0)
                time.sleep(params.settle_time_voltage_s)

                v_led, i_led = smu.smu1.measure()[0]
                v_pd, i_pd = smu.smu2.measure()[0]
                i_led_mA = i_led * 1000.0
                i_pd_uA = -i_pd * 1_000_000.0
                j_led = current_density_mA_cm2(i_led_mA, params.pixel_area_mm2)
                lum = luminance_cd_m2(i_pd_uA, params.luminance_cd_m2_per_uA)

                row = summary_header_row + idx
                spectra_col = idx + 1
                if i_led_mA >= params.current_limit_mA:
                    status = "CURRENT_LIMIT"
                    log(f"  Стоп: ток {i_led_mA:.3f} мА >= {params.current_limit_mA:.3f} мА")
                    summary_values = [idx, float(voltage), v_led, i_led_mA, j_led, v_pd, i_pd_uA, lum, "—", "—", "—", "—", status, get_column_letter(spectra_col), 0, "", "", ""]
                    for c, val in enumerate(summary_values, start=1):
                        ws_sum.cell(row=row, column=c, value=val)
                    ws_spec.cell(row=12, column=spectra_col, value=status)
                    final_status = status
                    wb.save(filename)
                    break

                opt = helper.optimize_integration_time(spec)
                if opt is None:
                    status = "FAILED"
                    t_int = params.t_int_initial_s
                    summary_values = [idx, float(voltage), v_led, i_led_mA, j_led, v_pd, i_pd_uA, lum, t_int * 1000, "—", "—", "—", status, get_column_letter(spectra_col), 0, "", "", ""]
                    for c, val in enumerate(summary_values, start=1):
                        ws_sum.cell(row=row, column=c, value=val)
                    final_status = status
                    wb.save(filename)
                    continue

                t_int, wavelengths, spectrum, dark = opt
                processed = helper.process_spectrum(wavelengths, spectrum, dark, t_int)
                spectrum_to_save = processed["baseline_corrected"]
                normalized_to_save = processed["normalized"]
                raw_to_save = processed["raw"]
                dark_to_save = processed["dark_corrected"]
                baseline_to_save = processed["baseline"]
                baseline_value = float(processed.get("baseline_value", 0.0))
                baseline_region = processed.get("baseline_region", (float("nan"), float("nan")))
                baseline_region_text = (
                    f"{float(baseline_region[0]):.2f}–{float(baseline_region[1]):.2f}"
                    if np.all(np.isfinite(np.asarray(baseline_region, dtype=np.float64)))
                    else ""
                )
                peaks = processed["peaks"]

                peak_int, peak_wl, fwhm, _, _, _, _, status = helper.analyze_quality(wavelengths, processed["baseline_corrected"], params.led_type)
                if peaks:
                    peak_int = peaks[0]["intensity"]
                    peak_wl = peaks[0]["wavelength_nm"]
                    fwhm = peaks[0]["fwhm_nm"]
                if status not in {"SATURATED", "FAILED", "NO_PEAK"} and (
                    helper.adaptive_initial_time_enabled or helper.last_optimization_started_saturated_at_10ms
                ):
                    previous_t = float(params.t_int_initial_s)
                    params.t_int_initial_s = float(t_int)
                    helper.adaptive_initial_time_enabled = True
                    if helper.last_optimization_started_saturated_at_10ms:
                        log(
                            f"  Первая проба на 10 мс была saturated; следующее начальное T_int: "
                            f"{params.t_int_initial_s*1000:.2f} мс вместо {previous_t*1000:.2f} мс"
                        )
                peaks_nm = ", ".join(f"{p['wavelength_nm']:.1f}" for p in peaks)
                if peak_int and (
                    best_spectrum_metrics["spectrum_max_intensity"] is None
                    or float(peak_int) > float(best_spectrum_metrics["spectrum_max_intensity"])
                ):
                    best_spectrum_metrics = {
                        "spectrum_peak_count": len(peaks),
                        "spectrum_peaks_nm": peaks_nm,
                        "spectrum_max_intensity": float(peak_int),
                    }
                summary_values = [
                    idx,
                    float(voltage),
                    round(float(v_led), 6),
                    round(float(i_led_mA), 6),
                    round(float(j_led), 6) if j_led is not None else "—",
                    round(float(v_pd), 6),
                    round(float(i_pd_uA), 6),
                    round(float(lum), 6) if lum is not None else "—",
                    round(float(t_int) * 1000, 3),
                    round(float(peak_wl), 3) if peak_wl else "—",
                    round(float(peak_int), 1) if peak_int else "—",
                    round(float(fwhm), 3) if fwhm else "—",
                    status,
                    get_column_letter(spectra_col),
                    len(peaks),
                    peaks_nm,
                    round(baseline_value, 3),
                    baseline_region_text,
                ]
                for c, val in enumerate(summary_values, start=1):
                    ws_sum.cell(row=row, column=c, value=val)

                meta_values = [
                    f"Point {idx}: {float(voltage):.3f} V",
                    idx,
                    float(voltage),
                    round(float(v_led), 6),
                    round(float(i_led_mA), 6),
                    round(float(j_led), 6) if j_led is not None else "—",
                    round(float(v_pd), 6),
                    round(float(i_pd_uA), 6),
                    round(float(lum), 6) if lum is not None else "—",
                    round(float(t_int) * 1000, 3),
                    round(float(peak_wl), 3) if peak_wl else "—",
                    round(float(peak_int), 1) if peak_int else "—",
                    round(float(fwhm), 3) if fwhm else "—",
                    status,
                    "OK: spectrum saved",
                    len(peaks),
                    peaks_nm,
                    round(baseline_value, 3),
                    baseline_region_text,
                ]
                for r, val in enumerate(meta_values, start=1):
                    ws_spec.cell(row=r, column=spectra_col, value=val)
                    ws_norm.cell(row=r, column=spectra_col, value=val)
                    ws_raw.cell(row=r, column=spectra_col, value=val)
                    ws_dark.cell(row=r, column=spectra_col, value=val)
                    ws_baseline.cell(row=r, column=spectra_col, value=val)

                for r, wl in enumerate(wavelengths, start=spectra_data_start):
                    if ws_spec.cell(row=r, column=1).value is None:
                        ws_spec.cell(row=r, column=1, value=round(float(wl), 2))
                    if ws_norm.cell(row=r, column=1).value is None:
                        ws_norm.cell(row=r, column=1, value=round(float(wl), 2))
                    if ws_raw.cell(row=r, column=1).value is None:
                        ws_raw.cell(row=r, column=1, value=round(float(wl), 2))
                    if ws_dark.cell(row=r, column=1).value is None:
                        ws_dark.cell(row=r, column=1, value=round(float(wl), 2))
                    if ws_baseline.cell(row=r, column=1).value is None:
                        ws_baseline.cell(row=r, column=1, value=round(float(wl), 2))
                for r, intensity in enumerate(spectrum_to_save, start=spectra_data_start):
                    ws_spec.cell(row=r, column=spectra_col, value=round(float(intensity), 3))
                for r, intensity in enumerate(normalized_to_save, start=spectra_data_start):
                    ws_norm.cell(row=r, column=spectra_col, value=round(float(intensity), 3))
                for r, intensity in enumerate(raw_to_save, start=spectra_data_start):
                    ws_raw.cell(row=r, column=spectra_col, value=round(float(intensity), 3))
                for r, intensity in enumerate(dark_to_save, start=spectra_data_start):
                    ws_dark.cell(row=r, column=spectra_col, value=round(float(intensity), 3))
                for r, intensity in enumerate(baseline_to_save, start=spectra_data_start):
                    ws_baseline.cell(row=r, column=spectra_col, value=round(float(intensity), 3))

                final_status = status
                if progress_callback is not None:
                    progress_callback(idx, float(voltage), float(t_int), wavelengths, processed["raw"], spectrum_to_save, peaks, status)
                log(f"  Сохранено: T_int={t_int*1000:.2f} мс, peak={peak_int:.0f} @ {peak_wl:.1f} нм, {status}")
                if idx % 3 == 0 or idx == len(voltage_array):
                    wb.save(filename)
                    gc.collect()
        finally:
            safe_shutdown_smu(smu)

    for ws in wb.worksheets:
        autosize_columns(ws, max_width=45)
    wb.save(filename)
    wb.close()
    return {"file": filename, "status": final_status, **best_spectrum_metrics}


# -----------------------------------------------------------------------------
# Измерение стабильности
# -----------------------------------------------------------------------------

@dataclass
class StabilityParams:
    com_port: str = "COM3"
    current_setpoint_mA: float = 3.5
    voltage_start: float = 3.5
    voltage_limit: float = 5.0
    current_limit_mA: float = 10.0
    voltage_step_max: float = 0.02
    current_control_kp: float = 0.01
    measurement_time_s: float = 86400.0
    sample_interval_s: float = 1.0
    autosave_interval_s: float = 600.0
    photodiode_bias_V: float = -5.0
    photodiode_threshold_uA: float = 0.1
    photodiode_range: int = 4
    pixel_area_mm2: float = 1.0
    luminance_cd_m2_per_uA: float = 1.0

    def as_dict(self) -> Dict:
        return self.__dict__.copy()


def find_ivl_data_columns(ws) -> Optional[Tuple[int, Dict[str, int]]]:
    wanted = {
        "Voltage OLED / LED measured (V)": None,
        "Current OLED / LED (mA)": None,
    }
    for r in range(1, min(ws.max_row, 30) + 1):
        headers = {}
        for c in range(1, ws.max_column + 1):
            val = ws.cell(row=r, column=c).value
            if val in wanted:
                headers[val] = c
        if all(k in headers for k in wanted):
            return r, headers
    return None


def interpolate_voltage_at_current_from_ivl(ivl_file: Path, target_current_mA: float) -> Optional[float]:
    ivl_file = Path(ivl_file)
    if not ivl_file.exists():
        return None
    wb = load_workbook(ivl_file, data_only=True)
    points: List[Tuple[float, float]] = []
    try:
        for sheet_name in wb.sheetnames:
            if not sheet_name.startswith("Cycle_"):
                continue
            ws = wb[sheet_name]
            found = find_ivl_data_columns(ws)
            if not found:
                continue
            header_row, cols = found
            v_col = cols["Voltage OLED / LED measured (V)"]
            i_col = cols["Current OLED / LED (mA)"]
            for r in range(header_row + 1, ws.max_row + 1):
                v = as_float_or_none(ws.cell(row=r, column=v_col).value)
                i = as_float_or_none(ws.cell(row=r, column=i_col).value)
                if v is not None and i is not None and math.isfinite(v) and math.isfinite(i):
                    points.append((i, v))
    finally:
        wb.close()

    if len(points) < 2:
        return None
    points = sorted(points, key=lambda x: x[0])

    # Если цель вне диапазона, возвращаем ближайшую точку, но только если она не совсем далеко.
    if target_current_mA <= points[0][0]:
        return points[0][1]
    if target_current_mA >= points[-1][0]:
        return points[-1][1]

    for (i1, v1), (i2, v2) in zip(points[:-1], points[1:]):
        if i1 <= target_current_mA <= i2 and i2 != i1:
            k = (target_current_mA - i1) / (i2 - i1)
            return v1 + k * (v2 - v1)
    return min(points, key=lambda x: abs(x[0] - target_current_mA))[1]


def create_stability_workbook(filename: Path, pixel_id: str, params: StabilityParams) -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    ws["A1"] = "OLED stability"
    ws["A1"].font = Font(bold=True, size=14)
    info = [
        ("Pixel", pixel_id),
        ("Created", now_str()),
        ("Mode", "constant current, software control"),
        ("Current setpoint (mA)", params.current_setpoint_mA),
        ("Voltage start (V)", params.voltage_start),
        ("Voltage limit (V)", params.voltage_limit),
        ("Current limit (mA)", params.current_limit_mA),
        ("Measurement time set (s)", params.measurement_time_s),
        ("Sample interval (s)", params.sample_interval_s),
        ("Autosave interval (s)", params.autosave_interval_s),
        ("Photodiode threshold (uA)", params.photodiode_threshold_uA),
        ("Pixel area (mm^2)", params.pixel_area_mm2),
        ("Luminance conversion (cd/m^2 per uA)", params.luminance_cd_m2_per_uA),
        ("Status", "IN_PROGRESS"),
        ("Max photodiode current (uA)", 0),
        ("Last saved elapsed time (s)", 0),
    ]
    for r, (k, v) in enumerate(info, start=3):
        ws.cell(row=r, column=1, value=k).font = Font(bold=True)
        ws.cell(row=r, column=2, value=v)
    headers = [
        "Point",
        "Date time",
        "Time (s)",
        "Current setpoint (mA)",
        "Voltage OLED / LED (V)",
        "Current OLED / LED (mA)",
        "Current density (mA/cm^2)",
        "Voltage photodiode (V)",
        "Photodiode current (uA)",
        "Luminance (cd/m^2)",
    ]
    header_row = 20
    for c, h in enumerate(headers, start=1):
        ws.cell(row=header_row, column=c, value=h)
    style_header_row(ws, header_row, 1, len(headers))
    ws.freeze_panes = f"A{header_row + 1}"
    autosize_columns(ws, max_width=36)
    wb.save(filename)
    return wb


def update_stability_status(ws, status: str, max_photo: float, elapsed: float):
    ws["B16"] = status
    ws["B17"] = max_photo
    ws["B18"] = elapsed
    fill = PatternFill("solid", fgColor="C6EFCE" if status == "WORKING" else "FFEB9C" if status == "IN_PROGRESS" else "FFC7CE")
    ws["B16"].fill = fill


def save_stability_chart(filename: Path, pixel_id: str):
    wb = load_workbook(filename)
    ws = wb["Data"]
    max_row = ws.max_row
    if max_row > 22:
        chart = ScatterChart()
        chart.title = f"Stability {pixel_id}"
        chart.x_axis.title = "Time (s)"
        chart.y_axis.title = "Current OLED (mA) / Photodiode (uA)"
        chart.x_axis.majorGridlines = ChartLines()
        xvalues = Reference(ws, min_col=3, min_row=21, max_row=max_row)
        y_current = Reference(ws, min_col=6, min_row=20, max_row=max_row)
        y_photo = Reference(ws, min_col=9, min_row=20, max_row=max_row)
        chart.series.append(Series(y_current, xvalues, title_from_data=True))
        chart.series.append(Series(y_photo, xvalues, title_from_data=True))
        ws.add_chart(chart, "J3")
    autosize_columns(ws, max_width=36)
    wb.save(filename)
    wb.close()


def run_stability_measurement(pixel_id: str, output_dir: Path, params: StabilityParams, log: Callable[[str], None], app_settings: Optional[Dict[str, Any]] = None) -> Dict:
    prepare_hardware_environment(pixel_id, app_settings, log)
    import xtralien

    filename = output_dir / f"STABILITY_{safe_filename(pixel_id)}_{params.current_setpoint_mA:g}mA_{timestamp_for_file()}.xlsx"
    wb = create_stability_workbook(filename, pixel_id, params)
    ws = wb["Data"]

    max_photo = 0.0
    point_number = 0
    last_autosave_elapsed = 0.0
    voltage_set = params.voltage_start
    current_limit_reached = False
    voltage_limit_reached = False
    last_elapsed = 0.0

    with xtralien.Device(params.com_port) as smu:
        try:
            smu.smu1.set.enabled(True, response=0)
            smu.smu2.set.enabled(True, response=0)
            try:
                smu.smu2.set.range(params.photodiode_range, response=0)
            except Exception:
                pass
            time.sleep(0.1)
            smu.smu1.set.voltage(0, response=0)
            smu.smu2.set.voltage(params.photodiode_bias_V, response=0)
            time.sleep(0.3)

            start = time.time()
            next_point = start
            log(f"Стабильность {pixel_id}: I={params.current_setpoint_mA:.3f} мА, старт V={voltage_set:.3f} В")
            while True:
                now = time.time()
                elapsed = now - start
                if elapsed > params.measurement_time_s:
                    break
                if now < next_point:
                    time.sleep(next_point - now)
                next_point += params.sample_interval_s
                now = time.time()
                elapsed = now - start
                last_elapsed = elapsed

                smu.smu1.set.voltage(float(voltage_set), response=0)
                smu.smu2.set.voltage(params.photodiode_bias_V, response=0)
                time.sleep(0.02)
                v_led, i_led = smu.smu1.measure()[0]
                v_pd, i_pd = smu.smu2.measure()[0]
                i_led_mA = i_led * 1000.0
                i_pd_uA = -i_pd * 1_000_000.0
                j_led = current_density_mA_cm2(i_led_mA, params.pixel_area_mm2)
                lum = luminance_cd_m2(i_pd_uA, params.luminance_cd_m2_per_uA)

                max_photo = max(max_photo, i_pd_uA)
                point_number += 1
                ws.append([
                    point_number,
                    now_str(),
                    elapsed,
                    params.current_setpoint_mA,
                    v_led,
                    i_led_mA,
                    j_led,
                    v_pd,
                    i_pd_uA,
                    lum,
                ])

                if i_led_mA > params.current_limit_mA:
                    current_limit_reached = True
                    log(f"Аварийный стоп: ток {i_led_mA:.3f} мА > {params.current_limit_mA:.3f} мА")
                    try:
                        smu.smu1.set.voltage(0, response=0)
                    except Exception:
                        pass

                if not current_limit_reached:
                    error_mA = params.current_setpoint_mA - i_led_mA
                    dV = params.current_control_kp * error_mA
                    dV = float(np.clip(dV, -params.voltage_step_max, params.voltage_step_max))
                    voltage_set += dV
                    voltage_set = max(0.0, voltage_set)
                    if voltage_set >= params.voltage_limit:
                        voltage_set = params.voltage_limit
                        voltage_limit_reached = True
                        log(f"Достигнут предел напряжения {params.voltage_limit:.3f} В")

                if point_number == 1 or point_number % 60 == 0:
                    log(f"  t={elapsed:.1f} c; V={v_led:.3f} В; I={i_led_mA:.3f} мА; PD={i_pd_uA:.3f} мкА")

                need_save = (elapsed - last_autosave_elapsed >= params.autosave_interval_s) or current_limit_reached or voltage_limit_reached
                if need_save:
                    status = "CURRENT_LIMIT_STOP" if current_limit_reached else "VOLTAGE_LIMIT_STOP" if voltage_limit_reached else "IN_PROGRESS"
                    update_stability_status(ws, status, max_photo, elapsed)
                    wb.save(filename)
                    last_autosave_elapsed = elapsed
                    gc.collect()
                    log(f"  Автосохранение: {filename.name}, t={elapsed:.1f} c")

                if current_limit_reached or voltage_limit_reached:
                    break
        finally:
            safe_shutdown_smu(smu)

    if current_limit_reached:
        status = "BURNED"
    elif voltage_limit_reached and max_photo < params.photodiode_threshold_uA:
        status = "NO_CONTACT"
    elif max_photo >= params.photodiode_threshold_uA:
        status = "WORKING"
    else:
        status = "NONWORKING"

    update_stability_status(ws, status, max_photo, last_elapsed)
    wb.save(filename)
    wb.close()
    save_stability_chart(filename, pixel_id)
    return {"file": filename, "status": status, "max_photo_uA": max_photo}


class IVLProgressWindow:
    def __init__(self, parent: tk.Misc, pixel_id: str, params: IVLParams):
        self.closed = False
        self.stop_requested = False
        self.pixel_id = pixel_id
        self.params = params
        self.current_cycle = 1
        self.points: List[Tuple[float, float, float, float, float]] = []
        self.win = tk.Toplevel(parent)
        self.win.title(f"ВАЯХ: {pixel_id}")
        self.win.geometry("980x700")
        self.win.minsize(680, 460)
        self.win.protocol("WM_DELETE_WINDOW", self.close)

        main = ttk.Frame(self.win, padding=10)
        main.pack(fill="both", expand=True)
        self.status_var = tk.StringVar(value=f"Пиксель {pixel_id}: ожидание старта")
        ttk.Label(main, textvariable=self.status_var, font=("Segoe UI", 10, "bold")).pack(anchor="w", pady=(0, 6))
        controls = ttk.Frame(main)
        controls.pack(fill="x", pady=(0, 8))
        ttk.Button(controls, text="Остановить измерение и поставить 0 В", command=self.request_stop).pack(side="left")
        ui_settings = getattr(parent, "app_settings", {}).get("ui", {}) if hasattr(parent, "app_settings") else {}
        self.graph_mode = tk.StringVar(value=str(ui_settings.get("last_ivl_graph_mode", "raw") or "raw"))
        ttk.Radiobutton(controls, text="I / ФД", variable=self.graph_mode, value="raw", command=self._on_graph_mode_changed).pack(side="left", padx=(14, 0))
        ttk.Radiobutton(controls, text="J / L", variable=self.graph_mode, value="converted", command=self._on_graph_mode_changed).pack(side="left", padx=(8, 0))
        ttk.Label(
            controls,
            text=f"S={params.pixel_area_mm2:g} мм^2; k={params.luminance_cd_m2_per_uA:g} кд/м^2/мкА",
            foreground="#555555",
        ).pack(side="left", padx=(12, 0))

        self.canvas = tk.Canvas(main, width=860, height=280, bg="white", highlightthickness=1, highlightbackground="#BFBFBF")
        self.canvas.pack(fill="x", pady=(0, 8))
        self.canvas.bind("<Configure>", lambda _event: self._redraw_plot())

        table_wrap = ttk.Frame(main)
        table_wrap.pack(fill="both", expand=True)
        columns = ("cycle", "point", "vset", "vled", "iled", "jled", "ipd", "lum")
        self.tree = ttk.Treeview(table_wrap, columns=columns, show="headings", height=12)
        headers = {
            "cycle": "Цикл",
            "point": "Точка",
            "vset": "V set, В",
            "vled": "V LED, В",
            "iled": "I LED, мА",
            "jled": "J, мА/см^2",
            "ipd": "I ФД, мкА",
            "lum": "L, кд/м^2",
        }
        widths = {"cycle": 60, "point": 70, "vset": 100, "vled": 100, "iled": 100, "jled": 115, "ipd": 105, "lum": 115}
        for col in columns:
            self.tree.heading(col, text=headers[col])
            self.tree.column(col, width=widths[col], minwidth=widths[col], stretch=(col == "ipd"))
        yscroll = ttk.Scrollbar(table_wrap, orient="vertical", command=self.tree.yview)
        xscroll = ttk.Scrollbar(table_wrap, orient="horizontal", command=self.tree.xview)
        self.tree.configure(yscrollcommand=yscroll.set, xscrollcommand=xscroll.set)
        self.tree.grid(row=0, column=0, sticky="nsew")
        yscroll.grid(row=0, column=1, sticky="ns")
        xscroll.grid(row=1, column=0, sticky="ew")
        table_wrap.rowconfigure(0, weight=1)
        table_wrap.columnconfigure(0, weight=1)
        fit_toplevel_to_content(self.win, 980, 700)
        self._redraw_plot()
        self._safe_update()

    def close(self):
        if self.closed:
            return
        self.closed = True
        try:
            self.win.destroy()
        except Exception:
            pass

    def request_stop(self):
        if self.closed:
            return
        self.stop_requested = True
        self.status_var.set("Остановка запрошена: на следующей точке будет отправлено 0 В")
        self._safe_update()

    def set_status(self, text: str):
        if self.closed:
            return
        self.status_var.set(text)
        self._safe_update()

    def add_point(self, cycle_number: int, row: Dict[str, Any]):
        if self.closed:
            return
        if self.stop_requested:
            raise MeasurementStopped("Измерение остановлено пользователем")
        if cycle_number != self.current_cycle:
            self.current_cycle = cycle_number
            self.points = []
            for item in self.tree.get_children():
                self.tree.delete(item)
        v = float(row.get("Voltage OLED / LED measured (V)", row.get("Voltage set (V)", 0.0)) or 0.0)
        i = float(row.get("Current OLED / LED (mA)", 0.0) or 0.0)
        pd = float(row.get("Photodiode current (uA)", 0.0) or 0.0)
        j = as_float_or_none(row.get("Current density (mA/cm^2)"))
        lum = as_float_or_none(row.get("Luminance (cd/m^2)"))
        j = float(j) if j is not None else 0.0
        lum = float(lum) if lum is not None else 0.0
        self.points.append((v, i, pd, j, lum))
        self.tree.insert(
            "",
            "end",
            values=(
                cycle_number,
                row.get("Point", ""),
                f"{float(row.get('Voltage set (V)', 0.0)):.4f}",
                f"{v:.4f}",
                f"{i:.4f}",
                f"{j:.4f}",
                f"{pd:.4f}",
                f"{lum:.4f}",
            ),
        )
        children = self.tree.get_children()
        if children:
            self.tree.see(children[-1])
        self.status_var.set(
            f"Пиксель {self.pixel_id} | цикл {cycle_number} | точка {row.get('Point', '')} | "
            f"V={v:.3f} В | I={i:.3f} мА | J={j:.3f} мА/см^2 | L={lum:.3f} кд/м^2"
        )
        self._redraw_plot()
        self._safe_update()

    def _on_graph_mode_changed(self):
        parent = self.win.master
        try:
            if hasattr(parent, "save_ui_preference"):
                parent.save_ui_preference("last_ivl_graph_mode", self.graph_mode.get())
        except Exception:
            pass
        self._redraw_plot()

    def _redraw_plot(self):
        if self.closed:
            return
        c = self.canvas
        c.delete("all")
        width = int(c.winfo_width() or 860)
        height = int(c.winfo_height() or 280)
        left, top, right, bottom = 68, 18, width - 72, height - 38
        c.create_rectangle(left, top, right, bottom, outline="#A0A0A0")
        converted = self.graph_mode.get() == "converted"
        left_axis = "J OLED\nмА/см^2" if converted else "I OLED\nмА"
        right_axis = "L\nкд/м^2" if converted else "I PD\nмкА"
        blue_label = "J OLED" if converted else "I OLED"
        red_label = "L" if converted else "I PD"
        title_suffix = "плотность тока и светимость" if converted else "ток OLED и фототок"
        c.create_text(width - 28, (top + bottom) / 2, text=right_axis, font=("Segoe UI", 9), justify="center", fill="#C43C30")
        c.create_line(right - 158, top + 14, right - 132, top + 14, fill="#0B61A4", width=2)
        c.create_text(right - 126, top + 14, text=blue_label, anchor="w", font=("Segoe UI", 8), fill="#0B61A4")
        c.create_line(right - 78, top + 14, right - 52, top + 14, fill="#C43C30", width=2)
        c.create_text(right - 46, top + 14, text=red_label, anchor="w", font=("Segoe UI", 8), fill="#C43C30")
        c.create_text((left + right) / 2, 8, text=f"ВАХ / ВАЯХ — {title_suffix}, {self.pixel_id}", font=("Segoe UI", 10, "bold"))
        c.create_text((left + right) / 2, height - 12, text="Напряжение, В", font=("Segoe UI", 9))
        c.create_text(18, (top + bottom) / 2, text=left_axis, font=("Segoe UI", 9), justify="center", fill="#0B61A4")

        if not self.points:
            c.create_text((left + right) / 2, (top + bottom) / 2, text="Данные появятся во время измерения", fill="#666666")
            return

        max_x = max(max(v for v, *_rest in self.points), 0.1)
        if converted:
            left_values = [j for _v, _i, _pd, j, _lum in self.points]
            right_values = [lum for _v, _i, _pd, _j, lum in self.points]
        else:
            left_values = [i for _v, i, _pd, _j, _lum in self.points]
            right_values = [pd for _v, _i, pd, _j, _lum in self.points]
        max_left = max(max(abs(v) for v in left_values), 0.1) * 1.08
        max_right = max(max(abs(v) for v in right_values), 0.1) * 1.08
        for frac in [0.0, 0.25, 0.5, 0.75, 1.0]:
            y = bottom - frac * (bottom - top)
            val_left = frac * max_left
            val_right = frac * max_right
            c.create_line(left, y, right, y, fill="#EEEEEE")
            c.create_text(left - 8, y, text=f"{val_left:.2f}", anchor="e", font=("Segoe UI", 8), fill="#0B61A4")
            c.create_text(right + 8, y, text=f"{val_right:.2f}", anchor="w", font=("Segoe UI", 8), fill="#C43C30")
        for frac in [0.0, 0.25, 0.5, 0.75, 1.0]:
            x = left + frac * (right - left)
            val = frac * max_x
            c.create_line(x, top, x, bottom, fill="#F4F4F4")
            c.create_text(x, bottom + 14, text=f"{val:.2f}", anchor="n", font=("Segoe UI", 8))

        led_coords: List[float] = []
        pd_coords: List[float] = []
        for v, i, pd, j, lum in self.points:
            y_left_value = j if converted else i
            y_right_value = lum if converted else abs(pd)
            x = left + (v / max_x) * (right - left)
            y_led = bottom - (y_left_value / max_left) * (bottom - top)
            y_pd = bottom - (y_right_value / max_right) * (bottom - top)
            led_coords.extend([x, y_led])
            pd_coords.extend([x, y_pd])
            c.create_oval(x - 2.5, y_led - 2.5, x + 2.5, y_led + 2.5, fill="#0B61A4", outline="#0B61A4")
            c.create_rectangle(x - 2.5, y_pd - 2.5, x + 2.5, y_pd + 2.5, fill="#C43C30", outline="#C43C30")
        if len(led_coords) >= 4:
            c.create_line(*led_coords, fill="#0B61A4", width=2)
        if len(pd_coords) >= 4:
            c.create_line(*pd_coords, fill="#C43C30", width=2)

    def _safe_update(self):
        if self.closed:
            return
        try:
            self.win.update_idletasks()
            self.win.update()
        except Exception:
            self.closed = True


class SpectrumProgressWindow:
    def __init__(self, parent: tk.Misc, pixel_id: str):
        self.closed = False
        self.pixel_id = pixel_id
        self.last: Optional[Dict[str, Any]] = None
        self.win = tk.Toplevel(parent)
        self.win.title(f"Спектр: {pixel_id}")
        self.win.geometry("980x700")
        self.win.minsize(680, 460)
        self.win.protocol("WM_DELETE_WINDOW", self.close)

        main = ttk.Frame(self.win, padding=10)
        main.pack(fill="both", expand=True)
        self.status_var = tk.StringVar(value=f"Пиксель {pixel_id}: ожидание спектра")
        ttk.Label(main, textvariable=self.status_var, font=("Segoe UI", 10, "bold")).pack(anchor="w", pady=(0, 6))
        self.canvas = tk.Canvas(main, width=820, height=360, bg="white", highlightthickness=1, highlightbackground="#BFBFBF")
        self.canvas.pack(fill="both", expand=True)
        self.canvas.bind("<Configure>", lambda _event: self._redraw())
        fit_toplevel_to_content(self.win, 980, 700)

    def close(self):
        if self.closed:
            return
        self.closed = True
        try:
            self.win.destroy()
        except Exception:
            pass

    def update_spectrum(self, point: int, voltage: float, t_int: float, wavelengths: np.ndarray, raw: np.ndarray, normalized: np.ndarray, peaks: List[Dict[str, float]], status: str):
        if self.closed:
            return
        wavelengths_arr = np.asarray(wavelengths, dtype=np.float64)
        normalized_arr = np.asarray(normalized, dtype=np.float64)
        if normalized_arr.size and np.any(np.isfinite(normalized_arr)):
            max_idx = int(np.nanargmax(normalized_arr))
            max_wavelength = float(wavelengths_arr[max_idx])
            max_intensity = float(normalized_arr[max_idx])
        else:
            max_wavelength = 0.0
            max_intensity = 0.0
        self.last = {
            "point": point,
            "voltage": voltage,
            "t_int": t_int,
            "wavelengths": wavelengths_arr,
            "raw": np.asarray(raw, dtype=np.float64),
            "normalized": normalized_arr,
            "peaks": peaks,
            "status": status,
            "max_wavelength": max_wavelength,
            "max_intensity": max_intensity,
        }
        peak_text = ", ".join(f"{p['wavelength_nm']:.1f}" for p in peaks[:5]) or "нет"
        self.status_var.set(
            f"Точка {point}, V={voltage:.3f} В, T_int={t_int*1000:.2f} мс, "
            f"max={max_wavelength:.1f} нм / {max_intensity:.0f}, пики: {peak_text}, {status}"
        )
        self._redraw()
        try:
            self.win.update_idletasks()
            self.win.update()
        except Exception:
            self.closed = True

    def _redraw(self):
        if self.closed:
            return
        c = self.canvas
        c.delete("all")
        width = int(c.winfo_width() or 820)
        height = int(c.winfo_height() or 360)
        left, top, right, bottom = 62, 24, width - 28, height - 38
        c.create_rectangle(left, top, right, bottom, outline="#A0A0A0")
        c.create_text((left + right) / 2, 10, text="Live spectrum: raw and corrected counts/s", font=("Segoe UI", 10, "bold"))
        c.create_text((left + right) / 2, height - 12, text="Wavelength, nm", font=("Segoe UI", 9))
        c.create_text(20, (top + bottom) / 2, text="a.u.", font=("Segoe UI", 9), justify="center")
        c.create_line(right - 210, top + 14, right - 184, top + 14, fill="#999999", width=2)
        c.create_text(right - 178, top + 14, text="raw", anchor="w", font=("Segoe UI", 8), fill="#666666")
        c.create_line(right - 145, top + 14, right - 119, top + 14, fill="#0B61A4", width=2)
        c.create_text(right - 113, top + 14, text="corrected", anchor="w", font=("Segoe UI", 8), fill="#0B61A4")

        if not self.last:
            c.create_text((left + right) / 2, (top + bottom) / 2, text="Спектр появится во время съемки", fill="#666666")
            return

        wl = self.last["wavelengths"]
        raw = self.last["raw"]
        norm = self.last["normalized"]
        if wl.size < 2:
            return
        x_min, x_max = float(np.nanmin(wl)), float(np.nanmax(wl))
        raw_scaled = raw / max(float(np.nanmax(raw)), 1e-9)
        norm_scaled = norm / max(float(np.nanmax(norm)), 1e-9)

        def to_xy(xv, yv):
            x = left + ((float(xv) - x_min) / max(x_max - x_min, 1e-9)) * (right - left)
            y = bottom - float(np.clip(yv, 0.0, 1.0)) * (bottom - top)
            return x, y

        for frac in [0.0, 0.25, 0.5, 0.75, 1.0]:
            y = bottom - frac * (bottom - top)
            c.create_line(left, y, right, y, fill="#EEEEEE")
            x = left + frac * (right - left)
            c.create_line(x, top, x, bottom, fill="#F4F4F4")
            c.create_text(x, bottom + 14, text=f"{x_min + frac*(x_max-x_min):.0f}", anchor="n", font=("Segoe UI", 8))

        raw_coords: List[float] = []
        norm_coords: List[float] = []
        step = max(1, int(wl.size / 700))
        for xw, yr, yn in zip(wl[::step], raw_scaled[::step], norm_scaled[::step]):
            raw_coords.extend(to_xy(xw, yr))
            norm_coords.extend(to_xy(xw, yn))
        if len(raw_coords) >= 4:
            c.create_line(*raw_coords, fill="#999999", width=1)
        if len(norm_coords) >= 4:
            c.create_line(*norm_coords, fill="#0B61A4", width=2)
        max_wavelength = float(self.last.get("max_wavelength") or 0.0)
        max_intensity = float(self.last.get("max_intensity") or 0.0)
        if max_wavelength > 0:
            x, y = to_xy(max_wavelength, max_intensity / max(float(np.nanmax(norm)), 1e-9))
            c.create_line(x, top, x, bottom, fill="#C43C30", dash=(4, 3))
            c.create_oval(x - 4, y - 4, x + 4, y + 4, fill="#C43C30", outline="#C43C30")
            c.create_text(x + 8, max(top + 30, y - 10), text=f"max {max_wavelength:.1f} нм", anchor="w", fill="#C43C30", font=("Segoe UI", 8, "bold"))

        for peak in self.last["peaks"][:8]:
            x, y = to_xy(peak["wavelength_nm"], peak["intensity"] / max(float(np.nanmax(norm)), 1e-9))
            c.create_oval(x - 4, y - 4, x + 4, y + 4, fill="#C43C30", outline="#C43C30")
            c.create_text(x, y - 10, text=f"{peak['wavelength_nm']:.0f}", fill="#C43C30", font=("Segoe UI", 8))


# -----------------------------------------------------------------------------
# GUI
# -----------------------------------------------------------------------------

class OLEDApp(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("OLED Measurement App")
        self._set_initial_window_geometry()
        self.minsize(640, 440)
        self.series: Optional[SeriesManager] = None
        self.log_widget: Optional[ScrolledText] = None
        self._ui_scale = 1.0
        self._base_tk_scaling = 1.0
        self._scale_after_id = None
        self._hardware_probe_running = False
        self._hardware_status_canvas: Optional[tk.Canvas] = None
        self._hardware_status_title: Optional[tk.StringVar] = None
        self._hardware_status_detail: Optional[tk.StringVar] = None
        self._setup_gui_style()
        self.bind("<Configure>", self._schedule_ui_scale_update)
        self.app_settings: Dict[str, Any] = load_app_settings()
        ensure_default_sim_config(Path(self.app_settings.get("simulator_config_path") or SCRIPT_DIR / SIM_CONFIG_FILE))
        self.show_start_screen()

    def save_ui_preference(self, key: str, value: Any) -> None:
        self.app_settings.setdefault("ui", {})[key] = value
        save_app_settings(self.app_settings)

    def measurement_defaults(self, section: str) -> Dict[str, Any]:
        defaults = deepcopy(DEFAULT_APP_SETTINGS.get("measurement_defaults", {}).get(section, {}))
        saved = self.app_settings.get("measurement_defaults", {}).get(section, {})
        return deep_update(defaults, saved) if isinstance(saved, dict) else defaults

    def save_measurement_defaults(self, section: str, values: Dict[str, Any]) -> None:
        self.app_settings.setdefault("measurement_defaults", {})
        current = self.measurement_defaults(section)
        current.update(values)
        self.app_settings["measurement_defaults"][section] = current
        save_app_settings(self.app_settings)

    def _set_initial_window_geometry(self):
        try:
            sw = max(int(self.winfo_screenwidth()), 1120)
            sh = max(int(self.winfo_screenheight()), 760)
            width = int(np.clip(sw * 0.68, 900, 1500))
            height = int(np.clip(sh * 0.68, 620, 920))
            x = max(0, (sw - width) // 2)
            y = max(0, (sh - height) // 2)
            self.geometry(f"{width}x{height}+{x}+{y}")
        except Exception:
            self.geometry("980x680")

    def _setup_gui_style(self):
        """Настройка размеров виджетов под Windows/DPI, чтобы строки Treeview не обрезались."""
        try:
            self._style = ttk.Style(self)
            self._font_base_sizes = {}
            self._base_tk_scaling = float(self.tk.call("tk", "scaling"))
            for name in ["TkDefaultFont", "TkTextFont", "TkHeadingFont", "TkMenuFont", "TkFixedFont"]:
                font = tkfont.nametofont(name)
                self._font_base_sizes[name] = abs(int(font.cget("size") or 10))
            self._ui_scale = 0.0
            self._apply_ui_scale()
        except Exception:
            pass

    def _schedule_ui_scale_update(self, event=None):
        if event is not None and event.widget is not self:
            return
        try:
            if self._scale_after_id is not None:
                self.after_cancel(self._scale_after_id)
            self._scale_after_id = self.after(120, self._apply_ui_scale)
        except Exception:
            pass

    def _apply_ui_scale(self):
        try:
            width = max(int(self.winfo_width()), 1)
            height = max(int(self.winfo_height()), 1)
            scale = float(np.clip(min(width / 1120.0, height / 760.0), 0.78, 1.03))
            if abs(scale - getattr(self, "_ui_scale", 1.0)) < 0.025:
                return
            self._ui_scale = scale
            self.tk.call("tk", "scaling", max(1.0, self._base_tk_scaling))
            for name, base_size in getattr(self, "_font_base_sizes", {}).items():
                tkfont.nametofont(name).configure(size=max(8, int(round(base_size * scale))))
            default_font = tkfont.nametofont("TkDefaultFont")
            rowheight = max(20, int(default_font.metrics("linespace") * 1.32))
            style = getattr(self, "_style", ttk.Style(self))
            style.configure("Treeview", rowheight=rowheight)
            style.configure("Treeview.Heading", padding=(max(2, int(3 * scale)), max(2, int(4 * scale))))
            self.option_add("*TCombobox*Listbox.font", default_font)
        except Exception:
            pass

    def _tree_with_scrollbars(self, parent, columns, height: int = 14):
        """Создает таблицу Treeview с вертикальной и горизонтальной прокруткой."""
        container = ttk.Frame(parent)
        tree = ttk.Treeview(container, columns=columns, show="headings", height=height, selectmode="browse")
        yscroll = ttk.Scrollbar(container, orient="vertical", command=tree.yview)
        xscroll = ttk.Scrollbar(container, orient="horizontal", command=tree.xview)
        tree.configure(yscrollcommand=yscroll.set, xscrollcommand=xscroll.set)
        tree.grid(row=0, column=0, sticky="nsew")
        yscroll.grid(row=0, column=1, sticky="ns")
        xscroll.grid(row=1, column=0, sticky="ew")
        container.rowconfigure(0, weight=1)
        container.columnconfigure(0, weight=1)
        return container, tree

    def _scrollable_frame(self, padding: int = 16) -> ttk.Frame:
        outer = ttk.Frame(self)
        outer.pack(fill="both", expand=True)
        return self._scrollable_child(outer, padding=padding)

    def _scrollable_child(self, parent, padding: int = 16) -> ttk.Frame:
        canvas = tk.Canvas(parent, highlightthickness=0)
        yscroll = ttk.Scrollbar(parent, orient="vertical", command=canvas.yview)
        xscroll = ttk.Scrollbar(parent, orient="horizontal", command=canvas.xview)
        canvas.configure(yscrollcommand=yscroll.set, xscrollcommand=xscroll.set)
        canvas.grid(row=0, column=0, sticky="nsew")
        yscroll.grid(row=0, column=1, sticky="ns")
        xscroll.grid(row=1, column=0, sticky="ew")
        parent.rowconfigure(0, weight=1)
        parent.columnconfigure(0, weight=1)

        frame = ttk.Frame(canvas, padding=padding)
        window_id = canvas.create_window((0, 0), window=frame, anchor="nw")

        def update_scroll_region(_event=None):
            canvas.configure(scrollregion=canvas.bbox("all"))

        def update_inner_width(event):
            requested = max(event.width, frame.winfo_reqwidth())
            canvas.itemconfigure(window_id, width=requested)

        def on_mousewheel(event):
            delta = -1 * int(event.delta / 120) if event.delta else 0
            canvas.yview_scroll(delta, "units")

        frame.bind("<Configure>", update_scroll_region)
        canvas.bind("<Configure>", update_inner_width)
        canvas.bind("<Enter>", lambda _event: canvas.bind_all("<MouseWheel>", on_mousewheel))
        canvas.bind("<Leave>", lambda _event: canvas.unbind_all("<MouseWheel>"))
        return frame

    def _scrollable_notebook_tab(self, notebook, title: str, padding: int = 12) -> ttk.Frame:
        outer = ttk.Frame(notebook)
        frame = self._scrollable_child(outer, padding=padding)
        notebook.add(outer, text=title)
        return frame

    def clear(self):
        for child in self.winfo_children():
            child.destroy()

    def log(self, text: str):
        print(text)
        if self.log_widget is not None:
            self.log_widget.configure(state="normal")
            self.log_widget.insert("end", str(text) + "\n")
            self.log_widget.see("end")
            self.log_widget.configure(state="disabled")
            self.update_idletasks()
            self.update()

    def _build_hardware_status_bar(self, parent):
        bar = ttk.Frame(parent)
        bar.pack(fill="x", pady=(6, 10))
        self._hardware_status_canvas = tk.Canvas(bar, width=18, height=18, highlightthickness=0)
        self._hardware_status_canvas.pack(side="left", padx=(0, 8))
        text_frame = ttk.Frame(bar)
        text_frame.pack(side="left", fill="x", expand=True)
        self._hardware_status_title = tk.StringVar(value="Оборудование: не проверено")
        self._hardware_status_detail = tk.StringVar(value="Нажмите 'Проверить оборудование', чтобы опросить SMU и спектрометр.")
        ttk.Label(text_frame, textvariable=self._hardware_status_title, font=("Segoe UI", 10, "bold")).pack(anchor="w")
        ttk.Label(text_frame, textvariable=self._hardware_status_detail, foreground="#555555", wraplength=900).pack(anchor="w")
        ttk.Button(bar, text="Проверить оборудование", command=self.check_hardware_status, width=24).pack(side="right", padx=(10, 0))
        self._set_hardware_status_indicator("unknown")

    def _set_hardware_status_indicator(self, level: str):
        canvas = self._hardware_status_canvas
        if canvas is None:
            return
        color = {
            "ok": "#2FA66A",
            "warning": "#D8A239",
            "error": "#C43C30",
            "checking": "#2F80ED",
            "unknown": "#A7B0B5",
        }.get(level, "#A7B0B5")
        canvas.delete("all")
        canvas.create_oval(2, 2, 16, 16, fill=color, outline=color)

    def check_hardware_status(self):
        if self._hardware_probe_running:
            return
        self.app_settings = load_app_settings()
        self._hardware_probe_running = True
        self._set_hardware_status_indicator("checking")
        if self._hardware_status_title is not None:
            self._hardware_status_title.set("Оборудование: идет проверка")
        if self._hardware_status_detail is not None:
            mode = hardware_mode_label(self.app_settings)
            self._hardware_status_detail.set(f"{mode}, COM: {self.app_settings.get('com_port', 'COM3')}")

        settings_snapshot = deepcopy(self.app_settings)

        def worker():
            try:
                result = probe_hardware(settings_snapshot)
            except Exception as exc:
                result = {
                    "level": "error",
                    "title": "Ошибка проверки оборудования",
                    "details": str(exc),
                    "smu": "",
                    "spectrometer": "",
                }
            self.after(0, lambda: self._finish_hardware_probe(result))

        threading.Thread(target=worker, daemon=True).start()

    def _finish_hardware_probe(self, result: Dict[str, Any]):
        self._hardware_probe_running = False
        level = str(result.get("level") or "unknown")
        self._set_hardware_status_indicator(level)
        if self._hardware_status_title is not None:
            self._hardware_status_title.set(f"Оборудование: {result.get('title', 'неизвестно')}")
        if self._hardware_status_detail is not None:
            detail = str(result.get("details") or "")
            smu = str(result.get("smu") or "")
            spectrometer = str(result.get("spectrometer") or "")
            parts = [part for part in [smu, spectrometer, detail] if part]
            self._hardware_status_detail.set(" | ".join(parts) if parts else "Нет деталей проверки.")
        try:
            self.log(f"Проверка оборудования: {result.get('title')}; {result.get('details')}")
        except Exception:
            pass

    def show_start_screen(self):
        self.clear()
        self.app_settings = load_app_settings()

        frame = self._scrollable_frame(padding=22)
        header = ttk.Frame(frame)
        header.pack(fill="x")
        ttk.Label(header, text="OLED Measurement App", font=("Segoe UI", 22, "bold")).pack(side="left")
        ttk.Button(header, text="Настройки", command=self.open_settings_window, width=16).pack(side="right")

        mode_text = f"Режим: {hardware_mode_label(self.app_settings)}"
        if self.app_settings.get("hardware_mode") == HARDWARE_MODE_SIM:
            mode_text += "  |  измерения идут на эмуляторе"
        ttk.Label(frame, text=mode_text, font=("Segoe UI", 11, "bold")).pack(anchor="w", pady=(4, 2))
        self._build_hardware_status_bar(frame)
        self.after(250, self.check_hardware_status)
        ttk.Label(frame, text="Выберите существующую серию или создайте новую.", font=("Segoe UI", 11)).pack(anchor="w", pady=(0, 14))

        root_bar = ttk.Frame(frame)
        root_bar.pack(fill="x", pady=(0, 10))
        ttk.Label(root_bar, text="Корневая папка серий:").grid(row=0, column=0, sticky="w", padx=(0, 8))
        root_var = tk.StringVar(value=str(self.app_settings.get("default_root") or SCRIPT_DIR / DEFAULT_ROOT))
        root_entry = ttk.Entry(root_bar, textvariable=root_var, width=78)
        root_entry.grid(row=0, column=1, sticky="we")
        root_bar.columnconfigure(1, weight=1)
        ttk.Button(root_bar, text="Обзор", command=lambda: self._browse_root_and_refresh(root_var)).grid(row=0, column=2, padx=(8, 0))
        ttk.Button(root_bar, text="Обновить", command=lambda: self.refresh_series_list(root_var.get())).grid(row=0, column=3, padx=(8, 0))

        series_table, self.series_tree = self._tree_with_scrollbars(
            frame,
            columns=("deposition", "keyword", "created", "measurements", "folder"),
            height=12,
        )
        self.series_tree.heading("deposition", text="Дата напыления")
        self.series_tree.heading("keyword", text="Кодовое слово")
        self.series_tree.heading("created", text="Создана")
        self.series_tree.heading("measurements", text="Измерений")
        self.series_tree.heading("folder", text="Папка")
        self.series_tree.column("deposition", width=150, minwidth=130, stretch=False)
        self.series_tree.column("keyword", width=190, minwidth=140, stretch=False)
        self.series_tree.column("created", width=210, minwidth=170, stretch=False)
        self.series_tree.column("measurements", width=110, minwidth=90, anchor="center", stretch=False)
        self.series_tree.column("folder", width=760, minwidth=420, stretch=True)
        series_table.pack(fill="both", expand=True)
        self.series_tree.bind("<Double-1>", lambda _event: self.open_selected_series())

        buttons = ttk.Frame(frame)
        buttons.pack(fill="x", pady=(12, 0))
        ttk.Button(buttons, text="Открыть выбранную серию", command=self.open_selected_series, width=28).pack(side="left")
        ttk.Button(buttons, text="Открыть папку вручную", command=self.open_existing_series, width=24).pack(side="left", padx=(10, 0))
        ttk.Button(buttons, text="Создать новую серию", command=self.show_new_series_screen, width=24).pack(side="right")

        self.log_widget = ScrolledText(frame, height=8, state="disabled")
        self.log_widget.pack(fill="x", pady=(14, 0))
        self.refresh_series_list(root_var.get())

    def _browse_root_and_refresh(self, root_var: tk.StringVar):
        folder = filedialog.askdirectory(title="Корневая папка для серий")
        if not folder:
            return
        root_var.set(folder)
        self.app_settings["default_root"] = folder
        save_app_settings(self.app_settings)
        self.refresh_series_list(folder)

    def find_existing_series(self, root_folder: Path) -> List[Dict[str, Any]]:
        root = Path(root_folder)
        if not root.exists():
            return []
        found: List[Dict[str, Any]] = []
        for cfg_path in sorted(root.rglob(CONFIG_FILE)):
            folder = cfg_path.parent
            try:
                cfg = json.loads(cfg_path.read_text(encoding="utf-8"))
            except Exception:
                cfg = {}
            measurements_count = ""
            journal_path = folder / JOURNAL_FILE
            if journal_path.exists():
                try:
                    wb = load_workbook(journal_path, data_only=True, read_only=True)
                    if MEASUREMENTS_SHEET in wb.sheetnames:
                        measurements_count = max(wb[MEASUREMENTS_SHEET].max_row - 1, 0)
                    wb.close()
                except Exception:
                    measurements_count = "?"
            found.append({
                "folder": folder,
                "deposition_date": cfg.get("deposition_date", ""),
                "keyword": cfg.get("keyword", ""),
                "created_at": cfg.get("created_at", ""),
                "measurements_count": measurements_count,
            })
        return found

    def refresh_series_list(self, root_folder: str):
        if not hasattr(self, "series_tree"):
            return
        self.app_settings["default_root"] = str(root_folder)
        save_app_settings(self.app_settings)
        for item in self.series_tree.get_children():
            self.series_tree.delete(item)
        series_list = self.find_existing_series(Path(root_folder))
        for item in series_list:
            folder = item["folder"]
            self.series_tree.insert(
                "",
                "end",
                iid=str(folder.resolve()),
                values=(
                    item.get("deposition_date", "") or "",
                    item.get("keyword", "") or "",
                    item.get("created_at", "") or "",
                    item.get("measurements_count", "") if item.get("measurements_count", "") is not None else "",
                    str(folder),
                ),
            )
        self.log(f"Найдено серий: {len(series_list)}. Режим оборудования: {hardware_mode_label(self.app_settings)}.")

    def open_selected_series(self):
        if not hasattr(self, "series_tree"):
            return
        selection = self.series_tree.selection()
        if not selection:
            messagebox.showwarning("Серия", "Выберите серию в списке.")
            return
        folder = Path(selection[0])
        try:
            self.series = SeriesManager(folder)
            self.show_measurement_menu()
        except Exception as exc:
            messagebox.showerror("Не удалось открыть серию", str(exc))

    def open_existing_series(self):
        folder = filedialog.askdirectory(title="Выберите папку серии, где лежит series_config.json")
        if not folder:
            return
        try:
            self.series = SeriesManager(Path(folder))
            self.show_measurement_menu()
        except Exception as exc:
            messagebox.showerror("Не удалось открыть серию", str(exc))

    def _add_settings_entry(self, parent, row: int, label: str, var: tk.StringVar, width: int = 18):
        ttk.Label(parent, text=label + ":").grid(row=row, column=0, sticky="e", pady=3, padx=(0, 8))
        ttk.Entry(parent, textvariable=var, width=width).grid(row=row, column=1, sticky="w", pady=3)

    def open_settings_window(self):
        self.app_settings = load_app_settings()
        win = tk.Toplevel(self)
        win.title("Настройки приложения")
        win.geometry("720x640")
        win.transient(self)

        main = ttk.Frame(win, padding=12)
        main.pack(fill="both", expand=True)
        nb = ttk.Notebook(main)
        nb.pack(fill="both", expand=True)

        general = self._scrollable_notebook_tab(nb, "Общие")
        sim_tab = self._scrollable_notebook_tab(nb, "Эмулятор")
        ivl_tab = self._scrollable_notebook_tab(nb, "ВАЯХ доп.")
        spec_tab = self._scrollable_notebook_tab(nb, "Спектры доп.")
        stab_tab = self._scrollable_notebook_tab(nb, "Стабильность доп.")

        root_var = tk.StringVar(value=str(self.app_settings.get("default_root", "")))
        mode_var = tk.StringVar(value=str(self.app_settings.get("hardware_mode", HARDWARE_MODE_SIM)))
        com_var = tk.StringVar(value=str(self.app_settings.get("com_port", "COM3")))
        units = self.app_settings.get("measurement_units", DEFAULT_APP_SETTINGS["measurement_units"])
        pixel_area_var = tk.StringVar(value=str(units.get("pixel_area_mm2", 1.0)))
        luminance_coeff_var = tk.StringVar(value=str(units.get("luminance_cd_m2_per_uA", 1.0)))

        ttk.Label(general, text="Корневая папка серий:").grid(row=0, column=0, sticky="e", pady=4, padx=(0, 8))
        ttk.Entry(general, textvariable=root_var, width=62).grid(row=0, column=1, sticky="we", pady=4)
        ttk.Button(general, text="Обзор", command=lambda: self._browse_root(root_var)).grid(row=0, column=2, padx=(8, 0))
        ttk.Label(general, text="Режим оборудования:").grid(row=1, column=0, sticky="e", pady=4, padx=(0, 8))
        ttk.Combobox(
            general,
            textvariable=mode_var,
            values=[HARDWARE_MODE_SIM, HARDWARE_MODE_REAL],
            state="readonly",
            width=18,
        ).grid(row=1, column=1, sticky="w", pady=4)
        self._add_settings_entry(general, 2, "COM port по умолчанию", com_var)
        self._add_settings_entry(general, 3, "Площадь пикселя, мм^2", pixel_area_var)
        self._add_settings_entry(general, 4, "Коэфф. мкА → кд/м^2", luminance_coeff_var)
        ttk.Label(
            general,
            text="simulator = встроенная эмуляция пикселя; real = настоящие xtralien/seabreeze из Python-среды.",
            foreground="#555555",
            wraplength=610,
            justify="left",
        ).grid(row=5, column=0, columnspan=3, sticky="w", pady=(12, 0))
        general.columnconfigure(1, weight=1)

        sim_cfg_var = tk.StringVar(value=str(self.app_settings.get("simulator_config_path") or SCRIPT_DIR / SIM_CONFIG_FILE))
        ttk.Label(sim_tab, text="JSON-конфиг пикселей:").grid(row=0, column=0, sticky="e", pady=4, padx=(0, 8))
        ttk.Entry(sim_tab, textvariable=sim_cfg_var, width=62).grid(row=0, column=1, sticky="we", pady=4)
        ttk.Button(sim_tab, text="Обзор", command=lambda: self._browse_file_for_var(sim_cfg_var)).grid(row=0, column=2, padx=(8, 0))
        ttk.Button(sim_tab, text="Создать/обновить пример JSON", command=lambda: self._write_default_sim_config_from_settings(sim_cfg_var)).grid(row=1, column=1, sticky="w", pady=(8, 4))
        ttk.Label(
            sim_tab,
            text="В этом JSON задаются режимы пикселя: working, weak, nonworking, no_contact, burned/short; напряжение открытия, ток, фототок, спектральные пики, деградация.",
            foreground="#555555",
            wraplength=620,
            justify="left",
        ).grid(row=2, column=0, columnspan=3, sticky="w", pady=(12, 0))
        sim_tab.columnconfigure(1, weight=1)

        def make_vars(section: str) -> Dict[str, tk.StringVar]:
            return {k: tk.StringVar(value=str(v)) for k, v in self.app_settings.get(section, {}).items() if not isinstance(v, bool)}

        ivl_vars = make_vars("ivl_advanced")
        ivl_bool_vars: Dict[str, tk.BooleanVar] = {}
        ivl_labels = [
            ("photodiode_bias_V", "Смещение фотодиода, В"),
            ("photodiode_range", "Диапазон фотодиода"),
            ("photodiode_threshold_uA", "Порог фототока, мкА"),
            ("burnout_current_threshold_mA", "Ток пробоя/сгорания, мА"),
            ("no_contact_max_led_current_mA", "Макс. ток при отсутствии контакта, мА"),
            ("burned_confirmation_cycles", "Доп. циклов после BURNED"),
        ]
        for row, (key, label) in enumerate(ivl_labels):
            self._add_settings_entry(ivl_tab, row, label, ivl_vars[key])
        ttk.Label(ivl_tab, text="BURNED ставится только при достижении тока пробоя/сгорания.", foreground="#555555").grid(row=len(ivl_labels), column=0, columnspan=2, sticky="w", pady=(8, 0))
        ttk.Label(ivl_tab, text="Эти параметры убраны из основного окна ВАЯХ, чтобы оно не было перегружено.", foreground="#555555").grid(row=len(ivl_labels)+1, column=0, columnspan=2, sticky="w", pady=(12, 0))

        spec_vars = make_vars("spectrum_advanced")
        spec_bool_vars = {
            "discard_first_scan_after_tint_change": tk.BooleanVar(value=bool(self.app_settings.get("spectrum_advanced", {}).get("discard_first_scan_after_tint_change", True))),
            "dark_spectrum_enabled": tk.BooleanVar(value=bool(self.app_settings.get("spectrum_advanced", {}).get("dark_spectrum_enabled", False))),
            "baseline_correction_enabled": tk.BooleanVar(value=bool(self.app_settings.get("spectrum_advanced", {}).get("baseline_correction_enabled", True))),
            "peak_detection_enabled": tk.BooleanVar(value=bool(self.app_settings.get("spectrum_advanced", {}).get("peak_detection_enabled", False))),
        }
        spec_labels = [
            ("photodiode_bias_V", "Смещение фотодиода, В"),
            ("photodiode_range", "Диапазон фотодиода"),
            ("target_intensity", "Целевая интенсивность, counts"),
            ("intensity_min", "Мин. интенсивность, counts"),
            ("intensity_max", "Макс. интенсивность, counts"),
            ("saturation_level", "Насыщение, counts"),
            ("min_peak_width_nm", "Мин. FWHM, нм"),
            ("max_peak_width_nm", "Макс. FWHM, нм"),
            ("t_int_initial_s", "Начальное T_int, с"),
            ("t_int_min_s", "Мин. T_int, с"),
            ("t_int_max_s", "Макс. T_int, с"),
            ("kp", "Kp подбора T_int"),
            ("ki", "Ki подбора T_int"),
            ("max_iterations", "Макс. итераций"),
            ("tolerance", "Допуск подбора"),
            ("peak_search_mode_for_tint", "Область поиска пика"),
            ("settle_time_voltage_s", "Пауза после напряжения, с"),
            ("settle_time_spectrum_s", "Пауза спектрометра, с"),
            ("dark_spectrum_scans", "Число dark-сканов"),
        ]
        for row, (key, label) in enumerate(spec_labels):
            self._add_settings_entry(spec_tab, row, label, spec_vars[key])
        ttk.Checkbutton(spec_tab, text="Сбрасывать первый спектр после смены T_int", variable=spec_bool_vars["discard_first_scan_after_tint_change"]).grid(row=len(spec_labels), column=0, columnspan=2, sticky="w", pady=(8, 0))
        ttk.Checkbutton(spec_tab, text="Снимать dark spectrum", variable=spec_bool_vars["dark_spectrum_enabled"]).grid(row=len(spec_labels)+1, column=0, columnspan=2, sticky="w", pady=(4, 0))
        ttk.Checkbutton(spec_tab, text="Вычитать средний фон из raw-спектра", variable=spec_bool_vars["baseline_correction_enabled"]).grid(row=len(spec_labels)+2, column=0, columnspan=2, sticky="w", pady=(4, 0))
        ttk.Checkbutton(spec_tab, text="Искать пики производными", variable=spec_bool_vars["peak_detection_enabled"]).grid(row=len(spec_labels)+3, column=0, columnspan=2, sticky="w", pady=(4, 0))

        stab_vars = make_vars("stability_advanced")
        stab_labels = [
            ("voltage_step_max", "Макс. шаг напряжения, В"),
            ("current_control_kp", "Kp удержания тока, В/мА"),
            ("photodiode_bias_V", "Смещение фотодиода, В"),
            ("photodiode_threshold_uA", "Порог фототока, мкА"),
            ("photodiode_range", "Диапазон фотодиода"),
        ]
        for row, (key, label) in enumerate(stab_labels):
            self._add_settings_entry(stab_tab, row, label, stab_vars[key])

        def cast_like(default_value, raw: str):
            if isinstance(default_value, int) and not isinstance(default_value, bool):
                return parse_int(raw, "настройка")
            if isinstance(default_value, float):
                return parse_float(raw, "настройка")
            return str(raw)

        def collect_section(section: str, vars_dict: Dict[str, tk.StringVar], bool_vars: Dict[str, tk.BooleanVar]) -> Dict[str, Any]:
            defaults = DEFAULT_APP_SETTINGS[section]
            out: Dict[str, Any] = {}
            for key, var in vars_dict.items():
                out[key] = cast_like(defaults.get(key, ""), var.get())
            for key, var in bool_vars.items():
                out[key] = bool(var.get())
            return out

        def save():
            try:
                settings = load_app_settings()
                settings["default_root"] = root_var.get().strip() or str(SCRIPT_DIR / DEFAULT_ROOT)
                settings["hardware_mode"] = mode_var.get().strip() or HARDWARE_MODE_REAL
                settings["com_port"] = com_var.get().strip() or "COM3"
                settings["measurement_units"] = {
                    "pixel_area_mm2": parse_float(pixel_area_var.get(), "Площадь пикселя"),
                    "luminance_cd_m2_per_uA": parse_float(luminance_coeff_var.get(), "Коэффициент яркости"),
                }
                settings["simulator_config_path"] = sim_cfg_var.get().strip() or str(SCRIPT_DIR / SIM_CONFIG_FILE)
                settings["ivl_advanced"] = collect_section("ivl_advanced", ivl_vars, ivl_bool_vars)
                settings["spectrum_advanced"] = collect_section("spectrum_advanced", spec_vars, spec_bool_vars)
                settings["stability_advanced"] = collect_section("stability_advanced", stab_vars, {})
                save_app_settings(settings)
                self.app_settings = settings
                if settings["hardware_mode"] == HARDWARE_MODE_SIM:
                    ensure_default_sim_config(Path(settings["simulator_config_path"]))
                messagebox.showinfo("Настройки", "Настройки сохранены.", parent=win)
                win.destroy()
                self.show_start_screen()
            except Exception as exc:
                messagebox.showerror("Ошибка настроек", str(exc), parent=win)

        bottom = ttk.Frame(main)
        bottom.pack(fill="x", pady=(12, 0))
        ttk.Button(bottom, text="Отмена", command=win.destroy).pack(side="left")
        ttk.Button(bottom, text="Сохранить", command=save).pack(side="right")
        fit_toplevel_to_content(win, 860, 760)

    def _browse_file_for_var(self, var: tk.StringVar):
        filename = filedialog.askopenfilename(title="Выберите JSON-конфиг", filetypes=[("JSON", "*.json"), ("All files", "*.*")])
        if filename:
            var.set(filename)

    def _write_default_sim_config_from_settings(self, var: tk.StringVar):
        try:
            path = ensure_default_sim_config(Path(var.get().strip() or SCRIPT_DIR / SIM_CONFIG_FILE))
            var.set(str(path))
            messagebox.showinfo("Эмулятор", f"Пример конфига создан/найден:\n{path}")
        except Exception as exc:
            messagebox.showerror("Эмулятор", str(exc))

    def show_new_series_screen(self):
        self.clear()
        main = self._scrollable_frame(padding=16)
        ttk.Label(main, text="Новая серия напыления", font=("Segoe UI", 18, "bold")).pack(anchor="w")

        top = ttk.Frame(main)
        top.pack(fill="x", pady=(12, 8))
        ttk.Label(top, text="Корневая папка:").grid(row=0, column=0, sticky="w", padx=(0, 6), pady=4)
        root_var = tk.StringVar(value=str(self.app_settings.get("default_root") or SCRIPT_DIR / DEFAULT_ROOT))
        ttk.Entry(top, textvariable=root_var, width=75).grid(row=0, column=1, sticky="we", pady=4)
        top.columnconfigure(1, weight=1)
        ttk.Button(top, text="Обзор", command=lambda: self._browse_root(root_var)).grid(row=0, column=2, padx=(6, 0), pady=4)

        ttk.Label(top, text="Дата напыления:").grid(row=1, column=0, sticky="w", padx=(0, 6), pady=4)
        dep_date_var = tk.StringVar(value=today_iso())
        ttk.Entry(top, textvariable=dep_date_var, width=20).grid(row=1, column=1, sticky="w", pady=4)

        ttk.Label(top, text="Кодовое слово, опционально:").grid(row=2, column=0, sticky="w", padx=(0, 6), pady=4)
        keyword_var = tk.StringVar()
        ttk.Entry(top, textvariable=keyword_var, width=30).grid(row=2, column=1, sticky="w", pady=4)

        quarter_vars = {str(q): tk.StringVar(value="Q") for q in range(1, 5)}

        holder_frame = ttk.LabelFrame(main, text="Схема подложкодержателя и названия четвертей")
        holder_frame.pack(fill="both", expand=True, pady=(10, 4))
        self._draw_series_creation_holder(holder_frame, quarter_vars)

        ttk.Label(
            main,
            text="Названия четвертей вводятся прямо рядом с соответствующей четвертью на схеме. "
                 "Например, для четверти 1: CR → пиксели CR1_1_1 ... CR1_3_4.",
            foreground="#555555",
            wraplength=1080,
            justify="left",
        ).pack(anchor="w", pady=(6, 0))

        bottom = ttk.Frame(main)
        bottom.pack(fill="x", pady=(10, 0))
        ttk.Button(bottom, text="Назад", command=self.show_start_screen).pack(side="left")
        ttk.Button(
            bottom,
            text="Создать серию и журнал",
            command=lambda: self.create_series(root_var, dep_date_var, keyword_var, quarter_vars),
        ).pack(side="right")

        self.log_widget = ScrolledText(main, height=8, state="disabled")
        self.log_widget.pack(fill="x", pady=(14, 0))

    def _browse_root(self, var: tk.StringVar):
        folder = filedialog.askdirectory(title="Корневая папка для серий")
        if folder:
            var.set(folder)

    def _draw_holder_base(self, canvas: tk.Canvas, width: int, height: int, title: str = ""):
        canvas.delete("all")
        if title:
            canvas.create_text(width / 2, 22, text=title, fill="#17345F", font=("Segoe UI", 10, "bold"))

    def _draw_series_creation_holder(self, parent, quarter_vars: Dict[str, tk.StringVar]):
        width, height = 930, 620
        canvas = tk.Canvas(parent, width=width, height=height, background="white", highlightthickness=0)
        canvas.pack(fill="both", expand=True, padx=8, pady=8)
        layout = build_holder_layout(width, height)
        self._draw_holder_base(canvas, width, height, "")

        for q in [2, 1, 3, 4]:
            info = layout[q]
            canvas.create_text(*info["number_xy"], text=str(q), font=("Segoe UI", 24, "bold"), fill="#17345F")
            entry = ttk.Entry(canvas, textvariable=quarter_vars[str(q)], width=9)
            canvas.create_window(*info["entry_xy"], anchor="w", window=entry)
            quarter_vars[str(q)].trace_add(
                "write",
                lambda *_args, cv=canvas, qv=quarter_vars: self._update_series_canvas_labels(cv, qv),
            )

            for substrate in info["substrates"]:
                x, y, w, h = substrate["x"], substrate["y"], substrate["w"], substrate["h"]
                canvas.create_rectangle(x, y, x + w, y + h, fill="#FFFFFF", outline="#17345F", width=2)
                for pix in range(1, 5):
                    px, py, pw, ph = self._pixel_rect_inside_substrate(x, y, w, h, pix)
                    canvas.create_rectangle(px, py, px + pw, py + ph, fill="#F7F7F7", outline="#A0A0A0")
                    canvas.create_text(px + pw / 2, py + ph / 2, text=str(pix), font=("Segoe UI", 7))

        self._update_series_canvas_labels(canvas, quarter_vars)

    def _update_series_canvas_labels(self, canvas: tk.Canvas, quarter_vars: Dict[str, tk.StringVar]):
        canvas.delete("quarter_dynamic")
        width, height = int(canvas.cget("width")), int(canvas.cget("height"))
        layout = build_holder_layout(width, height)
        for q in [2, 1, 3, 4]:
            code = safe_filename(quarter_vars[str(q)].get().strip() or f"Q{q}", fallback=f"Q{q}")
            info = layout[q]
            # Короткий итог рядом с полем ввода.
            label_x = info["entry_xy"][0] + 80
            canvas.create_text(
                label_x,
                info["entry_xy"][1] + 1,
                text=f"→ {code}{q}",
                anchor="w",
                tags="quarter_dynamic",
                font=("Segoe UI", 8, "bold"),
                fill="#0B61A4",
            )
            for substrate in info["substrates"]:
                x, y, w, _h = substrate["x"], substrate["y"], substrate["w"], substrate["h"]
                sid = f"{code}{q}_{substrate['substrate_number']}"
                canvas.create_text(
                    x + w / 2,
                    y - 8,
                    text=sid,
                    tags="quarter_dynamic",
                    font=("Segoe UI", 8, "bold"),
                    fill="#17345F",
                )

    def _pixel_rect_inside_substrate(self, x: float, y: float, w: float, h: float, pixel_number: int) -> Tuple[float, float, float, float]:
        pad_x = 10
        pad_top = 13
        pad_bottom = 8
        gap = 5
        inner_w = (w - 2 * pad_x - gap) / 2
        inner_h = (h - pad_top - pad_bottom - gap) / 2
        row = 0 if pixel_number in {1, 2} else 1
        col = 0 if pixel_number in {1, 3} else 1
        px = x + pad_x + col * (inner_w + gap)
        py = y + pad_top + row * (inner_h + gap)
        return px, py, inner_w, inner_h

    def _create_status_holder_canvas(self, parent):
        width, height = 900, 540
        canvas = tk.Canvas(parent, width=width, height=height, background="white", highlightthickness=0)
        canvas.pack(fill="x", expand=False, padx=8, pady=8)
        self.status_canvas = canvas
        self.status_canvas_layout = build_holder_layout(width, height)
        return canvas

    def _render_status_holder_canvas(self):
        if self.series is None or not hasattr(self, "status_canvas"):
            return
        canvas = self.status_canvas
        width = int(canvas.cget("width"))
        height = int(canvas.cget("height"))
        self._draw_holder_base(canvas, width, height, "")
        rows = {row.get("Pixel ID"): row for row in self.series.journal.list_pixels()}
        quarter_names = self.series.config.get("quarter_names", {})
        deposition_date = short_date_for_map(str(self.series.config.get("deposition_date", "") or ""))
        layout = getattr(self, "status_canvas_layout", build_holder_layout(width, height))

        for q in [2, 1, 3, 4]:
            code = safe_filename(quarter_names.get(str(q), f"Q{q}"), fallback=f"Q{q}")
            info = layout[q]
            canvas.create_text(*info["number_xy"], text=str(q), font=("Segoe UI", 24, "bold"), fill="#17345F")
            for substrate in info["substrates"]:
                x, y, w, h = substrate["x"], substrate["y"], substrate["w"], substrate["h"]
                sid = f"{code}{q}_{substrate['substrate_number']}"
                canvas.create_text(
                    x + w / 2,
                    y - 10,
                    text=deposition_date,
                    font=("Segoe UI", 7),
                    fill="#17345F",
                )
                canvas.create_rectangle(x, y, x + w, y + h, fill="#FFFFFF", outline="#17345F", width=2)
                for pix in range(1, 5):
                    pixel_id = f"{sid}_{pix}"
                    status = (rows.get(pixel_id, {}) or {}).get("Last status", "UNKNOWN")
                    color = pixel_status_color(status)
                    px, py, pw, ph = self._pixel_rect_inside_substrate(x, y, w, h, pix)
                    canvas.create_rectangle(px, py, px + pw, py + ph, fill=color, outline="#808080")
                    canvas.create_text(px + pw / 2, py + ph / 2, text=str(pix), font=("Segoe UI", 7))
                canvas.create_text(
                    x + w / 2,
                    y + h + 11,
                    text=sid,
                    font=("Segoe UI", 8, "bold"),
                    fill="#17345F",
                )

        legend_y = height - 25
        legend = [
            ("#8FD694", "рабочий"),
            ("#F2D96B", "нет контакта"),
            ("#F28B82", "нераб./пробой"),
            ("#D9D9D9", "не измерен"),
        ]
        lx = 135
        for color, label in legend:
            canvas.create_rectangle(lx, legend_y - 7, lx + 16, legend_y + 7, fill=color, outline="#808080")
            canvas.create_text(lx + 22, legend_y, text=label, anchor="w", font=("Segoe UI", 8))
            lx += 175

    def _create_ivl_history_tree(self, parent):
        table_frame, tree = self._tree_with_scrollbars(parent, columns=("pixel",), height=8)
        table_frame.pack(fill="both", expand=True, padx=8, pady=8)
        self.ivl_history_tree = tree
        return tree

    def _refresh_ivl_history_tree(self):
        if self.series is None or not hasattr(self, "ivl_history_tree"):
            return

        tree = self.ivl_history_tree
        for item in tree.get_children():
            tree.delete(item)

        pixels = [str(row.get("Pixel ID") or "") for row in self.series.journal.list_pixels()]
        measurements = [
            row for row in self.series.journal.list_measurements()
            if str(row.get("Type") or "").upper() == "IVL"
        ]
        dates = sorted({str(row.get("Measurement day") or "") for row in measurements if row.get("Measurement day")})
        date_columns = [f"d{i}" for i in range(len(dates))]
        columns = ["pixel"] + date_columns
        tree.configure(columns=columns)

        tree.heading("pixel", text="Pixel")
        tree.column("pixel", width=145, minwidth=120, stretch=False)
        for col, date_text in zip(date_columns, dates):
            tree.heading(col, text=date_text)
            tree.column(col, width=130, minwidth=105, anchor="center", stretch=False)

        latest_by_pixel_date: Dict[Tuple[str, str], str] = {}
        for row in measurements:
            pixel_id = str(row.get("Pixel ID") or "")
            day = str(row.get("Measurement day") or "")
            status = str(row.get("Status") or "")
            if pixel_id and day:
                latest_by_pixel_date[(pixel_id, day)] = status

        for pixel_id in pixels:
            values = [pixel_id]
            for day in dates:
                values.append(ivl_status_marker(latest_by_pixel_date.get((pixel_id, day), "")))
            tree.insert("", "end", iid=f"ivl_{pixel_id}", values=values)

    def create_series(self, root_var, dep_date_var, keyword_var, quarter_vars):
        dep_date = dep_date_var.get().strip()
        if not re.match(r"^\d{4}-\d{2}-\d{2}$", dep_date):
            messagebox.showerror("Дата", "Дата напыления должна быть в формате YYYY-MM-DD")
            return
        qnames = {str(q): quarter_vars[str(q)].get().strip() or f"Q{q}" for q in range(1, 5)}
        try:
            self.app_settings["default_root"] = str(Path(root_var.get()))
            save_app_settings(self.app_settings)
            self.series = SeriesManager.create_new(Path(root_var.get()), dep_date, keyword_var.get(), qnames)
            self.show_measurement_menu()
        except Exception as exc:
            messagebox.showerror("Не удалось создать серию", str(exc))

    def show_measurement_menu(self):
        assert self.series is not None
        self.clear()
        main = self._scrollable_frame(padding=14)

        header = ttk.Frame(main)
        header.pack(fill="x")
        ttk.Label(header, text="Измерения OLED", font=("Segoe UI", 18, "bold")).pack(side="left")
        ttk.Button(header, text="Открыть другую серию", command=self.show_start_screen).pack(side="right")
        ttk.Button(header, text="Настройки", command=self.open_settings_window).pack(side="right", padx=(0, 10))

        ttk.Label(main, text=f"Серия: {self.series.series_folder}").pack(anchor="w", pady=(4, 2))
        ttk.Label(main, text=f"Режим оборудования: {hardware_mode_label(self.app_settings)}", font=("Segoe UI", 10, "bold")).pack(anchor="w", pady=(0, 4))
        self._build_hardware_status_bar(main)
        self.after(250, self.check_hardware_status)

        btns = ttk.Frame(main)
        btns.pack(fill="x", pady=(0, 10))
        ttk.Button(btns, text="ВАЯХ", command=self.open_ivl_window, width=18).grid(row=0, column=0, padx=(0, 8))
        state_after_ivl = "normal" if self.series.journal.has_any_ivl() else "disabled"
        ttk.Button(btns, text="Спектры", command=self.open_spectrum_window, width=18, state=state_after_ivl).grid(row=0, column=1, padx=8)
        ttk.Button(btns, text="Стабильность", command=self.open_stability_window, width=18, state=state_after_ivl).grid(row=0, column=2, padx=8)
        ttk.Button(btns, text="Обновить", command=self.refresh_pixel_table, width=14).grid(row=0, column=3, padx=8)
        ttk.Button(btns, text="Журнал", command=lambda: messagebox.showinfo("Журнал", str(self.series.journal.path)), width=12).grid(row=0, column=4, padx=8)

        status_history = ttk.Frame(main)
        status_history.pack(fill="both", expand=True, pady=(0, 10))
        status_history.columnconfigure(0, weight=3)
        status_history.columnconfigure(1, weight=2)
        status_history.rowconfigure(0, weight=1)
        map_frame = ttk.LabelFrame(status_history, text="Подложкодержатель / карта статусов")
        map_frame.grid(row=0, column=0, sticky="nsew", padx=(0, 8))
        self._create_status_holder_canvas(map_frame)

        ivl_history_frame = ttk.LabelFrame(status_history, text="История ВАХ по датам")
        ivl_history_frame.grid(row=0, column=1, sticky="nsew", padx=(8, 0))
        self._create_ivl_history_tree(ivl_history_frame)

        latest_frame = ttk.LabelFrame(main, text="Последние даты и метрики")
        latest_frame.pack(fill="both", expand=True, pady=(0, 10))
        table_frame, self.tree = self._tree_with_scrollbars(
            latest_frame,
            columns=("pixel", "status", "opening", "ivl", "ivl_photo", "ivl_current", "spectrum", "spectrum_peaks", "spectrum_peak_nm", "stability"),
            height=14,
        )
        self.tree.heading("pixel", text="Пиксель")
        self.tree.heading("status", text="Статус")
        self.tree.heading("opening", text="V открытия, В")
        self.tree.heading("ivl", text="ВАЯХ")
        self.tree.heading("ivl_photo", text="Max PD, мкА")
        self.tree.heading("ivl_current", text="Max I, мА")
        self.tree.heading("spectrum", text="Спектры")
        self.tree.heading("spectrum_peaks", text="Пиков")
        self.tree.heading("spectrum_peak_nm", text="Пики, нм")
        self.tree.heading("stability", text="Стабильность")
        self.tree.column("pixel", width=150, minwidth=120, stretch=False)
        self.tree.column("status", width=140, minwidth=110, stretch=False)
        self.tree.column("opening", width=130, minwidth=110, stretch=False)
        self.tree.column("ivl", width=175, minwidth=145, stretch=False)
        self.tree.column("ivl_photo", width=105, minwidth=90, stretch=False)
        self.tree.column("ivl_current", width=95, minwidth=85, stretch=False)
        self.tree.column("spectrum", width=175, minwidth=145, stretch=False)
        self.tree.column("spectrum_peaks", width=70, minwidth=60, anchor="center", stretch=False)
        self.tree.column("spectrum_peak_nm", width=220, minwidth=150, stretch=False)
        self.tree.column("stability", width=190, minwidth=150, stretch=True)
        table_frame.pack(fill="both", expand=True)

        self.log_widget = None
        self.refresh_pixel_table()
        if state_after_ivl == "disabled":
            self.log("В журнале пока нет ВАЯХ: кнопки 'Спектры' и 'Стабильность' неактивны.")

    def refresh_pixel_table(self):
        if self.series is None or not hasattr(self, "tree"):
            return
        for item in self.tree.get_children():
            self.tree.delete(item)
        rows = self.series.journal.list_pixels()
        for r in rows:
            pid = r.get("Pixel ID")
            spectrum_peak_count = r.get("Last spectrum peak count", "")
            spectrum_peaks_nm = r.get("Last spectrum peaks nm", "")
            spectrum_max_intensity = r.get("Last spectrum max intensity (counts/s)", "")
            if (not spectrum_peak_count and not spectrum_peaks_nm) and r.get("Last spectrum file"):
                metrics = read_spectrum_metrics_from_workbook(resolve_series_file(self.series.series_folder, r.get("Last spectrum file")))
                spectrum_peak_count = metrics.get("peak_count", "")
                spectrum_peaks_nm = metrics.get("peaks_nm", "")
                spectrum_max_intensity = metrics.get("max_intensity", spectrum_max_intensity)
            spectrum_text = r.get("Last spectrum date", "") or ""
            if spectrum_max_intensity not in (None, ""):
                spectrum_text = f"{spectrum_text} | max {spectrum_max_intensity}" if spectrum_text else f"max {spectrum_max_intensity}"
            self.tree.insert(
                "",
                "end",
                iid=pid,
                values=(
                    pid or "",
                    r.get("Last status", "") or "",
                    r.get("Opening voltage (V)", "") or "",
                    r.get("Last IVL date", "") or "",
                    r.get("Last IVL max photodiode (uA)", "") or "",
                    r.get("Last IVL max current (mA)", "") or "",
                    spectrum_text,
                    spectrum_peak_count or "",
                    spectrum_peaks_nm or "",
                    r.get("Last stability date", "") or "",
                ),
            )
        self._render_status_holder_canvas()
        self._refresh_ivl_history_tree()

    def pixel_ids(self, require_ivl: bool = False, require_opening: bool = False) -> List[str]:
        assert self.series is not None
        rows = self.series.journal.list_pixels()
        out = []
        for row in rows:
            if require_ivl and not row.get("Last IVL file"):
                continue
            if require_opening and as_float_or_none(row.get("Opening voltage (V)")) is None:
                continue
            out.append(row["Pixel ID"])
        return out

    # -------------------------- окна ВАЯХ --------------------------
    def open_ivl_window(self):
        if self.series is None:
            return
        win = tk.Toplevel(self)
        win.title("ВАЯХ")
        win.geometry("560x520")
        frm = ttk.Frame(win, padding=14)
        frm.pack(fill="both", expand=True)

        mode_var = tk.StringVar(value="single")
        ttk.Radiobutton(frm, text="Конкретный пиксель", variable=mode_var, value="single").grid(row=0, column=0, sticky="w", pady=4)
        ttk.Radiobutton(frm, text="Вся серия последовательно", variable=mode_var, value="series").grid(row=0, column=1, sticky="w", pady=4)

        ttk.Label(frm, text="Пиксель:").grid(row=1, column=0, sticky="e", pady=5)
        pixel_var = tk.StringVar(value=self.pixel_ids()[0] if self.pixel_ids() else "")
        pixel_combo = ttk.Combobox(frm, textvariable=pixel_var, values=self.pixel_ids(), width=24, state="readonly")
        pixel_combo.grid(row=1, column=1, sticky="w", pady=5)

        saved_ivl = self.measurement_defaults("ivl")
        fields = [
            ("COM port", str(self.app_settings.get("com_port", "COM3"))),
            ("Sweep start, V", str(saved_ivl.get("sweep_start_V", "0"))),
            ("Sweep end, V", str(saved_ivl.get("sweep_end_V", "5"))),
            ("Step, V", str(saved_ivl.get("step_V", "0.02"))),
            ("Time per point, s", str(saved_ivl.get("time_per_point_s", "0.01"))),
            ("Cycles", str(saved_ivl.get("cycles", "1"))),
            ("Delay between cycles, s", str(saved_ivl.get("delay_between_cycles_s", "1"))),
            ("Current limit, mA", str(saved_ivl.get("current_limit_mA", "10"))),
        ]
        vars_ = {}
        for i, (label, default) in enumerate(fields, start=2):
            ttk.Label(frm, text=label + ":").grid(row=i, column=0, sticky="e", pady=3, padx=(0, 8))
            var = tk.StringVar(value=default)
            vars_[label] = var
            ttk.Entry(frm, textvariable=var, width=18).grid(row=i, column=1, sticky="w", pady=3)

        def start():
            progress = None
            try:
                adv = self.app_settings.get("ivl_advanced", DEFAULT_APP_SETTINGS["ivl_advanced"])
                units = self.app_settings.get("measurement_units", DEFAULT_APP_SETTINGS["measurement_units"])
                params = IVLParams(
                    com_port=vars_["COM port"].get().strip(),
                    sweep_start=parse_float(vars_["Sweep start, V"].get(), "Sweep start"),
                    sweep_end=parse_float(vars_["Sweep end, V"].get(), "Sweep end"),
                    sweep_increment=parse_float(vars_["Step, V"].get(), "Step"),
                    sweep_time_per_point=parse_float(vars_["Time per point, s"].get(), "Time per point"),
                    num_cycles=parse_int(vars_["Cycles"].get(), "Cycles"),
                    delay_between_cycles=parse_float(vars_["Delay between cycles, s"].get(), "Delay"),
                    current_limit_mA=parse_float(vars_["Current limit, mA"].get(), "Current limit"),
                    photodiode_bias_V=float(adv.get("photodiode_bias_V", -5.0)),
                    photodiode_range=int(adv.get("photodiode_range", 4)),
                    photodiode_threshold_uA=float(adv.get("photodiode_threshold_uA", 0.5)),
                    burnout_current_threshold_mA=float(adv.get("burnout_current_threshold_mA", 10.0)),
                    mark_current_limit_as_burnout=bool(adv.get("mark_current_limit_as_burnout", False)),
                    no_contact_max_led_current_mA=float(adv.get("no_contact_max_led_current_mA", 0.05)),
                    burned_confirmation_cycles=int(adv.get("burned_confirmation_cycles", 1)),
                    pixel_area_mm2=float(units.get("pixel_area_mm2", 1.0)),
                    luminance_cd_m2_per_uA=float(units.get("luminance_cd_m2_per_uA", 1.0)),
                )
                self.save_measurement_defaults("ivl", {
                    "sweep_start_V": vars_["Sweep start, V"].get(),
                    "sweep_end_V": vars_["Sweep end, V"].get(),
                    "step_V": vars_["Step, V"].get(),
                    "time_per_point_s": vars_["Time per point, s"].get(),
                    "cycles": vars_["Cycles"].get(),
                    "delay_between_cycles_s": vars_["Delay between cycles, s"].get(),
                    "current_limit_mA": vars_["Current limit, mA"].get(),
                })
                selected_pixel = pixel_var.get()
                win.destroy()
                if mode_var.get() == "single":
                    self.measure_one_ivl(selected_pixel, params)
                else:
                    self.measure_series_ivl(params, start_pixel=selected_pixel)
            except Exception as exc:
                messagebox.showerror("Ошибка параметров", str(exc))

        ttk.Label(frm, text="Дополнительные параметры ВАЯХ вынесены в Настройки → ВАЯХ доп.", foreground="#555555").grid(row=len(fields)+2, column=0, columnspan=2, sticky="w", pady=(10, 2))
        ttk.Button(frm, text="Открыть настройки", command=self.open_settings_window).grid(row=len(fields)+3, column=0, sticky="w", pady=12)
        ttk.Button(frm, text="Начать ВАЯХ", command=start).grid(row=len(fields)+3, column=1, sticky="w", pady=12)
        fit_toplevel_to_content(win, 620, 620)

    def measure_one_ivl(self, pixel_id: str, params: IVLParams, return_to_menu: bool = True) -> Optional[Dict[str, Any]]:
        assert self.series is not None
        if not pixel_id:
            messagebox.showwarning("Пиксель", "Пиксель не выбран")
            return None
        output_dir = ensure_measurement_folder(
            self.series.series_folder,
            "IVL",
            pixel_id,
            self.series.journal.get_pixel(pixel_id),
        )
        progress = IVLProgressWindow(self, pixel_id, params)
        try:
            progress.set_status(f"Пиксель {pixel_id}: идет съемка ВАЯХ")
            result = run_ivl_measurement(
                pixel_id,
                output_dir,
                params,
                self.log,
                self.app_settings,
                progress_callback=progress.add_point,
            )
            progress.set_status(f"Пиксель {pixel_id}: измерение завершено, статус {result.get('status', '')}")
            opening = result.get("opening_voltage")
            if result["status"] == "WORKING":
                opening = simpledialog.askfloat(
                    "Напряжение открытия",
                    f"Пиксель {pixel_id}\nЗадайте напряжение открытия, при котором пиксель начинает светить.",
                    initialvalue=round(float(opening), 3) if opening is not None else None,
                    parent=self,
                )
            self.series.journal.update_after_measurement(
                "IVL",
                pixel_id,
                result["status"],
                result["file"],
                params.as_dict(),
                notes=result.get("ivl_diagnosis", ""),
                opening_voltage=opening,
                max_current_mA=result.get("max_current_mA"),
                max_photo_uA=result.get("max_photo_uA"),
            )
            if result.get("ivl_diagnosis"):
                self.log(f"Первый промер {pixel_id}: {result['ivl_diagnosis']}")
            if result["status"] == "NO_CONTACT":
                self.log(f"Нет контакта на {pixel_id}: переставьте/проверьте эту подложку и снимите пиксель заново.")
                messagebox.showwarning(
                    "Нет контакта",
                    f"Пиксель {pixel_id}: ток почти нулевой, фототока нет.\n\n"
                    "Нужно переставить или проверить эту подложку и заново снять измерение. "
                    "Если проблема у всей подложки, при съемке серии можно перейти к следующей подложке.",
                    parent=self,
                )
            self.log(f"ВАЯХ завершена: {pixel_id}, файл {result['file'].name}")
            self.refresh_pixel_table()
            if return_to_menu:
                self.show_measurement_menu()
            progress.close()
            return result
        except MeasurementStopped as exc:
            self.log(str(exc))
            try:
                progress.set_status("Измерение остановлено. На SMU отправлено 0 В.")
            except Exception:
                pass
            messagebox.showinfo("Измерение остановлено", "Измерение остановлено. На каналы SMU отправлено 0 В.", parent=self)
            progress.close()
            if return_to_menu:
                self.show_measurement_menu()
            return None
        except Exception as exc:
            self.log(traceback.format_exc())
            try:
                progress.set_status(f"Ошибка: {exc}")
            except Exception:
                pass
            progress.close()
            messagebox.showerror("Ошибка ВАЯХ", str(exc))
            return None

    def pixel_info_from_journal(self, pixel_id: str) -> Optional[Dict[str, Any]]:
        if self.series is None:
            return None
        return self.series.journal.get_pixel(pixel_id)

    def remove_same_substrate_from_queue(self, remaining: List[str], pixel_id: str) -> List[str]:
        row = self.pixel_info_from_journal(pixel_id)
        if not row:
            return remaining
        q = row.get("Quarter number")
        substrate = row.get("Substrate number")
        out = []
        for pid in remaining:
            r = self.pixel_info_from_journal(pid)
            if r and r.get("Quarter number") == q and r.get("Substrate number") == substrate:
                continue
            out.append(pid)
        return out

    def measure_series_ivl(self, params: IVLParams, start_pixel: Optional[str] = None):
        assert self.series is not None
        all_pixels = self.pixel_ids()
        measured: List[str] = []
        remaining = all_pixels.copy()
        if start_pixel in remaining:
            start_idx = remaining.index(start_pixel)
            remaining = remaining[start_idx:]
        while remaining:
            next_pixel = remaining[0]
            msg = f"Следующий пиксель: {next_pixel}\n\nИзмерены:\n" + (", ".join(measured[-20:]) if measured else "пока нет")
            choice = messagebox.askyesnocancel(
                "Съем всей серии",
                msg + "\n\nДа — снять следующий.\nНет — выбрать произвольный пиксель.\nОтмена — остановить серию.",
                parent=self,
            )
            if choice is None:
                self.log("Съем всей серии отменен пользователем.")
                break
            if choice is False:
                chosen = self.ask_pixel("Выберите произвольный пиксель", values=all_pixels)
                if not chosen:
                    continue
                next_pixel = chosen
                if next_pixel in remaining:
                    remaining.remove(next_pixel)
            else:
                remaining.pop(0)

            while True:
                result = self.measure_one_ivl(next_pixel, params, return_to_menu=False)
                if result is None:
                    break
                status = result.get("status")
                if status != "NO_CONTACT":
                    measured.append(next_pixel)
                    break

                action = messagebox.askyesnocancel(
                    "Нет контакта на подложке",
                    f"Для {next_pixel} нет контакта.\n\n"
                    "Да — переставить/проверить эту подложку и переснять этот же пиксель.\n"
                    "Нет — пропустить оставшиеся пиксели этой подложки и перейти к следующей подложке.\n"
                    "Отмена — продолжить обычную очередь без пропуска.",
                    parent=self,
                )
                if action is True:
                    self.log(f"Повторная съемка {next_pixel} после перестановки/проверки подложки.")
                    continue
                if action is False:
                    remaining = self.remove_same_substrate_from_queue(remaining, next_pixel)
                    self.log(f"Оставшиеся пиксели подложки {next_pixel} пропущены; переход к следующей подложке.")
                measured.append(next_pixel)
                break

        self.show_measurement_menu()

    def ask_pixel(self, title: str, values: List[str]) -> Optional[str]:
        dlg = tk.Toplevel(self)
        dlg.title(title)
        dlg.geometry("360x120")
        dlg.transient(self)
        dlg.grab_set()
        var = tk.StringVar(value=values[0] if values else "")
        ttk.Label(dlg, text=title).pack(padx=12, pady=(12, 4))
        combo = ttk.Combobox(dlg, values=values, textvariable=var, state="readonly", width=30)
        combo.pack(padx=12, pady=4)
        result = {"value": None}
        def ok():
            result["value"] = var.get()
            dlg.destroy()
        ttk.Button(dlg, text="OK", command=ok).pack(pady=8)
        fit_toplevel_to_content(dlg, 420, 160)
        self.wait_window(dlg)
        return result["value"]

    # -------------------------- окно спектров --------------------------
    def open_spectrum_window(self):
        if self.series is None:
            return
        pixels = self.pixel_ids(require_ivl=True, require_opening=True)
        if not pixels:
            messagebox.showwarning("Спектры", "В журнале нет пикселей с ВАЯХ и заданным напряжением открытия.")
            return
        win = tk.Toplevel(self)
        win.title("Спектры")
        win.geometry("560x560")
        frm = ttk.Frame(win, padding=14)
        frm.pack(fill="both", expand=True)
        ttk.Label(frm, text="Пиксель:").grid(row=0, column=0, sticky="e", pady=5)
        pixel_var = tk.StringVar(value=pixels[0])
        pixel_combo = ttk.Combobox(frm, values=pixels, textvariable=pixel_var, state="readonly", width=26)
        pixel_combo.grid(row=0, column=1, sticky="w", pady=5)
        first_pixel = self.series.journal.get_pixel(pixels[0])
        first_opening = as_float_or_none(first_pixel.get("Opening voltage (V)")) if first_pixel else None
        saved_spectrum = self.measurement_defaults("spectrum")
        use_opening_var = tk.BooleanVar(value=bool(saved_spectrum.get("use_opening_voltage", True)))
        opening_info_var = tk.StringVar(value=f"V открытия: {first_opening:.3f} В" if first_opening is not None else "V открытия: нет")

        fields = [
            ("COM port", str(self.app_settings.get("com_port", "COM3"))),
            ("Voltage start, V", f"{first_opening:.3f}" if first_opening is not None else "2.0"),
            ("Voltage end, V", str(saved_spectrum.get("voltage_end_V", "5"))),
            ("Voltage step, V", str(saved_spectrum.get("voltage_step_V", "0.1"))),
            ("Current limit, mA", str(saved_spectrum.get("current_limit_mA", "6"))),
            ("LED type", str(saved_spectrum.get("led_type", "auto"))),
        ]
        vars_ = {}
        for i, (label, default) in enumerate(fields, start=1):
            ttk.Label(frm, text=label + ":").grid(row=i, column=0, sticky="e", pady=3, padx=(0, 8))
            var = tk.StringVar(value=default)
            vars_[label] = var
            ttk.Entry(frm, textvariable=var, width=18).grid(row=i, column=1, sticky="w", pady=3)
        ttk.Checkbutton(frm, text="Стартовать от V открытия из журнала", variable=use_opening_var).grid(row=len(fields)+1, column=0, columnspan=2, sticky="w", pady=(8, 2))
        ttk.Label(frm, textvariable=opening_info_var, foreground="#555555").grid(row=len(fields)+2, column=0, columnspan=2, sticky="w", pady=(0, 2))

        def update_opening_info(_event=None):
            pixel = self.series.journal.get_pixel(pixel_var.get())
            opening = as_float_or_none(pixel.get("Opening voltage (V)")) if pixel else None
            opening_info_var.set(f"V открытия: {opening:.3f} В" if opening is not None else "V открытия: нет")
            if opening is not None and use_opening_var.get():
                vars_["Voltage start, V"].set(f"{opening:.3f}")

        pixel_combo.bind("<<ComboboxSelected>>", update_opening_info)

        def start():
            progress = None
            try:
                pid = pixel_var.get()
                pixel = self.series.journal.get_pixel(pid)
                opening = as_float_or_none(pixel.get("Opening voltage (V)")) if pixel else None
                if opening is None:
                    raise ValueError("Для выбранного пикселя нет напряжения открытия")
                voltage_start = float(opening) if use_opening_var.get() else parse_float(vars_["Voltage start, V"].get(), "Voltage start")
                adv = self.app_settings.get("spectrum_advanced", DEFAULT_APP_SETTINGS["spectrum_advanced"])
                units = self.app_settings.get("measurement_units", DEFAULT_APP_SETTINGS["measurement_units"])
                params = SpectrumParams(
                    com_port=vars_["COM port"].get().strip(),
                    voltage_start=voltage_start,
                    voltage_end=parse_float(vars_["Voltage end, V"].get(), "Voltage end"),
                    voltage_step=parse_float(vars_["Voltage step, V"].get(), "Voltage step"),
                    opening_voltage=float(opening),
                    voltage_start_source="opening" if use_opening_var.get() else "manual",
                    current_limit_mA=parse_float(vars_["Current limit, mA"].get(), "Current limit"),
                    photodiode_bias_V=float(adv.get("photodiode_bias_V", -5.0)),
                    photodiode_range=int(adv.get("photodiode_range", 4)),
                    target_intensity=float(adv.get("target_intensity", 40000.0)),
                    intensity_min=float(adv.get("intensity_min", 20000.0)),
                    intensity_max=float(adv.get("intensity_max", 55000.0)),
                    saturation_level=float(adv.get("saturation_level", 60000.0)),
                    min_peak_width_nm=float(adv.get("min_peak_width_nm", 15.0)),
                    max_peak_width_nm=float(adv.get("max_peak_width_nm", 150.0)),
                    t_int_initial_s=float(adv.get("t_int_initial_s", 0.01)),
                    t_int_min_s=float(adv.get("t_int_min_s", 0.001)),
                    t_int_max_s=float(adv.get("t_int_max_s", 10.0)),
                    discard_first_scan_after_tint_change=bool(adv.get("discard_first_scan_after_tint_change", True)),
                    kp=float(adv.get("kp", 0.3)),
                    ki=float(adv.get("ki", 0.05)),
                    max_iterations=int(adv.get("max_iterations", 20)),
                    tolerance=float(adv.get("tolerance", 0.05)),
                    led_type=vars_["LED type"].get().strip() or "auto",
                    peak_search_mode_for_tint=str(adv.get("peak_search_mode_for_tint", "auto")),
                    settle_time_voltage_s=float(adv.get("settle_time_voltage_s", 0.1)),
                    settle_time_spectrum_s=float(adv.get("settle_time_spectrum_s", 0.05)),
                    dark_spectrum_enabled=bool(adv.get("dark_spectrum_enabled", False)),
                    dark_spectrum_scans=int(adv.get("dark_spectrum_scans", 3)),
                    baseline_correction_enabled=bool(adv.get("baseline_correction_enabled", True)),
                    peak_detection_enabled=bool(adv.get("peak_detection_enabled", False)),
                    pixel_area_mm2=float(units.get("pixel_area_mm2", 1.0)),
                    luminance_cd_m2_per_uA=float(units.get("luminance_cd_m2_per_uA", 1.0)),
                )
                self.save_measurement_defaults("spectrum", {
                    "voltage_end_V": vars_["Voltage end, V"].get(),
                    "voltage_step_V": vars_["Voltage step, V"].get(),
                    "current_limit_mA": vars_["Current limit, mA"].get(),
                    "led_type": vars_["LED type"].get(),
                    "use_opening_voltage": bool(use_opening_var.get()),
                })
                output_dir = ensure_measurement_folder(
                    self.series.series_folder,
                    "SPECTRUM",
                    pid,
                    self.series.journal.get_pixel(pid),
                )
                progress = SpectrumProgressWindow(self, pid)
                result = run_spectrum_measurement(pid, output_dir, params, self.log, self.app_settings, progress_callback=progress.update_spectrum)
                progress.close()
                self.series.journal.update_after_measurement(
                    "SPECTRUM",
                    pid,
                    result["status"],
                    result["file"],
                    params.as_dict(),
                    spectrum_peak_count=result.get("spectrum_peak_count"),
                    spectrum_peaks_nm=result.get("spectrum_peaks_nm", ""),
                    spectrum_max_intensity=result.get("spectrum_max_intensity"),
                )
                self.log(f"Спектры завершены: {pid}, файл {result['file'].name}")
                self.refresh_pixel_table()
                self.show_measurement_menu()
                win.destroy()
            except Exception as exc:
                if progress is not None:
                    progress.close()
                self.log(traceback.format_exc())
                messagebox.showerror("Ошибка спектров", str(exc))

        ttk.Label(frm, text="V открытия остается в журнале. Для спектра можно временно выбрать другое стартовое напряжение.", foreground="#555555", wraplength=500).grid(row=len(fields)+3, column=0, columnspan=2, sticky="w", pady=(8, 2))
        ttk.Button(frm, text="Открыть настройки", command=self.open_settings_window).grid(row=len(fields)+4, column=0, sticky="w", pady=16)
        ttk.Button(frm, text="Начать съемку спектров", command=start).grid(row=len(fields)+4, column=1, sticky="w", pady=16)
        fit_toplevel_to_content(win, 620, 650)

    # -------------------------- окно стабильности --------------------------
    def open_stability_window(self):
        if self.series is None:
            return
        pixels = self.pixel_ids(require_ivl=True)
        if not pixels:
            messagebox.showwarning("Стабильность", "В журнале нет пикселей с ВАЯХ.")
            return
        win = tk.Toplevel(self)
        win.title("Стабильность по току")
        win.geometry("600x560")
        frm = ttk.Frame(win, padding=14)
        frm.pack(fill="both", expand=True)
        ttk.Label(frm, text="Пиксель:").grid(row=0, column=0, sticky="e", pady=5)
        pixel_var = tk.StringVar(value=pixels[0])
        ttk.Combobox(frm, values=pixels, textvariable=pixel_var, state="readonly", width=26).grid(row=0, column=1, sticky="w", pady=5)

        saved_stability = self.measurement_defaults("stability")
        fields = [
            ("COM port", str(self.app_settings.get("com_port", "COM3"))),
            ("Current setpoint, mA", str(saved_stability.get("current_setpoint_mA", "3.5"))),
            ("Voltage limit, V", str(saved_stability.get("voltage_limit_V", "5"))),
            ("Current limit, mA", str(saved_stability.get("current_limit_mA", "10"))),
            ("Measurement time, s", str(saved_stability.get("measurement_time_s", "86400"))),
            ("Sample interval, s", str(saved_stability.get("sample_interval_s", "1"))),
            ("Autosave interval, s", str(saved_stability.get("autosave_interval_s", "600"))),
        ]
        vars_ = {}
        for i, (label, default) in enumerate(fields, start=1):
            ttk.Label(frm, text=label + ":").grid(row=i, column=0, sticky="e", pady=3, padx=(0, 8))
            var = tk.StringVar(value=default)
            vars_[label] = var
            ttk.Entry(frm, textvariable=var, width=18).grid(row=i, column=1, sticky="w", pady=3)

        start_voltage_label = ttk.Label(frm, text="Стартовое напряжение будет рассчитано как 0.9 × V(ВАЯХ) для заданного тока. Остальное — Настройки → Стабильность доп.", foreground="#555555", wraplength=540)
        start_voltage_label.grid(row=len(fields)+1, column=0, columnspan=2, sticky="w", pady=(8, 2))

        def start():
            try:
                pid = pixel_var.get()
                target_current = parse_float(vars_["Current setpoint, mA"].get(), "Current setpoint")
                pixel = self.series.journal.get_pixel(pid)
                ivl_rel = pixel.get("Last IVL file") if pixel else ""
                ivl_file = self.series.series_folder / ivl_rel if ivl_rel else None
                v_at_current = interpolate_voltage_at_current_from_ivl(ivl_file, target_current) if ivl_file else None
                if v_at_current is None:
                    manual = simpledialog.askfloat(
                        "Стартовое напряжение",
                        f"Не удалось найти V в ВАЯХ для {target_current:g} мА.\nВведите напряжение, соответствующее этому току по ВАХ.",
                        parent=self,
                    )
                    if manual is None:
                        return
                    v_at_current = manual
                voltage_start = 0.9 * float(v_at_current)
                if not messagebox.askokcancel(
                    "Проверка старта",
                    f"Для {pid}: V по ВАЯХ при {target_current:g} мА ≈ {v_at_current:.3f} В.\n"
                    f"Старт стабильности будет {voltage_start:.3f} В, то есть на 10% ниже.\n\nПродолжить?",
                    parent=self,
                ):
                    return
                adv = self.app_settings.get("stability_advanced", DEFAULT_APP_SETTINGS["stability_advanced"])
                units = self.app_settings.get("measurement_units", DEFAULT_APP_SETTINGS["measurement_units"])
                params = StabilityParams(
                    com_port=vars_["COM port"].get().strip(),
                    current_setpoint_mA=target_current,
                    voltage_start=voltage_start,
                    voltage_limit=parse_float(vars_["Voltage limit, V"].get(), "Voltage limit"),
                    current_limit_mA=parse_float(vars_["Current limit, mA"].get(), "Current limit"),
                    voltage_step_max=float(adv.get("voltage_step_max", 0.02)),
                    current_control_kp=float(adv.get("current_control_kp", 0.01)),
                    measurement_time_s=parse_float(vars_["Measurement time, s"].get(), "Measurement time"),
                    sample_interval_s=parse_float(vars_["Sample interval, s"].get(), "Sample interval"),
                    autosave_interval_s=parse_float(vars_["Autosave interval, s"].get(), "Autosave interval"),
                    photodiode_bias_V=float(adv.get("photodiode_bias_V", -5.0)),
                    photodiode_threshold_uA=float(adv.get("photodiode_threshold_uA", 0.1)),
                    photodiode_range=int(adv.get("photodiode_range", 4)),
                    pixel_area_mm2=float(units.get("pixel_area_mm2", 1.0)),
                    luminance_cd_m2_per_uA=float(units.get("luminance_cd_m2_per_uA", 1.0)),
                )
                self.save_measurement_defaults("stability", {
                    "current_setpoint_mA": vars_["Current setpoint, mA"].get(),
                    "voltage_limit_V": vars_["Voltage limit, V"].get(),
                    "current_limit_mA": vars_["Current limit, mA"].get(),
                    "measurement_time_s": vars_["Measurement time, s"].get(),
                    "sample_interval_s": vars_["Sample interval, s"].get(),
                    "autosave_interval_s": vars_["Autosave interval, s"].get(),
                })
                output_dir = ensure_measurement_folder(
                    self.series.series_folder,
                    "STABILITY",
                    pid,
                    self.series.journal.get_pixel(pid),
                )
                result = run_stability_measurement(pid, output_dir, params, self.log, self.app_settings)
                self.series.journal.update_after_measurement("STABILITY", pid, result["status"], result["file"], params.as_dict())
                self.log(f"Стабильность завершена: {pid}, файл {result['file'].name}")
                self.refresh_pixel_table()
                self.show_measurement_menu()
                win.destroy()
            except Exception as exc:
                self.log(traceback.format_exc())
                messagebox.showerror("Ошибка стабильности", str(exc))

        ttk.Button(frm, text="Открыть настройки", command=self.open_settings_window).grid(row=len(fields)+3, column=0, sticky="w", pady=16)
        ttk.Button(frm, text="Начать стабильность", command=start).grid(row=len(fields)+3, column=1, sticky="w", pady=16)
        fit_toplevel_to_content(win, 660, 650)


def main():
    os.chdir(SCRIPT_DIR)
    enable_windows_dpi_awareness()
    app = OLEDApp()
    app.mainloop()


if __name__ == "__main__":
    main()


"""Thread-safe bridge between the v2 API and the existing series contracts."""

from __future__ import annotations

import json
import threading
from datetime import date, datetime
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional

from openpyxl import load_workbook

from oled_app.constants import CONFIG_FILE, JOURNAL_FILE, MEASUREMENTS_SHEET
from oled_app.processing.ivl_preview import (
    create_ivl_thumbnail_from_workbook,
    ivl_thumbnail_needs_refresh,
    ivl_thumbnail_path,
)
from oled_app.series.manager import SeriesManager
from oled_app.series.metadata import (
    expand_descriptions_for_scope,
    led_color_label,
    normalize_description_scope,
    normalize_half_orientation,
    normalize_led_color,
    quarter_base,
    quarter_code,
    quarter_description,
    quarter_led_color,
    series_description_scope,
    series_half_orientation,
)
from oled_app.settings import load_app_settings
from oled_app.utils import resolve_series_file


class SeriesServiceError(RuntimeError):
    """Base class for operator-facing series errors."""


class SeriesValidationError(SeriesServiceError):
    pass


class SeriesNotFoundError(SeriesServiceError):
    pass


class SeriesConflictError(SeriesServiceError):
    pass


def _json_value(value: Any) -> Any:
    if value in (None, ""):
        return None
    if isinstance(value, (datetime, date)):
        return value.isoformat(sep=" ") if isinstance(value, datetime) else value.isoformat()
    if isinstance(value, (str, int, float, bool)):
        return value
    return str(value)


def _clean_text(value: Any, field: str, max_length: int, required: bool = False) -> str:
    text = str(value or "").strip()
    if required and not text:
        raise SeriesValidationError(f"Поле «{field}» обязательно.")
    if len(text) > max_length:
        raise SeriesValidationError(
            f"Поле «{field}» не должно быть длиннее {max_length} символов."
        )
    return text


def _existing_directory(value: Any, field: str) -> Path:
    text = _clean_text(value, field, 2048, required=True)
    try:
        path = Path(text).expanduser().resolve(strict=True)
    except (OSError, RuntimeError) as exc:
        raise SeriesNotFoundError(f"Папка не найдена: {text}") from exc
    if not path.is_dir():
        raise SeriesValidationError(f"Указанный путь не является папкой: {path}")
    return path


class SeriesService:
    """Own the active series for one isolated desktop backend session."""

    def __init__(self, default_root: Optional[Path] = None, logger=None) -> None:
        settings = load_app_settings()
        configured_root = default_root or Path(str(settings.get("default_root") or "OLED_series"))
        self._root = Path(configured_root).expanduser().resolve()
        self._active: Optional[SeriesManager] = None
        self._lock = threading.RLock()
        self._logger = logger

    @property
    def root(self) -> Path:
        with self._lock:
            return self._root

    @property
    def active_path(self) -> Optional[Path]:
        with self._lock:
            return self._active.series_folder if self._active is not None else None

    def app_summary(self) -> Dict[str, Any]:
        active_path = self.active_path
        return {
            "active": active_path is not None,
            "path": str(active_path) if active_path is not None else None,
            "root": str(self.root),
        }

    def state(self) -> Dict[str, Any]:
        with self._lock:
            return {
                "root": str(self._root),
                "recent": self._list_series_locked(),
                "active": self._active_payload_locked(),
            }

    def set_root(self, value: Any) -> Dict[str, Any]:
        root = _existing_directory(value, "Корневая папка")
        with self._lock:
            self._root = root
            self._log(f"Series root selected: {root}")
            return self.state()

    def open_series(self, value: Any) -> Dict[str, Any]:
        folder = _existing_directory(value, "Папка серии")
        if not (folder / CONFIG_FILE).is_file():
            raise SeriesNotFoundError(f"В папке нет {CONFIG_FILE}: {folder}")
        try:
            manager = SeriesManager(folder)
        except (OSError, ValueError, json.JSONDecodeError) as exc:
            raise SeriesValidationError(f"Не удалось открыть серию: {exc}") from exc
        with self._lock:
            self._active = manager
            self._root = folder.parent
            self._log(f"Series opened: {folder}")
            return self.state()

    def close_series(self) -> Dict[str, Any]:
        with self._lock:
            if self._active is not None:
                self._log(f"Series closed: {self._active.series_folder}")
            self._active = None
            return self.state()

    def create_series(self, payload: Dict[str, Any]) -> Dict[str, Any]:
        with self._lock:
            root_value = payload.get("root") or str(self._root)
        root_text = _clean_text(root_value, "Корневая папка", 2048, required=True)
        root = Path(root_text).expanduser().resolve()
        try:
            root.mkdir(parents=True, exist_ok=True)
        except OSError as exc:
            raise SeriesValidationError(f"Не удалось создать корневую папку: {exc}") from exc
        if not root.is_dir():
            raise SeriesValidationError(f"Корневая папка недоступна: {root}")

        values = self._validated_config_payload(payload)
        try:
            manager = SeriesManager.create_new(
                root,
                values["deposition_date"],
                values["keyword"],
                values["quarter_bases"],
                values["quarter_descriptions"],
                values["quarter_led_colors"],
                values["description_scope"],
                values["half_orientation"],
            )
        except OSError as exc:
            raise SeriesConflictError(f"Не удалось создать серию: {exc}") from exc
        with self._lock:
            self._root = root
            self._active = manager
            self._log(f"Series created: {manager.series_folder}")
            return self.state()

    def update_active(self, payload: Dict[str, Any]) -> Dict[str, Any]:
        values = self._validated_config_payload(payload)
        with self._lock:
            manager = self._require_active_locked()
            try:
                manager.update_config(
                    values["deposition_date"],
                    values["keyword"],
                    values["quarter_bases"],
                    values["quarter_descriptions"],
                    values["quarter_led_colors"],
                    values["description_scope"],
                    values["half_orientation"],
                )
            except OSError as exc:
                raise SeriesValidationError(f"Не удалось сохранить серию: {exc}") from exc
            self._log(f"Series settings updated: {manager.series_folder}")
            return self.state()

    def refresh_active(self, generate_thumbnails: bool = True) -> Dict[str, Any]:
        with self._lock:
            manager = self._require_active_locked()
            folder = manager.series_folder
            self._active = SeriesManager(folder)
            refreshed = self._refresh_thumbnails_locked() if generate_thumbnails else 0
            payload = self.state()
            payload["refreshed_thumbnails"] = refreshed
            self._log(f"Series refreshed: {folder}; thumbnails={refreshed}")
            return payload

    def set_spectrum_priority(
        self,
        pixel_id: Any,
        enabled: Any,
        scope: Any = "pixel",
    ) -> Dict[str, Any]:
        selected = _clean_text(pixel_id, "Пиксель", 160, required=True)
        requested_scope = str(scope or "pixel").strip().lower()
        if requested_scope not in {"pixel", "substrate"}:
            raise SeriesValidationError("Область очереди должна быть pixel или substrate.")
        if not isinstance(enabled, bool):
            raise SeriesValidationError("Состояние очереди должно быть логическим значением.")
        with self._lock:
            manager = self._require_active_locked()
            rows = manager.journal.list_pixels()
            by_id = {str(row.get("Pixel ID") or ""): row for row in rows}
            if selected not in by_id:
                raise SeriesNotFoundError(f"Пиксель не найден: {selected}")
            if requested_scope == "substrate":
                prefix = selected.rsplit("_", 1)[0]
                pixel_ids = [
                    pixel
                    for pixel, row in by_id.items()
                    if pixel.rsplit("_", 1)[0] == prefix and not row.get("Last spectrum file")
                ]
            else:
                pixel_ids = [selected]
            changed = manager.journal.set_spectrum_priorities(pixel_ids, enabled)
            result = self.state()
            result["queue_update"] = {
                "scope": requested_scope,
                "requested": len(pixel_ids),
                "changed": changed,
                "enabled": enabled,
            }
            return result

    def thumbnail_for_pixel(self, pixel_id: Any) -> Path:
        selected = _clean_text(pixel_id, "Пиксель", 160, required=True)
        with self._lock:
            manager = self._require_active_locked()
            row = manager.journal.get_pixel(selected)
            if row is None:
                raise SeriesNotFoundError(f"Пиксель не найден: {selected}")
            workbook = resolve_series_file(manager.series_folder, row.get("Last IVL file"))
            if workbook is None:
                raise SeriesNotFoundError(f"Для {selected} нет файла ВАЯХ.")
            thumbnail = ivl_thumbnail_path(workbook, selected).resolve()
            try:
                thumbnail.relative_to(manager.series_folder.resolve())
            except ValueError as exc:
                raise SeriesValidationError("Миниатюра находится вне папки серии.") from exc
            if not thumbnail.is_file():
                raise SeriesNotFoundError(f"Миниатюра ВАЯХ для {selected} не найдена.")
            return thumbnail

    def _validated_config_payload(self, payload: Dict[str, Any]) -> Dict[str, Any]:
        deposition_date = _clean_text(payload.get("deposition_date"), "Дата напыления", 32, True)
        try:
            date.fromisoformat(deposition_date)
        except ValueError as exc:
            raise SeriesValidationError("Дата напыления должна иметь формат ГГГГ-ММ-ДД.") from exc
        keyword = _clean_text(payload.get("keyword"), "Кодовое слово", 120)
        source_bases = payload.get("quarter_bases")
        source_descriptions = payload.get("quarter_descriptions")
        if not isinstance(source_bases, dict) or not isinstance(source_descriptions, dict):
            raise SeriesValidationError("Не заданы параметры четырёх четвертей.")
        color = normalize_led_color(payload.get("series_led_color"))
        description_scope = normalize_description_scope(payload.get("description_scope"))
        half_orientation = normalize_half_orientation(payload.get("half_orientation"))
        bases: Dict[str, str] = {}
        descriptions: Dict[str, str] = {}
        for number in range(1, 5):
            key = str(number)
            bases[key] = _clean_text(source_bases.get(key) or "Q", f"База четверти {key}", 32, True)
            descriptions[key] = _clean_text(
                source_descriptions.get(key),
                f"Описание четверти {key}",
                180,
            )
        descriptions = expand_descriptions_for_scope(
            descriptions,
            description_scope,
            half_orientation,
        )
        return {
            "deposition_date": deposition_date,
            "keyword": keyword,
            "quarter_bases": bases,
            "quarter_descriptions": descriptions,
            "quarter_led_colors": {str(number): color for number in range(1, 5)},
            "description_scope": description_scope,
            "half_orientation": half_orientation,
        }

    def _require_active_locked(self) -> SeriesManager:
        if self._active is None:
            raise SeriesConflictError("Серия не открыта.")
        return self._active

    def _list_series_locked(self) -> List[Dict[str, Any]]:
        root = self._root
        if not root.is_dir():
            return []
        found: List[Dict[str, Any]] = []
        try:
            config_paths: Iterable[Path] = root.rglob(CONFIG_FILE)
            for config_path in config_paths:
                if len(found) >= 200:
                    break
                folder = config_path.parent
                try:
                    config = json.loads(config_path.read_text(encoding="utf-8"))
                except (OSError, ValueError, json.JSONDecodeError):
                    continue
                found.append(
                    {
                        "path": str(folder.resolve()),
                        "folder_name": folder.name,
                        "deposition_date": _json_value(config.get("deposition_date")),
                        "keyword": _json_value(config.get("keyword")),
                        "created_at": _json_value(config.get("created_at")),
                        "measurements_count": self._measurement_count(folder / JOURNAL_FILE),
                    }
                )
        except OSError:
            return []
        found.sort(
            key=lambda item: str(item.get("created_at") or item.get("deposition_date") or ""),
            reverse=True,
        )
        return found

    @staticmethod
    def _measurement_count(journal_path: Path) -> Any:
        if not journal_path.is_file():
            return 0
        try:
            workbook = load_workbook(journal_path, data_only=True, read_only=True)
            try:
                if MEASUREMENTS_SHEET not in workbook.sheetnames:
                    return 0
                return max(workbook[MEASUREMENTS_SHEET].max_row - 1, 0)
            finally:
                workbook.close()
        except Exception:
            return None

    def _active_payload_locked(self) -> Optional[Dict[str, Any]]:
        if self._active is None:
            return None
        manager = self._active
        config = manager.config
        rows = manager.journal.list_pixels()
        measurements = manager.journal.list_measurements()
        pixels = [self._pixel_payload(manager, row) for row in rows]
        history = [
            {
                "date_time": _json_value(row.get("Date time")),
                "measurement_day": _json_value(row.get("Measurement day")),
                "type": str(row.get("Type") or ""),
                "pixel_id": str(row.get("Pixel ID") or ""),
                "status": str(row.get("Status") or ""),
                "file": _json_value(row.get("File")),
                "notes": _json_value(row.get("Notes")),
            }
            for row in measurements[-300:]
        ]
        measured = sum(
            1
            for row in rows
            if row.get("Last IVL file")
            or row.get("Last spectrum file")
            or row.get("Last stability file")
            or str(row.get("Last status") or "").upper() not in {"", "UNKNOWN"}
        )
        return {
            "path": str(manager.series_folder.resolve()),
            "folder_name": manager.series_folder.name,
            "deposition_date": str(config.get("deposition_date") or ""),
            "keyword": str(config.get("keyword") or ""),
            "created_at": _json_value(config.get("created_at")),
            "series_led_color": quarter_led_color(config, 1),
            "description_scope": series_description_scope(config),
            "half_orientation": series_half_orientation(config),
            "quarters": [
                {
                    "number": number,
                    "code": quarter_code(config, number),
                    "base": quarter_base(config, number),
                    "description": quarter_description(config, number),
                    "led_color": quarter_led_color(config, number),
                    "led_color_label": led_color_label(quarter_led_color(config, number)),
                }
                for number in range(1, 5)
            ],
            "pixels": pixels,
            "history": history,
            "metrics": {
                "substrates": 12,
                "pixels": len(pixels),
                "measured": measured,
                "ivl": sum(1 for row in rows if row.get("Last IVL file")),
                "spectra": sum(1 for row in rows if row.get("Last spectrum file")),
                "stability": sum(1 for row in rows if row.get("Last stability file")),
                "spectrum_queue": sum(
                    1 for row in rows if bool(row.get("Spectrum priority")) and not row.get("Last spectrum file")
                ),
                "history": len(measurements),
            },
        }

    @staticmethod
    def _pixel_payload(manager: SeriesManager, row: Dict[str, Any]) -> Dict[str, Any]:
        pixel_id = str(row.get("Pixel ID") or "")
        workbook = resolve_series_file(manager.series_folder, row.get("Last IVL file"))
        thumbnail_available = bool(
            workbook is not None and ivl_thumbnail_path(workbook, pixel_id).is_file()
        )
        return {
            "pixel_id": pixel_id,
            "quarter_code": str(row.get("Quarter code") or ""),
            "quarter_number": int(row.get("Quarter number") or 0),
            "quarter_description": str(row.get("Quarter description") or ""),
            "led_color": str(row.get("LED color") or ""),
            "substrate_number": int(row.get("Substrate number") or 0),
            "pixel_number": int(row.get("Pixel number") or 0),
            "status": str(row.get("Last status") or "UNKNOWN"),
            "opening_voltage_V": _json_value(row.get("Opening voltage (V)")),
            "last_ivl_date": _json_value(row.get("Last IVL date")),
            "last_ivl_file": _json_value(row.get("Last IVL file")),
            "last_ivl_max_current_mA": _json_value(row.get("Last IVL max current (mA)")),
            "last_ivl_max_photodiode_uA": _json_value(row.get("Last IVL max photodiode (uA)")),
            "spectrum_priority": bool(row.get("Spectrum priority")),
            "last_spectrum_date": _json_value(row.get("Last spectrum date")),
            "last_spectrum_file": _json_value(row.get("Last spectrum file")),
            "last_spectrum_peak_count": _json_value(row.get("Last spectrum peak count")),
            "last_spectrum_peaks_nm": _json_value(row.get("Last spectrum peaks nm")),
            "last_spectrum_max_intensity": _json_value(
                row.get("Last spectrum max intensity (counts/s)")
            ),
            "last_stability_date": _json_value(row.get("Last stability date")),
            "last_stability_file": _json_value(row.get("Last stability file")),
            "last_updated": _json_value(row.get("Last updated")),
            "thumbnail_available": thumbnail_available,
        }

    def _refresh_thumbnails_locked(self) -> int:
        manager = self._require_active_locked()
        refreshed = 0
        for row in manager.journal.list_pixels():
            pixel_id = str(row.get("Pixel ID") or "")
            workbook = resolve_series_file(manager.series_folder, row.get("Last IVL file"))
            if not pixel_id or workbook is None:
                continue
            output = ivl_thumbnail_path(workbook, pixel_id)
            try:
                if ivl_thumbnail_needs_refresh(output, workbook):
                    create_ivl_thumbnail_from_workbook(workbook, output)
                    refreshed += 1
            except Exception as exc:
                self._log(f"IVL thumbnail refresh failed pixel={pixel_id}: {exc}")
        return refreshed

    def _log(self, message: str) -> None:
        if self._logger is not None:
            self._logger.info(message)

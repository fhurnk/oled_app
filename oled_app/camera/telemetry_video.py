"""Build an annotated stability video without covering the camera image."""

from __future__ import annotations

import json
import math
import os
import re
import shutil
import subprocess
import tempfile
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Sequence, Tuple

from openpyxl import load_workbook


TELEMETRY_BAR_MIN_HEIGHT = 80
TELEMETRY_BAR_HEIGHT_RATIO = 0.12
TELEMETRY_VIDEO_SUFFIX = "_telemetry"


@dataclass(frozen=True)
class StabilityTelemetrySample:
    """One measured stability point used by the video telemetry renderer."""

    measurement_time_s: float
    voltage_V: float
    current_mA: float


def find_ffmpeg_executable() -> Path:
    """Return a system FFmpeg or the binary supplied by imageio-ffmpeg."""

    system_ffmpeg = shutil.which("ffmpeg")
    if system_ffmpeg:
        return Path(system_ffmpeg)
    try:
        import imageio_ffmpeg
    except ImportError as exc:
        raise RuntimeError(
            "Для создания видео с показаниями установите зависимости из requirements.txt "
            "(пакет imageio-ffmpeg)."
        ) from exc
    executable = Path(imageio_ffmpeg.get_ffmpeg_exe())
    if not executable.exists():
        raise RuntimeError("Исполняемый файл FFmpeg не найден.")
    return executable


def read_stability_telemetry(workbook_path: Path | str) -> List[StabilityTelemetrySample]:
    """Read elapsed time, measured voltage and measured current from a stability workbook."""

    path = Path(workbook_path)
    if not path.exists():
        raise RuntimeError(f"Файл стабильности не найден: {path}")
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        worksheet = workbook["Data"] if "Data" in workbook.sheetnames else workbook.active
        required = {
            "Time (s)": "time",
            "Voltage OLED / LED (V)": "voltage",
            "Current OLED / LED (mA)": "current",
        }
        header_row: Optional[int] = None
        columns: Dict[str, int] = {}
        for row_index in range(1, min(worksheet.max_row, 60) + 1):
            row_columns: Dict[str, int] = {}
            for column_index in range(1, worksheet.max_column + 1):
                value = worksheet.cell(row=row_index, column=column_index).value
                if value in required:
                    row_columns[required[str(value)]] = column_index
            if len(row_columns) == len(required):
                header_row = row_index
                columns = row_columns
                break
        if header_row is None:
            raise RuntimeError("В XLSX стабильности не найдены столбцы времени, напряжения и тока.")

        samples: List[StabilityTelemetrySample] = []
        for row_index in range(header_row + 1, worksheet.max_row + 1):
            values = (
                worksheet.cell(row=row_index, column=columns["time"]).value,
                worksheet.cell(row=row_index, column=columns["voltage"]).value,
                worksheet.cell(row=row_index, column=columns["current"]).value,
            )
            try:
                measurement_time_s, voltage_V, current_mA = (float(value) for value in values)
            except (TypeError, ValueError):
                continue
            if not all(math.isfinite(value) for value in (measurement_time_s, voltage_V, current_mA)):
                continue
            samples.append(
                StabilityTelemetrySample(
                    measurement_time_s=measurement_time_s,
                    voltage_V=voltage_V,
                    current_mA=current_mA,
                )
            )
    finally:
        workbook.close()

    samples.sort(key=lambda sample: sample.measurement_time_s)
    if not samples:
        raise RuntimeError("В XLSX стабильности нет измеренных точек для видео.")
    return samples


def load_video_sync(video_path: Path | str) -> Dict[str, Any]:
    """Load the sidecar written when a camera video is downloaded."""

    video = Path(video_path)
    sync_path = video.with_name(f"{video.stem}_sync.json")
    if not sync_path.exists():
        raise RuntimeError(f"Не найден файл синхронизации видео: {sync_path.name}")
    try:
        metadata = json.loads(sync_path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError) as exc:
        raise RuntimeError(f"Не удалось прочитать файл синхронизации {sync_path.name}.") from exc
    if not metadata.get("measurement_sync"):
        raise RuntimeError(str(metadata.get("sync_error") or "Видео не связано с измерением стабильности."))
    if metadata.get("measurement_time_at_video_start_s") is None:
        raise RuntimeError("В файле синхронизации отсутствует смещение времени измерения.")
    return metadata


def _probe_video(ffmpeg_executable: Path, video_path: Path) -> Tuple[int, int, float]:
    result = subprocess.run(
        [str(ffmpeg_executable), "-hide_banner", "-i", str(video_path)],
        capture_output=True,
        text=True,
        encoding="utf-8",
        errors="replace",
        check=False,
    )
    details = "\n".join((result.stdout, result.stderr))
    video_line = next((line for line in details.splitlines() if "Video:" in line), "")
    size_match = re.search(r"\b(\d{2,5})x(\d{2,5})\b", video_line)
    duration_match = re.search(r"Duration:\s*(\d+):(\d+):(\d+(?:\.\d+)?)", details)
    if not size_match or not duration_match:
        raise RuntimeError("FFmpeg не смог определить размер или длительность исходного видео.")
    width, height = (int(value) for value in size_match.groups())
    hours, minutes = (int(value) for value in duration_match.groups()[:2])
    seconds = float(duration_match.group(3))
    duration_s = hours * 3600.0 + minutes * 60.0 + seconds
    if width <= 0 or height <= 0 or duration_s <= 0:
        raise RuntimeError("Исходное видео имеет некорректный размер или длительность.")
    return width, height, duration_s


def _ass_timestamp(seconds: float) -> str:
    centiseconds = max(0, int(round(float(seconds) * 100.0)))
    hours, remainder = divmod(centiseconds, 360_000)
    minutes, remainder = divmod(remainder, 6_000)
    whole_seconds, fraction = divmod(remainder, 100)
    return f"{hours}:{minutes:02d}:{whole_seconds:02d}.{fraction:02d}"


def _ass_text(value: str) -> str:
    return str(value).replace("\\", r"\\").replace("{", r"\{").replace("}", r"\}")


def build_telemetry_intervals(
    samples: Sequence[StabilityTelemetrySample],
    measurement_time_at_video_start_s: float,
    video_duration_s: float,
) -> List[Tuple[float, float, Optional[StabilityTelemetrySample]]]:
    """Map measurement samples to video-time intervals using last-value hold."""

    duration = max(0.0, float(video_duration_s))
    offset = float(measurement_time_at_video_start_s)
    points = sorted(
        ((sample.measurement_time_s - offset, sample) for sample in samples),
        key=lambda item: item[0],
    )
    intervals: List[Tuple[float, float, Optional[StabilityTelemetrySample]]] = []
    active: Optional[StabilityTelemetrySample] = None
    cursor = 0.0
    for video_time, sample in points:
        if video_time <= 0.0:
            active = sample
            continue
        if video_time >= duration:
            break
        if video_time > cursor:
            intervals.append((cursor, video_time, active))
        active = sample
        cursor = video_time
    if cursor < duration:
        intervals.append((cursor, duration, active))
    return [(start, end, sample) for start, end, sample in intervals if end - start >= 0.005]


def _write_ass_file(
    path: Path,
    width: int,
    output_height: int,
    bar_height: int,
    pixel_id: str,
    intervals: Iterable[Tuple[float, float, Optional[StabilityTelemetrySample]]],
    current_limit_video_time_s: Optional[float] = None,
) -> None:
    font_size = max(18, min(34, int(round(bar_height * 0.27))))
    lines = [
        "[Script Info]",
        "ScriptType: v4.00+",
        f"PlayResX: {width}",
        f"PlayResY: {output_height}",
        "ScaledBorderAndShadow: yes",
        "",
        "[V4+ Styles]",
        (
            "Format: Name, Fontname, Fontsize, PrimaryColour, SecondaryColour, OutlineColour, "
            "BackColour, Bold, Italic, Underline, StrikeOut, ScaleX, ScaleY, Spacing, Angle, "
            "BorderStyle, Outline, Shadow, Alignment, MarginL, MarginR, MarginV, Encoding"
        ),
        (
            f"Style: Telemetry,Arial,{font_size},&H00FFFFFF,&H00FFFFFF,&H00000000,&H00000000,"
            f"-1,0,0,0,100,100,0,0,1,0,0,8,12,12,{max(6, int(bar_height * 0.08))},1"
        ),
        "",
        "[Events]",
        "Format: Layer, Start, End, Style, Name, MarginL, MarginR, MarginV, Effect, Text",
    ]
    safe_pixel = _ass_text(pixel_id or "—")
    limit_time = float(current_limit_video_time_s) if current_limit_video_time_s is not None else None
    for start, end, sample in intervals:
        if sample is None:
            text = f"Стабильность · {safe_pixel}\\NОжидание первой точки измерения"
        else:
            warning = " · ЛИМИТ ТОКА" if limit_time is not None and end > limit_time else ""
            text = (
                f"Стабильность · {safe_pixel}{warning}\\N"
                f"t = {sample.measurement_time_s:.1f} с    "
                f"U = {sample.voltage_V:.3f} В    I = {sample.current_mA:.3f} мА"
            )
        lines.append(
            f"Dialogue: 0,{_ass_timestamp(start)},{_ass_timestamp(end)},Telemetry,,0,0,0,,{text}"
        )
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def create_stability_telemetry_video(
    video_path: Path | str,
    stability_workbook_path: Path | str,
) -> Path:
    """Create ``*_telemetry.mp4`` with a separate top bar and preserve the source MP4."""

    source = Path(video_path)
    if not source.exists():
        raise RuntimeError(f"Исходное видео не найдено: {source}")
    sync = load_video_sync(source)
    samples = read_stability_telemetry(stability_workbook_path)
    ffmpeg_executable = find_ffmpeg_executable()
    width, height, actual_duration_s = _probe_video(ffmpeg_executable, source)
    bar_height = max(TELEMETRY_BAR_MIN_HEIGHT, int(round(height * TELEMETRY_BAR_HEIGHT_RATIO)))
    if bar_height % 2:
        bar_height += 1
    output_height = height + bar_height
    intervals = build_telemetry_intervals(
        samples,
        float(sync["measurement_time_at_video_start_s"]),
        actual_duration_s,
    )
    if not intervals:
        raise RuntimeError("Не удалось сопоставить точки стабильности со временем видео.")

    current_limit_video_time_s = next(
        (
            float(event["video_time_s"])
            for event in sync.get("events") or []
            if event.get("event") == "current_limit_or_breakdown" and event.get("video_time_s") is not None
        ),
        None,
    )
    output = source.with_name(f"{source.stem}{TELEMETRY_VIDEO_SUFFIX}.mp4")
    temporary_output = output.with_name(f"{output.stem}.part.mp4")
    with tempfile.TemporaryDirectory(prefix="oled-telemetry-") as temporary_folder:
        ass_path = Path(temporary_folder) / "telemetry.ass"
        _write_ass_file(
            ass_path,
            width,
            output_height,
            bar_height,
            str(sync.get("pixel_id") or ""),
            intervals,
            current_limit_video_time_s=current_limit_video_time_s,
        )
        filter_graph = (
            f"pad=width=iw:height=ih+{bar_height}:x=0:y={bar_height}:color=0x111827,"
            "ass=filename=telemetry.ass"
        )
        command = [
            str(ffmpeg_executable),
            "-y",
            "-hide_banner",
            "-loglevel",
            "error",
            "-i",
            str(source.resolve()),
            "-vf",
            filter_graph,
            "-map",
            "0:v:0",
            "-map",
            "0:a?",
            "-c:v",
            "libx264",
            "-preset",
            "veryfast",
            "-crf",
            "20",
            "-pix_fmt",
            "yuv420p",
            "-fps_mode",
            "vfr",
            "-c:a",
            "copy",
            "-movflags",
            "+faststart",
            str(temporary_output.resolve()),
        ]
        result = subprocess.run(
            command,
            cwd=temporary_folder,
            capture_output=True,
            text=True,
            encoding="utf-8",
            errors="replace",
            check=False,
        )
    if result.returncode != 0 or not temporary_output.exists() or temporary_output.stat().st_size <= 0:
        temporary_output.unlink(missing_ok=True)
        details = (result.stderr or result.stdout or "неизвестная ошибка FFmpeg").strip()
        raise RuntimeError(f"Не удалось создать видео с показаниями:\n{details[-2000:]}")
    os.replace(temporary_output, output)
    return output

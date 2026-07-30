"""Series journal workbook and generated pixel list."""

from __future__ import annotations

import json
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, List, Optional

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font

from oled_app.constants import (
    APP_VERSION,
    JOURNAL_FILE,
    MEASUREMENT_HEADERS,
    MEASUREMENTS_SHEET,
    PIXEL_HEADERS,
    PIXELS_SHEET,
    QUARTERS_SHEET,
    SERIES_SHEET,
)
from oled_app.series.metadata import led_color_label, quarter_code, quarter_description, quarter_led_color
from oled_app.utils import autosize_columns, now_str, relative_to_or_abs, style_header_row, today_iso


@dataclass
class PixelInfo:
    pixel_id: str
    quarter_code: str
    quarter_number: int
    quarter_description: str
    led_color: str
    substrate_number: int
    pixel_number: int


def generate_pixels(config_or_quarter_names: Dict[str, str]) -> List[PixelInfo]:
    pixels: List[PixelInfo] = []
    if any(key in config_or_quarter_names for key in ("quarter_names", "quarter_bases", "quarter_led_colors")):
        config = config_or_quarter_names
    else:
        config = {"quarter_names": config_or_quarter_names}
    for q in range(1, 5):
        code = quarter_code(config, q)
        description = quarter_description(config, q)
        led_color = quarter_led_color(config, q)
        for substrate in range(1, 4):
            for pix in range(1, 5):
                pixel_id = f"{code}{q}_{substrate}_{pix}"
                pixels.append(PixelInfo(pixel_id, code, q, description, led_color, substrate, pix))
    return pixels


class SeriesJournal:
    def __init__(self, series_folder: Path, config: Dict):
        self.series_folder = Path(series_folder)
        self.config = config
        self.path = self.series_folder / JOURNAL_FILE

    def initialize_or_update(self):
        if self.path.exists():
            wb = load_workbook(self.path)
        else:
            wb = Workbook()
            wb.remove(wb.active)

        self._ensure_series_sheet(wb)
        self._ensure_quarters_sheet(wb)
        self._ensure_pixels_sheet(wb)
        self._ensure_measurements_sheet(wb)

        wb.save(self.path)
        wb.close()

    def _ensure_series_sheet(self, wb: Workbook):
        if SERIES_SHEET in wb.sheetnames:
            ws = wb[SERIES_SHEET]
            ws.delete_rows(1, ws.max_row)
        else:
            ws = wb.create_sheet(SERIES_SHEET, 0)

        ws["A1"] = "OLED series journal"
        ws["A1"].font = Font(bold=True, size=14)
        rows = [
            ("App version", APP_VERSION),
            ("Created at", self.config.get("created_at", "")),
            ("Deposition date", self.config.get("deposition_date", "")),
            ("Keyword", self.config.get("keyword", "")),
            ("Series folder", str(self.series_folder.resolve())),
            ("Naming rule", "{quarter_code}{quarter_number}_{substrate_number}_{pixel_number}"),
            ("Example", "CR1_2_3"),
        ]
        for idx, (k, v) in enumerate(rows, start=3):
            ws.cell(row=idx, column=1, value=k).font = Font(bold=True)
            ws.cell(row=idx, column=2, value=v)
        autosize_columns(ws)

    def _ensure_quarters_sheet(self, wb: Workbook):
        if QUARTERS_SHEET in wb.sheetnames:
            ws = wb[QUARTERS_SHEET]
            ws.delete_rows(1, ws.max_row)
        else:
            ws = wb.create_sheet(QUARTERS_SHEET)

        headers = ["Quarter number", "Quarter code/name", "LED color", "Short description", "Generated pixel prefix example"]
        ws.append(headers)
        style_header_row(ws, 1, 1, len(headers))
        for q in range(1, 5):
            code = quarter_code(self.config, q)
            color = quarter_led_color(self.config, q)
            ws.append([q, code, led_color_label(color), quarter_description(self.config, q), f"{code}{q}_1_1"])
        autosize_columns(ws)

    def _ensure_pixels_sheet(self, wb: Workbook):
        if PIXELS_SHEET in wb.sheetnames:
            ws = wb[PIXELS_SHEET]
            existing = self._read_sheet_as_dicts(ws)
            existing_by_id = {row.get("Pixel ID"): row for row in existing if row.get("Pixel ID")}
            ws.delete_rows(1, ws.max_row)
        else:
            ws = wb.create_sheet(PIXELS_SHEET)
            existing_by_id = {}

        ws.append(PIXEL_HEADERS)
        style_header_row(ws, 1, 1, len(PIXEL_HEADERS))

        pixels = generate_pixels(self.config)
        for p in pixels:
            old = existing_by_id.get(p.pixel_id, {})
            ws.append([
                p.pixel_id,
                p.quarter_code,
                p.quarter_number,
                p.quarter_description,
                led_color_label(p.led_color),
                p.substrate_number,
                p.pixel_number,
                old.get("Last status", "UNKNOWN"),
                old.get("Opening voltage (V)", ""),
                old.get("Last IVL date", ""),
                old.get("Last IVL file", ""),
                old.get("Last IVL max current (mA)", ""),
                old.get("Last IVL max photodiode (uA)", ""),
                bool(old.get("Spectrum priority", False)),
                old.get("Last spectrum date", ""),
                old.get("Last spectrum file", ""),
                old.get("Last spectrum peak count", ""),
                old.get("Last spectrum peaks nm", ""),
                old.get("Last spectrum max intensity (counts/s)", ""),
                old.get("Last stability date", ""),
                old.get("Last stability file", ""),
                old.get("Last updated", ""),
            ])
        ws.freeze_panes = "A2"
        autosize_columns(ws, max_width=38)

    def _ensure_measurements_sheet(self, wb: Workbook):
        if MEASUREMENTS_SHEET not in wb.sheetnames:
            ws = wb.create_sheet(MEASUREMENTS_SHEET)
            ws.append(MEASUREMENT_HEADERS)
            style_header_row(ws, 1, 1, len(MEASUREMENT_HEADERS))
            ws.freeze_panes = "A2"
            autosize_columns(ws, max_width=55)
        else:
            ws = wb[MEASUREMENTS_SHEET]
            if ws.max_row == 0 or ws.cell(row=1, column=1).value != MEASUREMENT_HEADERS[0]:
                ws.delete_rows(1, ws.max_row)
                ws.append(MEASUREMENT_HEADERS)
                style_header_row(ws, 1, 1, len(MEASUREMENT_HEADERS))

    @staticmethod
    def _read_sheet_as_dicts(ws) -> List[Dict]:
        headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
        rows = []
        for r in range(2, ws.max_row + 1):
            row = {}
            has_value = False
            for c, header in enumerate(headers, start=1):
                val = ws.cell(row=r, column=c).value
                if val not in (None, ""):
                    has_value = True
                row[header] = val
            if has_value:
                rows.append(row)
        return rows

    def list_pixels(self) -> List[Dict]:
        wb = load_workbook(self.path, data_only=True)
        ws = wb[PIXELS_SHEET]
        rows = self._read_sheet_as_dicts(ws)
        wb.close()
        return rows

    def list_measurements(self) -> List[Dict]:
        if not self.path.exists():
            return []
        wb = load_workbook(self.path, data_only=True)
        if MEASUREMENTS_SHEET not in wb.sheetnames:
            wb.close()
            return []
        ws = wb[MEASUREMENTS_SHEET]
        rows = self._read_sheet_as_dicts(ws)
        wb.close()
        return rows

    def get_pixel(self, pixel_id: str) -> Optional[Dict]:
        for row in self.list_pixels():
            if row.get("Pixel ID") == pixel_id:
                return row
        return None

    def set_spectrum_priority(self, pixel_id: str, enabled: bool) -> int:
        return self.set_spectrum_priorities([pixel_id], enabled)

    def set_spectrum_priorities(self, pixel_ids: List[str], enabled: bool) -> int:
        """Update the spectrum queue for unmeasured pixels and return the changed count."""

        selected_ids = {str(pixel_id) for pixel_id in pixel_ids if pixel_id}
        if not selected_ids:
            return 0
        wb = load_workbook(self.path)
        try:
            ws = wb[PIXELS_SHEET]
            headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
            columns = {header: index + 1 for index, header in enumerate(headers)}
            if "Spectrum priority" not in columns:
                column = ws.max_column + 1
                ws.cell(row=1, column=column, value="Spectrum priority")
                columns["Spectrum priority"] = column
                style_header_row(ws, 1, column, column)
            changed = 0
            for row in range(2, ws.max_row + 1):
                pixel_id = str(ws.cell(row=row, column=columns["Pixel ID"]).value or "")
                if pixel_id not in selected_ids:
                    continue
                if "Last spectrum file" in columns and ws.cell(
                    row=row,
                    column=columns["Last spectrum file"],
                ).value:
                    continue
                priority_cell = ws.cell(row=row, column=columns["Spectrum priority"])
                if bool(priority_cell.value) == bool(enabled):
                    continue
                priority_cell.value = bool(enabled)
                changed += 1
            if changed:
                wb.save(self.path)
            return changed
        finally:
            wb.close()

    def has_any_ivl(self) -> bool:
        if not self.path.exists():
            return False
        wb = load_workbook(self.path, data_only=True)
        if MEASUREMENTS_SHEET not in wb.sheetnames:
            wb.close()
            return False
        ws = wb[MEASUREMENTS_SHEET]
        headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
        try:
            type_col = headers.index("Type") + 1
        except ValueError:
            wb.close()
            return False
        for r in range(2, ws.max_row + 1):
            if ws.cell(row=r, column=type_col).value == "IVL":
                wb.close()
                return True
        wb.close()
        return False

    def update_after_measurement(
        self,
        measurement_type: str,
        pixel_id: str,
        status: str,
        file_path: Optional[Path],
        params: Dict,
        notes: str = "",
        opening_voltage: Optional[float] = None,
        max_current_mA: Optional[float] = None,
        max_photo_uA: Optional[float] = None,
        spectrum_peak_count: Optional[int] = None,
        spectrum_peaks_nm: str = "",
        spectrum_max_intensity: Optional[float] = None,
    ):
        wb = load_workbook(self.path)
        ws_pixels = wb[PIXELS_SHEET]
        ws_meas = wb[MEASUREMENTS_SHEET]

        rel_file = relative_to_or_abs(file_path, self.series_folder) if file_path else ""
        date_text = now_str()
        day_text = today_iso()

        ws_meas.append([
            date_text,
            day_text,
            measurement_type,
            pixel_id,
            status,
            rel_file,
            json.dumps(params, ensure_ascii=False),
            notes,
        ])

        headers = [ws_pixels.cell(row=1, column=c).value for c in range(1, ws_pixels.max_column + 1)]
        col = {h: i + 1 for i, h in enumerate(headers)}
        row_idx = None
        for r in range(2, ws_pixels.max_row + 1):
            if ws_pixels.cell(row=r, column=col["Pixel ID"]).value == pixel_id:
                row_idx = r
                break

        if row_idx:
            if measurement_type in {"IVL", "STABILITY"}:
                ws_pixels.cell(row=row_idx, column=col["Last status"], value=status)
            elif measurement_type == "SPECTRUM" and str(status).upper() in {
                "NEEDS_REVIEW",
                "CURRENT_LIMIT",
                "NO_CONTACT",
                "BURNED",
                "FAILED",
            }:
                ws_pixels.cell(row=row_idx, column=col["Last status"], value=status)
            ws_pixels.cell(row=row_idx, column=col["Last updated"], value=date_text)
            if opening_voltage is not None:
                ws_pixels.cell(row=row_idx, column=col["Opening voltage (V)"], value=float(opening_voltage))

            if measurement_type == "IVL":
                ws_pixels.cell(row=row_idx, column=col["Last IVL date"], value=date_text)
                ws_pixels.cell(row=row_idx, column=col["Last IVL file"], value=rel_file)
                if max_current_mA is not None:
                    ws_pixels.cell(row=row_idx, column=col["Last IVL max current (mA)"], value=float(max_current_mA))
                if max_photo_uA is not None:
                    ws_pixels.cell(row=row_idx, column=col["Last IVL max photodiode (uA)"], value=float(max_photo_uA))
            elif measurement_type == "SPECTRUM":
                if "Spectrum priority" in col:
                    ws_pixels.cell(row=row_idx, column=col["Spectrum priority"], value=False)
                if rel_file:
                    ws_pixels.cell(row=row_idx, column=col["Last spectrum date"], value=date_text)
                    ws_pixels.cell(row=row_idx, column=col["Last spectrum file"], value=rel_file)
                    if spectrum_peak_count is not None and "Last spectrum peak count" in col:
                        ws_pixels.cell(row=row_idx, column=col["Last spectrum peak count"], value=int(spectrum_peak_count))
                    if spectrum_peaks_nm and "Last spectrum peaks nm" in col:
                        ws_pixels.cell(row=row_idx, column=col["Last spectrum peaks nm"], value=str(spectrum_peaks_nm))
                    if spectrum_max_intensity is not None and "Last spectrum max intensity (counts/s)" in col:
                        ws_pixels.cell(row=row_idx, column=col["Last spectrum max intensity (counts/s)"], value=float(spectrum_max_intensity))
            elif measurement_type == "STABILITY":
                ws_pixels.cell(row=row_idx, column=col["Last stability date"], value=date_text)
                ws_pixels.cell(row=row_idx, column=col["Last stability file"], value=rel_file)

        for ws in [ws_pixels, ws_meas]:
            autosize_columns(ws, max_width=55)
        wb.save(self.path)
        wb.close()

"""Spectrum measurement workflow for the modular OLED application."""

from __future__ import annotations

import gc
import time
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Callable, Dict, Iterable, List, Optional, Tuple

import numpy as np
from openpyxl import Workbook
from openpyxl.chart import Reference, ScatterChart, Series
from openpyxl.chart.axis import ChartLines
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from oled_app.hardware import prepare_hardware_environment, safe_shutdown_smu
from oled_app.utils import (
    autosize_columns,
    current_density_mA_cm2,
    luminance_cd_m2,
    now_str,
    safe_filename,
    style_header_row,
    timestamp_for_file,
)


@dataclass
class SpectrumParams:
    com_port: str = "COM3"
    voltage_start: float = 2.0
    voltage_end: float = 5.0
    voltage_step: float = 0.1
    opening_voltage: Optional[float] = None
    voltage_start_source: str = "opening"
    current_limit_mA: float = 6.0
    photodiode_bias_V: float = -5.0
    photodiode_range: int = 4
    target_intensity: float = 40000.0
    intensity_min: float = 20000.0
    intensity_max: float = 55000.0
    saturation_level: float = 60000.0
    min_peak_width_nm: float = 15.0
    max_peak_width_nm: float = 150.0
    t_int_initial_s: float = 0.01
    t_int_min_s: float = 0.001
    t_int_max_s: float = 10.0
    discard_first_scan_after_tint_change: bool = True
    kp: float = 0.3
    ki: float = 0.05
    max_iterations: int = 20
    tolerance: float = 0.05
    led_type: str = "auto"
    peak_search_mode_for_tint: str = "auto"
    settle_time_voltage_s: float = 0.1
    settle_time_spectrum_s: float = 0.05
    dark_spectrum_enabled: bool = False
    dark_spectrum_scans: int = 3
    baseline_correction_enabled: bool = True
    peak_detection_enabled: bool = False
    pixel_area_mm2: float = 1.0
    luminance_cd_m2_per_uA: float = 1.0

    def as_dict(self) -> Dict[str, Any]:
        return self.__dict__.copy()


class SpectrumHelper:
    def __init__(self, params: SpectrumParams, log: Callable[[str], None]):
        self.params = params
        self.log = log
        self._last_integration_time_us: Optional[int] = None
        self.last_optimization_started_saturated = False
        self.last_optimization_started_saturated_at_10ms = False
        self.adaptive_initial_time_enabled = False

    def init_spectrometer(self):
        import seabreeze

        seabreeze.use("cseabreeze")
        from seabreeze.spectrometers import Spectrometer, list_devices

        devices = list_devices()
        if not devices:
            raise RuntimeError("Спектрометр не найден. Проверь USB и драйвер SeaBreeze.")
        spec = Spectrometer(devices[0])
        wavelengths = spec.wavelengths()
        self.log(f"Спектрометр: {spec.model}; диапазон {wavelengths[0]:.1f}-{wavelengths[-1]:.1f} нм")
        return spec

    def set_integration_time(self, spec, integration_time_s: float) -> Tuple[float, bool]:
        p = self.params
        integration_time_s = float(np.clip(integration_time_s, p.t_int_min_s, p.t_int_max_s))
        integration_time_us = max(1, int(round(integration_time_s * 1e6)))
        changed = self._last_integration_time_us != integration_time_us
        spec.integration_time_micros(integration_time_us)
        self._last_integration_time_us = integration_time_us
        return integration_time_us / 1e6, changed

    def get_spectrum(self, spec, integration_time_s: float, discard_stale_after_change: bool = True):
        actual_t, changed = self.set_integration_time(spec, integration_time_s)
        wavelengths = spec.wavelengths()
        if discard_stale_after_change and self.params.discard_first_scan_after_tint_change and changed:
            time.sleep(self.params.settle_time_spectrum_s)
            try:
                _ = spec.intensities()
            except Exception as exc:
                self.log(f"  Не удалось сбросить первый спектр после смены T_int: {exc}")
        time.sleep(self.params.settle_time_spectrum_s)
        intensities = spec.intensities().astype(np.float64)
        return wavelengths, intensities, actual_t

    def get_dark_spectrum(self, spec, integration_time_s: float):
        spectra = []
        wavelengths = None
        actual_t = integration_time_s
        for i in range(self.params.dark_spectrum_scans):
            wavelengths, intensities, actual_t = self.get_spectrum(
                spec,
                integration_time_s,
                discard_stale_after_change=(i == 0),
            )
            spectra.append(intensities)
            time.sleep(0.05)
        return wavelengths, np.mean(spectra, axis=0), actual_t

    @staticmethod
    def smooth_array(values: np.ndarray, window: int = 9) -> np.ndarray:
        values = np.asarray(values, dtype=np.float64)
        if values.size < 3:
            return values.copy()
        window = int(max(3, min(window, values.size if values.size % 2 == 1 else values.size - 1)))
        if window < 3:
            return values.copy()
        kernel = np.ones(window, dtype=np.float64) / window
        padded = np.pad(values, (window // 2, window // 2), mode="edge")
        return np.convolve(padded, kernel, mode="valid")

    def estimate_baseline(self, wavelengths: np.ndarray, intensities: np.ndarray) -> Tuple[np.ndarray, Tuple[float, float], float]:
        wavelengths = np.asarray(wavelengths, dtype=np.float64)
        intensities = np.asarray(intensities, dtype=np.float64)
        if intensities.size < 3:
            value = float(np.nanmean(intensities)) if intensities.size else 0.0
            return np.full_like(intensities, value), (float("nan"), float("nan")), value

        finite = np.where(np.isfinite(wavelengths) & np.isfinite(intensities))[0]
        if finite.size < 3:
            value = float(np.nanmean(intensities[finite])) if finite.size else 0.0
            return np.full_like(intensities, value), (float("nan"), float("nan")), value

        n = int(finite.size)
        window = int(np.clip(round(n * 0.06), 8, 40))
        window = min(window, n)
        best_start = 0
        best_score = float("inf")
        for start in range(0, n - window + 1):
            idx = finite[start:start + window]
            x = wavelengths[idx]
            y = intensities[idx]
            if np.any(~np.isfinite(x)) or np.any(~np.isfinite(y)):
                continue
            span = float(max(x[-1] - x[0], 1e-9))
            slope = float(abs(y[-1] - y[0]) / span)
            scatter = float(np.nanstd(y))
            level_penalty = float(max(np.nanmean(y), 0.0)) * 1e-6
            score = scatter + slope * span * 0.25 + level_penalty
            if score < best_score:
                best_score = score
                best_start = start

        baseline_idx = finite[best_start:best_start + window]
        baseline_value = float(np.nanmean(intensities[baseline_idx]))
        baseline = np.full_like(intensities, baseline_value)
        wl_range = (float(wavelengths[baseline_idx[0]]), float(wavelengths[baseline_idx[-1]]))
        return baseline, wl_range, baseline_value

    def find_peaks_by_derivatives(self, wavelengths: np.ndarray, intensities: np.ndarray) -> List[Dict[str, float]]:
        wavelengths = np.asarray(wavelengths, dtype=np.float64)
        intensities = np.asarray(intensities, dtype=np.float64)
        if intensities.size < 5:
            return []

        smooth = self.smooth_array(intensities, window=9)
        d1 = np.gradient(smooth, wavelengths)
        d2 = np.gradient(d1, wavelengths)
        noise = float(np.nanmedian(np.abs(smooth - self.smooth_array(smooth, window=21)))) if smooth.size >= 21 else 0.0
        threshold = max(float(np.nanmax(smooth)) * 0.05, noise * 5.0, 1e-9)
        peaks: List[Dict[str, float]] = []

        for i in range(1, smooth.size - 1):
            sign_change = d1[i - 1] > 0 and d1[i + 1] < 0
            local_max = smooth[i] >= smooth[i - 1] and smooth[i] >= smooth[i + 1]
            concave_down = d2[i] < 0
            if sign_change and local_max and concave_down and smooth[i] >= threshold:
                half = smooth[i] / 2.0
                left = i
                while left > 0 and smooth[left] >= half:
                    left -= 1
                right = i
                while right < smooth.size - 1 and smooth[right] >= half:
                    right += 1
                peaks.append(
                    {
                        "wavelength_nm": float(wavelengths[i]),
                        "intensity": float(smooth[i]),
                        "fwhm_nm": float(max(wavelengths[right] - wavelengths[left], 0.0)),
                    }
                )

        peaks.sort(key=lambda item: item["intensity"], reverse=True)
        return peaks[:8]

    def process_spectrum(
        self,
        wavelengths: np.ndarray,
        spectrum: np.ndarray,
        dark: Optional[np.ndarray],
        integration_time_s: float,
    ) -> Dict[str, Any]:
        raw = np.asarray(spectrum, dtype=np.float64)
        dark_corrected = raw.copy()
        if dark is not None and self.params.dark_spectrum_enabled:
            dark_corrected = dark_corrected - np.asarray(dark, dtype=np.float64)
        if self.params.baseline_correction_enabled:
            baseline, baseline_region, baseline_value = self.estimate_baseline(wavelengths, raw)
        else:
            baseline = np.zeros_like(raw)
            baseline_region = (float("nan"), float("nan"))
            baseline_value = 0.0
        baseline_corrected = raw - baseline
        normalized = baseline_corrected / max(float(integration_time_s), 1e-9)
        peaks = self.find_peaks_by_derivatives(wavelengths, normalized) if self.params.peak_detection_enabled else []
        return {
            "raw": raw,
            "dark_corrected": dark_corrected,
            "baseline": baseline,
            "baseline_region": baseline_region,
            "baseline_value": baseline_value,
            "baseline_corrected": baseline_corrected,
            "normalized": normalized,
            "peaks": peaks,
        }

    @staticmethod
    def peak_range(wavelengths, mode: str) -> Tuple[float, float]:
        ranges = {
            "red": (580, 700),
            "green": (480, 580),
            "blue": (400, 500),
            "other": (300, 1000),
            "auto": (380, 780),
            "visible": (380, 780),
            "all": (float(wavelengths[0]), float(wavelengths[-1])),
        }
        return ranges.get(mode, ranges["auto"])

    def find_peak_region(self, wavelengths, intensities, mode: str):
        wl_min, wl_max = self.peak_range(wavelengths, mode)
        mask = (wavelengths >= wl_min) & (wavelengths <= wl_max)
        if not np.any(mask):
            return None, None, None
        roi_wl = wavelengths[mask]
        roi_int = intensities[mask]
        if len(roi_int) == 0 or np.all(~np.isfinite(roi_int)):
            return None, None, None
        idx = int(np.nanargmax(roi_int))
        peak_wl = float(roi_wl[idx])
        peak_int = float(roi_int[idx])
        half = peak_int / 2.0
        above = roi_int >= half
        if np.any(above):
            inds = np.where(above)[0]
            fwhm = float(roi_wl[inds[-1]] - roi_wl[inds[0]])
        else:
            fwhm = 0.0
        return peak_wl, peak_int, fwhm

    def analyze_quality(self, wavelengths, intensities, mode: str):
        p = self.params
        peak_wl, peak_int, fwhm = self.find_peak_region(wavelengths, intensities, mode)
        if peak_int is None:
            return 0.0, 0.0, 0.0, False, True, False, False, "NO_PEAK"
        is_sat = bool(np.any(intensities >= p.saturation_level))
        is_weak = bool(peak_int < p.intensity_min)
        is_wide = bool(fwhm > p.max_peak_width_nm)
        is_narrow = bool(fwhm < p.min_peak_width_nm)
        if is_sat:
            status = "SATURATED"
        elif is_weak:
            status = "TOO_WEAK"
        elif is_wide:
            status = "TOO_WIDE"
        elif is_narrow:
            status = "TOO_NARROW"
        elif p.intensity_min <= peak_int <= p.intensity_max:
            status = "GOOD"
        else:
            status = "OK"
        return peak_int, peak_wl, fwhm, is_sat, is_weak, is_wide, is_narrow, status

    def optimize_integration_time(self, spec):
        p = self.params
        t_int = p.t_int_initial_s
        integral = 0.0
        best = None
        best_score = float("inf")
        self.last_optimization_started_saturated = False
        self.last_optimization_started_saturated_at_10ms = False

        self.log(f"  Подбор T_int: цель {p.target_intensity:.0f} counts, область {p.peak_search_mode_for_tint}")
        for iteration in range(1, p.max_iterations + 1):
            wl, inten, actual_t = self.get_spectrum(spec, t_int)
            peak_int, peak_wl, fwhm, is_sat, is_weak, _, _, status = self.analyze_quality(wl, inten, p.peak_search_mode_for_tint)
            if iteration == 1 and is_sat:
                self.last_optimization_started_saturated = True
                self.last_optimization_started_saturated_at_10ms = actual_t >= 0.0095
            if is_sat:
                score = float("inf")
            elif is_weak:
                score = abs(peak_int - p.target_intensity) * 10.0
            else:
                score = abs(peak_int - p.target_intensity)
            if score < best_score and not is_sat:
                best_score = score
                best = (actual_t, wl.copy(), inten.copy(), peak_int, peak_wl, fwhm, status)

            self.log(f"    {iteration}: T={actual_t*1000:.2f} мс, peak={peak_int:.0f} @ {peak_wl:.1f} нм, {status}")

            if status == "GOOD" or (status == "OK" and abs(peak_int - p.target_intensity) / p.target_intensity < p.tolerance):
                best = (actual_t, wl.copy(), inten.copy(), peak_int, peak_wl, fwhm, status)
                break
            if is_sat and actual_t <= p.t_int_min_s + 1e-4:
                best = (actual_t, wl.copy(), inten.copy(), peak_int, peak_wl, fwhm, status)
                break
            if is_weak and actual_t >= p.t_int_max_s - 1e-3:
                best = (actual_t, wl.copy(), inten.copy(), peak_int, peak_wl, fwhm, status)
                break

            current_int = max(peak_int, 1.0)
            error = (p.target_intensity - current_int) / p.target_intensity
            if abs(error) > 0.9:
                kp = p.kp * 5
                ki = p.ki * 3
            elif abs(error) > 0.7:
                kp = p.kp * 3
                ki = p.ki * 2
            elif abs(error) > 0.5:
                kp = p.kp * 2
                ki = p.ki * 1.5
            else:
                kp = p.kp
                ki = p.ki
            integral += ki * error
            integral = float(np.clip(integral, -2.0, 2.0))
            adjustment = float(np.clip(kp * error + integral, -0.9, 5.0))
            t_int = float(np.clip(actual_t * (1.0 + adjustment), p.t_int_min_s, p.t_int_max_s))

        if best is None:
            return None
        t, wl, inten, _, _, _, _ = best
        dark = None
        if p.dark_spectrum_enabled:
            _, dark, _ = self.get_dark_spectrum(spec, t)
        return t, wl, inten, dark


def create_spectrum_workbook(filename: Path, pixel_id: str, params: SpectrumParams, voltage_array: Iterable[float]) -> Workbook:
    filename.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws_sum = wb.active
    ws_sum.title = "Сводка"
    ws_spec = wb.create_sheet("Спектры")
    ws_norm = wb.create_sheet("Processed counts per s")
    ws_raw = wb.create_sheet("Raw spectra")
    ws_dark = wb.create_sheet("Dark corrected")
    ws_baseline = wb.create_sheet("Baseline")
    ws_desc = wb.create_sheet("Описание полей")

    ws_sum["A1"] = "Спектро-электронное сканирование OLED"
    ws_sum["A1"].font = Font(bold=True, size=14)
    meta = [
        ("Pixel", pixel_id),
        ("Created", now_str()),
        ("Voltage range", f"{params.voltage_start}-{params.voltage_end} В, step {params.voltage_step} В"),
        ("Opening voltage stored (V)", params.opening_voltage if params.opening_voltage is not None else ""),
        ("Voltage start source", params.voltage_start_source),
        ("Current limit", f"{params.current_limit_mA} мА"),
        ("Pixel area (mm^2)", params.pixel_area_mm2),
        ("Luminance conversion (cd/m^2 per uA)", params.luminance_cd_m2_per_uA),
        ("LED_TYPE final", params.led_type),
        ("Peak search for T_int", params.peak_search_mode_for_tint),
        ("Derivative peak detection", "YES" if params.peak_detection_enabled else "NO"),
        ("T_int range", f"{params.t_int_min_s*1000:.2f}-{params.t_int_max_s*1000:.2f} мс"),
        ("Discard first scan after T_int change", "YES" if params.discard_first_scan_after_tint_change else "NO"),
        ("Baseline correction", "YES" if params.baseline_correction_enabled else "NO"),
        ("Saved intensity units", "Спектры: raw counts минус фон; Processed counts per s: то же деленное на T_int"),
    ]
    for idx, (key, value) in enumerate(meta, start=3):
        ws_sum.cell(row=idx, column=1, value=key).font = Font(bold=True)
        ws_sum.cell(row=idx, column=2, value=value)

    headers = [
        "Point",
        "V set (V)",
        "V LED measured (V)",
        "I LED (mA)",
        "J LED (mA/cm^2)",
        "V photodiode measured (V)",
        "I photodiode (uA)",
        "Luminance (cd/m^2)",
        "T_int saved spectrum (ms)",
        "Peak (nm)",
        "Max intensity processed (counts)",
        "FWHM (nm)",
        "Status",
        "Spectra column",
        "Peaks detected",
        "Peaks nm",
        "Background mean raw counts",
        "Background region nm",
    ]
    header_row = 21
    for column, header in enumerate(headers, start=1):
        ws_sum.cell(row=header_row, column=column, value=header)
    style_header_row(ws_sum, header_row, 1, len(headers))
    ws_sum.freeze_panes = f"A{header_row + 1}"

    meta_labels = [
        "Название столбца",
        "Point",
        "V set (V)",
        "V LED measured (V)",
        "I LED (mA)",
        "J LED (mA/cm^2)",
        "V photodiode measured (V)",
        "I photodiode (uA)",
        "Luminance (cd/m^2)",
        "T_int saved spectrum (ms)",
        "Peak (nm)",
        "Max intensity processed (counts/s)",
        "FWHM (nm)",
        "Status",
        "Comment",
        "Peaks detected",
        "Peaks nm",
        "Background mean raw counts",
        "Background region nm",
    ]
    for row, label in enumerate(meta_labels, start=1):
        ws_spec.cell(row=row, column=1, value=label).font = Font(bold=True)
        ws_spec.cell(row=row, column=1).fill = PatternFill("solid", fgColor="E2F0D9")

    data_header_row = 20
    ws_spec.cell(row=data_header_row, column=1, value="Wavelength (nm)").font = Font(bold=True)
    ws_spec.cell(row=data_header_row, column=1).fill = PatternFill("solid", fgColor="D9E1F2")
    for idx, voltage in enumerate(voltage_array, start=1):
        col = idx + 1
        ws_spec.cell(row=1, column=col, value=f"Point {idx}: {float(voltage):.3f} V")
        ws_spec.cell(row=2, column=col, value=idx)
        ws_spec.cell(row=3, column=col, value=float(voltage))
        ws_spec.cell(row=data_header_row, column=col, value=f"Intensity point {idx}, counts corrected")
        ws_spec.cell(row=data_header_row, column=col).font = Font(bold=True)
        ws_spec.cell(row=data_header_row, column=col).fill = PatternFill("solid", fgColor="D9E1F2")

    ws_spec.freeze_panes = f"B{data_header_row + 1}"
    extra_sheets = [
        (ws_norm, "Processed counts per s"),
        (ws_raw, "Raw counts"),
        (ws_dark, "Dark-corrected counts"),
        (ws_baseline, "Background mean counts"),
    ]
    for extra_ws, data_label in extra_sheets:
        for row, label in enumerate(meta_labels, start=1):
            extra_ws.cell(row=row, column=1, value=label).font = Font(bold=True)
            extra_ws.cell(row=row, column=1).fill = PatternFill("solid", fgColor="E2F0D9")
        extra_ws.cell(row=data_header_row, column=1, value="Wavelength (nm)").font = Font(bold=True)
        extra_ws.cell(row=data_header_row, column=1).fill = PatternFill("solid", fgColor="D9E1F2")
        for idx, voltage in enumerate(voltage_array, start=1):
            col = idx + 1
            extra_ws.cell(row=1, column=col, value=f"Point {idx}: {float(voltage):.3f} V")
            extra_ws.cell(row=2, column=col, value=idx)
            extra_ws.cell(row=3, column=col, value=float(voltage))
            extra_ws.cell(row=data_header_row, column=col, value=f"{data_label} point {idx}")
            extra_ws.cell(row=data_header_row, column=col).font = Font(bold=True)
            extra_ws.cell(row=data_header_row, column=col).fill = PatternFill("solid", fgColor="D9E1F2")
        extra_ws.freeze_panes = f"B{data_header_row + 1}"

    ws_desc.append(["Лист", "Поле", "Что означает"])
    style_header_row(ws_desc, 1, 1, 3)
    desc_rows = [
        ("Сводка", "T_int saved spectrum", "Время интегрирования именно того спектра, который записан в столбец."),
        ("Сводка", "Peaks detected / Peaks nm", "Заполняется только если включен флажок поиска пиков производными."),
        ("Спектры", "Intensity", "Raw counts минус среднее значение фона из найденного плоского участка raw-спектра, counts."),
        ("Processed counts per s", "Processed counts/s", "Копия обработанного спектра, деленная на время интегрирования."),
        ("Raw spectra", "Raw counts", "Сырые counts напрямую со спектрометра."),
        ("Dark corrected", "Dark-corrected counts", "Диагностический лист: Raw минус dark spectrum, если dark включен; основная обработка его не использует."),
        ("Baseline", "Background mean counts", "Константный уровень фона: среднее значение найденного плоского участка raw-спектра."),
        ("Спектры", "Строки 1-15", "Метаданные каждого спектра."),
        ("Спектры", "Строка 21+", "Длины волн и интенсивности."),
        ("Важно", "Первый спектр после смены T_int", "Сбрасывается, чтобы не записать старый буфер спектрометра."),
    ]
    for row in desc_rows:
        ws_desc.append(row)

    for ws in wb.worksheets:
        autosize_columns(ws, max_width=45)
    wb.save(filename)
    return wb


def run_spectrum_measurement(
    pixel_id: str,
    output_dir: Path,
    params: SpectrumParams,
    log: Callable[[str], None],
    app_settings: Optional[Dict[str, Any]] = None,
    progress_callback: Optional[
        Callable[[int, float, float, np.ndarray, np.ndarray, np.ndarray, List[Dict[str, float]], str], None]
    ] = None,
) -> Dict[str, Any]:
    prepare_hardware_environment(pixel_id, app_settings, log)
    import xtralien

    helper = SpectrumHelper(params, log)
    spec = helper.init_spectrometer()

    output_dir.mkdir(parents=True, exist_ok=True)
    voltage_array = np.arange(params.voltage_start, params.voltage_end + params.voltage_step / 2, params.voltage_step)
    voltage_array = np.round(voltage_array, 6)
    filename = output_dir / f"SPECTRUM_{safe_filename(pixel_id)}_{timestamp_for_file()}.xlsx"
    wb = create_spectrum_workbook(filename, pixel_id, params, voltage_array)
    ws_sum = wb["Сводка"]
    ws_spec = wb["Спектры"]
    ws_norm = wb["Processed counts per s"]
    ws_raw = wb["Raw spectra"]
    ws_dark = wb["Dark corrected"]
    ws_baseline = wb["Baseline"]
    summary_header_row = 21
    spectra_data_start = 21

    final_status = "FAILED"
    best_spectrum_metrics = {
        "spectrum_peak_count": None,
        "spectrum_peaks_nm": "",
        "spectrum_max_intensity": None,
    }

    with xtralien.Device(params.com_port) as smu:
        try:
            smu.smu1.set.enabled(True, response=0)
            smu.smu2.set.enabled(True, response=0)
            try:
                smu.smu2.set.range(params.photodiode_range, response=0)
            except Exception:
                pass
            smu.smu1.set.voltage(0, response=0)
            smu.smu2.set.voltage(params.photodiode_bias_V, response=0)
            time.sleep(0.3)

            for idx, voltage in enumerate(voltage_array, start=1):
                log(f"\nСпектр {pixel_id}: точка {idx}/{len(voltage_array)}, V={voltage:.3f} В")
                smu.smu1.set.voltage(float(voltage), response=0)
                smu.smu2.set.voltage(params.photodiode_bias_V, response=0)
                time.sleep(params.settle_time_voltage_s)

                v_led, i_led = smu.smu1.measure()[0]
                v_pd, i_pd = smu.smu2.measure()[0]
                i_led_mA = i_led * 1000.0
                i_pd_uA = -i_pd * 1_000_000.0
                j_led = current_density_mA_cm2(i_led_mA, params.pixel_area_mm2)
                lum = luminance_cd_m2(i_pd_uA, params.luminance_cd_m2_per_uA)

                row = summary_header_row + idx
                spectra_col = idx + 1
                if i_led_mA >= params.current_limit_mA:
                    status = "NEEDS_REVIEW"
                    log(f"  Стоп: ток {i_led_mA:.3f} мА >= {params.current_limit_mA:.3f} мА")
                    summary_values = [
                        idx,
                        float(voltage),
                        v_led,
                        i_led_mA,
                        j_led,
                        v_pd,
                        i_pd_uA,
                        lum,
                        "-",
                        "-",
                        "-",
                        "-",
                        status,
                        get_column_letter(spectra_col),
                        0,
                        "",
                        "",
                        "",
                    ]
                    for column, value in enumerate(summary_values, start=1):
                        ws_sum.cell(row=row, column=column, value=value)
                    ws_spec.cell(row=12, column=spectra_col, value=status)
                    final_status = status
                    wb.save(filename)
                    break

                opt = helper.optimize_integration_time(spec)
                if opt is None:
                    status = "FAILED"
                    t_int = params.t_int_initial_s
                    summary_values = [
                        idx,
                        float(voltage),
                        v_led,
                        i_led_mA,
                        j_led,
                        v_pd,
                        i_pd_uA,
                        lum,
                        t_int * 1000,
                        "-",
                        "-",
                        "-",
                        status,
                        get_column_letter(spectra_col),
                        0,
                        "",
                        "",
                        "",
                    ]
                    for column, value in enumerate(summary_values, start=1):
                        ws_sum.cell(row=row, column=column, value=value)
                    final_status = status
                    wb.save(filename)
                    continue

                t_int, wavelengths, spectrum, dark = opt
                processed = helper.process_spectrum(wavelengths, spectrum, dark, t_int)
                spectrum_to_save = processed["baseline_corrected"]
                normalized_to_save = processed["normalized"]
                raw_to_save = processed["raw"]
                dark_to_save = processed["dark_corrected"]
                baseline_to_save = processed["baseline"]
                baseline_value = float(processed.get("baseline_value", 0.0))
                baseline_region = processed.get("baseline_region", (float("nan"), float("nan")))
                baseline_region_text = (
                    f"{float(baseline_region[0]):.2f}-{float(baseline_region[1]):.2f}"
                    if np.all(np.isfinite(np.asarray(baseline_region, dtype=np.float64)))
                    else ""
                )
                peaks = processed["peaks"]

                peak_int, peak_wl, fwhm, _, _, _, _, status = helper.analyze_quality(
                    wavelengths,
                    processed["baseline_corrected"],
                    params.led_type,
                )
                if peaks:
                    peak_int = peaks[0]["intensity"]
                    peak_wl = peaks[0]["wavelength_nm"]
                    fwhm = peaks[0]["fwhm_nm"]
                if status not in {"SATURATED", "FAILED", "NO_PEAK"} and (
                    helper.adaptive_initial_time_enabled or helper.last_optimization_started_saturated_at_10ms
                ):
                    previous_t = float(params.t_int_initial_s)
                    params.t_int_initial_s = float(t_int)
                    helper.adaptive_initial_time_enabled = True
                    if helper.last_optimization_started_saturated_at_10ms:
                        log(
                            f"  Первая проба на 10 мс была saturated; следующее начальное T_int: "
                            f"{params.t_int_initial_s*1000:.2f} мс вместо {previous_t*1000:.2f} мс"
                        )
                peaks_nm = ", ".join(f"{peak['wavelength_nm']:.1f}" for peak in peaks)
                if peak_int and (
                    best_spectrum_metrics["spectrum_max_intensity"] is None
                    or float(peak_int) > float(best_spectrum_metrics["spectrum_max_intensity"])
                ):
                    best_spectrum_metrics = {
                        "spectrum_peak_count": len(peaks),
                        "spectrum_peaks_nm": peaks_nm,
                        "spectrum_max_intensity": float(peak_int),
                    }
                summary_values = [
                    idx,
                    float(voltage),
                    round(float(v_led), 6),
                    round(float(i_led_mA), 6),
                    round(float(j_led), 6) if j_led is not None else "-",
                    round(float(v_pd), 6),
                    round(float(i_pd_uA), 6),
                    round(float(lum), 6) if lum is not None else "-",
                    round(float(t_int) * 1000, 3),
                    round(float(peak_wl), 3) if peak_wl else "-",
                    round(float(peak_int), 1) if peak_int else "-",
                    round(float(fwhm), 3) if fwhm else "-",
                    status,
                    get_column_letter(spectra_col),
                    len(peaks),
                    peaks_nm,
                    round(baseline_value, 3),
                    baseline_region_text,
                ]
                for column, value in enumerate(summary_values, start=1):
                    ws_sum.cell(row=row, column=column, value=value)

                meta_values = [
                    f"Point {idx}: {float(voltage):.3f} V",
                    idx,
                    float(voltage),
                    round(float(v_led), 6),
                    round(float(i_led_mA), 6),
                    round(float(j_led), 6) if j_led is not None else "-",
                    round(float(v_pd), 6),
                    round(float(i_pd_uA), 6),
                    round(float(lum), 6) if lum is not None else "-",
                    round(float(t_int) * 1000, 3),
                    round(float(peak_wl), 3) if peak_wl else "-",
                    round(float(peak_int), 1) if peak_int else "-",
                    round(float(fwhm), 3) if fwhm else "-",
                    status,
                    "OK: spectrum saved",
                    len(peaks),
                    peaks_nm,
                    round(baseline_value, 3),
                    baseline_region_text,
                ]
                for meta_row, value in enumerate(meta_values, start=1):
                    ws_spec.cell(row=meta_row, column=spectra_col, value=value)
                    ws_norm.cell(row=meta_row, column=spectra_col, value=value)
                    ws_raw.cell(row=meta_row, column=spectra_col, value=value)
                    ws_dark.cell(row=meta_row, column=spectra_col, value=value)
                    ws_baseline.cell(row=meta_row, column=spectra_col, value=value)

                for data_row, wavelength in enumerate(wavelengths, start=spectra_data_start):
                    if ws_spec.cell(row=data_row, column=1).value is None:
                        ws_spec.cell(row=data_row, column=1, value=round(float(wavelength), 2))
                    if ws_norm.cell(row=data_row, column=1).value is None:
                        ws_norm.cell(row=data_row, column=1, value=round(float(wavelength), 2))
                    if ws_raw.cell(row=data_row, column=1).value is None:
                        ws_raw.cell(row=data_row, column=1, value=round(float(wavelength), 2))
                    if ws_dark.cell(row=data_row, column=1).value is None:
                        ws_dark.cell(row=data_row, column=1, value=round(float(wavelength), 2))
                    if ws_baseline.cell(row=data_row, column=1).value is None:
                        ws_baseline.cell(row=data_row, column=1, value=round(float(wavelength), 2))
                for data_row, intensity in enumerate(spectrum_to_save, start=spectra_data_start):
                    ws_spec.cell(row=data_row, column=spectra_col, value=round(float(intensity), 3))
                for data_row, intensity in enumerate(normalized_to_save, start=spectra_data_start):
                    ws_norm.cell(row=data_row, column=spectra_col, value=round(float(intensity), 3))
                for data_row, intensity in enumerate(raw_to_save, start=spectra_data_start):
                    ws_raw.cell(row=data_row, column=spectra_col, value=round(float(intensity), 3))
                for data_row, intensity in enumerate(dark_to_save, start=spectra_data_start):
                    ws_dark.cell(row=data_row, column=spectra_col, value=round(float(intensity), 3))
                for data_row, intensity in enumerate(baseline_to_save, start=spectra_data_start):
                    ws_baseline.cell(row=data_row, column=spectra_col, value=round(float(intensity), 3))

                final_status = status
                if progress_callback is not None:
                    progress_callback(idx, float(voltage), float(t_int), wavelengths, processed["raw"], spectrum_to_save, peaks, status)
                log(f"  Сохранено: T_int={t_int*1000:.2f} мс, peak={peak_int:.0f} @ {peak_wl:.1f} нм, {status}")
                if idx % 3 == 0 or idx == len(voltage_array):
                    wb.save(filename)
                    gc.collect()
        finally:
            safe_shutdown_smu(smu)

    for ws in wb.worksheets:
        autosize_columns(ws, max_width=45)
    wb.save(filename)
    wb.close()
    return {"file": filename, "status": final_status, **best_spectrum_metrics}

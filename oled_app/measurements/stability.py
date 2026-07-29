"""Stability measurement workflow for the modular OLED application."""

from __future__ import annotations

import math
import threading
import time
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Callable, Dict, List, Optional, Tuple

import numpy as np
from openpyxl import load_workbook

from oled_app.hardware import prepare_hardware_environment, shutdown_smu_with_reconnect
from oled_app.measurements.raw_io import RawCsvWriter, cleanup_raw_files, raw_csv_path
from oled_app.processing.stability_results import (
    STABILITY_RAW_HEADERS,
    build_stability_workbook_from_raw_csv,
    create_stability_workbook as write_stability_workbook,
    save_stability_chart as write_stability_chart,
    update_stability_status as write_stability_status,
)
from oled_app.utils import (
    as_float_or_none,
    current_density_mA_cm2,
    luminance_cd_m2_at_voltage,
    now_str,
    safe_filename,
    timestamp_for_file,
)


@dataclass
class StabilityParams:
    com_port: str = "COM3"
    control_mode: str = "current"
    current_setpoint_mA: float = 3.5
    voltage_setpoint_V: float = 3.5
    voltage_start: float = 3.5
    voltage_limit: float = 5.0
    current_limit_mA: float = 10.0
    voltage_step_max: float = 0.02
    current_control_kp: float = 0.01
    measurement_time_s: float = 86400.0
    sample_interval_s: float = 1.0
    autosave_interval_s: float = 600.0
    photodiode_bias_V: float = -5.0
    photodiode_threshold_uA: float = 0.1
    photodiode_range: int = 4
    pixel_area_mm2: float = 1.0
    luminance_cd_m2_per_uA: float = 1.0
    geometric_coefficient: float = 1.0
    luminance_calibration_model: Optional[Dict[str, Any]] = None

    def as_dict(self) -> Dict[str, Any]:
        return self.__dict__.copy()


class StabilitySetpointController:
    """Thread-safe mutable target used by the live stability controls."""

    MODES = {"current", "voltage"}

    def __init__(self, mode: str, target: float, maximum: Optional[float] = None):
        normalized_mode = str(mode or "").strip().lower()
        if normalized_mode not in self.MODES:
            raise ValueError("Режим стабильности должен быть current или voltage.")
        self.mode = normalized_mode
        self.maximum = float(maximum) if maximum is not None else None
        self._lock = threading.Lock()
        self._target = 0.0
        self._revision = 0
        self._stop_requested = False
        self.set_target(target)

    def _normalize_target(self, value: float) -> float:
        target = float(value)
        if not math.isfinite(target):
            raise ValueError("Уставка должна быть конечным числом.")
        target = max(0.0, target)
        if self.maximum is not None:
            target = min(target, self.maximum)
        return target

    def set_target(self, value: float) -> float:
        target = self._normalize_target(value)
        with self._lock:
            if not math.isclose(target, self._target, rel_tol=0.0, abs_tol=1e-12):
                self._target = target
                self._revision += 1
            return self._target

    def add(self, delta: float) -> float:
        with self._lock:
            current = self._target
        return self.set_target(current + float(delta))

    def request_stop(self) -> None:
        with self._lock:
            self._stop_requested = True

    def snapshot(self) -> Tuple[float, int, bool]:
        with self._lock:
            return self._target, self._revision, self._stop_requested


def next_stability_voltage(
    control_mode: str,
    voltage_set_V: float,
    target_setpoint: float,
    measured_current_mA: float,
    params: StabilityParams,
) -> Tuple[float, bool]:
    """Return the next applied voltage and whether current mode hit its voltage limit."""

    if control_mode == "current":
        error_mA = float(target_setpoint) - float(measured_current_mA)
        delta = float(
            np.clip(
                params.current_control_kp * error_mA,
                -params.voltage_step_max,
                params.voltage_step_max,
            )
        )
        next_voltage = max(0.0, float(voltage_set_V) + delta)
        limit_reached = next_voltage >= params.voltage_limit
        return min(next_voltage, params.voltage_limit), limit_reached
    if control_mode == "voltage":
        desired = min(max(0.0, float(target_setpoint)), params.voltage_limit)
        return desired, False
    raise ValueError("Неизвестный режим стабильности.")


def find_ivl_data_columns(ws) -> Optional[Tuple[int, Dict[str, int]]]:
    wanted = {
        "Voltage OLED / LED measured (V)": None,
        "Current OLED / LED (mA)": None,
    }
    for row in range(1, min(ws.max_row, 30) + 1):
        headers = {}
        for column in range(1, ws.max_column + 1):
            value = ws.cell(row=row, column=column).value
            if value in wanted:
                headers[value] = column
        if all(key in headers for key in wanted):
            return row, headers
    return None


def interpolate_voltage_at_current_from_ivl(ivl_file: Path, target_current_mA: float) -> Optional[float]:
    ivl_file = Path(ivl_file)
    if not ivl_file.exists():
        return None

    wb = load_workbook(ivl_file, data_only=True)
    points: List[Tuple[float, float]] = []
    try:
        for sheet_name in wb.sheetnames:
            if not sheet_name.startswith("Cycle_"):
                continue
            ws = wb[sheet_name]
            found = find_ivl_data_columns(ws)
            if not found:
                continue
            header_row, cols = found
            v_col = cols["Voltage OLED / LED measured (V)"]
            i_col = cols["Current OLED / LED (mA)"]
            for row in range(header_row + 1, ws.max_row + 1):
                voltage = as_float_or_none(ws.cell(row=row, column=v_col).value)
                current = as_float_or_none(ws.cell(row=row, column=i_col).value)
                if voltage is not None and current is not None and math.isfinite(voltage) and math.isfinite(current):
                    points.append((current, voltage))
    finally:
        wb.close()

    if len(points) < 2:
        return None
    points = sorted(points, key=lambda item: item[0])

    if target_current_mA <= points[0][0]:
        return points[0][1]
    if target_current_mA >= points[-1][0]:
        return points[-1][1]

    for (i1, v1), (i2, v2) in zip(points[:-1], points[1:]):
        if i1 <= target_current_mA <= i2 and i2 != i1:
            ratio = (target_current_mA - i1) / (i2 - i1)
            return v1 + ratio * (v2 - v1)
    return min(points, key=lambda item: abs(item[0] - target_current_mA))[1]


def create_stability_workbook(filename: Path, pixel_id: str, params: StabilityParams):
    return write_stability_workbook(filename, pixel_id, params)


def update_stability_status(ws, status: str, max_photo: float, elapsed: float) -> None:
    write_stability_status(ws, status, max_photo, elapsed)


def save_stability_chart(filename: Path, pixel_id: str) -> None:
    write_stability_chart(filename, pixel_id)


def run_stability_measurement(
    pixel_id: str,
    output_dir: Path,
    params: StabilityParams,
    log: Callable[[str], None],
    app_settings: Optional[Dict[str, Any]] = None,
    control: Optional[StabilitySetpointController] = None,
    progress: Optional[Callable[[Dict[str, Any]], None]] = None,
    measurement_started_monotonic: Optional[float] = None,
) -> Dict[str, Any]:
    prepare_hardware_environment(pixel_id, app_settings, log)
    import xtralien

    output_dir.mkdir(parents=True, exist_ok=True)
    measurement_timestamp = timestamp_for_file()
    control_mode = str(params.control_mode or "current").strip().lower()
    if control_mode not in StabilitySetpointController.MODES:
        raise ValueError("Неизвестный режим стабильности.")
    initial_target = params.current_setpoint_mA if control_mode == "current" else params.voltage_setpoint_V
    target_unit = "mA" if control_mode == "current" else "V"
    if control is None:
        maximum = params.voltage_limit if control_mode == "voltage" else params.current_limit_mA
        control = StabilitySetpointController(control_mode, initial_target, maximum=maximum)
    elif control.mode != control_mode:
        raise ValueError("Режим контроллера уставки не совпадает с параметрами стабильности.")
    file_stem = (
        f"STABILITY_{safe_filename(pixel_id)}_{control_mode}_{initial_target:g}{target_unit}_{measurement_timestamp}"
    )
    filename = output_dir / f"{file_stem}.xlsx"
    raw_file = raw_csv_path(output_dir, f"{file_stem}_raw.csv", app_settings)
    log(f"Raw CSV стабильности: {raw_file}")

    max_photo = 0.0
    point_number = 0
    last_autosave_elapsed = 0.0
    voltage_set = params.voltage_start
    current_limit_reached = False
    current_limit_elapsed_s: Optional[float] = None
    voltage_limit_reached = False
    stopped_by_user = False
    last_elapsed = 0.0
    last_control_revision = -1
    last_progress_payload: Optional[Dict[str, Any]] = None
    measurement_error: Optional[Exception] = None
    shutdown_confirmed = True

    with RawCsvWriter(raw_file, STABILITY_RAW_HEADERS) as raw_writer:
        with xtralien.Device(params.com_port) as smu:
            try:
                smu.smu1.set.enabled(True, response=0)
                smu.smu2.set.enabled(True, response=0)
                try:
                    smu.smu2.set.range(params.photodiode_range, response=0)
                except Exception:
                    pass
                time.sleep(0.1)
                smu.smu1.set.voltage(0, response=0)
                smu.smu2.set.voltage(params.photodiode_bias_V, response=0)
                time.sleep(0.3)

                start = time.monotonic()
                if measurement_started_monotonic is not None:
                    start = float(measurement_started_monotonic)
                next_point = time.monotonic()
                log(
                    f"Стабильность {pixel_id}: режим {control_mode}, "
                    f"цель={initial_target:.3f} {target_unit}, старт V={voltage_set:.3f} В"
                )
                while True:
                    target_setpoint, control_revision, stop_requested = control.snapshot()
                    if stop_requested:
                        stopped_by_user = True
                        log("Стабильность остановлена пользователем.")
                        break
                    if control_revision != last_control_revision:
                        log(f"  Новая уставка: {target_setpoint:.3f} {target_unit}")
                        last_control_revision = control_revision

                    now = time.monotonic()
                    elapsed = now - start
                    if elapsed > params.measurement_time_s:
                        break
                    while now < next_point:
                        if progress is not None and last_progress_payload is not None:
                            progress(last_progress_payload)
                        if control.snapshot()[2]:
                            stopped_by_user = True
                            break
                        time.sleep(min(0.1, next_point - now))
                        now = time.monotonic()
                    if stopped_by_user:
                        log("Стабильность остановлена пользователем.")
                        break
                    next_point += params.sample_interval_s
                    now = time.monotonic()
                    elapsed = now - start
                    last_elapsed = elapsed

                    voltage_set_for_point = float(voltage_set)
                    smu.smu1.set.voltage(voltage_set_for_point, response=0)
                    smu.smu2.set.voltage(params.photodiode_bias_V, response=0)
                    time.sleep(0.02)
                    v_led, i_led = smu.smu1.measure()[0]
                    v_pd, i_pd = smu.smu2.measure()[0]
                    i_led_mA = i_led * 1000.0
                    i_pd_uA = -i_pd * 1_000_000.0
                    j_led = current_density_mA_cm2(i_led_mA, params.pixel_area_mm2)
                    lum = luminance_cd_m2_at_voltage(
                        i_pd_uA,
                        params.luminance_cd_m2_per_uA,
                        voltage_set_for_point,
                        params.luminance_calibration_model,
                    )

                    max_photo = max(max_photo, i_pd_uA)
                    point_number += 1
                    raw_writer.writerow(
                        {
                            "point": point_number,
                            "date_time": now_str(),
                            "elapsed_s": elapsed,
                            "control_mode": control_mode,
                            "target_setpoint": target_setpoint,
                            "target_unit": target_unit,
                            "control_revision": control_revision,
                            "current_setpoint_mA": target_setpoint if control_mode == "current" else "",
                            "voltage_setpoint_V": target_setpoint if control_mode == "voltage" else "",
                            "voltage_set_V": voltage_set_for_point,
                            "voltage_led_measured_V": float(v_led),
                            "current_led_A": float(i_led),
                            "voltage_photodiode_measured_V": float(v_pd),
                            "current_photodiode_A": float(i_pd),
                            "current_led_mA": float(i_led_mA),
                            "current_density_mA_cm2": j_led,
                            "current_photodiode_uA": float(i_pd_uA),
                            "luminance_cd_m2": lum,
                        }
                    )

                    if i_led_mA > params.current_limit_mA:
                        current_limit_reached = True
                        current_limit_elapsed_s = float(elapsed)
                        log(f"Аварийный стоп: ток {i_led_mA:.3f} мА > {params.current_limit_mA:.3f} мА")
                        try:
                            smu.smu1.set.voltage(0, response=0)
                        except Exception:
                            pass

                    last_progress_payload = {
                        "point": point_number,
                        "elapsed_s": elapsed,
                        "control_mode": control_mode,
                        "target_setpoint": target_setpoint,
                        "target_unit": target_unit,
                        "voltage_set_V": voltage_set_for_point,
                        "voltage_measured_V": float(v_led),
                        "current_measured_mA": float(i_led_mA),
                        "photodiode_uA": float(i_pd_uA),
                        "luminance_cd_m2": lum,
                    }
                    if progress is not None:
                        progress(last_progress_payload)

                    target_setpoint, _control_revision, stop_requested = control.snapshot()
                    if stop_requested:
                        stopped_by_user = True
                    if not current_limit_reached and not stopped_by_user:
                        voltage_set, voltage_limit_reached = next_stability_voltage(
                            control_mode,
                            voltage_set,
                            target_setpoint,
                            i_led_mA,
                            params,
                        )
                        if voltage_limit_reached:
                            log(f"Достигнут предел напряжения {params.voltage_limit:.3f} В")

                    if point_number == 1 or point_number % 60 == 0:
                        log(f"  t={elapsed:.1f} c; V={v_led:.3f} В; I={i_led_mA:.3f} мА; PD={i_pd_uA:.3f} мкА")

                    need_save = (
                        (elapsed - last_autosave_elapsed >= params.autosave_interval_s)
                        or current_limit_reached
                        or voltage_limit_reached
                    )
                    if need_save:
                        last_autosave_elapsed = elapsed
                        log(f"  Raw CSV сохранен: {raw_file.name}, t={elapsed:.1f} c")

                    if current_limit_reached or voltage_limit_reached or stopped_by_user:
                        break
            except Exception as exc:
                measurement_error = exc
            finally:
                shutdown_confirmed = shutdown_smu_with_reconnect(smu, params.com_port, log=log)

    if measurement_error is not None or not shutdown_confirmed:
        if shutdown_confirmed:
            safety_note = "Команды безопасного сброса выходов (0 В и/или отключение) приняты прибором."
        else:
            safety_note = (
                "КРИТИЧЕСКИ: подтвердить отключение выходов не удалось. "
                "Немедленно отключите выход SMU вручную."
            )
        if measurement_error is None:
            raise RuntimeError(
                "Измерение стабильности завершилось, но безопасное выключение SMU не подтверждено.\n"
                f"{safety_note}"
            )
        raise RuntimeError(
            f"Измерение стабильности остановлено из-за ошибки связи с SMU: {measurement_error}\n"
            f"{safety_note}"
        ) from measurement_error

    if current_limit_reached:
        status = "BURNED"
    elif voltage_limit_reached and max_photo < params.photodiode_threshold_uA:
        status = "NO_CONTACT"
    elif max_photo >= params.photodiode_threshold_uA:
        status = "WORKING"
    else:
        status = "NONWORKING"

    filename = build_stability_workbook_from_raw_csv(
        raw_file,
        filename,
        pixel_id,
        params,
        status,
        max_photo,
        last_elapsed,
    )
    kept_raw_files = cleanup_raw_files([raw_file], app_settings, log)
    return {
        "file": filename,
        "raw_file": kept_raw_files[0] if kept_raw_files else None,
        "status": status,
        "max_photo_uA": max_photo,
        "control_mode": control_mode,
        "final_setpoint": control.snapshot()[0],
        "stopped_by_user": stopped_by_user,
        "events": (
            [
                {
                    "event": "current_limit_or_breakdown",
                    "label": "Лимит тока / возможный пробой или шунт",
                    "measurement_time_s": current_limit_elapsed_s,
                    "point": point_number,
                }
            ]
            if current_limit_elapsed_s is not None
            else []
        ),
    }

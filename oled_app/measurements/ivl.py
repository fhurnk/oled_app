"""IVL measurement workflow for the modular OLED application."""

from __future__ import annotations

import time
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Callable, Dict, List, Optional, Tuple

import numpy as np
from openpyxl import Workbook
from openpyxl.chart import ScatterChart, Series, Reference
from openpyxl.chart.axis import ChartLines
from openpyxl.styles import Font

from oled_app.hardware import prepare_hardware_environment, safe_shutdown_smu
from oled_app.utils import (
    autosize_columns,
    current_density_mA_cm2,
    luminance_cd_m2,
    now_str,
    safe_filename,
    style_header_row,
    timestamp_for_file,
)


class MeasurementStopped(Exception):
    """Raised by progress callbacks when a measurement should stop."""


@dataclass
class IVLParams:
    com_port: str = "COM3"
    sweep_start: float = 0.0
    sweep_end: float = 5.0
    sweep_increment: float = 0.02
    sweep_time_per_point: float = 0.01
    num_cycles: int = 1
    delay_between_cycles: float = 1.0
    current_limit_mA: float = 10.0
    photodiode_bias_V: float = -5.0
    photodiode_range: int = 4
    photodiode_threshold_uA: float = 0.5
    burnout_current_threshold_mA: float = 10.0
    mark_current_limit_as_burnout: bool = False
    no_contact_max_led_current_mA: float = 0.05
    burned_confirmation_cycles: int = 1
    pixel_area_mm2: float = 1.0
    luminance_cd_m2_per_uA: float = 1.0

    def as_dict(self) -> Dict[str, Any]:
        return self.__dict__.copy()


def define_ivl_pixel_status(
    max_photo_uA: float,
    max_led_current_mA: float,
    current_limit_reached: bool,
    params: IVLParams,
) -> Tuple[str, str]:
    light_detected = max_photo_uA >= params.photodiode_threshold_uA
    burnout_by_high_current = max_led_current_mA >= params.burnout_current_threshold_mA

    if burnout_by_high_current:
        return "BURNED", "Пробой / сгорание по току"
    if light_detected:
        return "WORKING", "Рабочий: фототок выше порога"
    if max_led_current_mA <= params.no_contact_max_led_current_mA:
        return "NO_CONTACT", "Нет контакта: ток почти нулевой"
    return "NONWORKING", "Нерабочий: ток есть, фототока нет"


def detect_opening_voltage(cycle_data: List[Dict[str, Any]], threshold_uA: float) -> Optional[float]:
    for row in cycle_data:
        if row.get("Photodiode current (uA)", 0) >= threshold_uA:
            return float(row.get("Voltage OLED / LED measured (V)", row.get("Voltage set (V)", 0)))
    return None


def describe_ivl_first_measurement(cycles: List[Dict[str, Any]]) -> str:
    if not cycles:
        return "ВАЯХ не выполнена"
    burned = next((cycle for cycle in cycles if cycle.get("status") == "BURNED"), None)
    if burned is not None:
        cycle_number = int(burned.get("cycle", 1) or 1)
        if cycle_number == 1:
            return "Светодиод сгорел/пробился на первом цикле ВАЯХ"
        return f"Светодиод сгорел/пробился на цикле {cycle_number} ВАЯХ"

    first = str(cycles[0].get("status", "") or "").upper()
    if first == "WORKING":
        return "На первом промере светодиод рабочий"
    if first == "NONWORKING":
        return "Светодиод сразу нерабочий: ток есть, фототока нет"
    if first == "NO_CONTACT":
        return "На первом промере нет контакта с подложкой"
    if first == "BURNED":
        return "Светодиод сгорел/пробился на первом цикле ВАЯХ"
    return f"Первый промер завершился статусом {first or 'UNKNOWN'}"


def run_ivl_cycle(
    smu,
    pixel_id: str,
    cycle_number: int,
    params: IVLParams,
    log: Callable[[str], None],
    progress_callback: Optional[Callable[[int, Dict[str, Any]], None]] = None,
) -> Dict[str, Any]:
    log(f"\nВАЯХ {pixel_id}, цикл {cycle_number}: до {params.sweep_end:.3f} В, лимит {params.current_limit_mA:.3f} мА")

    smu.smu1.set.enabled(True, response=0)
    smu.smu2.set.enabled(True, response=0)
    try:
        smu.smu2.set.range(params.photodiode_range, response=0)
    except Exception:
        pass

    time.sleep(0.1)
    smu.smu1.set.voltage(0, response=0)
    smu.smu2.set.voltage(params.photodiode_bias_V, response=0)
    time.sleep(0.2)

    data: List[Dict[str, Any]] = []
    current_limit_reached = False

    voltage_values = np.arange(
        params.sweep_start,
        params.sweep_end + params.sweep_increment / 2,
        params.sweep_increment,
    )
    voltage_values = np.round(voltage_values, 6)

    for idx, set_v in enumerate(voltage_values, start=1):
        smu.smu1.set.voltage(float(set_v), response=0)
        smu.smu2.set.voltage(params.photodiode_bias_V, response=0)
        time.sleep(params.sweep_time_per_point)

        voltage_led, current_led = smu.smu1.measure()[0]
        voltage_pd, current_pd = smu.smu2.measure()[0]
        current_led_mA = current_led * 1000.0
        current_pd_uA = -current_pd * 1_000_000.0

        point_row = {
            "Point": idx,
            "Voltage set (V)": float(set_v),
            "Voltage OLED / LED measured (V)": float(voltage_led),
            "Current OLED / LED (mA)": float(current_led_mA),
            "Current density (mA/cm^2)": current_density_mA_cm2(current_led_mA, params.pixel_area_mm2),
            "Voltage photodiode measured (V)": float(voltage_pd),
            "Photodiode current (uA)": float(current_pd_uA),
            "Luminance (cd/m^2)": luminance_cd_m2(current_pd_uA, params.luminance_cd_m2_per_uA),
        }
        data.append(point_row)
        if progress_callback is not None:
            try:
                progress_callback(cycle_number, point_row)
            except MeasurementStopped:
                safe_shutdown_smu(smu)
                raise

        if idx == 1 or idx % 25 == 0:
            log(f"  {idx}/{len(voltage_values)}: V={voltage_led:.3f} В, I={current_led_mA:.3f} мА, PD={current_pd_uA:.3f} мкА")

        if current_led_mA >= params.current_limit_mA:
            current_limit_reached = True
            log(f"  Аварийный стоп: ток {current_led_mA:.3f} мА >= {params.current_limit_mA:.3f} мА")
            try:
                smu.smu1.set.voltage(0, response=0)
            except Exception:
                pass
            break

    safe_shutdown_smu(smu)

    max_photo = max([row["Photodiode current (uA)"] for row in data], default=0.0)
    max_current = max([row["Current OLED / LED (mA)"] for row in data], default=0.0)
    status, status_desc = define_ivl_pixel_status(max_photo, max_current, current_limit_reached, params)
    opening = detect_opening_voltage(data, params.photodiode_threshold_uA)

    log(f"  Цикл {cycle_number}: статус {status}, max I={max_current:.3f} мА, max PD={max_photo:.3f} мкА")
    return {
        "cycle": cycle_number,
        "status": status,
        "status_desc": status_desc,
        "current_limit_reached": current_limit_reached,
        "max_photo_uA": max_photo,
        "max_current_mA": max_current,
        "opening_voltage": opening,
        "data": data,
    }


def save_ivl_workbook(pixel_id: str, output_dir: Path, params: IVLParams, cycles: List[Dict[str, Any]]) -> Path:
    filename = output_dir / f"IVL_{safe_filename(pixel_id)}_{timestamp_for_file()}.xlsx"
    ivl_diagnosis = describe_ivl_first_measurement(cycles)
    output_dir.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws_sum = wb.active
    ws_sum.title = "Summary"

    ws_sum["A1"] = "IVL / ВАЯХ"
    ws_sum["A1"].font = Font(bold=True, size=14)
    meta = [
        ("Pixel", pixel_id),
        ("Created", now_str()),
        ("COM port", params.com_port),
        ("Sweep", f"{params.sweep_start}-{params.sweep_end} В, step {params.sweep_increment} В"),
        ("Cycles requested", params.num_cycles),
        ("Current limit (mA)", params.current_limit_mA),
        ("Photodiode threshold (uA)", params.photodiode_threshold_uA),
        ("Pixel area (mm^2)", params.pixel_area_mm2),
        ("Luminance conversion (cd/m^2 per uA)", params.luminance_cd_m2_per_uA),
        ("Burned confirmation cycles", params.burned_confirmation_cycles),
        ("Первый промер / диагноз", ivl_diagnosis),
        ("Naming rule", "{quarter_code}{quarter_number}_{substrate_number}_{pixel_number}"),
    ]
    for idx, (key, value) in enumerate(meta, start=3):
        ws_sum.cell(row=idx, column=1, value=key).font = Font(bold=True)
        ws_sum.cell(row=idx, column=2, value=value)

    summary_headers = [
        "Cycle",
        "Status",
        "Status description",
        "Max current (mA)",
        "Max photodiode (uA)",
        "Opening voltage detected (V)",
        "Current limit reached",
        "First measurement diagnosis",
    ]
    start = 13
    ws_sum.append([])
    for column, header in enumerate(summary_headers, start=1):
        ws_sum.cell(row=start, column=column, value=header)
    style_header_row(ws_sum, start, 1, len(summary_headers))

    for row_idx, cycle in enumerate(cycles, start=start + 1):
        ws_sum.cell(row=row_idx, column=1, value=cycle["cycle"])
        ws_sum.cell(row=row_idx, column=2, value=cycle["status"])
        ws_sum.cell(row=row_idx, column=3, value=cycle["status_desc"])
        ws_sum.cell(row=row_idx, column=4, value=cycle["max_current_mA"])
        ws_sum.cell(row=row_idx, column=5, value=cycle["max_photo_uA"])
        ws_sum.cell(row=row_idx, column=6, value=cycle["opening_voltage"])
        ws_sum.cell(row=row_idx, column=7, value="YES" if cycle["current_limit_reached"] else "NO")
        ws_sum.cell(row=row_idx, column=8, value=ivl_diagnosis if cycle["cycle"] == 1 else "")

    headers = [
        "Point",
        "Voltage set (V)",
        "Voltage OLED / LED measured (V)",
        "Current OLED / LED (mA)",
        "Current density (mA/cm^2)",
        "Voltage photodiode measured (V)",
        "Photodiode current (uA)",
        "Luminance (cd/m^2)",
    ]
    for cycle in cycles:
        ws = wb.create_sheet(f"Cycle_{cycle['cycle']}")
        ws["A1"] = f"Pixel {pixel_id} | Cycle {cycle['cycle']} | {cycle['status']}"
        ws["A1"].font = Font(bold=True, size=13)
        header_row = 4
        for column, header in enumerate(headers, start=1):
            ws.cell(row=header_row, column=column, value=header)
        style_header_row(ws, header_row, 1, len(headers))

        for row_idx, row in enumerate(cycle["data"], start=header_row + 1):
            for col_idx, header in enumerate(headers, start=1):
                ws.cell(row=row_idx, column=col_idx, value=row.get(header))
        ws.freeze_panes = f"A{header_row + 1}"

        if len(cycle["data"]) >= 2:
            chart = ScatterChart()
            chart.title = f"{pixel_id} | Cycle {cycle['cycle']} | {cycle['status']}"
            chart.x_axis.title = "Voltage OLED / LED (V)"
            chart.y_axis.title = "Current OLED / LED (mA) / Photodiode (uA)"
            chart.x_axis.majorGridlines = ChartLines()
            min_row = header_row + 1
            max_row = header_row + len(cycle["data"])
            xvalues = Reference(ws, min_col=2, min_row=min_row, max_row=max_row)
            y_current = Reference(ws, min_col=4, min_row=header_row, max_row=max_row)
            y_photo = Reference(ws, min_col=7, min_row=header_row, max_row=max_row)
            chart.series.append(Series(y_current, xvalues, title_from_data=True))
            chart.series.append(Series(y_photo, xvalues, title_from_data=True))
            ws.add_chart(chart, "H4")
        autosize_columns(ws, max_width=34)

    for ws in wb.worksheets:
        autosize_columns(ws, max_width=42)
    wb.save(filename)
    wb.close()
    return filename


def run_ivl_measurement(
    pixel_id: str,
    output_dir: Path,
    params: IVLParams,
    log: Callable[[str], None],
    app_settings: Optional[Dict[str, Any]] = None,
    progress_callback: Optional[Callable[[int, Dict[str, Any]], None]] = None,
) -> Dict[str, Any]:
    prepare_hardware_environment(pixel_id, app_settings, log)
    import xtralien

    cycles: List[Dict[str, Any]] = []
    cycles_to_run = max(1, int(params.num_cycles))
    burned_confirmations_left = max(0, int(params.burned_confirmation_cycles))
    with xtralien.Device(params.com_port) as smu:
        cycle = 1
        while cycle <= cycles_to_run:
            cycle_result = run_ivl_cycle(smu, pixel_id, cycle, params, log, progress_callback=progress_callback)
            cycles.append(cycle_result)
            if cycle_result["status"] == "BURNED" and burned_confirmations_left > 0:
                burned_confirmations_left -= 1
                cycles_to_run = max(cycles_to_run, cycle + 1)
                log("  BURNED: запускается дополнительный подтверждающий цикл.")
            elif cycle_result["status"] in {"BURNED", "NO_CONTACT", "NONWORKING"}:
                log(f"  Дальнейшие циклы остановлены: {cycle_result['status']}")
                break
            if cycle < cycles_to_run:
                time.sleep(params.delay_between_cycles)
            cycle += 1

    filename = save_ivl_workbook(pixel_id, output_dir, params, cycles)
    best_opening = next((cycle.get("opening_voltage") for cycle in cycles if cycle.get("opening_voltage") is not None), None)
    max_current = max([cycle["max_current_mA"] for cycle in cycles], default=0.0)
    max_photo = max([cycle["max_photo_uA"] for cycle in cycles], default=0.0)
    ivl_diagnosis = describe_ivl_first_measurement(cycles)
    burned_cycle = next(
        (int(cycle.get("cycle", 1) or 1) for cycle in cycles if cycle.get("status") == "BURNED"),
        None,
    )
    final_status = "BURNED" if burned_cycle is not None else cycles[-1]["status"] if cycles else "FAILED"
    return {
        "file": filename,
        "status": final_status,
        "opening_voltage": best_opening,
        "max_current_mA": max_current,
        "max_photo_uA": max_photo,
        "ivl_diagnosis": ivl_diagnosis,
        "burned_cycle": burned_cycle,
        "first_cycle_status": cycles[0]["status"] if cycles else "FAILED",
    }

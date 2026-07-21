"""Build a report-preparation Origin project from OLED IVL and spectra measurements.

This module can write directly to an Origin OPJU project through OriginPro
Python, or write an xlsx preview workbook for debugging.
"""

from __future__ import annotations

import argparse
import re
import sys
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Iterable

from openpyxl import Workbook, load_workbook
from openpyxl.chart import ScatterChart, Series
from openpyxl.chart.reference import Reference
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


TIMESTAMP_RE = re.compile(
    r"_(?P<day>\d{2})-(?P<month>\d{2})-(?P<year>\d{4})_"
    r"(?P<hour>\d{1,2})h(?P<minute>\d{2})m(?P<second>\d{2})s\.xlsx$",
    re.IGNORECASE,
)

POINTS_PER_PT = 12700
LINE_WIDTH_2PT = 2 * POINTS_PER_PT
LINE_WIDTH_3PT = 3 * POINTS_PER_PT

PALETTE = [
    "1F77B4",
    "FF7F0E",
    "2CA02C",
    "D62728",
    "9467BD",
    "8C564B",
    "E377C2",
    "7F7F7F",
    "BCBD22",
    "17BECF",
]

COLOR4LINE_RGB = [
    (81, 81, 81),
    (241, 64, 64),
    (26, 111, 223),
    (55, 173, 107),
    (177, 119, 222),
    (204, 153, 0),
    (0, 203, 204),
    (125, 78, 78),
    (142, 142, 0),
    (251, 101, 1),
    (102, 153, 204),
    (111, 184, 2),
]

ORIGIN_REPORT_ROOT = ""

REPORT_MODE_FULL = "full"
REPORT_MODE_IVL = "ivl"
REPORT_MODE_SPECTRA = "spectra"
REPORT_MODES = (REPORT_MODE_FULL, REPORT_MODE_IVL, REPORT_MODE_SPECTRA)


def report_mode(args: argparse.Namespace) -> str:
    """Return the requested report composition, preserving old callers as full reports."""

    return getattr(args, "report_mode", REPORT_MODE_FULL)


def report_includes_ivl(args: argparse.Namespace) -> bool:
    return report_mode(args) in {REPORT_MODE_FULL, REPORT_MODE_IVL}


def report_includes_spectra(args: argparse.Namespace) -> bool:
    return report_mode(args) in {REPORT_MODE_FULL, REPORT_MODE_SPECTRA}


@dataclass(frozen=True)
class MeasurementPath:
    path: Path
    date_dir: str
    series: str
    subseries: str
    pixel: str
    timestamp: datetime


@dataclass
class IvRecord:
    path: Path
    date_dir: str
    series: str
    subseries: str
    pixel: str
    cycle: str
    status: str
    rows: list[dict[str, float]]


@dataclass
class IvBlock:
    record: IvRecord
    sheet_name: str
    first_data_row: int
    last_data_row: int
    voltage_col: int | None
    current_col: int
    density_col: int
    photodiode_col: int
    luminance_col: int


@dataclass
class SpectrumRecord:
    path: Path
    date_dir: str
    series: str
    subseries: str
    pixel: str
    voltages: list[float]
    wavelengths: list[float]
    intensities: list[list[float]]


@dataclass
class SpectrumBlock:
    record: SpectrumRecord
    sheet_name: str
    first_data_row: int
    last_data_row: int
    wavelength_col: int
    first_intensity_col: int


@dataclass
class ReportData:
    iv_records: list[IvRecord]
    spectrum_records: list[SpectrumRecord]
    common_voltage: float | None
    warnings: list[str]


def parse_timestamp(path: Path) -> datetime:
    match = TIMESTAMP_RE.search(path.name)
    if not match:
        return datetime.fromtimestamp(path.stat().st_mtime)
    parts = {key: int(value) for key, value in match.groupdict().items()}
    return datetime(
        parts["year"],
        parts["month"],
        parts["day"],
        parts["hour"],
        parts["minute"],
        parts["second"],
    )


def split_measurement_path(root: Path, path: Path) -> MeasurementPath:
    rel = path.relative_to(root)
    parts = rel.parts
    if len(parts) < 5:
        pixel = path.stem.split("_", 1)[-1]
        tokens = pixel.split("_")
        series = tokens[0] if tokens else "UNKNOWN"
        subseries = "_".join(tokens[:2]) if len(tokens) >= 2 else series
        return MeasurementPath(path, "UNKNOWN", series, subseries, pixel, parse_timestamp(path))
    return MeasurementPath(path, parts[0], parts[1], parts[2], parts[3], parse_timestamp(path))


def latest_by_pixel(root: Path, pattern: str, date_filter: str | None = None) -> dict[str, MeasurementPath]:
    latest: dict[str, MeasurementPath] = {}
    for path in root.rglob(pattern):
        try:
            meta = split_measurement_path(root, path)
        except Exception:
            continue
        if date_filter and meta.date_dir != date_filter:
            continue
        prev = latest.get(meta.pixel)
        if prev is None or meta.timestamp > prev.timestamp:
            latest[meta.pixel] = meta
    return latest


def find_header_row(ws, required: str) -> tuple[int, dict[str, int]]:
    for row in range(1, min(ws.max_row, 40) + 1):
        values = [ws.cell(row=row, column=col).value for col in range(1, ws.max_column + 1)]
        if required in values:
            return row, {str(value): idx + 1 for idx, value in enumerate(values) if value is not None}
    raise ValueError(f"Header row with {required!r} was not found on sheet {ws.title!r}")


def cycle_number(sheet_name: str) -> int:
    try:
        return int(sheet_name.split("_", 1)[1])
    except Exception:
        return -1


def parse_status(cell_value: object) -> str:
    text = str(cell_value or "")
    if "|" not in text:
        return "UNKNOWN"
    return text.rsplit("|", 1)[-1].strip().upper()


def natural_key(value: object) -> tuple:
    text = str(value or "")
    parts = re.split(r"(\d+)", text)
    return tuple(int(part) if part.isdigit() else part.lower() for part in parts)


def pixel_position_key(pixel: str) -> tuple:
    parts = str(pixel or "").split("_")
    series = parts[0] if parts else ""
    quarter_match = re.search(r"(\d+)$", series)
    quarter = int(quarter_match.group(1)) if quarter_match else 9999
    substrate = int(parts[1]) if len(parts) > 1 and parts[1].isdigit() else 9999
    pixel_number = int(parts[2]) if len(parts) > 2 and parts[2].isdigit() else 9999
    prefix = series[: quarter_match.start(1)] if quarter_match else series
    return (quarter, substrate, pixel_number, natural_key(prefix), natural_key(pixel))


def measurement_sort_key(record) -> tuple:
    return (*pixel_position_key(record.pixel), natural_key(record.series), natural_key(record.subseries))


def read_iv_record(meta: MeasurementPath, warnings: list[str]) -> IvRecord | None:
    try:
        wb = load_workbook(meta.path, read_only=False, data_only=True)
    except Exception as exc:
        warnings.append(f"IVL file skipped, cannot read: {meta.path} ({exc})")
        return None

    try:
        cycle_sheets = [ws for ws in wb.worksheets if ws.title.startswith("Cycle_")]
        if not cycle_sheets:
            warnings.append(f"IVL file skipped, no Cycle_N sheets: {meta.path}")
            return None
        ws = sorted(cycle_sheets, key=lambda sheet: cycle_number(sheet.title))[-1]
        status = parse_status(ws.cell(row=1, column=1).value)

        try:
            header_row, headers = find_header_row(ws, "Point")
            required = {
                "voltage": "Voltage OLED / LED measured (V)",
                "current": "Current OLED / LED (mA)",
                "density": "Current density (mA/cm^2)",
                "photodiode": "Photodiode current (uA)",
                "luminance": "Luminance (cd/m^2)",
            }
            missing = [name for name in required.values() if name not in headers]
            if missing:
                warnings.append(f"IVL file skipped, missing columns {missing}: {meta.path}")
                return None

            rows: list[dict[str, float]] = []
            for row in range(header_row + 1, ws.max_row + 1):
                voltage = ws.cell(row=row, column=headers[required["voltage"]]).value
                if voltage is None:
                    continue
                rows.append(
                    {
                        "voltage": voltage,
                        "current": ws.cell(row=row, column=headers[required["current"]]).value,
                        "density": ws.cell(row=row, column=headers[required["density"]]).value,
                        "photodiode": ws.cell(row=row, column=headers[required["photodiode"]]).value,
                        "luminance": ws.cell(row=row, column=headers[required["luminance"]]).value,
                    }
                )
        except Exception as exc:
            warnings.append(f"IVL file skipped, cannot parse last cycle: {meta.path} ({exc})")
            return None

        return IvRecord(
            path=meta.path,
            date_dir=meta.date_dir,
            series=meta.series,
            subseries=meta.subseries,
            pixel=meta.pixel,
            cycle=ws.title,
            status=status,
            rows=rows,
        )
    finally:
        wb.close()


def collect_iv_records(iv_root: Path, warnings: list[str], date_filter: str | None = None) -> list[IvRecord]:
    records: list[IvRecord] = []
    for meta in latest_by_pixel(iv_root, "IVL_*.xlsx", date_filter).values():
        record = read_iv_record(meta, warnings)
        if record is None:
            continue
        if record.status != "WORKING":
            continue
        records.append(record)
    return sorted(records, key=measurement_sort_key)


def read_metadata_value(ws, key: str) -> object | None:
    for row in range(1, min(ws.max_row, 40) + 1):
        if ws.cell(row=row, column=1).value == key:
            return ws.cell(row=row, column=2).value
    return None


def read_spectrum_record(meta: MeasurementPath, sheet_name: str, warnings: list[str]) -> SpectrumRecord | None:
    try:
        wb = load_workbook(meta.path, read_only=False, data_only=True)
    except Exception as exc:
        warnings.append(f"Spectrum file skipped, cannot read: {meta.path} ({exc})")
        return None

    try:
        if sheet_name not in wb.sheetnames:
            warnings.append(f"Spectrum file skipped, no sheet {sheet_name!r}: {meta.path}")
            return None
        ws = wb[sheet_name]
        summary = wb.worksheets[0]
        pixel = read_metadata_value(summary, "Pixel") or meta.pixel

        try:
            header_row, headers = find_header_row(ws, "Wavelength (nm)")
            voltage_row = None
            for row in range(1, header_row):
                if ws.cell(row=row, column=1).value == "V set (V)":
                    voltage_row = row
                    break
            if voltage_row is None:
                raise ValueError("V set row was not found")

            voltages: list[float] = []
            data_columns: list[int] = []
            for col in range(2, ws.max_column + 1):
                value = ws.cell(row=voltage_row, column=col).value
                if isinstance(value, (int, float)):
                    voltages.append(float(value))
                    data_columns.append(col)

            wavelengths: list[float] = []
            intensities: list[list[float]] = [[] for _ in data_columns]
            for row in range(header_row + 1, ws.max_row + 1):
                wavelength = ws.cell(row=row, column=1).value
                if not isinstance(wavelength, (int, float)):
                    continue
                wavelengths.append(float(wavelength))
                for idx, col in enumerate(data_columns):
                    value = ws.cell(row=row, column=col).value
                    intensities[idx].append(value if isinstance(value, (int, float)) else None)
        except Exception as exc:
            warnings.append(f"Spectrum file skipped, cannot parse data: {meta.path} ({exc})")
            return None

        return SpectrumRecord(
            path=meta.path,
            date_dir=meta.date_dir,
            series=meta.series,
            subseries=meta.subseries,
            pixel=str(pixel),
            voltages=voltages,
            wavelengths=wavelengths,
            intensities=intensities,
        )
    finally:
        wb.close()


def parse_spectrum_pixels(values: Iterable[str]) -> dict[str, str]:
    result: dict[str, str] = {}
    for value in values:
        if "=" not in value:
            raise ValueError(f"Expected GROUP=PIXEL, got {value!r}")
        subseries, pixel = value.split("=", 1)
        result[subseries.strip()] = pixel.strip()
    return result


def collect_spectrum_records(
    spectra_root: Path,
    explicit_pixels: dict[str, str],
    sheet_name: str,
    require_explicit_selection: bool,
    warnings: list[str],
    date_filter: str | None = None,
    explicit_series_pixels: dict[str, str] | None = None,
) -> list[SpectrumRecord]:
    explicit_series_pixels = explicit_series_pixels or {}
    if explicit_series_pixels:
        candidates_by_series: dict[str, list[MeasurementPath]] = {}
        for meta in latest_by_pixel(spectra_root, "SPECTRUM_*.xlsx", date_filter).values():
            candidates_by_series.setdefault(meta.series, []).append(meta)

        selected_by_series: dict[str, MeasurementPath] = {}
        for series, candidates in candidates_by_series.items():
            selected_pixel = explicit_series_pixels.get(series)
            if selected_pixel:
                matching = [item for item in candidates if item.pixel == selected_pixel]
                if not matching:
                    warnings.append(f"Selected spectrum pixel not found for series {series}: {selected_pixel}")
                    continue
                selected_by_series[series] = sorted(matching, key=lambda item: item.timestamp)[-1]
                continue

            pixels = sorted({item.pixel for item in candidates}, key=pixel_position_key)
            if require_explicit_selection:
                warnings.append(
                    f"Spectrum series {series} requires explicit substrate and pixel selection; "
                    f"available pixels: {pixels}"
                )
                continue
            if len(pixels) > 1:
                warnings.append(
                    f"Spectrum series {series} has multiple pixels {pixels}; "
                    "using the latest file because no explicit selection was passed."
                )
            selected_by_series[series] = sorted(candidates, key=lambda item: item.timestamp)[-1]

        records: list[SpectrumRecord] = []
        for meta in selected_by_series.values():
            record = read_spectrum_record(meta, sheet_name, warnings)
            if record is not None:
                records.append(record)
        return sorted(records, key=measurement_sort_key)

    latest: dict[str, MeasurementPath] = {}
    candidates_by_subseries: dict[str, list[MeasurementPath]] = {}
    for meta in latest_by_pixel(spectra_root, "SPECTRUM_*.xlsx", date_filter).values():
        candidates_by_subseries.setdefault(meta.subseries, []).append(meta)

    for subseries, candidates in candidates_by_subseries.items():
        selected_pixel = explicit_pixels.get(subseries)
        if selected_pixel:
            matching = [item for item in candidates if item.pixel == selected_pixel]
            if not matching:
                warnings.append(f"Selected spectrum pixel not found for {subseries}: {selected_pixel}")
                continue
            latest[subseries] = sorted(matching, key=lambda item: item.timestamp)[-1]
            continue

        pixels = sorted({item.pixel for item in candidates})
        if require_explicit_selection:
            warnings.append(
                f"Spectrum subseries {subseries} requires explicit pixel selection; "
                f"available pixels: {pixels}"
            )
            continue
        if len(pixels) > 1:
            warnings.append(
                f"Spectrum subseries {subseries} has multiple pixels {pixels}; "
                "using the latest file because no explicit selection was passed."
            )
        latest[subseries] = sorted(candidates, key=lambda item: item.timestamp)[-1]

    records: list[SpectrumRecord] = []
    for meta in latest.values():
        record = read_spectrum_record(meta, sheet_name, warnings)
        if record is not None:
            records.append(record)
    return sorted(records, key=measurement_sort_key)


def validate_voltage_grids(
    records: list[SpectrumRecord],
    expected_start: float | None,
    expected_stop: float | None,
    expected_step: float | None,
    warnings: list[str],
) -> None:
    if not records:
        return
    baseline = [round(value, 6) for value in records[0].voltages]
    for record in records:
        grid = [round(value, 6) for value in record.voltages]
        if grid != baseline:
            warnings.append(f"Spectrum voltage grid differs for {record.pixel}: {record.voltages}")
        if expected_start is not None and record.voltages and abs(record.voltages[0] - expected_start) > 1e-6:
            warnings.append(f"Spectrum {record.pixel}: start voltage {record.voltages[0]} != {expected_start}")
        if expected_stop is not None and record.voltages and abs(record.voltages[-1] - expected_stop) > 1e-6:
            warnings.append(f"Spectrum {record.pixel}: stop voltage {record.voltages[-1]} != {expected_stop}")
        if expected_step is not None and len(record.voltages) > 1:
            for prev, current in zip(record.voltages, record.voltages[1:]):
                if abs((current - prev) - expected_step) > 1e-6:
                    warnings.append(f"Spectrum {record.pixel}: voltage step is not {expected_step}")
                    break


def build_voltage_grid(start: float, stop: float, step: float) -> list[float]:
    if step <= 0:
        raise ValueError("Voltage step must be positive")
    if stop < start:
        raise ValueError("Voltage stop must be greater than or equal to voltage start")
    values: list[float] = []
    current = float(start)
    # The half-step margin keeps common decimal grids stable despite float math.
    while current <= stop + step / 2:
        values.append(round(current, 6))
        current += step
        if len(values) > 10000:
            raise ValueError("Voltage grid is too large")
    return values


def parse_spectrum_voltage_grids(values: Iterable[str]) -> dict[str, list[float]]:
    result: dict[str, list[float]] = {}
    for value in values:
        if "=" not in value:
            raise ValueError(f"Expected PIXEL=START:STOP:STEP, got {value!r}")
        pixel, grid_text = value.split("=", 1)
        parts = [item.strip() for item in grid_text.split(":")]
        if len(parts) != 3:
            raise ValueError(f"Expected START:STOP:STEP for {pixel.strip()}, got {grid_text!r}")
        start, stop, step = (float(item) for item in parts)
        result[pixel.strip()] = build_voltage_grid(start, stop, step)
    return result


def filter_spectrum_record_to_grid(
    record: SpectrumRecord,
    requested_voltages: list[float],
    warnings: list[str],
) -> SpectrumRecord | None:
    indices: list[int] = []
    missing: list[float] = []
    available = {round(value, 6): idx for idx, value in enumerate(record.voltages)}
    for voltage in requested_voltages:
        rounded = round(voltage, 6)
        idx = available.get(rounded)
        if idx is None:
            missing.append(voltage)
        else:
            indices.append(idx)
    if missing:
        warnings.append(f"Spectrum {record.pixel}: requested voltages are absent: {missing}")
    if not indices:
        warnings.append(f"Spectrum {record.pixel}: no requested voltages are available")
        return None
    return SpectrumRecord(
        path=record.path,
        date_dir=record.date_dir,
        series=record.series,
        subseries=record.subseries,
        pixel=record.pixel,
        voltages=[record.voltages[idx] for idx in indices],
        wavelengths=record.wavelengths,
        intensities=[record.intensities[idx] for idx in indices],
    )


def apply_voltage_filters(
    records: list[SpectrumRecord],
    global_start: float | None,
    global_stop: float | None,
    global_step: float | None,
    per_pixel_grids: dict[str, list[float]],
    warnings: list[str],
) -> list[SpectrumRecord]:
    global_grid = None
    if global_start is not None or global_stop is not None or global_step is not None:
        if global_start is not None and global_stop is not None and global_step is not None:
            global_grid = build_voltage_grid(global_start, global_stop, global_step)

    filtered: list[SpectrumRecord] = []
    for record in records:
        requested = per_pixel_grids.get(record.pixel) or global_grid
        if requested is None:
            filtered.append(record)
            continue
        next_record = filter_spectrum_record_to_grid(record, requested, warnings)
        if next_record is not None:
            filtered.append(next_record)
    return filtered


def setup_workbook() -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "README"
    return wb


def write_readme(ws, args: argparse.Namespace, iv_count: int, spectra_count: int) -> None:
    includes_ivl = report_includes_ivl(args)
    includes_spectra = report_includes_spectra(args)
    rows = [
        ("OLED report origin-preparation workbook", None),
        ("Created by", "scripts/build_report_origin_workbook.py"),
        ("Report mode", report_mode(args)),
        ("Measurements dir", str(args.measurements_dir)),
        ("IVL date", (args.ivl_date or "latest") if includes_ivl else "excluded"),
        ("Spectrum date", (args.spectrum_date or "latest") if includes_spectra else "excluded"),
        ("IVL working pixels", iv_count),
        ("Selected spectrum pixels", spectra_count),
        ("Spectrum source sheet", args.spectrum_sheet),
        ("Expected voltage start", args.voltage_start),
        ("Expected voltage stop", args.voltage_stop),
        ("Expected voltage step", args.voltage_step),
    ]
    for row in rows:
        ws.append(row)
    ws["A1"].font = Font(bold=True, size=14)
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 80


def style_header(ws, row: int) -> None:
    fill = PatternFill("solid", fgColor="D9EAF7")
    for cell in ws[row]:
        if cell.value is not None:
            cell.font = Font(bold=True)
            cell.fill = fill
            cell.alignment = Alignment(horizontal="center")


def autosize(ws, max_width: int = 42) -> None:
    for column in range(1, ws.max_column + 1):
        letter = get_column_letter(column)
        width = 10
        for row in range(1, min(ws.max_row, 200) + 1):
            value = ws.cell(row=row, column=column).value
            if value is not None:
                width = max(width, min(max_width, len(str(value)) + 2))
        ws.column_dimensions[letter].width = width


def write_iv_sheet(ws, records: list[IvRecord], include_voltage: bool) -> list[IvBlock]:
    blocks: list[IvBlock] = []
    row = 1
    for record in records:
        ws.cell(row=row, column=1, value=f"{record.series} / {record.subseries} / {record.pixel}")
        ws.cell(row=row, column=1).font = Font(bold=True, size=12)
        ws.cell(row=row, column=2, value=str(record.path))
        row += 1

        headers = ["Series", "Subseries", "Pixel", "Source file", "Cycle"]
        if include_voltage:
            headers.append("Voltage OLED / LED measured (V)")
        headers.extend(
            [
                "Current OLED / LED (mA)",
                "Current density (mA/cm^2)",
                "Photodiode current (uA)",
                "Luminance (cd/m^2)",
            ]
        )
        ws.append(headers)
        style_header(ws, row)
        row += 1

        first_data_row = row
        for item in record.rows:
            values = [record.series, record.subseries, record.pixel, str(record.path), record.cycle]
            if include_voltage:
                values.append(item["voltage"])
            values.extend([item["current"], item["density"], item["photodiode"], item["luminance"]])
            ws.append(values)
            row += 1
        last_data_row = row - 1

        base_col = 6 if include_voltage else 5
        blocks.append(
            IvBlock(
                record=record,
                sheet_name=ws.title,
                first_data_row=first_data_row,
                last_data_row=last_data_row,
                voltage_col=6 if include_voltage else None,
                current_col=base_col + 1 if include_voltage else base_col + 1,
                density_col=base_col + 2 if include_voltage else base_col + 2,
                photodiode_col=base_col + 3 if include_voltage else base_col + 3,
                luminance_col=base_col + 4 if include_voltage else base_col + 4,
            )
        )
        row += 3
    autosize(ws)
    ws.freeze_panes = "A3"
    return blocks


def chart_with_iv_series(
    title: str,
    sheet_name: str,
    blocks: list[IvBlock],
    x_col_getter,
    y_col_getter,
    y_title: str,
    focus_series: str | None = None,
) -> ScatterChart:
    chart = ScatterChart()
    chart.title = title
    chart.style = 13
    chart.x_axis.title = "Voltage OLED / LED measured (V)" if focus_series != "JL" else "Current density (mA/cm^2)"
    chart.y_axis.title = y_title
    chart.width = 14
    chart.height = 8
    for idx, block in enumerate(blocks):
        x_col = x_col_getter(block)
        y_col = y_col_getter(block)
        if x_col is None or y_col is None:
            continue
        xvalues = Reference(
            chart.parent[sheet_name] if hasattr(chart, "parent") and chart.parent else None,
            min_col=x_col,
            min_row=block.first_data_row,
            max_row=block.last_data_row,
        )
        yvalues = Reference(
            chart.parent[sheet_name] if hasattr(chart, "parent") and chart.parent else None,
            min_col=y_col,
            min_row=block.first_data_row,
            max_row=block.last_data_row,
        )
        series = Series(yvalues, xvalues, title=block.record.pixel)
        if focus_series is None or block.record.series == focus_series:
            series.graphicalProperties.line.width = LINE_WIDTH_3PT if focus_series else LINE_WIDTH_2PT
            series.graphicalProperties.line.solidFill = PALETTE[idx % len(PALETTE)]
        else:
            series.graphicalProperties.line.width = LINE_WIDTH_2PT
            series.graphicalProperties.line.solidFill = "C7C7C7"
        series.marker.symbol = "none"
        chart.series.append(series)
    return chart


def add_iv_charts(wb: Workbook, blocks: list[IvBlock]) -> None:
    ws = wb.create_sheet("Charts_IVL")
    source_ws = wb["IVL_U_I_PD"]

    def add_chart(title: str, anchor: str, x_col, y_col, y_title: str, focus: str | None = None, x_title: str | None = None):
        chart = ScatterChart()
        chart.title = title
        chart.style = 13
        chart.x_axis.title = x_title or "Voltage OLED / LED measured (V)"
        chart.y_axis.title = y_title
        chart.width = 14
        chart.height = 8
        for idx, block in enumerate(blocks):
            xv_col = x_col(block)
            yv_col = y_col(block)
            if xv_col is None or yv_col is None:
                continue
            xvalues = Reference(source_ws, min_col=xv_col, min_row=block.first_data_row, max_row=block.last_data_row)
            yvalues = Reference(source_ws, min_col=yv_col, min_row=block.first_data_row, max_row=block.last_data_row)
            series = Series(yvalues, xvalues, title=block.record.pixel)
            if focus is None or block.record.series == focus:
                series.graphicalProperties.line.width = LINE_WIDTH_3PT if focus else LINE_WIDTH_2PT
                series.graphicalProperties.line.solidFill = PALETTE[idx % len(PALETTE)]
            else:
                series.graphicalProperties.line.width = LINE_WIDTH_2PT
                series.graphicalProperties.line.solidFill = "C7C7C7"
            series.marker.symbol = "none"
            chart.series.append(series)
        ws.add_chart(chart, anchor)

    add_chart(
        "All working pixels: V-I",
        "A1",
        lambda block: block.voltage_col,
        lambda block: block.current_col,
        "Current OLED / LED (mA)",
    )
    add_chart(
        "All working pixels: J-L",
        "J1",
        lambda block: block.density_col,
        lambda block: block.luminance_col,
        "Luminance (cd/m^2)",
        x_title="Current density (mA/cm^2)",
    )

    row_anchor = 18
    for series_name in sorted({block.record.series for block in blocks}, key=pixel_position_key):
        add_chart(
            f"Focus {series_name}: V-I",
            f"A{row_anchor}",
            lambda block: block.voltage_col,
            lambda block: block.current_col,
            "Current OLED / LED (mA)",
            focus=series_name,
        )
        add_chart(
            f"Focus {series_name}: J-L",
            f"J{row_anchor}",
            lambda block: block.density_col,
            lambda block: block.luminance_col,
            "Luminance (cd/m^2)",
            focus=series_name,
            x_title="Current density (mA/cm^2)",
        )
        row_anchor += 17


def write_spectra_by_voltage(ws, records: list[SpectrumRecord]) -> list[SpectrumBlock]:
    blocks: list[SpectrumBlock] = []
    col = 1
    for record in records:
        ws.cell(row=1, column=col, value=f"{record.series} / {record.subseries} / {record.pixel}")
        ws.cell(row=1, column=col).font = Font(bold=True, size=12)
        ws.cell(row=2, column=col, value="Wavelength (nm)")
        for idx, voltage in enumerate(record.voltages):
            ws.cell(row=2, column=col + idx + 1, value=f"{voltage:g} V")
        style_header(ws, 2)
        first_data_row = 3
        for row_idx, wavelength in enumerate(record.wavelengths, start=first_data_row):
            ws.cell(row=row_idx, column=col, value=wavelength)
            for idx, values in enumerate(record.intensities):
                ws.cell(row=row_idx, column=col + idx + 1, value=values[row_idx - first_data_row])
        last_data_row = first_data_row + len(record.wavelengths) - 1
        blocks.append(
            SpectrumBlock(
                record=record,
                sheet_name=ws.title,
                first_data_row=first_data_row,
                last_data_row=last_data_row,
                wavelength_col=col,
                first_intensity_col=col + 1,
            )
        )
        col += len(record.voltages) + 3
    autosize(ws, max_width=18)
    ws.freeze_panes = "B3"
    return blocks


def choose_common_voltage(records: list[SpectrumRecord], requested_max: float | None) -> float | None:
    if not records:
        return None
    common = set(round(value, 6) for value in records[0].voltages)
    for record in records[1:]:
        common &= set(round(value, 6) for value in record.voltages)
    if requested_max is not None:
        requested = round(requested_max, 6)
        return requested if requested in common else None
    return max(common) if common else None


def voltage_index(record: SpectrumRecord, voltage: float) -> int | None:
    rounded = round(voltage, 6)
    for idx, item in enumerate(record.voltages):
        if round(item, 6) == rounded:
            return idx
    return None


def write_common_spectra(
    ws,
    records: list[SpectrumRecord],
    common_voltage: float | None,
    normalized: bool,
    warnings: list[str],
) -> None:
    ws.cell(row=1, column=1, value=f"Common voltage: {common_voltage:g} V" if common_voltage is not None else "No common voltage")
    ws.cell(row=2, column=1, value="Wavelength (nm)")
    for col, record in enumerate(records, start=2):
        ws.cell(row=2, column=col, value=record.pixel)
    style_header(ws, 2)
    if common_voltage is None or not records:
        return
    wavelengths = records[0].wavelengths
    for row_idx, wavelength in enumerate(wavelengths, start=3):
        ws.cell(row=row_idx, column=1, value=wavelength)
    for col, record in enumerate(records, start=2):
        idx = voltage_index(record, common_voltage)
        if idx is None:
            warnings.append(f"Spectrum {record.pixel}: common voltage {common_voltage} V not found")
            continue
        values = record.intensities[idx]
        max_value = max([value for value in values if isinstance(value, (int, float))], default=None)
        for row_idx, value in enumerate(values, start=3):
            if normalized:
                ws.cell(row=row_idx, column=col, value=(value / max_value) if max_value else None)
            else:
                ws.cell(row=row_idx, column=col, value=value)
    autosize(ws, max_width=18)
    ws.freeze_panes = "B3"


def write_origin_plot_specs(
    wb: Workbook,
    iv_blocks: list[IvBlock],
    spectrum_records: list[SpectrumRecord],
    common_voltage: float | None,
) -> None:
    ws = wb.create_sheet("Origin_plot_specs")
    ws.append(
        [
            "Plot group",
            "Plot title",
            "Curve",
            "Source sheet",
            "X data",
            "Y data",
            "Line width pt",
            "Transparency percent",
            "Note",
        ]
    )
    style_header(ws, 1)

    def col_range(sheet_name: str, col: int, first_row: int, last_row: int) -> str:
        letter = get_column_letter(col)
        return f"'{sheet_name}'!${letter}${first_row}:${letter}${last_row}"

    for block in iv_blocks:
        curve = f"{block.record.series}/{block.record.subseries}/{block.record.pixel}"
        ws.append(
            [
                "IVL",
                "All working pixels: V-I",
                curve,
                block.sheet_name,
                col_range(block.sheet_name, block.voltage_col, block.first_data_row, block.last_data_row)
                if block.voltage_col
                else None,
                col_range(block.sheet_name, block.current_col, block.first_data_row, block.last_data_row),
                2,
                0,
                "Base V-I chart.",
            ]
        )
        ws.append(
            [
                "IVL",
                "All working pixels: J-L",
                curve,
                block.sheet_name,
                col_range(block.sheet_name, block.density_col, block.first_data_row, block.last_data_row),
                col_range(block.sheet_name, block.luminance_col, block.first_data_row, block.last_data_row),
                2,
                0,
                "Base J-L chart.",
            ]
        )

    for focus_series in sorted({block.record.series for block in iv_blocks}, key=pixel_position_key):
        for block in iv_blocks:
            is_focus = block.record.series == focus_series
            curve = f"{block.record.series}/{block.record.subseries}/{block.record.pixel}"
            for plot_name, x_col, y_col in (
                ("V-I", block.voltage_col, block.current_col),
                ("J-L", block.density_col, block.luminance_col),
            ):
                ws.append(
                    [
                        "IVL focus",
                        f"Focus {focus_series}: {plot_name}",
                        curve,
                        block.sheet_name,
                        col_range(block.sheet_name, x_col, block.first_data_row, block.last_data_row) if x_col else None,
                        col_range(block.sheet_name, y_col, block.first_data_row, block.last_data_row),
                        3 if is_focus else 2,
                        0 if is_focus else 80,
                        "Use real transparency in Origin; Excel preview uses muted gray for non-focus curves.",
                    ]
                )

    for record in spectrum_records:
        for voltage in record.voltages:
            ws.append(
                [
                    "Spectrum by voltage",
                    f"{record.pixel}: spectra by voltage",
                    f"{record.pixel} / {voltage:g} V",
                    "Spectra_by_voltage",
                    "Wavelength (nm)",
                    f"Processed counts per s at {voltage:g} V",
                    2,
                    0,
                    "One chart per selected pixel.",
                ]
            )

    if common_voltage is not None:
        for record in spectrum_records:
            for plot_title, y_label in (
                (f"All selected pixels at {common_voltage:g} V", "Processed counts per s"),
                (f"Normalized spectra at {common_voltage:g} V", "Normalized intensity"),
            ):
                ws.append(
                    [
                        "Spectrum common voltage",
                        plot_title,
                        record.pixel,
                        "Spectra_max_common" if y_label == "Processed counts per s" else "Spectra_max_norm",
                        "Wavelength (nm)",
                        y_label,
                        2,
                        0,
                        "Common voltage must be present for every selected spectrum.",
                    ]
                )
    autosize(ws, max_width=72)
    ws.freeze_panes = "A2"


def add_spectra_charts(wb: Workbook, blocks: list[SpectrumBlock], common_voltage: float | None, records: list[SpectrumRecord]) -> None:
    ws = wb.create_sheet("Charts_Spectra")
    spectra_ws = wb["Spectra_by_voltage"]

    row_anchor = 1
    for block in blocks:
        chart = ScatterChart()
        chart.title = f"{block.record.pixel}: spectra by voltage"
        chart.style = 13
        chart.x_axis.title = "Wavelength (nm)"
        chart.y_axis.title = "Processed counts per s"
        chart.width = 14
        chart.height = 8
        xvalues = Reference(
            spectra_ws,
            min_col=block.wavelength_col,
            min_row=block.first_data_row,
            max_row=block.last_data_row,
        )
        for idx, voltage in enumerate(block.record.voltages):
            yvalues = Reference(
                spectra_ws,
                min_col=block.first_intensity_col + idx,
                min_row=block.first_data_row,
                max_row=block.last_data_row,
            )
            series = Series(yvalues, xvalues, title=f"{voltage:g} V")
            series.graphicalProperties.line.width = LINE_WIDTH_2PT
            series.graphicalProperties.line.solidFill = PALETTE[idx % len(PALETTE)]
            series.marker.symbol = "none"
            chart.series.append(series)
        ws.add_chart(chart, f"A{row_anchor}")
        row_anchor += 17

    def add_common_chart(sheet_name: str, title: str, anchor: str, y_title: str):
        source_ws = wb[sheet_name]
        chart = ScatterChart()
        chart.title = title
        chart.style = 13
        chart.x_axis.title = "Wavelength (nm)"
        chart.y_axis.title = y_title
        chart.width = 14
        chart.height = 8
        max_row = source_ws.max_row
        xvalues = Reference(source_ws, min_col=1, min_row=3, max_row=max_row)
        for idx, record in enumerate(records, start=2):
            yvalues = Reference(source_ws, min_col=idx, min_row=3, max_row=max_row)
            series = Series(yvalues, xvalues, title=record.pixel)
            series.graphicalProperties.line.width = LINE_WIDTH_2PT
            series.graphicalProperties.line.solidFill = PALETTE[(idx - 2) % len(PALETTE)]
            series.marker.symbol = "none"
            chart.series.append(series)
        ws.add_chart(chart, anchor)

    if common_voltage is not None and records:
        add_common_chart(
            "Spectra_max_common",
            f"All selected pixels at {common_voltage:g} V",
            f"J1",
            "Processed counts per s",
        )
        add_common_chart(
            "Spectra_max_norm",
            f"Normalized spectra at {common_voltage:g} V",
            f"J18",
            "Normalized intensity",
        )


def write_warnings(wb: Workbook, warnings: list[str]) -> None:
    ws = wb.create_sheet("Warnings")
    ws.append(["Level", "Message"])
    style_header(ws, 1)
    if warnings:
        for message in warnings:
            ws.append(["WARN", message])
    else:
        ws.append(["OK", "No warnings"])
    autosize(ws, max_width=120)


def collect_report_data(args: argparse.Namespace) -> ReportData:
    warnings: list[str] = []
    measurements_dir = Path(args.measurements_dir)
    iv_root = measurements_dir / "01_IVL_VAH"
    spectra_root = measurements_dir / "02_SPECTRA"

    includes_ivl = report_includes_ivl(args)
    includes_spectra = report_includes_spectra(args)

    if includes_ivl and not iv_root.exists():
        warnings.append(f"IVL root not found: {iv_root}")
    if includes_spectra and not spectra_root.exists():
        warnings.append(f"Spectrum root not found: {spectra_root}")

    iv_records = collect_iv_records(iv_root, warnings, args.ivl_date) if includes_ivl and iv_root.exists() else []
    explicit_pixels = parse_spectrum_pixels(args.spectrum_pixel or [])
    explicit_series_pixels = parse_spectrum_pixels(getattr(args, "spectrum_series_pixel", None) or [])
    spectrum_records = (
        collect_spectrum_records(
            spectra_root,
            explicit_pixels,
            args.spectrum_sheet,
            args.require_spectrum_pixel_selection,
            warnings,
            args.spectrum_date,
            explicit_series_pixels,
        )
        if includes_spectra and spectra_root.exists()
        else []
    )
    if includes_spectra:
        per_pixel_voltage_grids = parse_spectrum_voltage_grids(args.spectrum_voltage_grid or [])
        spectrum_records = apply_voltage_filters(
            spectrum_records,
            args.voltage_start,
            args.voltage_stop,
            args.voltage_step,
            per_pixel_voltage_grids,
            warnings,
        )
        if not per_pixel_voltage_grids:
            validate_voltage_grids(
                spectrum_records,
                args.voltage_start,
                args.voltage_stop,
                args.voltage_step,
                warnings,
            )
        common_voltage = choose_common_voltage(spectrum_records, args.max_voltage)
        if args.max_voltage is not None and common_voltage is None:
            warnings.append(f"Requested common max voltage is absent from selected spectra: {args.max_voltage}")
        if spectrum_records and common_voltage is None:
            warnings.append("Selected spectra have no common voltage for all-pixel spectrum charts")
    else:
        common_voltage = None

    if args.strict and warnings:
        raise RuntimeError("Strict mode failed:\n" + "\n".join(warnings))

    return ReportData(
        iv_records=iv_records,
        spectrum_records=spectrum_records,
        common_voltage=common_voltage,
        warnings=warnings,
    )


def build_workbook(args: argparse.Namespace, data: ReportData) -> tuple[Workbook, list[str]]:
    wb = setup_workbook()
    write_readme(wb["README"], args, len(data.iv_records), len(data.spectrum_records))

    iv_blocks: list[IvBlock] = []
    spectra_blocks: list[SpectrumBlock] = []
    if report_includes_ivl(args):
        iv_sheet = wb.create_sheet("IVL_U_I_PD")
        iv_blocks = write_iv_sheet(iv_sheet, data.iv_records, include_voltage=True)

        iv_no_voltage_sheet = wb.create_sheet("IVL_I_PD")
        write_iv_sheet(iv_no_voltage_sheet, data.iv_records, include_voltage=False)

        if iv_blocks and not args.no_charts:
            add_iv_charts(wb, iv_blocks)

    if report_includes_spectra(args):
        spectra_sheet = wb.create_sheet("Spectra_by_voltage")
        spectra_blocks = write_spectra_by_voltage(spectra_sheet, data.spectrum_records)

        write_common_spectra(
            wb.create_sheet("Spectra_max_common"),
            data.spectrum_records,
            data.common_voltage,
            normalized=False,
            warnings=data.warnings,
        )
        write_common_spectra(
            wb.create_sheet("Spectra_max_norm"),
            data.spectrum_records,
            data.common_voltage,
            normalized=True,
            warnings=data.warnings,
        )
    write_origin_plot_specs(wb, iv_blocks, data.spectrum_records, data.common_voltage)
    if report_includes_spectra(args) and spectra_blocks and not args.no_charts:
        add_spectra_charts(wb, spectra_blocks, data.common_voltage, data.spectrum_records)

    write_warnings(wb, data.warnings)
    return wb, data.warnings


def safe_origin_name(value: str, max_len: int = 24) -> str:
    cleaned = re.sub(r"[^0-9A-Za-z_]+", "_", value).strip("_")
    if not cleaned:
        cleaned = "Sheet"
    if cleaned[0].isdigit():
        cleaned = f"S_{cleaned}"
    return cleaned[:max_len]


def import_originpro():
    try:
        import originpro as op  # type: ignore[import-not-found]
    except Exception as exc:
        raise RuntimeError(
            "Origin output requires Origin/OriginPro Python package 'originpro'. "
            "Install or expose OriginPro's Python integration in this environment, "
            "or run with '--format xlsx' for the debug workbook."
        ) from exc
    return op


def origin_new_sheet(op, name: str):
    try:
        return op.new_sheet("w", lname=name)
    except TypeError:
        return op.new_sheet(lname=name)


def origin_new_graph(op, name: str, short_name: str | None = None):
    try:
        graph = op.new_graph(template="line")
    except TypeError:
        graph = op.new_graph(template="line")
    if short_name:
        try:
            graph.name = safe_origin_name(short_name)
        except Exception:
            pass
    try:
        graph.lname = name
    except Exception:
        pass
    return graph

def origin_new_book(op, name: str):
    try:
        return op.new_book("w", lname=name)
    except TypeError:
        book = op.new_book("w")
        try:
            book.lname = name
        except Exception:
            pass
        return book


def origin_set_folder(op, folder_path: str) -> str:
    try:
        from originpro import pe

        base_path = ORIGIN_REPORT_ROOT or pe.cd()
        pe.cd(base_path)
        current = base_path.rstrip("/")
        for part in [item for item in folder_path.strip("/").split("/") if item]:
            current = pe.mkdir(part, True).rstrip("/")
            try:
                pe.cd(current)
            except Exception:
                pass
        if not folder_path.strip("/"):
            pe.cd(base_path)
            return base_path
        target = current + "/"
        pe.cd(target)
        return target
    except Exception:
        return folder_path


def origin_move_page_to_folder(page, folder_path: str) -> None:
    try:
        from originpro import pe

        target = origin_set_folder(None, folder_path)
        current = pe.search(page.name, 0)
        if current and current.rstrip("/").lower() == target.rstrip("/").lower():
            return
        pe.move(page.name, target)
        pe.cd(target)
    except Exception:
        pass


def origin_write_column(
    wks,
    col: int,
    values: list,
    long_name: str,
    units: str = "",
    axis: str = "Y",
    comments: str = "",
) -> None:
    try:
        wks.from_list(col, values, lname=long_name, units=units, comments=comments, axis=axis)
    except TypeError:
        try:
            wks.from_list(col, values, lname=long_name, units=units, comments=comments)
        except TypeError:
            wks.from_list(col, values, lname=long_name)
            if comments:
                try:
                    wks.set_label(col, comments, "C")
                except Exception:
                    pass


def origin_add_line_plot(layer, wks, y_col: int, x_col: int, title: str):
    for plot_type in ("line", "l"):
        try:
            plot = layer.add_plot(wks, y_col, x_col, type=plot_type)
            break
        except TypeError:
            plot = None
        except Exception:
            plot = None
    if plot is None:
        plot = layer.add_plot(wks, y_col, x_col)
    try:
        plot.lname = title
    except Exception:
        pass
    return plot


def origin_color(index: int) -> tuple[int, int, int]:
    return COLOR4LINE_RGB[index % len(COLOR4LINE_RGB)]


def origin_set_plot_style(plot, width_pt: float, color, transparency_percent: int, warnings: list[str]) -> None:
    style_attempts = (
        ("line.width", width_pt),
        ("line.transparency", transparency_percent),
    )
    for prop, value in style_attempts:
        try:
            if isinstance(value, float):
                plot.set_float(prop, value)
            else:
                plot.set_int(prop, value)
            continue
        except Exception:
            pass
        try:
            plot.set_int(prop, int(value))
        except Exception as exc:
            prefix = f"Origin plot style not applied ({prop}"
            if not any(message.startswith(prefix) for message in warnings):
                warnings.append(f"Origin plot style not applied ({prop}={value}): {exc}")
    try:
        plot.transparency = transparency_percent
    except Exception:
        pass
    try:
        plot.color = color
    except Exception as exc:
        prefix = "Origin plot style not applied (line.color"
        if not any(message.startswith(prefix) for message in warnings):
            warnings.append(f"Origin plot style not applied (line.color={color}): {exc}")


def origin_group_layer(layer, grouped: bool, warnings: list[str]) -> None:
    try:
        layer.group(grouped)
    except Exception as exc:
        prefix = "Origin plot grouping not applied"
        if not any(message.startswith(prefix) for message in warnings):
            warnings.append(f"{prefix}: {exc}")


def origin_rescale(layer) -> None:
    try:
        layer.rescale()
    except Exception:
        try:
            layer.lt_exec("layer -a")
        except Exception:
            pass


def origin_refresh_legend(layer, warnings: list[str]) -> None:
    for command in ("legend -r", "legend -u"):
        try:
            layer.lt_exec(command)
            return
        except Exception:
            pass
    prefix = "Origin legend refresh not applied"
    if not any(message.startswith(prefix) for message in warnings):
        warnings.append(f"{prefix}: update the graph legend in Origin if labels are incomplete")


def write_origin_readme(op, args: argparse.Namespace, data: ReportData) -> None:
    origin_set_folder(op, "/")
    wks = origin_new_sheet(op, "README")
    origin_write_column(
        wks,
        0,
        [
            "OLED report Origin project",
            "Created by scripts/build_report_origin_workbook.py",
            f"Report mode: {report_mode(args)}",
            f"Measurements dir: {args.measurements_dir}",
            f"IVL date: {(args.ivl_date or 'latest') if report_includes_ivl(args) else 'excluded'}",
            f"Spectrum date: {(args.spectrum_date or 'latest') if report_includes_spectra(args) else 'excluded'}",
            f"IVL working pixels: {len(data.iv_records)}",
            f"Selected spectrum pixels: {len(data.spectrum_records)}",
            f"Spectrum source sheet: {args.spectrum_sheet}",
            f"Expected voltage start: {args.voltage_start}",
            f"Expected voltage stop: {args.voltage_stop}",
            f"Expected voltage step: {args.voltage_step}",
        ],
        "Parameter",
        axis="X",
    )


def series_names(records: list[IvRecord]) -> list[str]:
    return sorted({record.series for record in records}, key=pixel_position_key)


def origin_series_values(records: list[IvRecord], series: str, field: str) -> list[float | None]:
    values: list[float | None] = []
    for record in sorted([item for item in records if item.series == series], key=measurement_sort_key):
        values.extend(row[field] for row in record.rows)
        values.extend([None, None, None])
    while values and values[-1] is None:
        values.pop()
    return values


def create_origin_iv_book(op, records: list[IvRecord]):
    origin_set_folder(op, "/IVL")
    try:
        book = op.find_book("w", "Book1")
    except Exception:
        book = None
    if book is None:
        book = origin_new_book(op, "Book1")
    origin_move_page_to_folder(book, "/IVL")
    sheet_iv = book[0]
    sheet_iv.name = "Sheet1"
    try:
        sheet_jl = book.add_sheet("Sheet2")
    except TypeError:
        sheet_jl = book.add_sheet()
        sheet_jl.name = "Sheet2"

    for idx, series in enumerate(series_names(records)):
        base = idx * 3
        origin_write_column(
            sheet_iv,
            base,
            origin_series_values(records, series, "voltage"),
            "V",
            "V",
            "X",
            comments=series,
        )
        origin_write_column(
            sheet_iv,
            base + 1,
            origin_series_values(records, series, "density"),
            "j",
            "mA/cm^2",
            comments=series,
        )
        origin_write_column(
            sheet_iv,
            base + 2,
            origin_series_values(records, series, "luminance"),
            "I\\-(lum)",
            "cd/m^2",
            comments=series,
        )

        jl_base = idx * 2
        density_values = origin_series_values(records, series, "density")
        luminance_values = origin_series_values(records, series, "luminance")
        origin_write_column(sheet_jl, jl_base, density_values, "j", "mA/cm^2", "X", comments=series)
        origin_write_column(sheet_jl, jl_base + 1, luminance_values, "I\\-(lum)", "cd/m^2", comments=series)
        try:
            sheet_jl.set_formula(jl_base, f"Sheet1!{origin_col_name(base + 1)}")
            sheet_jl.set_formula(jl_base + 1, f"Sheet1!{origin_col_name(base + 2)}")
        except Exception:
            pass
    return {"book": book, "iv": sheet_iv, "jl": sheet_jl}


def origin_col_name(index: int) -> str:
    name = ""
    value = index + 1
    while value:
        value, rem = divmod(value - 1, 26)
        name = chr(65 + rem) + name
    return name


def spectrum_book_name(record: SpectrumRecord) -> str:
    return safe_origin_name(record.pixel.replace("_", ""), max_len=20)


def create_origin_spectrum_sheets(op, records: list[SpectrumRecord]) -> dict[str, object]:
    origin_set_folder(op, "/Spectra")
    sheets: dict[str, object] = {}
    for record in records:
        name = spectrum_book_name(record)
        wks = origin_new_sheet(op, name)
        try:
            origin_move_page_to_folder(wks.get_book(), "/Spectra")
        except Exception:
            pass
        origin_write_column(wks, 0, record.wavelengths, "\\g(l)", "nm", "X")
        for idx, voltage in enumerate(record.voltages, start=1):
            origin_write_column(
                wks,
                idx,
                record.intensities[idx - 1],
                "I",
                "cts/ms",
                comments=f"{record.pixel} {voltage:g}V",
            )
        sheets[record.pixel] = wks
    return sheets


def create_origin_common_spectrum_sheets(
    op,
    records: list[SpectrumRecord],
    common_voltage: float | None,
    warnings: list[str],
) -> tuple[object | None, object | None]:
    if common_voltage is None or not records:
        return None, None
    origin_set_folder(op, "/Spectra")
    book = origin_new_book(op, "All")
    origin_move_page_to_folder(book, "/Spectra")
    common_wks = book[0]
    common_wks.name = "Sheet1"
    try:
        norm_wks = book.add_sheet("Sheet2")
    except TypeError:
        norm_wks = book.add_sheet()
        norm_wks.name = "Sheet2"
    origin_write_column(common_wks, 0, records[0].wavelengths, "\\g(l)", "nm", "X")
    origin_write_column(norm_wks, 0, records[0].wavelengths, "\\g(l)", "nm", "X")
    for col, record in enumerate(records, start=1):
        idx = voltage_index(record, common_voltage)
        if idx is None:
            warnings.append(f"Spectrum {record.pixel}: common voltage {common_voltage} V not found")
            continue
        values = record.intensities[idx]
        max_value = max([value for value in values if isinstance(value, (int, float))], default=None)
        normalized = [(value / max_value) if max_value else None for value in values]
        comment = f"{record.pixel} {common_voltage:g}V"
        origin_write_column(common_wks, col, values, "I", "cts/ms", comments=comment)
        origin_write_column(norm_wks, col, normalized, "I\\-(rel)", "", comments=comment)
    return common_wks, norm_wks


def create_origin_iv_graphs(op, records: list[IvRecord], iv_book: dict[str, object], warnings: list[str]) -> None:
    names = series_names(records)

    def add_iv_graph(name: str, folder: str, sheet, x_col_for_idx, y_col_for_idx, focus_series: str | None = None) -> None:
        origin_set_folder(op, folder)
        graph = origin_new_graph(op, name, short_name=name if focus_series is None else None)
        origin_move_page_to_folder(graph, folder)
        layer = graph[0]
        for idx, series in enumerate(names):
            plot = origin_add_line_plot(layer, sheet, y_col_for_idx(idx), x_col_for_idx(idx), series)
            is_focus = focus_series is None or series == focus_series
            width = 3.0 if focus_series and is_focus else 2.0
            transparency = 0 if is_focus else 80
            origin_set_plot_style(plot, width, origin_color(idx), transparency, warnings)
        origin_group_layer(layer, focus_series is None, warnings)
        if focus_series is None:
            for idx, plot in enumerate(layer.plot_list()):
                origin_set_plot_style(plot, 2.0, origin_color(idx), 0, warnings)
        origin_rescale(layer)
        origin_refresh_legend(layer, warnings)

    if not records:
        return
    add_iv_graph("AllIV", "/IVL/IV", iv_book["iv"], lambda idx: idx * 3, lambda idx: idx * 3 + 1)
    add_iv_graph("AlljL", "/IVL/JL", iv_book["jl"], lambda idx: idx * 2, lambda idx: idx * 2 + 1)
    for series_name in names:
        add_iv_graph(series_name, "/IVL/IV", iv_book["iv"], lambda idx: idx * 3, lambda idx: idx * 3 + 1, series_name)
        add_iv_graph(series_name, "/IVL/JL", iv_book["jl"], lambda idx: idx * 2, lambda idx: idx * 2 + 1, series_name)


def create_origin_spectrum_graphs(
    op,
    records: list[SpectrumRecord],
    spectrum_sheets: dict[str, object],
    common_wks,
    norm_wks,
    common_voltage: float | None,
    warnings: list[str],
) -> None:
    for record in records:
        origin_set_folder(op, "/Spectra")
        graph_name = f"{spectrum_book_name(record)}spctr"
        graph = origin_new_graph(op, graph_name, short_name=graph_name)
        origin_move_page_to_folder(graph, "/Spectra")
        layer = graph[0]
        wks = spectrum_sheets[record.pixel]
        for idx, voltage in enumerate(record.voltages, start=1):
            plot = origin_add_line_plot(layer, wks, idx, 0, f"{voltage:g} V")
            origin_set_plot_style(plot, 2.0, origin_color(idx - 1), 0, warnings)
        origin_group_layer(layer, True, warnings)
        for idx, plot in enumerate(layer.plot_list()):
            origin_set_plot_style(plot, 2.0, origin_color(idx), 0, warnings)
        origin_rescale(layer)
        origin_refresh_legend(layer, warnings)

    if common_voltage is None or common_wks is None or norm_wks is None:
        return
    for title, wks, units in (
        ("Allspctr", common_wks, "processed counts/s"),
        ("Allrelspctr", norm_wks, "normalized"),
    ):
        origin_set_folder(op, "/Spectra")
        graph = origin_new_graph(op, title, short_name=title)
        origin_move_page_to_folder(graph, "/Spectra")
        layer = graph[0]
        for idx, record in enumerate(records, start=1):
            plot = origin_add_line_plot(layer, wks, idx, 0, f"{record.pixel} {units}")
            origin_set_plot_style(plot, 2.0, origin_color(idx - 1), 0, warnings)
        origin_group_layer(layer, True, warnings)
        for idx, plot in enumerate(layer.plot_list()):
            origin_set_plot_style(plot, 2.0, origin_color(idx), 0, warnings)
        origin_rescale(layer)
        origin_refresh_legend(layer, warnings)


def write_origin_warnings(op, warnings: list[str]) -> None:
    origin_set_folder(op, "/")
    wks = origin_new_sheet(op, "Warnings")
    levels = ["WARN" for _ in warnings] or ["OK"]
    messages = warnings or ["No warnings"]
    origin_write_column(wks, 0, levels, "Level")
    origin_write_column(wks, 1, messages, "Message")


def build_origin_project(args: argparse.Namespace, data: ReportData) -> list[str]:
    global ORIGIN_REPORT_ROOT
    op = import_originpro()
    origin_warnings = data.warnings
    try:
        try:
            op.set_show(bool(args.show_origin))
        except Exception:
            pass
        try:
            op.new()
        except Exception:
            pass
        try:
            from originpro import pe

            ORIGIN_REPORT_ROOT = pe.cd()
        except Exception:
            ORIGIN_REPORT_ROOT = ""

        if report_includes_ivl(args):
            iv_sheets = create_origin_iv_book(op, data.iv_records)
            create_origin_iv_graphs(op, data.iv_records, iv_sheets, origin_warnings)
        if report_includes_spectra(args):
            spectrum_sheets = create_origin_spectrum_sheets(op, data.spectrum_records)
            common_wks, norm_wks = create_origin_common_spectrum_sheets(
                op,
                data.spectrum_records,
                data.common_voltage,
                origin_warnings,
            )
            create_origin_spectrum_graphs(
                op,
                data.spectrum_records,
                spectrum_sheets,
                common_wks,
                norm_wks,
                data.common_voltage,
                origin_warnings,
            )
        write_origin_readme(op, args, data)
        write_origin_warnings(op, origin_warnings)

        output_path = args.output.resolve()
        output_path.parent.mkdir(parents=True, exist_ok=True)
        try:
            saved = op.save(str(output_path))
        except Exception as exc:
            raise RuntimeError(f"Origin project save failed: {output_path} ({exc})") from exc
        if not saved or not output_path.exists():
            raise RuntimeError(f"Origin project save failed: {output_path}")
        return origin_warnings
    finally:
        ORIGIN_REPORT_ROOT = ""
        if args.keep_origin_open:
            try:
                op.detach()
            except Exception:
                pass
        else:
            try:
                op.exit()
            except Exception:
                pass


def parse_args(argv: list[str]) -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Build an OLED report Origin project from IVL and spectrum xlsx files.",
    )
    parser.add_argument(
        "--measurements-dir",
        type=Path,
        default=Path("report") / "measurements",
        help="Path to report measurements directory.",
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=None,
        help="Output path. Defaults to report/origin_report.opju for Origin mode.",
    )
    parser.add_argument(
        "--format",
        choices=("origin", "xlsx"),
        default=None,
        help="Output format. Defaults to origin unless --output ends with .xlsx.",
    )
    parser.add_argument(
        "--spectrum-sheet",
        default="Processed counts per s",
        help="Spectrum worksheet to use.",
    )
    parser.add_argument(
        "--report-mode",
        choices=REPORT_MODES,
        default=REPORT_MODE_FULL,
        help="Report composition: full, IVL only, or spectra only.",
    )
    parser.add_argument("--ivl-date", default=None, help="Use IVL files only from this YYYY-MM-DD measurement date.")
    parser.add_argument("--spectrum-date", default=None, help="Use spectrum files only from this YYYY-MM-DD measurement date.")
    parser.add_argument(
        "--spectrum-pixel",
        action="append",
        default=[],
        help="Explicit spectrum pixel selection in SUBSERIES=PIXEL form. Can be passed multiple times.",
    )
    parser.add_argument(
        "--spectrum-series-pixel",
        action="append",
        default=[],
        help=(
            "Explicit one-pixel selection in SERIES=PIXEL form. This selects one substrate through its pixel "
            "when a series contains spectra from multiple substrates. Can be passed multiple times."
        ),
    )
    parser.add_argument(
        "--spectrum-voltage-grid",
        action="append",
        default=[],
        help="Per-pixel spectrum voltage filter in PIXEL=START:STOP:STEP form. Can be passed multiple times.",
    )
    parser.add_argument(
        "--require-spectrum-pixel-selection",
        action="store_true",
        help="Skip spectrum groups without an explicit pixel selection.",
    )
    parser.add_argument("--voltage-start", type=float, default=None)
    parser.add_argument("--voltage-stop", type=float, default=None)
    parser.add_argument("--voltage-step", type=float, default=None)
    parser.add_argument(
        "--max-voltage",
        type=float,
        default=None,
        help="Common voltage for all-pixel spectrum charts. Defaults to maximum common voltage.",
    )
    parser.add_argument(
        "--strict",
        action="store_true",
        help="Fail if warnings are produced during selection or validation.",
    )
    parser.add_argument(
        "--no-charts",
        action="store_true",
        help="In xlsx mode, write data and Origin plot specifications without creating Excel preview charts.",
    )
    parser.add_argument(
        "--show-origin",
        action="store_true",
        help="Show Origin while creating the .opju project.",
    )
    parser.add_argument(
        "--keep-origin-open",
        action="store_true",
        help="Do not close Origin after saving the .opju project.",
    )
    args = parser.parse_args(argv)
    if args.format is None:
        args.format = "xlsx" if args.output and args.output.suffix.lower() == ".xlsx" else "origin"
    if args.output is None:
        args.output = Path("report") / ("origin_report.opju" if args.format == "origin" else "origin_report_data.xlsx")
    return args


def main(argv: list[str] | None = None) -> int:
    args = parse_args(argv or sys.argv[1:])
    try:
        data = collect_report_data(args)
        if args.format == "origin":
            warnings = build_origin_project(args, data)
        else:
            wb, warnings = build_workbook(args, data)
            args.output.parent.mkdir(parents=True, exist_ok=True)
            wb.save(args.output)
    except Exception as exc:
        print(f"ERROR: {exc}", file=sys.stderr)
        return 1

    print(f"Saved {args.output}")
    if warnings:
        print(f"Warnings: {len(warnings)}")
        for message in warnings[:10]:
            print(f"- {message}")
        if len(warnings) > 10:
            print(f"- ... {len(warnings) - 10} more warnings in workbook")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

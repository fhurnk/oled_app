"""Start-screen views for the modular OLED GUI."""

from __future__ import annotations

import json
import tkinter as tk
from pathlib import Path
from tkinter import filedialog, messagebox, ttk
from tkinter.scrolledtext import ScrolledText
from typing import Any, Dict, List

from openpyxl import load_workbook

from oled_app.constants import CONFIG_FILE, DEFAULT_ROOT, JOURNAL_FILE, MEASUREMENTS_SHEET, SCRIPT_DIR
from oled_app.series import (
    LED_COLOR_LABELS,
    SeriesManager,
    build_holder_layout,
    led_color_from_label,
    led_color_label,
    quarter_base,
    quarter_code,
    quarter_description,
    quarter_led_color,
)
from oled_app.settings import hardware_mode_label, load_app_settings, save_app_settings
from oled_app.utils import today_iso

from .measurement_menu import build_hardware_status_bar, check_hardware_status
from .widgets import create_scrollable_frame, create_tree_with_scrollbars


def show_start_screen(app) -> None:
    app.clear()
    app.app_settings = load_app_settings()

    outer, frame = create_scrollable_frame(app, padding=22)
    outer.pack(fill="both", expand=True)
    header = ttk.Frame(frame)
    header.pack(fill="x")
    ttk.Label(header, text="OLED Measurement App", font=("Segoe UI", 22, "bold")).pack(side="left")
    ttk.Button(header, text="Настройки", command=app.open_settings_window, width=16).pack(side="right")

    mode_text = f"Режим: {hardware_mode_label(app.app_settings)}"
    if app.app_settings.get("hardware_mode") == "simulator":
        mode_text += "  |  измерения идут на эмуляторе"
    ttk.Label(frame, text=mode_text, font=("Segoe UI", 11, "bold")).pack(anchor="w", pady=(4, 2))
    build_hardware_status_bar(app, frame)
    app.after(250, lambda: check_hardware_status(app))
    ttk.Label(frame, text="Выберите существующую серию или создайте новую.", font=("Segoe UI", 11)).pack(anchor="w", pady=(0, 14))

    root_bar = ttk.Frame(frame)
    root_bar.pack(fill="x", pady=(0, 10))
    ttk.Label(root_bar, text="Корневая папка серий:").grid(row=0, column=0, sticky="w", padx=(0, 8))
    root_var = tk.StringVar(value=str(app.app_settings.get("default_root") or SCRIPT_DIR / DEFAULT_ROOT))
    root_entry = ttk.Entry(root_bar, textvariable=root_var, width=78)
    root_entry.grid(row=0, column=1, sticky="we")
    root_bar.columnconfigure(1, weight=1)
    ttk.Button(root_bar, text="Обзор", command=lambda: browse_root_and_refresh(app, root_var)).grid(row=0, column=2, padx=(8, 0))
    ttk.Button(root_bar, text="Обновить", command=lambda: refresh_series_list(app, root_var.get())).grid(row=0, column=3, padx=(8, 0))

    table, app.series_tree = create_tree_with_scrollbars(
        frame,
        columns=("deposition", "keyword", "created", "measurements", "folder"),
        height=12,
    )
    app.series_tree.heading("deposition", text="Дата напыления")
    app.series_tree.heading("keyword", text="Кодовое слово")
    app.series_tree.heading("created", text="Создана")
    app.series_tree.heading("measurements", text="Измерений")
    app.series_tree.heading("folder", text="Папка")
    app.series_tree.column("deposition", width=150, minwidth=130, stretch=False)
    app.series_tree.column("keyword", width=190, minwidth=140, stretch=False)
    app.series_tree.column("created", width=210, minwidth=170, stretch=False)
    app.series_tree.column("measurements", width=110, minwidth=90, anchor="center", stretch=False)
    app.series_tree.column("folder", width=760, minwidth=420, stretch=True)
    table.pack(fill="both", expand=True)
    app.series_tree.bind("<Double-1>", lambda _event: open_selected_series(app))

    buttons = ttk.Frame(frame)
    buttons.pack(fill="x", pady=(12, 0))
    ttk.Button(buttons, text="Открыть выбранную серию", command=lambda: open_selected_series(app), width=28).pack(side="left")
    ttk.Button(buttons, text="Открыть папку вручную", command=lambda: open_existing_series(app), width=24).pack(side="left", padx=(10, 0))
    ttk.Button(buttons, text="Создать новую серию", command=app.show_new_series_screen, width=24).pack(side="right")

    app.log_widget = ScrolledText(frame, height=8, state="disabled")
    app.log_widget.pack(fill="x", pady=(14, 0))
    refresh_series_list(app, root_var.get())


def browse_root_and_refresh(app, root_var: tk.StringVar) -> None:
    folder = filedialog.askdirectory(title="Корневая папка для серий")
    if not folder:
        return
    root_var.set(folder)
    app.app_settings["default_root"] = folder
    save_app_settings(app.app_settings)
    refresh_series_list(app, folder)


def find_existing_series(root_folder: Path) -> List[Dict[str, Any]]:
    root = Path(root_folder)
    if not root.exists():
        return []
    found: List[Dict[str, Any]] = []
    for cfg_path in sorted(root.rglob(CONFIG_FILE)):
        folder = cfg_path.parent
        try:
            cfg = json.loads(cfg_path.read_text(encoding="utf-8"))
        except Exception:
            cfg = {}
        measurements_count: Any = ""
        journal_path = folder / JOURNAL_FILE
        if journal_path.exists():
            try:
                wb = load_workbook(journal_path, data_only=True, read_only=True)
                if MEASUREMENTS_SHEET in wb.sheetnames:
                    measurements_count = max(wb[MEASUREMENTS_SHEET].max_row - 1, 0)
                wb.close()
            except Exception:
                measurements_count = "?"
        found.append(
            {
                "folder": folder,
                "deposition_date": cfg.get("deposition_date", ""),
                "keyword": cfg.get("keyword", ""),
                "created_at": cfg.get("created_at", ""),
                "measurements_count": measurements_count,
            }
        )
    return found


def refresh_series_list(app, root_folder: str) -> None:
    if not hasattr(app, "series_tree"):
        return
    app.app_settings["default_root"] = str(root_folder)
    save_app_settings(app.app_settings)
    for item in app.series_tree.get_children():
        app.series_tree.delete(item)
    series_list = find_existing_series(Path(root_folder))
    for item in series_list:
        folder = item["folder"]
        app.series_tree.insert(
            "",
            "end",
            iid=str(folder.resolve()),
            values=(
                item.get("deposition_date", "") or "",
                item.get("keyword", "") or "",
                item.get("created_at", "") or "",
                item.get("measurements_count", "") if item.get("measurements_count", "") is not None else "",
                str(folder),
            ),
        )
    app.log(f"Найдено серий: {len(series_list)}. Режим оборудования: {hardware_mode_label(app.app_settings)}.")


def open_selected_series(app) -> None:
    if not hasattr(app, "series_tree"):
        return
    selection = app.series_tree.selection()
    if not selection:
        messagebox.showwarning("Серия", "Выберите серию в списке.")
        return
    open_series_folder(app, Path(selection[0]))


def open_existing_series(app) -> None:
    folder = filedialog.askdirectory(title="Выберите папку серии, где лежит series_config.json")
    if not folder:
        return
    open_series_folder(app, Path(folder))


def open_series_folder(app, folder: Path) -> None:
    try:
        app.series = SeriesManager(folder)
        app.show_measurement_menu()
    except Exception as exc:
        messagebox.showerror("Не удалось открыть серию", str(exc))


def show_new_series_screen(app) -> None:
    show_series_settings_screen(app, edit_mode=False)


def show_edit_series_screen(app) -> None:
    if app.series is None:
        return
    show_series_settings_screen(app, edit_mode=True)


def show_series_settings_screen(app, edit_mode: bool = False) -> None:
    app.clear()
    outer, main = create_scrollable_frame(app, padding=22)
    outer.pack(fill="both", expand=True)
    title = "Настройки серии" if edit_mode else "Новая серия напыления"
    ttk.Label(main, text=title, font=("Segoe UI", 18, "bold")).pack(anchor="w")

    top = ttk.Frame(main)
    top.pack(fill="x", pady=(14, 10))
    row_offset = 0
    if edit_mode and app.series is not None:
        ttk.Label(top, text="Папка серии:").grid(row=0, column=0, sticky="w", padx=(0, 6), pady=4)
        ttk.Label(top, text=str(app.series.series_folder), foreground="#555555").grid(row=0, column=1, sticky="w", pady=4)
        root_var = tk.StringVar(value=str(app.series.series_folder.parent))
        row_offset = 1
    else:
        ttk.Label(top, text="Корневая папка:").grid(row=0, column=0, sticky="w", padx=(0, 6), pady=4)
        root_var = tk.StringVar(value=str(app.app_settings.get("default_root") or SCRIPT_DIR / DEFAULT_ROOT))
        ttk.Entry(top, textvariable=root_var, width=75).grid(row=0, column=1, sticky="we", pady=4)
        ttk.Button(top, text="Обзор", command=lambda: browse_root(root_var)).grid(row=0, column=2, padx=(6, 0), pady=4)
    top.columnconfigure(1, weight=1)

    config = app.series.config if edit_mode and app.series is not None else {}
    ttk.Label(top, text="Дата напыления:").grid(row=row_offset + 1, column=0, sticky="w", padx=(0, 6), pady=4)
    dep_date_var = tk.StringVar(value=str(config.get("deposition_date") or today_iso()))
    ttk.Entry(top, textvariable=dep_date_var, width=20).grid(row=row_offset + 1, column=1, sticky="w", pady=4)

    ttk.Label(top, text="Кодовое слово:").grid(row=row_offset + 2, column=0, sticky="w", padx=(0, 6), pady=4)
    keyword_var = tk.StringVar(value=str(config.get("keyword") or ""))
    ttk.Entry(top, textvariable=keyword_var, width=30).grid(row=row_offset + 2, column=1, sticky="w", pady=4)

    setup_frame = ttk.LabelFrame(main, text="Журнал серии: четверти, цвет и короткое описание")
    setup_frame.pack(fill="x", pady=(4, 10))
    setup_frame.columnconfigure(0, weight=1)
    series_color_var = tk.StringVar(value=led_color_label(quarter_led_color(config, 1)))
    color_bar = ttk.Frame(setup_frame)
    color_bar.grid(row=0, column=0, sticky="w", padx=10, pady=(10, 0))
    ttk.Label(color_bar, text="Цвет светодиодов серии:").pack(side="left", padx=(0, 8))
    ttk.Combobox(
        color_bar,
        textvariable=series_color_var,
        values=list(LED_COLOR_LABELS.values()),
        state="readonly",
        width=16,
    ).pack(side="left")
    holder_canvas = tk.Canvas(setup_frame, width=930, height=560, background="white", highlightthickness=1, highlightbackground="#D0D7DE")
    holder_canvas.grid(row=1, column=0, sticky="ew", padx=10, pady=(8, 8))

    quarter_vars = build_quarter_input_vars(config)
    layout = build_holder_layout(930, 560)
    for q in [2, 1, 3, 4]:
        info = layout[q]
        x, y = info["entry_xy"]
        control_y = y if q in {1, 2} else y + 54
        entry = ttk.Entry(holder_canvas, textvariable=quarter_vars[str(q)]["base"], width=9)
        holder_canvas.create_window(x, control_y, window=entry, anchor="w", tags=("controls",))
        desc = ttk.Entry(holder_canvas, textvariable=quarter_vars[str(q)]["description"], width=18)
        holder_canvas.create_window(x, control_y + 28, window=desc, anchor="w", tags=("controls",))

    def refresh_holder(*_args) -> None:
        render_series_setup_holder(holder_canvas, quarter_vars, series_color_var)

    for q_vars in quarter_vars.values():
        q_vars["base"].trace_add("write", refresh_holder)
        q_vars["description"].trace_add("write", refresh_holder)
    series_color_var.trace_add("write", refresh_holder)
    refresh_holder()

    ttk.Label(
        setup_frame,
        text="В поле четверти задается короткая база, например C. Цвет добавляет последнюю букву: C + красный = CR.",
        foreground="#555555",
        wraplength=880,
    ).grid(row=2, column=0, sticky="w", padx=10, pady=(0, 10))

    bottom = ttk.Frame(main)
    bottom.pack(fill="x", pady=(12, 0))
    back_command = app.show_measurement_menu if edit_mode else app.show_start_screen
    ttk.Button(bottom, text="Назад", command=back_command).pack(side="left")
    button_text = "Сохранить настройки серии" if edit_mode else "Создать серию"
    ttk.Button(
        bottom,
        text=button_text,
        command=lambda: save_series_settings(app, edit_mode, root_var, dep_date_var, keyword_var, quarter_vars, series_color_var),
    ).pack(side="right")


def browse_root(root_var: tk.StringVar) -> None:
    folder = filedialog.askdirectory(title="Корневая папка для серий")
    if folder:
        root_var.set(folder)


def build_quarter_input_vars(config: Dict[str, Any]) -> Dict[str, Dict[str, tk.StringVar]]:
    result: Dict[str, Dict[str, tk.StringVar]] = {}
    for q in range(1, 5):
        result[str(q)] = {
            "base": tk.StringVar(value=quarter_base(config, q) if config else "Q"),
            "description": tk.StringVar(value=quarter_description(config, q)),
        }
    return result


def collect_quarter_payload(quarter_vars: Dict[str, Dict[str, tk.StringVar]], series_color_var: tk.StringVar):
    quarter_bases = {str(q): quarter_vars[str(q)]["base"].get().strip() or "Q" for q in range(1, 5)}
    quarter_descriptions = {str(q): quarter_vars[str(q)]["description"].get().strip() for q in range(1, 5)}
    series_color = led_color_from_label(series_color_var.get())
    quarter_led_colors = {str(q): series_color for q in range(1, 5)}
    return quarter_bases, quarter_descriptions, quarter_led_colors


def render_series_setup_holder(canvas: tk.Canvas, quarter_vars: Dict[str, Dict[str, tk.StringVar]], series_color_var: tk.StringVar) -> None:
    canvas.delete("drawing")
    width = int(canvas.cget("width"))
    height = int(canvas.cget("height"))
    layout = build_holder_layout(width, height)
    canvas.create_text(width / 2, 22, text="Предпросмотр имен пикселей в журнале серии", fill="#17345F", font=("Segoe UI", 10, "bold"), tags=("drawing",))

    config = {
        "quarter_bases": {str(q): quarter_vars[str(q)]["base"].get().strip() or "Q" for q in range(1, 5)},
        "series_led_color": led_color_from_label(series_color_var.get()),
        "quarter_led_colors": {str(q): led_color_from_label(series_color_var.get()) for q in range(1, 5)},
        "quarter_descriptions": {str(q): quarter_vars[str(q)]["description"].get().strip() for q in range(1, 5)},
    }
    for q in [2, 1, 3, 4]:
        info = layout[q]
        code = quarter_code(config, q)
        desc = quarter_description(config, q)
        canvas.create_text(*info["number_xy"], text=str(q), font=("Segoe UI", 24, "bold"), fill="#17345F", tags=("drawing",))
        ex, ey = info["entry_xy"]
        label_y = ey if q in {1, 2} else ey + 54
        canvas.create_text(ex + 138, label_y, text=f"-> {code}", anchor="w", font=("Segoe UI", 8, "bold"), fill="#0B61A4", tags=("drawing",))
        if desc:
            canvas.create_text(ex + 138, label_y + 18, text=desc, anchor="w", font=("Segoe UI", 8), fill="#555555", tags=("drawing",))
        for substrate in info["substrates"]:
            x, y, w, h = substrate["x"], substrate["y"], substrate["w"], substrate["h"]
            substrate_id = f"{code}{q}_{substrate['substrate_number']}"
            canvas.create_text(x + w / 2, y - 10, text=substrate_id, font=("Segoe UI", 8, "bold"), fill="#17345F", tags=("drawing",))
            canvas.create_rectangle(x, y, x + w, y + h, fill="#FFFFFF", outline="#17345F", width=2, tags=("drawing",))
            for pix in range(1, 5):
                px, py, pw, ph = setup_pixel_rect(x, y, w, h, pix)
                canvas.create_rectangle(px, py, px + pw, py + ph, fill="#FFFFFF", outline="#808080", tags=("drawing",))
                canvas.create_text(px + pw / 2, py + ph / 2, text=str(pix), font=("Segoe UI", 7), tags=("drawing",))


def setup_pixel_rect(x: float, y: float, w: float, h: float, pixel_number: int):
    pad_x = 10
    pad_top = 13
    pad_bottom = 8
    gap = 5
    inner_w = (w - 2 * pad_x - gap) / 2
    inner_h = (h - pad_top - pad_bottom - gap) / 2
    row = 0 if pixel_number in {1, 2} else 1
    col = 0 if pixel_number in {1, 3} else 1
    return x + pad_x + col * (inner_w + gap), y + pad_top + row * (inner_h + gap), inner_w, inner_h


def save_series_settings(
    app,
    edit_mode: bool,
    root_var: tk.StringVar,
    dep_date_var: tk.StringVar,
    keyword_var: tk.StringVar,
    quarter_vars: Dict[str, Dict[str, tk.StringVar]],
    series_color_var: tk.StringVar,
) -> None:
    try:
        quarter_bases, quarter_descriptions, quarter_led_colors = collect_quarter_payload(quarter_vars, series_color_var)
        deposition_date = dep_date_var.get().strip() or today_iso()
        keyword = keyword_var.get().strip()
        if edit_mode:
            if app.series is None:
                raise ValueError("Серия не открыта")
            app.series.update_config(deposition_date, keyword, quarter_bases, quarter_descriptions, quarter_led_colors)
            app.log(f"Обновлены настройки серии: {app.series.series_folder}")
        else:
            app.app_settings["default_root"] = root_var.get()
            save_app_settings(app.app_settings)
            app.series = SeriesManager.create_new(
                Path(root_var.get()),
                deposition_date,
                keyword,
                quarter_bases,
                quarter_descriptions,
                quarter_led_colors,
            )
            app.log(f"Создана серия: {app.series.series_folder}")
        app.show_measurement_menu()
    except Exception as exc:
        title = "Не удалось сохранить серию" if edit_mode else "Не удалось создать серию"
        messagebox.showerror(title, str(exc))

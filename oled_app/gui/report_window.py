"""Origin report window for the modular OLED app."""

from __future__ import annotations

import subprocess
import sys
import threading
import tkinter as tk
from pathlib import Path
from tkinter import filedialog, messagebox, ttk
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import load_workbook

from oled_app.constants import MEASUREMENT_FOLDER_NAMES, SCRIPT_DIR
from oled_app.reports.origin_report import (
    REPORT_MODE_FULL,
    REPORT_MODE_IVL,
    REPORT_MODE_SPECTRA,
    series_quarter_number,
)
from oled_app.utils import (
    build_report_voltage_grid,
    format_voltage,
    parse_float,
    voltage_grid_missing,
)

from .widgets import fit_toplevel_to_content


def read_report_spectrum_voltages(path: Path) -> List[float]:
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
    except Exception:
        return []
    try:
        sheet_name = "Processed counts per s"
        if sheet_name not in wb.sheetnames:
            return []
        ws = wb[sheet_name]
        voltage_row = None
        for row in range(1, min(ws.max_row, 30) + 1):
            if ws.cell(row=row, column=1).value == "V set (V)":
                voltage_row = row
                break
        if voltage_row is None:
            return []
        voltages = []
        for column in range(2, ws.max_column + 1):
            value = ws.cell(row=voltage_row, column=column).value
            if isinstance(value, (int, float)):
                voltages.append(round(float(value), 6))
        return voltages
    finally:
        wb.close()


def collect_report_spectrum_candidates(
    app,
    date_filter: Optional[str] = None,
    excluded_quarters: Optional[set[int]] = None,
) -> Dict[str, Dict[str, Dict[str, Dict[str, Any]]]]:
    assert app.series is not None
    excluded_quarters = excluded_quarters or set()
    spectra_root = app.series.series_folder / "measurements" / MEASUREMENT_FOLDER_NAMES["SPECTRUM"]
    if not spectra_root.exists():
        return {}
    latest_by_pixel: Dict[str, Tuple[float, str, str, Path]] = {}
    for path in spectra_root.rglob("SPECTRUM_*.xlsx"):
        try:
            rel = path.relative_to(spectra_root)
        except Exception:
            continue
        parts = rel.parts
        if len(parts) < 5:
            continue
        if date_filter and parts[0] != date_filter:
            continue
        series = parts[1]
        if series_quarter_number(series) in excluded_quarters:
            continue
        substrate = parts[2]
        pixel = parts[3]
        mtime = path.stat().st_mtime
        prev = latest_by_pixel.get(pixel)
        if prev is None or mtime > prev[0]:
            latest_by_pixel[pixel] = (mtime, series, substrate, path)

    candidates: Dict[str, Dict[str, Dict[str, Dict[str, Any]]]] = {}
    for pixel, (_mtime, series, substrate, path) in latest_by_pixel.items():
        voltages = read_report_spectrum_voltages(path)
        if voltages:
            candidates.setdefault(series, {}).setdefault(substrate, {})[pixel] = {
                "file": path,
                "voltages": voltages,
            }
    return candidates


def measurement_dates_for_report(app, measurement_type: str) -> List[str]:
    assert app.series is not None
    folder = app.series.series_folder / "measurements" / MEASUREMENT_FOLDER_NAMES[measurement_type]
    if not folder.exists():
        return []
    return sorted([path.name for path in folder.iterdir() if path.is_dir()])


def selected_report_candidates(
    candidates: Dict[str, Dict[str, Dict[str, Dict[str, Any]]]],
    substrate_selection_vars: Dict[str, tk.StringVar],
    pixel_selection_vars: Dict[str, tk.StringVar],
) -> Dict[str, Dict[str, Any]]:
    selected: Dict[str, Dict[str, Any]] = {}
    for series, substrate_var in substrate_selection_vars.items():
        substrate = substrate_var.get().strip()
        pixel_var = pixel_selection_vars.get(series)
        pixel = pixel_var.get().strip() if pixel_var is not None else ""
        info = candidates.get(series, {}).get(substrate, {}).get(pixel)
        if info:
            selected[pixel] = {"series": series, "subseries": substrate, **info}
    return selected


def default_report_step(voltages: List[float]) -> float:
    values = sorted({round(float(value), 6) for value in voltages})
    diffs = [round(b - a, 6) for a, b in zip(values, values[1:]) if b > a]
    return min(diffs) if diffs else 0.1


def common_report_voltages(selected: Dict[str, Dict[str, Any]]) -> List[float]:
    common: Optional[set] = None
    for info in selected.values():
        values = {round(float(value), 6) for value in info["voltages"]}
        common = values if common is None else common & values
    return sorted(common or [])


def report_output_name(
    ivl_date: str,
    spectrum_date: str,
    suffix: str = ".opju",
    report_mode: str = REPORT_MODE_FULL,
) -> str:
    if report_mode == REPORT_MODE_IVL:
        return f"report_IVL_{ivl_date}{suffix}"
    if report_mode == REPORT_MODE_SPECTRA:
        return f"report_Spctr_{spectrum_date}{suffix}"
    if ivl_date == spectrum_date:
        stem = ivl_date
    else:
        stem = f"IVL_{ivl_date}_Spctr_{spectrum_date}"
    return f"report_{stem}{suffix}"


def open_report_window(app) -> None:
    if app.series is None:
        return
    ivl_dates = measurement_dates_for_report(app, "IVL")
    spectrum_dates = measurement_dates_for_report(app, "SPECTRUM")
    if not ivl_dates and not spectrum_dates:
        messagebox.showwarning("Отчет", "В серии не найдены ВАЯХ и спектры для отчета.", parent=app)
        return

    available_modes = []
    if ivl_dates and spectrum_dates:
        available_modes.append(REPORT_MODE_FULL)
    if ivl_dates:
        available_modes.append(REPORT_MODE_IVL)
    if spectrum_dates:
        available_modes.append(REPORT_MODE_SPECTRA)
    report_mode_var = tk.StringVar(value=available_modes[0])
    ivl_date_var = tk.StringVar(value=ivl_dates[-1] if ivl_dates else "")
    spectrum_date_var = tk.StringVar(value=spectrum_dates[-1] if spectrum_dates else "")
    excluded_quarter_vars = {quarter: tk.BooleanVar(value=False) for quarter in range(1, 5)}

    def selected_excluded_quarters() -> set[int]:
        return {quarter for quarter, var in excluded_quarter_vars.items() if var.get()}

    candidates = (
        collect_report_spectrum_candidates(app, spectrum_date_var.get(), selected_excluded_quarters())
        if spectrum_dates
        else {}
    )

    win = tk.Toplevel(app)
    win.title("Составить отчет")
    win.geometry("760x680")
    win.transient(app)
    main = ttk.Frame(win, padding=14)
    main.pack(fill="both", expand=True)

    mode_frame = ttk.LabelFrame(main, text="Состав отчета")
    mode_frame.pack(fill="x", pady=(0, 10))
    for column, (value, label) in enumerate(
        (
            (REPORT_MODE_FULL, "Полный отчет"),
            (REPORT_MODE_IVL, "Только ВАЯХ"),
            (REPORT_MODE_SPECTRA, "Только спектры"),
        )
    ):
        ttk.Radiobutton(
            mode_frame,
            text=label,
            variable=report_mode_var,
            value=value,
            state="normal" if value in available_modes else "disabled",
        ).grid(row=0, column=column, sticky="w", padx=10, pady=6)

    date_frame = ttk.LabelFrame(main, text="Даты измерений")
    date_frame.pack(fill="x", pady=(0, 10))
    ttk.Label(date_frame, text="ВАЯХ:").grid(row=0, column=0, sticky="e", padx=(8, 4), pady=6)
    ivl_date_combo = ttk.Combobox(date_frame, values=ivl_dates, textvariable=ivl_date_var, state="readonly", width=14)
    ivl_date_combo.grid(row=0, column=1, sticky="w", padx=(0, 16), pady=6)
    ttk.Label(date_frame, text="Спектры:").grid(row=0, column=2, sticky="e", padx=(8, 4), pady=6)
    spectrum_date_combo = ttk.Combobox(date_frame, values=spectrum_dates, textvariable=spectrum_date_var, state="readonly", width=14)
    spectrum_date_combo.grid(row=0, column=3, sticky="w", padx=(0, 8), pady=6)

    quarter_frame = ttk.LabelFrame(main, text="Исключить четверти из отчёта")
    quarter_frame.pack(fill="x", pady=(0, 10))
    ttk.Label(quarter_frame, text="Не включать:").pack(side="left", padx=(8, 4), pady=6)
    for quarter in range(1, 5):
        ttk.Checkbutton(
            quarter_frame,
            text=str(quarter),
            variable=excluded_quarter_vars[quarter],
            command=lambda: change_excluded_quarters(),
        ).pack(side="left", padx=6, pady=6)

    spectrum_options_frame = ttk.Frame(main)
    spectrum_options_frame.pack(fill="both", expand=True)
    ttk.Label(spectrum_options_frame, text="Выбор спектров для отчета", font=("Segoe UI", 12, "bold")).pack(anchor="w")
    selection_frame = ttk.LabelFrame(spectrum_options_frame, text="Подложка и пиксель на подсерии")
    selection_frame.pack(fill="x", pady=(8, 10))
    substrate_selection_vars: Dict[str, tk.StringVar] = {}
    pixel_selection_vars: Dict[str, tk.StringVar] = {}

    def rebuild_selection() -> None:
        for widget in selection_frame.winfo_children():
            widget.destroy()
        substrate_selection_vars.clear()
        pixel_selection_vars.clear()
        if not candidates:
            ttk.Label(selection_frame, text="За выбранную дату спектры не найдены.", foreground="#555555").grid(row=0, column=0, sticky="w", padx=8, pady=6)
            return

        for column, header in enumerate(("Подсерия", "Подложка", "Пиксель", "")):
            ttk.Label(selection_frame, text=header, font=("Segoe UI", 9, "bold")).grid(
                row=0,
                column=column,
                sticky="w",
                padx=(8, 6),
                pady=(5, 3),
            )

        for row, series in enumerate(sorted(candidates), start=1):
            substrates = sorted(candidates[series])
            substrate_selection_vars[series] = tk.StringVar(value=substrates[0])
            pixels = sorted(candidates[series][substrates[0]])
            pixel_selection_vars[series] = tk.StringVar(value=pixels[0])
            pixel_selection_vars[series].trace_add("write", refresh_defaults)

            ttk.Label(selection_frame, text=series + ":").grid(row=row, column=0, sticky="e", padx=(8, 6), pady=3)
            substrate_combo = ttk.Combobox(
                selection_frame,
                values=substrates,
                textvariable=substrate_selection_vars[series],
                state="readonly",
                width=18,
            )
            substrate_combo.grid(row=row, column=1, sticky="w", padx=(0, 8), pady=3)
            pixel_combo = ttk.Combobox(
                selection_frame,
                values=pixels,
                textvariable=pixel_selection_vars[series],
                state="readonly",
                width=18,
            )
            pixel_combo.grid(row=row, column=2, sticky="w", padx=(0, 8), pady=3)

            def update_pixels(_event=None, *, selected_series=series, combo=pixel_combo) -> None:
                substrate = substrate_selection_vars[selected_series].get()
                available_pixels = sorted(candidates.get(selected_series, {}).get(substrate, {}))
                combo.configure(values=available_pixels)
                pixel_selection_vars[selected_series].set(available_pixels[0] if available_pixels else "")

            substrate_combo.bind("<<ComboboxSelected>>", update_pixels)
            if len(substrates) > 1:
                note = "несколько снятых подложек — выберите одну"
            elif len(pixels) > 1:
                note = "выберите пиксель"
            else:
                note = "выбрано автоматически"
            ttk.Label(selection_frame, text=note, foreground="#555555").grid(row=row, column=3, sticky="w", padx=(0, 8), pady=3)

    output_manual = {"value": False}
    output_var = tk.StringVar(
        value=str(
            app.series.series_folder
            / report_output_name(ivl_date_var.get(), spectrum_date_var.get(), report_mode=report_mode_var.get())
        )
    )

    def refresh_output_name(force: bool = False) -> None:
        if output_manual["value"] and not force:
            return
        current = Path(output_var.get().strip() or "report.opju")
        suffix = current.suffix if current.suffix.lower() in {".opju", ".xlsx"} else ".opju"
        output_var.set(
            str(
                app.series.series_folder
                / report_output_name(
                    ivl_date_var.get(),
                    spectrum_date_var.get(),
                    suffix,
                    report_mode_var.get(),
                )
            )
        )
    out_frame = ttk.Frame(main)
    out_frame.pack(fill="x", pady=(0, 10))
    ttk.Label(out_frame, text="Файл отчета:").pack(side="left")
    output_entry = ttk.Entry(out_frame, textvariable=output_var, width=68)
    output_entry.pack(side="left", padx=8, fill="x", expand=True)
    output_entry.bind("<KeyRelease>", lambda _event: output_manual.__setitem__("value", True))

    def browse_output() -> None:
        filename = filedialog.asksaveasfilename(
            parent=win,
            title="Куда сохранить отчет",
            initialdir=str(app.series.series_folder),
            initialfile=Path(output_var.get()).name,
            defaultextension=".opju",
            filetypes=[("Origin project", "*.opju"), ("Debug Excel", "*.xlsx"), ("All files", "*.*")],
        )
        if filename:
            output_manual["value"] = True
            output_var.set(filename)

    ttk.Button(out_frame, text="Выбрать", command=browse_output).pack(side="left")

    same_grid_var = tk.BooleanVar(value=True)
    ttk.Checkbutton(
        spectrum_options_frame,
        text="Одинаковый диапазон и шаг напряжения для всех выбранных пикселей",
        variable=same_grid_var,
    ).pack(anchor="w", pady=(0, 6))

    grid_frame = ttk.LabelFrame(spectrum_options_frame, text="Диапазон напряжения")
    grid_frame.pack(fill="x", pady=(0, 10))
    global_vars = {"start": tk.StringVar(value=""), "stop": tk.StringVar(value=""), "step": tk.StringVar(value="")}
    for col, (key, label) in enumerate((("start", "Начало, В"), ("stop", "Конец, В"), ("step", "Шаг, В"))):
        ttk.Label(grid_frame, text=label).grid(row=0, column=col * 2, sticky="e", padx=(8, 4), pady=6)
        ttk.Entry(grid_frame, textvariable=global_vars[key], width=10).grid(row=0, column=col * 2 + 1, sticky="w", padx=(0, 8), pady=6)

    per_pixel_frame = ttk.LabelFrame(spectrum_options_frame, text="Индивидуальные диапазоны")
    per_pixel_vars: Dict[str, Dict[str, tk.StringVar]] = {}

    def update_per_pixel_visibility(*_args) -> None:
        if same_grid_var.get():
            per_pixel_frame.pack_forget()
        else:
            per_pixel_frame.pack(fill="both", expand=True, pady=(0, 10))

    status_var = tk.StringVar(value="")
    status_label = ttk.Label(main, textvariable=status_var, foreground="#555555", wraplength=700)
    status_label.pack(anchor="w", pady=(0, 8))
    same_grid_var.trace_add("write", update_per_pixel_visibility)

    def refresh_defaults(*_args) -> None:
        if report_mode_var.get() == REPORT_MODE_IVL:
            status_var.set("В отчет войдет только раздел ВАЯХ.")
            return
        selected = selected_report_candidates(candidates, substrate_selection_vars, pixel_selection_vars)
        common = common_report_voltages(selected)
        if common:
            global_vars["start"].set(format_voltage(common[0]))
            global_vars["stop"].set(format_voltage(common[-1]))
            global_vars["step"].set(format_voltage(default_report_step(common)))
            status_var.set(
                f"Общие доступные напряжения: {format_voltage(common[0])}...{format_voltage(common[-1])} В; "
                f"максимум для общего графика {format_voltage(common[-1])} В."
            )
        else:
            status_var.set("У выбранных спектров нет общего напряжения для общего графика.")

        for widget in per_pixel_frame.winfo_children():
            widget.destroy()
        per_pixel_vars.clear()
        for col, header in enumerate(["Пиксель", "Начало, В", "Конец, В", "Шаг, В", "Доступно"]):
            ttk.Label(per_pixel_frame, text=header, font=("Segoe UI", 9, "bold")).grid(row=0, column=col, sticky="w", padx=6, pady=(6, 3))
        for row, (pixel, info) in enumerate(sorted(selected.items()), start=1):
            voltages = sorted(info["voltages"])
            per_pixel_vars[pixel] = {
                "start": tk.StringVar(value=format_voltage(voltages[0])),
                "stop": tk.StringVar(value=format_voltage(voltages[-1])),
                "step": tk.StringVar(value=format_voltage(default_report_step(voltages))),
            }
            ttk.Label(per_pixel_frame, text=pixel).grid(row=row, column=0, sticky="w", padx=6, pady=3)
            for col, key in enumerate(("start", "stop", "step"), start=1):
                ttk.Entry(per_pixel_frame, textvariable=per_pixel_vars[pixel][key], width=10).grid(row=row, column=col, sticky="w", padx=6, pady=3)
            ttk.Label(
                per_pixel_frame,
                text=f"{format_voltage(voltages[0])}...{format_voltage(voltages[-1])} В",
                foreground="#555555",
            ).grid(row=row, column=4, sticky="w", padx=6, pady=3)

    def change_spectrum_date(*_args) -> None:
        nonlocal candidates
        candidates = collect_report_spectrum_candidates(
            app,
            spectrum_date_var.get(),
            selected_excluded_quarters(),
        )
        refresh_output_name()
        rebuild_selection()
        refresh_defaults()

    def change_excluded_quarters() -> None:
        nonlocal candidates
        candidates = (
            collect_report_spectrum_candidates(
                app,
                spectrum_date_var.get(),
                selected_excluded_quarters(),
            )
            if spectrum_dates
            else {}
        )
        rebuild_selection()
        refresh_defaults()

    def update_report_mode(*_args) -> None:
        mode = report_mode_var.get()
        ivl_date_combo.configure(state="readonly" if mode in {REPORT_MODE_FULL, REPORT_MODE_IVL} else "disabled")
        spectrum_date_combo.configure(
            state="readonly" if mode in {REPORT_MODE_FULL, REPORT_MODE_SPECTRA} else "disabled"
        )
        if mode == REPORT_MODE_IVL:
            spectrum_options_frame.pack_forget()
        elif not spectrum_options_frame.winfo_manager():
            spectrum_options_frame.pack(fill="both", expand=True, before=out_frame)
        refresh_output_name()
        update_per_pixel_visibility()
        refresh_defaults()

    ivl_date_var.trace_add("write", lambda *_args: refresh_output_name())
    spectrum_date_var.trace_add("write", change_spectrum_date)
    report_mode_var.trace_add("write", update_report_mode)
    rebuild_selection()
    update_report_mode()

    def build_command() -> List[str]:
        mode = report_mode_var.get()
        includes_ivl = mode in {REPORT_MODE_FULL, REPORT_MODE_IVL}
        includes_spectra = mode in {REPORT_MODE_FULL, REPORT_MODE_SPECTRA}
        output_text = output_var.get().strip()
        if not output_text:
            raise ValueError("Не задан файл отчета")
        output = Path(output_text)

        cmd = [
            sys.executable,
            str(SCRIPT_DIR / "scripts" / "build_report_origin_workbook.py"),
            "--measurements-dir",
            str(app.series.series_folder / "measurements"),
            "--output",
            str(output),
            "--report-mode",
            mode,
            "--strict",
        ]
        for quarter in sorted(selected_excluded_quarters()):
            cmd.extend(["--exclude-quarter", str(quarter)])
        if includes_ivl:
            if not ivl_date_var.get():
                raise ValueError("Не выбрана дата ВАЯХ")
            cmd.extend(["--ivl-date", ivl_date_var.get()])
        if not includes_spectra:
            return cmd

        if not spectrum_date_var.get():
            raise ValueError("Не выбрана дата спектров")
        selected = selected_report_candidates(candidates, substrate_selection_vars, pixel_selection_vars)
        if not selected:
            raise ValueError("Не выбран ни один спектральный пиксель")
        cmd.extend(
            [
                "--spectrum-date",
                spectrum_date_var.get(),
                "--require-spectrum-pixel-selection",
            ]
        )
        for pixel, info in sorted(selected.items()):
            cmd.extend(["--spectrum-series-pixel", f"{info['series']}={pixel}"])

        if same_grid_var.get():
            start = parse_float(global_vars["start"].get(), "Начало напряжения")
            stop = parse_float(global_vars["stop"].get(), "Конец напряжения")
            step = parse_float(global_vars["step"].get(), "Шаг напряжения")
            requested = build_report_voltage_grid(start, stop, step)
            for pixel, info in selected.items():
                missing = voltage_grid_missing(requested, info["voltages"])
                if missing:
                    raise ValueError(f"{pixel}: в спектре нет напряжений {', '.join(format_voltage(value) for value in missing[:8])}")
            cmd.extend(["--voltage-start", str(start), "--voltage-stop", str(stop), "--voltage-step", str(step)])
        else:
            selected_grids: Dict[str, List[float]] = {}
            for pixel, info in selected.items():
                vars_for_pixel = per_pixel_vars[pixel]
                start = parse_float(vars_for_pixel["start"].get(), f"{pixel}: начало напряжения")
                stop = parse_float(vars_for_pixel["stop"].get(), f"{pixel}: конец напряжения")
                step = parse_float(vars_for_pixel["step"].get(), f"{pixel}: шаг напряжения")
                requested = build_report_voltage_grid(start, stop, step)
                missing = voltage_grid_missing(requested, info["voltages"])
                if missing:
                    raise ValueError(f"{pixel}: в спектре нет напряжений {', '.join(format_voltage(value) for value in missing[:8])}")
                selected_grids[pixel] = requested
                cmd.extend(["--spectrum-voltage-grid", f"{pixel}={start}:{stop}:{step}"])
            common = set(round(value, 6) for value in next(iter(selected_grids.values())))
            for grid in list(selected_grids.values())[1:]:
                common &= {round(value, 6) for value in grid}
            if not common:
                raise ValueError("У индивидуальных диапазонов нет общего напряжения для общего графика")
        return cmd

    def run_report() -> None:
        try:
            cmd = build_command()
        except Exception as exc:
            messagebox.showerror("Параметры отчета", str(exc), parent=win)
            return
        status_var.set("Отчет создается...")
        run_button.configure(state="disabled")

        def finish(returncode: int, stdout: str, stderr: str) -> None:
            run_button.configure(state="normal")
            if returncode == 0:
                app.log(stdout.strip() or f"Отчет создан: {output_var.get()}")
                messagebox.showinfo("Отчет", f"Отчет создан:\n{output_var.get()}", parent=win)
                win.destroy()
            else:
                details = (stderr or stdout or "Неизвестная ошибка").strip()
                app.log(details)
                status_var.set("Ошибка создания отчета.")
                messagebox.showerror("Ошибка отчета", details, parent=win)

        def worker() -> None:
            try:
                completed = subprocess.run(cmd, cwd=str(SCRIPT_DIR), text=True, capture_output=True)
                app.after(0, finish, completed.returncode, completed.stdout, completed.stderr)
            except Exception as exc:
                app.after(0, finish, 1, "", str(exc))

        threading.Thread(target=worker, daemon=True).start()

    buttons = ttk.Frame(main)
    buttons.pack(fill="x")
    ttk.Button(buttons, text="Закрыть", command=win.destroy).pack(side="right")
    run_button = ttk.Button(buttons, text="Составить отчет", command=run_report)
    run_button.pack(side="right", padx=(0, 8))
    fit_toplevel_to_content(win, 800, 720)

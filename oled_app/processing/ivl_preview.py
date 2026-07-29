"""Small value-free IVL thumbnails for hover previews on the pixel map."""

from __future__ import annotations

import math
import re
from pathlib import Path
from typing import Any, Dict, Iterable, List, Sequence

from openpyxl import load_workbook
from PIL import Image, ImageDraw
from PIL.PngImagePlugin import PngInfo

from oled_app.utils import as_float_or_none, safe_filename


THUMBNAIL_SUFFIX = "_thumbnail.png"
THUMBNAIL_FOLDER = "thumbnails"
OLED_CURRENT_COLOR = "#0B61A4"
PHOTODIODE_CURRENT_COLOR = "#C43C30"
THUMBNAIL_RENDER_VERSION = "2"
THUMBNAIL_RENDER_VERSION_KEY = "OLED IVL thumbnail version"
IVL_TIMESTAMP_SUFFIX_RE = re.compile(
    r"(?:_\d{2}-\d{2}-\d{4}_\d{1,2}h\d{2}m\d{2}s|_\d{8}_\d{6})$",
    re.IGNORECASE,
)


def _pixel_id_from_workbook(workbook_path: Path) -> str:
    stem = Path(workbook_path).stem
    if stem.upper().startswith("IVL_"):
        stem = stem[4:]
    return IVL_TIMESTAMP_SUFFIX_RE.sub("", stem)


def ivl_thumbnail_path(
    workbook_path: Path,
    pixel_id: str | None = None,
) -> Path:
    workbook_path = Path(workbook_path)
    pixel = safe_filename(
        pixel_id or _pixel_id_from_workbook(workbook_path),
        fallback=workbook_path.stem,
    )
    return workbook_path.parent / THUMBNAIL_FOLDER / f"{pixel}{THUMBNAIL_SUFFIX}"


def ivl_thumbnail_needs_refresh(thumbnail_path: Path) -> bool:
    """Return true for missing, legacy, or unreadable thumbnail files."""

    thumbnail_path = Path(thumbnail_path)
    if not thumbnail_path.exists():
        return True
    try:
        with Image.open(thumbnail_path) as image:
            return (
                str(image.info.get(THUMBNAIL_RENDER_VERSION_KEY) or "")
                != THUMBNAIL_RENDER_VERSION
            )
    except Exception:
        return True


def _representative_cycles(
    cycles: Iterable[Dict[str, Any]],
) -> List[Dict[str, Any]]:
    """Exclude dead confirmation sweeps that otherwise cross out the preview."""

    candidates = [cycle for cycle in cycles if cycle.get("data")]
    for index, cycle in enumerate(candidates):
        if str(cycle.get("status") or "").upper() != "BURNED":
            continue
        if index + 1 >= len(candidates):
            return [cycle]
        next_status = str(candidates[index + 1].get("status") or "").upper()
        if next_status in {"BURNED", "NONWORKING", "NO_CONTACT"}:
            return [cycle]
    working = [
        cycle
        for cycle in candidates
        if str(cycle.get("status") or "").upper() == "WORKING"
    ]
    if working:
        return [working[-1]]
    return candidates[-1:] if candidates else []


def _series_points(
    voltages: Sequence[float],
    values: Sequence[float],
    left: int,
    top: int,
    width: int,
    height: int,
    y_min: float,
    y_max: float,
) -> List[tuple[int, int]]:
    if len(voltages) < 2 or len(values) != len(voltages):
        return []
    x_min, x_max = min(voltages), max(voltages)
    if math.isclose(x_min, x_max):
        return []
    if math.isclose(y_min, y_max):
        y_max = y_min + 1.0
    return [
        (
            int(left + (voltage - x_min) / (x_max - x_min) * width),
            int(top + height - (value - y_min) / (y_max - y_min) * height),
        )
        for voltage, value in zip(voltages, values)
    ]


def create_ivl_thumbnail(
    output_path: Path,
    cycles: Iterable[Dict[str, Any]],
    size: tuple[int, int] = (340, 220),
) -> Path:
    """Draw current and photodiode curves without numeric axes or values."""

    width, height = size
    image = Image.new("RGB", size, "white")
    draw = ImageDraw.Draw(image)
    left, top, right, bottom = 28, 18, width - 18, height - 24
    draw.rectangle((left, top, right, bottom), outline="#B7C2CC", width=1)

    any_points = False
    for cycle in _representative_cycles(cycles):
        rows = list(cycle.get("data") or [])
        voltages = [
            as_float_or_none(
                row.get("Voltage OLED / LED measured (V)", row.get("Voltage set (V)"))
            )
            for row in rows
        ]
        currents = [as_float_or_none(row.get("Current OLED / LED (mA)")) for row in rows]
        photo = [as_float_or_none(row.get("Photodiode current (uA)")) for row in rows]
        valid = [
            (float(v), float(i), float(p))
            for v, i, p in zip(voltages, currents, photo)
            if (
                v is not None
                and i is not None
                and p is not None
                and math.isfinite(float(v))
                and math.isfinite(float(i))
                and math.isfinite(float(p))
            )
        ]
        if len(valid) < 2:
            continue
        any_points = True
        v_values = [item[0] for item in valid]
        i_values = [item[1] for item in valid]
        p_values = [item[2] for item in valid]
        combined_values = i_values + p_values
        y_min = min(min(combined_values), 0.0)
        y_max = max(max(combined_values), 0.0)
        y_span = y_max - y_min
        if math.isclose(y_span, 0.0):
            y_span = 1.0
        y_padding = y_span * 0.04
        y_min -= y_padding
        y_max += y_padding
        current_points = _series_points(
            v_values,
            i_values,
            left,
            top,
            right - left,
            bottom - top,
            y_min,
            y_max,
        )
        photo_points = _series_points(
            v_values,
            p_values,
            left,
            top,
            right - left,
            bottom - top,
            y_min,
            y_max,
        )
        if current_points:
            draw.line(current_points, fill=OLED_CURRENT_COLOR, width=3)
        if photo_points:
            draw.line(photo_points, fill=PHOTODIODE_CURRENT_COLOR, width=3)

    if any_points:
        draw.line(
            (left + 8, top + 10, left + 34, top + 10),
            fill=OLED_CURRENT_COLOR,
            width=3,
        )
        draw.text((left + 40, top + 4), "I OLED", fill="#45515C")
        draw.line(
            (left + 100, top + 10, left + 126, top + 10),
            fill=PHOTODIODE_CURRENT_COLOR,
            width=3,
        )
        draw.text((left + 132, top + 4), "PD", fill="#45515C")
    else:
        draw.text((width // 2 - 42, height // 2 - 8), "Нет данных ВАЯХ", fill="#6B7280")

    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    temp_path = output_path.with_name(f".{output_path.stem}.tmp.png")
    metadata = PngInfo()
    metadata.add_text(THUMBNAIL_RENDER_VERSION_KEY, THUMBNAIL_RENDER_VERSION)
    image.save(temp_path, format="PNG", optimize=True, pnginfo=metadata)
    temp_path.replace(output_path)
    return output_path


def create_ivl_thumbnail_from_workbook(workbook_path: Path, output_path: Path | None = None) -> Path:
    workbook_path = Path(workbook_path)
    cycles: List[Dict[str, Any]] = []
    wb = load_workbook(workbook_path, data_only=True, read_only=True)
    try:
        for sheet_name in wb.sheetnames:
            if not sheet_name.startswith("Cycle_"):
                continue
            ws = wb[sheet_name]
            status = str(ws.cell(1, 1).value or "").rsplit("|", 1)[-1].strip().upper()
            header_row = None
            headers: Dict[str, int] = {}
            for row in range(1, min(ws.max_row, 20) + 1):
                candidate = {
                    str(ws.cell(row, column).value or ""): column
                    for column in range(1, ws.max_column + 1)
                }
                if "Current OLED / LED (mA)" in candidate and "Photodiode current (uA)" in candidate:
                    header_row = row
                    headers = candidate
                    break
            if header_row is None:
                continue
            data = []
            for row in range(header_row + 1, ws.max_row + 1):
                point = {
                    "Voltage OLED / LED measured (V)": ws.cell(
                        row,
                        headers.get("Voltage OLED / LED measured (V)", headers.get("Voltage set (V)", 2)),
                    ).value,
                    "Current OLED / LED (mA)": ws.cell(
                        row,
                        headers["Current OLED / LED (mA)"],
                    ).value,
                    "Photodiode current (uA)": ws.cell(
                        row,
                        headers["Photodiode current (uA)"],
                    ).value,
                }
                if any(value not in (None, "") for value in point.values()):
                    data.append(point)
            cycles.append({"status": status, "data": data})
    finally:
        wb.close()
    return create_ivl_thumbnail(output_path or ivl_thumbnail_path(workbook_path), cycles)

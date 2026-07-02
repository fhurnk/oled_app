"""Build spectrum workbooks from raw CSV measurement data."""

from __future__ import annotations

from pathlib import Path
from typing import Any, Dict, Iterable, List, Mapping, Optional, Tuple

import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from oled_app.measurements.raw_io import read_csv_dicts
from oled_app.utils import autosize_columns, now_str, style_header_row


SPECTRUM_SUMMARY_RAW_HEADERS = [
    "point",
    "date_time",
    "voltage_set_V",
    "voltage_led_measured_V",
    "current_led_A",
    "current_led_mA",
    "current_density_mA_cm2",
    "voltage_photodiode_measured_V",
    "current_photodiode_A",
    "current_photodiode_uA",
    "luminance_cd_m2",
    "integration_time_s",
    "status",
    "peak_nm",
    "peak_intensity_processed_counts",
    "fwhm_nm",
    "peaks_detected",
    "peaks_nm",
    "baseline_value_raw_counts",
    "baseline_region_nm",
]

SPECTRUM_SPECTRA_RAW_HEADERS = [
    "point",
    "voltage_set_V",
    "integration_time_s",
    "wavelength_nm",
    "raw_counts",
    "dark_counts",
]


def _float_or_none(value: Any) -> Optional[float]:
    if value in (None, ""):
        return None
    try:
        return float(value)
    except Exception:
        return None


def _int_or_default(value: Any, default: int = 0) -> int:
    parsed = _float_or_none(value)
    if parsed is None:
        return default
    return int(parsed)


def _params_value(params: Any, key: str, default: Any = None) -> Any:
    if isinstance(params, Mapping):
        return params.get(key, default)
    return getattr(params, key, default)


def _round_or_dash(value: Any, digits: int) -> Any:
    parsed = _float_or_none(value)
    if parsed is None:
        return "-"
    return round(float(parsed), digits)


def _float_or_dash(value: Any) -> Any:
    parsed = _float_or_none(value)
    return float(parsed) if parsed is not None else "-"


def _voltage_array_from_params(params: Any) -> np.ndarray:
    start = float(_params_value(params, "voltage_start", 0.0))
    end = float(_params_value(params, "voltage_end", start))
    step = float(_params_value(params, "voltage_step", 1.0))
    values = np.arange(start, end + step / 2, step)
    return np.round(values, 6)


def create_spectrum_workbook(filename: Path, pixel_id: str, params: Any, voltage_array: Iterable[float]) -> Workbook:
    filename.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws_sum = wb.active
    ws_sum.title = "Сводка"
    ws_spec = wb.create_sheet("Спектры")
    ws_norm = wb.create_sheet("Processed counts per s")
    ws_raw = wb.create_sheet("Raw spectra")
    ws_dark = wb.create_sheet("Dark corrected")
    ws_baseline = wb.create_sheet("Baseline")
    ws_desc = wb.create_sheet("Описание полей")

    ws_sum["A1"] = "Спектро-электронное сканирование OLED"
    ws_sum["A1"].font = Font(bold=True, size=14)
    meta = [
        ("Pixel", pixel_id),
        ("Created", now_str()),
        (
            "Voltage range",
            f"{_params_value(params, 'voltage_start', '')}-{_params_value(params, 'voltage_end', '')} В, "
            f"step {_params_value(params, 'voltage_step', '')} В",
        ),
        (
            "Opening voltage stored (V)",
            _params_value(params, "opening_voltage", "") if _params_value(params, "opening_voltage", None) is not None else "",
        ),
        ("Voltage start source", _params_value(params, "voltage_start_source", "")),
        ("Current limit", f"{_params_value(params, 'current_limit_mA', '')} мА"),
        ("Pixel area (mm^2)", _params_value(params, "pixel_area_mm2", "")),
        ("Luminance conversion (cd/m^2 per uA)", _params_value(params, "luminance_cd_m2_per_uA", "")),
        ("LED_TYPE final", _params_value(params, "led_type", "")),
        ("Peak search for T_int", _params_value(params, "peak_search_mode_for_tint", "")),
        ("Derivative peak detection", "YES" if _params_value(params, "peak_detection_enabled", False) else "NO"),
        (
            "T_int range",
            f"{float(_params_value(params, 't_int_min_s', 0))*1000:.2f}-{float(_params_value(params, 't_int_max_s', 0))*1000:.2f} мс",
        ),
        ("Discard first scan after T_int change", "YES" if _params_value(params, "discard_first_scan_after_tint_change", False) else "NO"),
        ("Baseline correction", "YES" if _params_value(params, "baseline_correction_enabled", False) else "NO"),
        ("Saved intensity units", "Спектры: raw counts минус фон; Processed counts per s: то же деленное на T_int"),
    ]
    for idx, (key, value) in enumerate(meta, start=3):
        ws_sum.cell(row=idx, column=1, value=key).font = Font(bold=True)
        ws_sum.cell(row=idx, column=2, value=value)

    headers = [
        "Point",
        "V set (V)",
        "V LED measured (V)",
        "I LED (mA)",
        "J LED (mA/cm^2)",
        "V photodiode measured (V)",
        "I photodiode (uA)",
        "Luminance (cd/m^2)",
        "T_int saved spectrum (ms)",
        "Peak (nm)",
        "Max intensity processed (counts)",
        "FWHM (nm)",
        "Status",
        "Spectra column",
        "Peaks detected",
        "Peaks nm",
        "Background mean raw counts",
        "Background region nm",
    ]
    header_row = 21
    for column, header in enumerate(headers, start=1):
        ws_sum.cell(row=header_row, column=column, value=header)
    style_header_row(ws_sum, header_row, 1, len(headers))
    ws_sum.freeze_panes = f"A{header_row + 1}"

    meta_labels = [
        "Название столбца",
        "Point",
        "V set (V)",
        "V LED measured (V)",
        "I LED (mA)",
        "J LED (mA/cm^2)",
        "V photodiode measured (V)",
        "I photodiode (uA)",
        "Luminance (cd/m^2)",
        "T_int saved spectrum (ms)",
        "Peak (nm)",
        "Max intensity processed (counts/s)",
        "FWHM (nm)",
        "Status",
        "Comment",
        "Peaks detected",
        "Peaks nm",
        "Background mean raw counts",
        "Background region nm",
    ]
    for row, label in enumerate(meta_labels, start=1):
        ws_spec.cell(row=row, column=1, value=label).font = Font(bold=True)
        ws_spec.cell(row=row, column=1).fill = PatternFill("solid", fgColor="E2F0D9")

    data_header_row = 20
    ws_spec.cell(row=data_header_row, column=1, value="Wavelength (nm)").font = Font(bold=True)
    ws_spec.cell(row=data_header_row, column=1).fill = PatternFill("solid", fgColor="D9E1F2")
    for idx, voltage in enumerate(voltage_array, start=1):
        col = idx + 1
        ws_spec.cell(row=1, column=col, value=f"Point {idx}: {float(voltage):.3f} V")
        ws_spec.cell(row=2, column=col, value=idx)
        ws_spec.cell(row=3, column=col, value=float(voltage))
        ws_spec.cell(row=data_header_row, column=col, value=f"Intensity point {idx}, counts corrected")
        ws_spec.cell(row=data_header_row, column=col).font = Font(bold=True)
        ws_spec.cell(row=data_header_row, column=col).fill = PatternFill("solid", fgColor="D9E1F2")

    ws_spec.freeze_panes = f"B{data_header_row + 1}"
    extra_sheets = [
        (ws_norm, "Processed counts per s"),
        (ws_raw, "Raw counts"),
        (ws_dark, "Dark-corrected counts"),
        (ws_baseline, "Background mean counts"),
    ]
    for extra_ws, data_label in extra_sheets:
        for row, label in enumerate(meta_labels, start=1):
            extra_ws.cell(row=row, column=1, value=label).font = Font(bold=True)
            extra_ws.cell(row=row, column=1).fill = PatternFill("solid", fgColor="E2F0D9")
        extra_ws.cell(row=data_header_row, column=1, value="Wavelength (nm)").font = Font(bold=True)
        extra_ws.cell(row=data_header_row, column=1).fill = PatternFill("solid", fgColor="D9E1F2")
        for idx, voltage in enumerate(voltage_array, start=1):
            col = idx + 1
            extra_ws.cell(row=1, column=col, value=f"Point {idx}: {float(voltage):.3f} V")
            extra_ws.cell(row=2, column=col, value=idx)
            extra_ws.cell(row=3, column=col, value=float(voltage))
            extra_ws.cell(row=data_header_row, column=col, value=f"{data_label} point {idx}")
            extra_ws.cell(row=data_header_row, column=col).font = Font(bold=True)
            extra_ws.cell(row=data_header_row, column=col).fill = PatternFill("solid", fgColor="D9E1F2")
        extra_ws.freeze_panes = f"B{data_header_row + 1}"

    ws_desc.append(["Лист", "Поле", "Что означает"])
    style_header_row(ws_desc, 1, 1, 3)
    desc_rows = [
        ("Сводка", "T_int saved spectrum", "Время интегрирования именно того спектра, который записан в столбец."),
        ("Сводка", "Peaks detected / Peaks nm", "Заполняется только если включен флажок поиска пиков производными."),
        ("Спектры", "Intensity", "Raw counts минус среднее значение фона из найденного плоского участка raw-спектра, counts."),
        ("Processed counts per s", "Processed counts/s", "Копия обработанного спектра, деленная на время интегрирования."),
        ("Raw spectra", "Raw counts", "Сырые counts напрямую со спектрометра."),
        ("Dark corrected", "Dark-corrected counts", "Диагностический лист: Raw минус dark spectrum, если dark включен; основная обработка его не использует."),
        ("Baseline", "Background mean counts", "Константный уровень фона: среднее значение найденного плоского участка raw-спектра."),
        ("Спектры", "Строки 1-15", "Метаданные каждого спектра."),
        ("Спектры", "Строка 21+", "Длины волн и интенсивности."),
        ("Важно", "Первый спектр после смены T_int", "Сбрасывается, чтобы не записать старый буфер спектрометра."),
    ]
    for row in desc_rows:
        ws_desc.append(row)

    for ws in wb.worksheets:
        autosize_columns(ws, max_width=45)
    wb.save(filename)
    return wb


def _summary_row_values(summary: Dict[str, str], spectra_col: int) -> List[Any]:
    point = _int_or_default(summary.get("point"), 0)
    voltage = _float_or_none(summary.get("voltage_set_V")) or 0.0
    t_int = _float_or_none(summary.get("integration_time_s"))
    return [
        point,
        float(voltage),
        _round_or_dash(summary.get("voltage_led_measured_V"), 6),
        _round_or_dash(summary.get("current_led_mA"), 6),
        _round_or_dash(summary.get("current_density_mA_cm2"), 6),
        _round_or_dash(summary.get("voltage_photodiode_measured_V"), 6),
        _round_or_dash(summary.get("current_photodiode_uA"), 6),
        _round_or_dash(summary.get("luminance_cd_m2"), 6),
        round(float(t_int) * 1000, 3) if t_int is not None else "-",
        _round_or_dash(summary.get("peak_nm"), 3),
        _round_or_dash(summary.get("peak_intensity_processed_counts"), 1),
        _round_or_dash(summary.get("fwhm_nm"), 3),
        summary.get("status") or "FAILED",
        get_column_letter(spectra_col),
        _int_or_default(summary.get("peaks_detected"), 0),
        summary.get("peaks_nm") or "",
        _round_or_dash(summary.get("baseline_value_raw_counts"), 3),
        summary.get("baseline_region_nm") or "",
    ]


def _meta_values(summary: Dict[str, str], spectra_col: int, has_spectrum: bool) -> List[Any]:
    values = _summary_row_values(summary, spectra_col)
    return [
        f"Point {values[0]}: {float(values[1]):.3f} V",
        values[0],
        values[1],
        values[2],
        values[3],
        values[4],
        values[5],
        values[6],
        values[7],
        values[8],
        values[9],
        values[10],
        values[11],
        values[12],
        "OK: spectrum saved" if has_spectrum else "",
        values[14],
        values[15],
        values[16],
        values[17],
    ]


def _read_summary_rows(path: Path) -> List[Dict[str, str]]:
    rows = read_csv_dicts(path)
    rows.sort(key=lambda item: _int_or_default(item.get("point"), 0))
    return rows


def _read_spectra_rows(path: Path) -> Dict[int, List[Dict[str, str]]]:
    by_point: Dict[int, List[Dict[str, str]]] = {}
    for row in read_csv_dicts(path):
        point = _int_or_default(row.get("point"), 0)
        by_point.setdefault(point, []).append(row)
    for rows in by_point.values():
        rows.sort(key=lambda item: _float_or_none(item.get("wavelength_nm")) or 0.0)
    return by_point


def _arrays_from_spectra_rows(
    rows: List[Dict[str, str]],
    baseline_value: float,
    t_int: float,
) -> Tuple[np.ndarray, np.ndarray, np.ndarray, np.ndarray, np.ndarray, np.ndarray]:
    wavelengths = np.asarray([_float_or_none(row.get("wavelength_nm")) or 0.0 for row in rows], dtype=np.float64)
    raw = np.asarray([_float_or_none(row.get("raw_counts")) or 0.0 for row in rows], dtype=np.float64)
    dark_values = [_float_or_none(row.get("dark_counts")) for row in rows]
    if any(value is not None for value in dark_values):
        dark = np.asarray([value if value is not None else 0.0 for value in dark_values], dtype=np.float64)
        dark_corrected = raw - dark
    else:
        dark_corrected = raw.copy()
    baseline = np.full_like(raw, float(baseline_value))
    corrected = raw - baseline
    normalized = corrected / max(float(t_int), 1e-9)
    return wavelengths, corrected, normalized, raw, dark_corrected, baseline


def build_spectrum_workbook_from_raw_csv(
    summary_raw_csv: Path,
    spectra_raw_csv: Path,
    filename: Path,
    pixel_id: str,
    params: Any,
) -> Path:
    voltage_array = _voltage_array_from_params(params)
    wb = create_spectrum_workbook(filename, pixel_id, params, voltage_array)
    ws_sum = wb["Сводка"]
    ws_spec = wb["Спектры"]
    ws_norm = wb["Processed counts per s"]
    ws_raw = wb["Raw spectra"]
    ws_dark = wb["Dark corrected"]
    ws_baseline = wb["Baseline"]
    summary_header_row = 21
    spectra_data_start = 21

    summary_rows = _read_summary_rows(summary_raw_csv)
    spectra_by_point = _read_spectra_rows(spectra_raw_csv) if Path(spectra_raw_csv).exists() else {}

    for summary in summary_rows:
        point = _int_or_default(summary.get("point"), 0)
        if point <= 0:
            continue
        row = summary_header_row + point
        spectra_col = point + 1
        summary_values = _summary_row_values(summary, spectra_col)
        for column, value in enumerate(summary_values, start=1):
            ws_sum.cell(row=row, column=column, value=value)

        spectra_rows = spectra_by_point.get(point, [])
        if not spectra_rows:
            ws_spec.cell(row=12, column=spectra_col, value=summary.get("status") or "FAILED")
            continue

        meta_values = _meta_values(summary, spectra_col, has_spectrum=True)
        for meta_row, value in enumerate(meta_values, start=1):
            ws_spec.cell(row=meta_row, column=spectra_col, value=value)
            ws_norm.cell(row=meta_row, column=spectra_col, value=value)
            ws_raw.cell(row=meta_row, column=spectra_col, value=value)
            ws_dark.cell(row=meta_row, column=spectra_col, value=value)
            ws_baseline.cell(row=meta_row, column=spectra_col, value=value)

        baseline_value = _float_or_none(summary.get("baseline_value_raw_counts")) or 0.0
        t_int = _float_or_none(summary.get("integration_time_s")) or 1e-9
        wavelengths, corrected, normalized, raw, dark_corrected, baseline = _arrays_from_spectra_rows(
            spectra_rows,
            baseline_value,
            t_int,
        )

        for data_row, wavelength in enumerate(wavelengths, start=spectra_data_start):
            value = round(float(wavelength), 2)
            if ws_spec.cell(row=data_row, column=1).value is None:
                ws_spec.cell(row=data_row, column=1, value=value)
            if ws_norm.cell(row=data_row, column=1).value is None:
                ws_norm.cell(row=data_row, column=1, value=value)
            if ws_raw.cell(row=data_row, column=1).value is None:
                ws_raw.cell(row=data_row, column=1, value=value)
            if ws_dark.cell(row=data_row, column=1).value is None:
                ws_dark.cell(row=data_row, column=1, value=value)
            if ws_baseline.cell(row=data_row, column=1).value is None:
                ws_baseline.cell(row=data_row, column=1, value=value)
        for data_row, intensity in enumerate(corrected, start=spectra_data_start):
            ws_spec.cell(row=data_row, column=spectra_col, value=round(float(intensity), 3))
        for data_row, intensity in enumerate(normalized, start=spectra_data_start):
            ws_norm.cell(row=data_row, column=spectra_col, value=round(float(intensity), 3))
        for data_row, intensity in enumerate(raw, start=spectra_data_start):
            ws_raw.cell(row=data_row, column=spectra_col, value=round(float(intensity), 3))
        for data_row, intensity in enumerate(dark_corrected, start=spectra_data_start):
            ws_dark.cell(row=data_row, column=spectra_col, value=round(float(intensity), 3))
        for data_row, intensity in enumerate(baseline, start=spectra_data_start):
            ws_baseline.cell(row=data_row, column=spectra_col, value=round(float(intensity), 3))

    for ws in wb.worksheets:
        autosize_columns(ws, max_width=45)
    wb.save(filename)
    wb.close()
    return Path(filename)

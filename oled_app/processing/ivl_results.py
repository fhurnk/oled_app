"""Build IVL workbooks from raw CSV measurement data."""

from __future__ import annotations

from pathlib import Path
from typing import Any, Dict, List, Mapping, Optional

from openpyxl import Workbook
from openpyxl.chart import ScatterChart, Series, Reference
from openpyxl.chart.axis import ChartLines
from openpyxl.styles import Font

from oled_app.measurements.raw_io import read_csv_dicts
from oled_app.utils import (
    autosize_columns,
    current_density_mA_cm2,
    luminance_cd_m2_at_voltage,
    now_str,
    style_header_row,
)


IVL_RAW_HEADERS = [
    "cycle",
    "point",
    "date_time",
    "elapsed_s",
    "voltage_set_V",
    "voltage_led_measured_V",
    "current_led_A",
    "voltage_photodiode_measured_V",
    "current_photodiode_A",
    "current_led_mA",
    "current_photodiode_uA",
]

IVL_WORKBOOK_HEADERS = [
    "Point",
    "Voltage set (V)",
    "Voltage OLED / LED measured (V)",
    "Current OLED / LED (mA)",
    "Current density (mA/cm^2)",
    "Voltage photodiode measured (V)",
    "Photodiode current (uA)",
    "Luminance (cd/m^2)",
    "Measurement time (s)",
]


def _float_or_none(value: Any) -> Optional[float]:
    if value in (None, ""):
        return None
    try:
        return float(value)
    except Exception:
        return None


def _int_or_default(value: Any, default: int = 0) -> int:
    parsed = _float_or_none(value)
    if parsed is None:
        return default
    return int(parsed)


def _params_value(params: Any, key: str, default: Any = None) -> Any:
    if isinstance(params, Mapping):
        return params.get(key, default)
    return getattr(params, key, default)


def raw_row_to_workbook_point(row: Dict[str, str], params: Any) -> Dict[str, Any]:
    current_led_mA = _float_or_none(row.get("current_led_mA"))
    if current_led_mA is None:
        current_led_A = _float_or_none(row.get("current_led_A")) or 0.0
        current_led_mA = current_led_A * 1000.0

    current_pd_uA = _float_or_none(row.get("current_photodiode_uA"))
    if current_pd_uA is None:
        current_pd_A = _float_or_none(row.get("current_photodiode_A")) or 0.0
        current_pd_uA = -current_pd_A * 1_000_000.0

    pixel_area = _params_value(params, "pixel_area_mm2", 1.0)
    luminance_coeff = _params_value(params, "luminance_cd_m2_per_uA", 1.0)
    calibration_voltage = _float_or_none(row.get("voltage_set_V"))
    if calibration_voltage is None:
        calibration_voltage = _float_or_none(row.get("voltage_led_measured_V"))
    return {
        "Point": _int_or_default(row.get("point"), 0),
        "Voltage set (V)": _float_or_none(row.get("voltage_set_V")) or 0.0,
        "Voltage OLED / LED measured (V)": _float_or_none(row.get("voltage_led_measured_V")) or 0.0,
        "Current OLED / LED (mA)": float(current_led_mA),
        "Current density (mA/cm^2)": current_density_mA_cm2(current_led_mA, pixel_area),
        "Voltage photodiode measured (V)": _float_or_none(row.get("voltage_photodiode_measured_V")) or 0.0,
        "Photodiode current (uA)": float(current_pd_uA),
        "Luminance (cd/m^2)": luminance_cd_m2_at_voltage(
            current_pd_uA,
            luminance_coeff,
            calibration_voltage,
            _params_value(params, "luminance_calibration_model"),
        ),
        "Measurement time (s)": _float_or_none(row.get("elapsed_s")),
    }


def read_ivl_cycles_from_raw_csv(raw_csv_path: Path, params: Any) -> Dict[int, List[Dict[str, Any]]]:
    cycles: Dict[int, List[Dict[str, Any]]] = {}
    for row in read_csv_dicts(raw_csv_path):
        cycle_number = _int_or_default(row.get("cycle"), 1)
        cycles.setdefault(cycle_number, []).append(raw_row_to_workbook_point(row, params))
    for points in cycles.values():
        points.sort(key=lambda item: int(item.get("Point") or 0))
    return cycles


def merge_ivl_raw_cycles(
    raw_csv_path: Path,
    params: Any,
    cycle_summaries: List[Dict[str, Any]],
) -> List[Dict[str, Any]]:
    raw_cycles = read_ivl_cycles_from_raw_csv(raw_csv_path, params)
    merged: List[Dict[str, Any]] = []
    summary_by_cycle = {int(item.get("cycle", idx + 1)): item for idx, item in enumerate(cycle_summaries)}
    cycle_numbers = sorted(set(raw_cycles) | set(summary_by_cycle))
    for cycle_number in cycle_numbers:
        summary = dict(summary_by_cycle.get(cycle_number, {}))
        summary.setdefault("cycle", cycle_number)
        summary.setdefault("status", "PARTIAL" if raw_cycles.get(cycle_number) else "FAILED")
        summary.setdefault("status_desc", "Данные восстановлены из raw CSV")
        summary.setdefault("current_limit_reached", False)
        summary.setdefault("max_photo_uA", max([row["Photodiode current (uA)"] for row in raw_cycles.get(cycle_number, [])], default=0.0))
        summary.setdefault("max_current_mA", max([row["Current OLED / LED (mA)"] for row in raw_cycles.get(cycle_number, [])], default=0.0))
        summary.setdefault("opening_voltage", None)
        summary["data"] = raw_cycles.get(cycle_number, summary.get("data", []))
        merged.append(summary)
    return merged


def confirmed_burned_cycle(cycles: List[Dict[str, Any]]) -> Optional[int]:
    """Return a burnout cycle only when the next cycle confirms loss of function."""

    for index, cycle in enumerate(cycles):
        if str(cycle.get("status") or "").upper() != "BURNED":
            continue
        cycle_number = int(cycle.get("cycle", index + 1) or index + 1)
        if index + 1 >= len(cycles):
            return cycle_number
        confirmation_status = str(
            cycles[index + 1].get("status") or ""
        ).upper()
        if confirmation_status in {"BURNED", "NONWORKING", "NO_CONTACT"}:
            return cycle_number
    return None


def final_ivl_status(cycles: List[Dict[str, Any]]) -> str:
    if not cycles:
        return "FAILED"
    if confirmed_burned_cycle(cycles) is not None:
        return "BURNED"
    return str(cycles[-1].get("status") or "FAILED").upper()


def describe_ivl_first_measurement(cycles: List[Dict[str, Any]]) -> str:
    if not cycles:
        return "ВАЯХ не выполнена"
    burned_cycle = confirmed_burned_cycle(cycles)
    if burned_cycle is not None:
        cycle_number = burned_cycle
        if cycle_number == 1:
            return "Светодиод сгорел/пробился на первом цикле ВАЯХ"
        return f"Светодиод сгорел/пробился на цикле {cycle_number} ВАЯХ"

    first = str(cycles[0].get("status", "") or "").upper()
    if first == "BURNED" and final_ivl_status(cycles) == "WORKING":
        return "Превышение тока не подтвердилось: на контрольном цикле светодиод рабочий"
    if first == "WORKING":
        return "На первом промере светодиод рабочий"
    if first == "NONWORKING":
        return "Светодиод сразу нерабочий: ток есть, фототока нет"
    if first == "NO_CONTACT":
        return "На первом промере нет контакта с подложкой"
    if first == "BURNED":
        return "Светодиод сгорел/пробился на первом цикле ВАЯХ"
    return f"Первый промер завершился статусом {first or 'UNKNOWN'}"


def save_ivl_workbook(pixel_id: str, filename: Path, params: Any, cycles: List[Dict[str, Any]]) -> Path:
    filename = Path(filename)
    filename.parent.mkdir(parents=True, exist_ok=True)
    ivl_diagnosis = describe_ivl_first_measurement(cycles)

    wb = Workbook()
    ws_sum = wb.active
    ws_sum.title = "Summary"

    ws_sum["A1"] = "IVL / ВАЯХ"
    ws_sum["A1"].font = Font(bold=True, size=14)
    meta = [
        ("Pixel", pixel_id),
        ("Created", now_str()),
        ("COM port", _params_value(params, "com_port", "")),
        (
            "Sweep",
            f"{_params_value(params, 'sweep_start', '')}-{_params_value(params, 'sweep_end', '')} В, "
            f"step {_params_value(params, 'sweep_increment', '')} В",
        ),
        ("Cycles requested", _params_value(params, "num_cycles", "")),
        ("Current limit (mA)", _params_value(params, "current_limit_mA", "")),
        ("Working photodiode threshold (uA)", _params_value(params, "photodiode_threshold_uA", "")),
        (
            "Opening photodiode threshold (uA)",
            _params_value(params, "opening_photodiode_threshold_uA", ""),
        ),
        (
            "Opening following confirmation points",
            _params_value(params, "opening_confirmation_points", ""),
        ),
        ("Pixel area (mm^2)", _params_value(params, "pixel_area_mm2", "")),
        ("Geometric coefficient", _params_value(params, "geometric_coefficient", 1.0)),
        ("Luminance conversion (cd/m^2 per uA)", _params_value(params, "luminance_cd_m2_per_uA", "")),
        ("Burned confirmation cycles", _params_value(params, "burned_confirmation_cycles", "")),
        ("Первый промер / диагноз", ivl_diagnosis),
        ("Naming rule", "{quarter_code}{quarter_number}_{substrate_number}_{pixel_number}"),
    ]
    for idx, (key, value) in enumerate(meta, start=3):
        ws_sum.cell(row=idx, column=1, value=key).font = Font(bold=True)
        ws_sum.cell(row=idx, column=2, value=value)

    summary_headers = [
        "Cycle",
        "Status",
        "Status description",
        "Max current (mA)",
        "Max photodiode (uA)",
        "Opening voltage detected (V)",
        "Current limit reached",
        "First measurement diagnosis",
    ]
    start = 3 + len(meta) + 1
    ws_sum.append([])
    for column, header in enumerate(summary_headers, start=1):
        ws_sum.cell(row=start, column=column, value=header)
    style_header_row(ws_sum, start, 1, len(summary_headers))

    for row_idx, cycle in enumerate(cycles, start=start + 1):
        ws_sum.cell(row=row_idx, column=1, value=cycle["cycle"])
        ws_sum.cell(row=row_idx, column=2, value=cycle["status"])
        ws_sum.cell(row=row_idx, column=3, value=cycle["status_desc"])
        ws_sum.cell(row=row_idx, column=4, value=cycle["max_current_mA"])
        ws_sum.cell(row=row_idx, column=5, value=cycle["max_photo_uA"])
        ws_sum.cell(row=row_idx, column=6, value=cycle["opening_voltage"])
        ws_sum.cell(row=row_idx, column=7, value="YES" if cycle["current_limit_reached"] else "NO")
        ws_sum.cell(row=row_idx, column=8, value=ivl_diagnosis if cycle["cycle"] == 1 else "")

    for cycle in cycles:
        ws = wb.create_sheet(f"Cycle_{cycle['cycle']}")
        ws["A1"] = f"Pixel {pixel_id} | Cycle {cycle['cycle']} | {cycle['status']}"
        ws["A1"].font = Font(bold=True, size=13)
        header_row = 4
        for column, header in enumerate(IVL_WORKBOOK_HEADERS, start=1):
            ws.cell(row=header_row, column=column, value=header)
        style_header_row(ws, header_row, 1, len(IVL_WORKBOOK_HEADERS))

        for row_idx, row in enumerate(cycle["data"], start=header_row + 1):
            for col_idx, header in enumerate(IVL_WORKBOOK_HEADERS, start=1):
                ws.cell(row=row_idx, column=col_idx, value=row.get(header))
        ws.freeze_panes = f"A{header_row + 1}"

        if len(cycle["data"]) >= 2:
            chart = ScatterChart()
            chart.title = f"{pixel_id} | Cycle {cycle['cycle']} | {cycle['status']}"
            chart.x_axis.title = "Voltage OLED / LED (V)"
            chart.y_axis.title = "Current OLED / LED (mA) / Photodiode (uA)"
            chart.x_axis.majorGridlines = ChartLines()
            min_row = header_row + 1
            max_row = header_row + len(cycle["data"])
            xvalues = Reference(ws, min_col=2, min_row=min_row, max_row=max_row)
            y_current = Reference(ws, min_col=4, min_row=header_row, max_row=max_row)
            y_photo = Reference(ws, min_col=7, min_row=header_row, max_row=max_row)
            chart.series.append(Series(y_current, xvalues, title_from_data=True))
            chart.series.append(Series(y_photo, xvalues, title_from_data=True))
            ws.add_chart(chart, "H4")
        autosize_columns(ws, max_width=34)

    for ws in wb.worksheets:
        autosize_columns(ws, max_width=42)
    wb.save(filename)
    wb.close()
    return filename


def build_ivl_workbook_from_raw_csv(
    raw_csv_path: Path,
    filename: Path,
    pixel_id: str,
    params: Any,
    cycle_summaries: List[Dict[str, Any]],
) -> Path:
    cycles = merge_ivl_raw_cycles(raw_csv_path, params, cycle_summaries)
    return save_ivl_workbook(pixel_id, filename, params, cycles)

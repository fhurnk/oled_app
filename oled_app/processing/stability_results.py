"""Build stability workbooks from raw CSV measurement data."""

from __future__ import annotations

from pathlib import Path
from typing import Any, Dict, List, Mapping, Optional

from openpyxl import Workbook, load_workbook
from openpyxl.chart import Reference, ScatterChart, Series
from openpyxl.chart.axis import ChartLines
from openpyxl.styles import Font, PatternFill

from oled_app.measurements.raw_io import read_csv_dicts
from oled_app.utils import (
    autosize_columns,
    current_density_mA_cm2,
    luminance_cd_m2,
    now_str,
    style_header_row,
)


STABILITY_RAW_HEADERS = [
    "point",
    "date_time",
    "elapsed_s",
    "control_mode",
    "target_setpoint",
    "target_unit",
    "control_revision",
    "current_setpoint_mA",
    "voltage_setpoint_V",
    "voltage_set_V",
    "voltage_led_measured_V",
    "current_led_A",
    "voltage_photodiode_measured_V",
    "current_photodiode_A",
    "current_led_mA",
    "current_density_mA_cm2",
    "current_photodiode_uA",
    "luminance_cd_m2",
]

STABILITY_WORKBOOK_HEADERS = [
    "Point",
    "Date time",
    "Time (s)",
    "Control mode",
    "Current setpoint (mA)",
    "Voltage setpoint (V)",
    "Applied voltage (V)",
    "Voltage OLED / LED (V)",
    "Current OLED / LED (mA)",
    "Current density (mA/cm^2)",
    "Voltage photodiode (V)",
    "Photodiode current (uA)",
    "Luminance (cd/m^2)",
]


def _float_or_none(value: Any) -> Optional[float]:
    if value in (None, ""):
        return None
    try:
        return float(value)
    except Exception:
        return None


def _int_or_default(value: Any, default: int = 0) -> int:
    parsed = _float_or_none(value)
    if parsed is None:
        return default
    return int(parsed)


def _params_value(params: Any, key: str, default: Any = None) -> Any:
    if isinstance(params, Mapping):
        return params.get(key, default)
    return getattr(params, key, default)


def raw_row_to_workbook_row(row: Dict[str, str], params: Any) -> List[Any]:
    current_led_mA = _float_or_none(row.get("current_led_mA"))
    if current_led_mA is None:
        current_led_A = _float_or_none(row.get("current_led_A")) or 0.0
        current_led_mA = current_led_A * 1000.0

    current_pd_uA = _float_or_none(row.get("current_photodiode_uA"))
    if current_pd_uA is None:
        current_pd_A = _float_or_none(row.get("current_photodiode_A")) or 0.0
        current_pd_uA = -current_pd_A * 1_000_000.0

    current_density = _float_or_none(row.get("current_density_mA_cm2"))
    if current_density is None:
        current_density = current_density_mA_cm2(current_led_mA, _params_value(params, "pixel_area_mm2", 1.0))

    luminance = _float_or_none(row.get("luminance_cd_m2"))
    if luminance is None:
        luminance = luminance_cd_m2(current_pd_uA, _params_value(params, "luminance_cd_m2_per_uA", 1.0))

    return [
        _int_or_default(row.get("point"), 0),
        row.get("date_time") or "",
        _float_or_none(row.get("elapsed_s")) or 0.0,
        row.get("control_mode") or _params_value(params, "control_mode", "current"),
        _float_or_none(row.get("current_setpoint_mA")),
        _float_or_none(row.get("voltage_setpoint_V")),
        _float_or_none(row.get("voltage_set_V")) or 0.0,
        _float_or_none(row.get("voltage_led_measured_V")) or 0.0,
        float(current_led_mA),
        current_density,
        _float_or_none(row.get("voltage_photodiode_measured_V")) or 0.0,
        float(current_pd_uA),
        luminance,
    ]


def read_stability_rows_from_raw_csv(raw_csv_path: Path, params: Any) -> List[List[Any]]:
    rows = [raw_row_to_workbook_row(row, params) for row in read_csv_dicts(raw_csv_path)]
    rows.sort(key=lambda item: int(item[0] or 0))
    return rows


def create_stability_workbook(filename: Path, pixel_id: str, params: Any) -> Workbook:
    filename.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    ws["A1"] = "OLED stability"
    ws["A1"].font = Font(bold=True, size=14)
    info = [
        ("Pixel", pixel_id),
        ("Created", now_str()),
        (
            "Mode",
            "constant current, software control"
            if _params_value(params, "control_mode", "current") == "current"
            else "constant voltage, ramped setpoint",
        ),
        ("Current setpoint (mA)", _params_value(params, "current_setpoint_mA", "")),
        ("Voltage setpoint (V)", _params_value(params, "voltage_setpoint_V", "")),
        ("Voltage start (V)", _params_value(params, "voltage_start", "")),
        ("Voltage limit (V)", _params_value(params, "voltage_limit", "")),
        ("Current limit (mA)", _params_value(params, "current_limit_mA", "")),
        ("Measurement time set (s)", _params_value(params, "measurement_time_s", "")),
        ("Sample interval (s)", _params_value(params, "sample_interval_s", "")),
        ("Autosave interval (s)", _params_value(params, "autosave_interval_s", "")),
        ("Photodiode threshold (uA)", _params_value(params, "photodiode_threshold_uA", "")),
        ("Pixel area (mm^2)", _params_value(params, "pixel_area_mm2", "")),
        ("Luminance conversion (cd/m^2 per uA)", _params_value(params, "luminance_cd_m2_per_uA", "")),
        ("Status", "IN_PROGRESS"),
        ("Max photodiode current (uA)", 0),
        ("Last saved elapsed time (s)", 0),
    ]
    for row, (key, value) in enumerate(info, start=3):
        ws.cell(row=row, column=1, value=key).font = Font(bold=True)
        ws.cell(row=row, column=2, value=value)

    header_row = 21
    for column, header in enumerate(STABILITY_WORKBOOK_HEADERS, start=1):
        ws.cell(row=header_row, column=column, value=header)
    style_header_row(ws, header_row, 1, len(STABILITY_WORKBOOK_HEADERS))
    ws.freeze_panes = f"A{header_row + 1}"
    autosize_columns(ws, max_width=36)
    wb.save(filename)
    return wb


def update_stability_status(ws, status: str, max_photo: float, elapsed: float) -> None:
    ws["B17"] = status
    ws["B18"] = max_photo
    ws["B19"] = elapsed
    fill = PatternFill(
        "solid",
        fgColor="C6EFCE" if status == "WORKING" else "FFEB9C" if status == "IN_PROGRESS" else "FFC7CE",
    )
    ws["B17"].fill = fill


def add_stability_chart(ws, pixel_id: str) -> None:
    max_row = ws.max_row
    if max_row > 23:
        chart = ScatterChart()
        chart.title = f"Stability {pixel_id}"
        chart.x_axis.title = "Time (s)"
        chart.y_axis.title = "Current OLED (mA) / Photodiode (uA)"
        chart.x_axis.majorGridlines = ChartLines()
        xvalues = Reference(ws, min_col=3, min_row=22, max_row=max_row)
        y_current = Reference(ws, min_col=9, min_row=21, max_row=max_row)
        y_photo = Reference(ws, min_col=12, min_row=21, max_row=max_row)
        chart.series.append(Series(y_current, xvalues, title_from_data=True))
        chart.series.append(Series(y_photo, xvalues, title_from_data=True))
        ws.add_chart(chart, "J3")


def save_stability_chart(filename: Path, pixel_id: str) -> None:
    wb = load_workbook(filename)
    ws = wb["Data"]
    add_stability_chart(ws, pixel_id)
    autosize_columns(ws, max_width=36)
    wb.save(filename)
    wb.close()


def build_stability_workbook_from_raw_csv(
    raw_csv_path: Path,
    filename: Path,
    pixel_id: str,
    params: Any,
    status: str,
    max_photo_uA: float,
    last_elapsed_s: float,
) -> Path:
    rows = read_stability_rows_from_raw_csv(raw_csv_path, params)
    wb = create_stability_workbook(filename, pixel_id, params)
    ws = wb["Data"]
    for row in rows:
        ws.append(row)
    update_stability_status(ws, status, max_photo_uA, last_elapsed_s)
    add_stability_chart(ws, pixel_id)
    autosize_columns(ws, max_width=36)
    wb.save(filename)
    wb.close()
    return Path(filename)

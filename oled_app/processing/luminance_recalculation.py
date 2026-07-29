"""Recalculate luminance in an existing series and restore missing raw CSV."""

from __future__ import annotations

import csv
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Callable, Dict, Iterable, List, Optional, Sequence, Tuple

from openpyxl import load_workbook

from oled_app.measurements.raw_io import RawCsvWriter, raw_data_settings
from oled_app.processing.ivl_preview import create_ivl_thumbnail_from_workbook
from oled_app.processing.ivl_results import IVL_RAW_HEADERS
from oled_app.processing.spectrum_results import (
    SPECTRUM_SPECTRA_RAW_HEADERS,
    SPECTRUM_SUMMARY_RAW_HEADERS,
)
from oled_app.processing.stability_results import STABILITY_RAW_HEADERS
from oled_app.utils import (
    as_float_or_none,
    current_density_mA_cm2,
    luminance_cd_m2_at_voltage,
    resolve_series_file,
)


@dataclass
class LuminanceRecalculationReport:
    workbooks_updated: int = 0
    raw_files_updated: int = 0
    raw_files_restored: int = 0
    thumbnails_created: int = 0
    skipped: int = 0
    errors: int = 0


def _find_header_row(ws, required: str, limit: int = 80) -> Tuple[int, Dict[str, int]]:
    for row in range(1, min(ws.max_row, limit) + 1):
        headers = {
            str(ws.cell(row, column).value or "").strip(): column
            for column in range(1, ws.max_column + 1)
        }
        if required in headers:
            return row, headers
    raise ValueError(f"{ws.title}: не найден столбец {required!r}")


def _set_meta_value(ws, label: str, value: Any, limit: int = 40) -> None:
    for row in range(1, min(ws.max_row, limit) + 1):
        if str(ws.cell(row, 1).value or "").strip() == label:
            ws.cell(row, 2, value)
            return


def _atomic_save_workbook(wb, workbook_path: Path) -> None:
    temp_path = workbook_path.with_name(f".{workbook_path.stem}.recalibrating.xlsx")
    wb.save(temp_path)
    temp_path.replace(workbook_path)


def _existing_or_default_raw_path(workbook_path: Path, suffix: str, folder_name: str) -> Tuple[Path, bool]:
    filename = f"{workbook_path.stem}{suffix}"
    direct = workbook_path.parent / folder_name / filename
    if direct.exists():
        return direct, True
    matches = list(workbook_path.parent.rglob(filename))
    if matches:
        return matches[0], True
    return direct, False


def _write_raw_rows(path: Path, fieldnames: Sequence[str], rows: Iterable[Dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    temp_path = path.with_name(f".{path.name}.recalibrating")
    with RawCsvWriter(temp_path, fieldnames) as writer:
        writer.writerows(rows)
    temp_path.replace(path)


def _read_raw_rows(path: Path) -> List[Dict[str, str]]:
    with path.open("r", encoding="utf-8", newline="") as handle:
        return list(csv.DictReader(handle))


def _recalculate_existing_raw(
    path: Path,
    fieldnames: Sequence[str],
    luminance_coefficient: float,
    pixel_area_mm2: float,
    calibration_model: Optional[Dict[str, Any]] = None,
) -> None:
    rows = _read_raw_rows(path)
    for row in rows:
        current_mA = as_float_or_none(row.get("current_led_mA"))
        if current_mA is None:
            current_A = as_float_or_none(row.get("current_led_A"))
            current_mA = None if current_A is None else float(current_A) * 1000.0
        photo_uA = as_float_or_none(row.get("current_photodiode_uA"))
        if photo_uA is None:
            photo_A = as_float_or_none(row.get("current_photodiode_A"))
            photo_uA = None if photo_A is None else -float(photo_A) * 1_000_000.0
        if "current_density_mA_cm2" in fieldnames:
            row["current_density_mA_cm2"] = current_density_mA_cm2(current_mA, pixel_area_mm2)
        if "luminance_cd_m2" in fieldnames:
            voltage = as_float_or_none(row.get("voltage_set_V"))
            if voltage is None:
                voltage = as_float_or_none(row.get("voltage_led_measured_V"))
            row["luminance_cd_m2"] = luminance_cd_m2_at_voltage(
                photo_uA,
                luminance_coefficient,
                voltage,
                calibration_model,
            )
    _write_raw_rows(path, fieldnames, rows)


def _ivl_rows_from_workbook(workbook_path: Path) -> List[Dict[str, Any]]:
    rows: List[Dict[str, Any]] = []
    wb = load_workbook(workbook_path, data_only=True, read_only=True)
    try:
        for sheet_name in wb.sheetnames:
            if not sheet_name.startswith("Cycle_"):
                continue
            ws = wb[sheet_name]
            header_row, headers = _find_header_row(ws, "Current OLED / LED (mA)")
            try:
                cycle = int(sheet_name.rsplit("_", 1)[1])
            except (ValueError, IndexError):
                cycle = 1
            for row in range(header_row + 1, ws.max_row + 1):
                point = as_float_or_none(ws.cell(row, headers.get("Point", 1)).value)
                if point is None:
                    continue
                current_mA = as_float_or_none(ws.cell(row, headers["Current OLED / LED (mA)"]).value) or 0.0
                photo_uA = as_float_or_none(ws.cell(row, headers["Photodiode current (uA)"]).value) or 0.0
                rows.append(
                    {
                        "cycle": cycle,
                        "point": int(point),
                        "date_time": "",
                        "elapsed_s": ws.cell(
                            row,
                            headers.get("Measurement time (s)", headers.get("Point", 1)),
                        ).value,
                        "voltage_set_V": ws.cell(row, headers.get("Voltage set (V)", 2)).value,
                        "voltage_led_measured_V": ws.cell(
                            row,
                            headers.get("Voltage OLED / LED measured (V)", headers.get("Voltage set (V)", 2)),
                        ).value,
                        "current_led_A": float(current_mA) / 1000.0,
                        "voltage_photodiode_measured_V": ws.cell(
                            row,
                            headers.get("Voltage photodiode measured (V)", 6),
                        ).value,
                        "current_photodiode_A": -float(photo_uA) / 1_000_000.0,
                        "current_led_mA": float(current_mA),
                        "current_photodiode_uA": float(photo_uA),
                    }
                )
    finally:
        wb.close()
    return rows


def _update_ivl_workbook(
    workbook_path: Path,
    luminance_coefficient: float,
    geometric_coefficient: float,
    calibration_model: Optional[Dict[str, Any]] = None,
) -> None:
    wb = load_workbook(workbook_path)
    try:
        ws_summary = wb["Summary"]
        _set_meta_value(ws_summary, "Geometric coefficient", geometric_coefficient)
        _set_meta_value(
            ws_summary,
            "Luminance conversion (cd/m^2 per uA)",
            luminance_coefficient,
        )
        for sheet_name in wb.sheetnames:
            if not sheet_name.startswith("Cycle_"):
                continue
            ws = wb[sheet_name]
            header_row, headers = _find_header_row(ws, "Photodiode current (uA)")
            lum_col = headers.get("Luminance (cd/m^2)")
            voltage_col = headers.get(
                "Voltage set (V)",
                headers.get("Voltage OLED / LED measured (V)"),
            )
            if lum_col is None:
                continue
            for row in range(header_row + 1, ws.max_row + 1):
                photo = ws.cell(row, headers["Photodiode current (uA)"]).value
                if photo not in (None, ""):
                    voltage = (
                        ws.cell(row, voltage_col).value
                        if voltage_col is not None
                        else None
                    )
                    ws.cell(
                        row,
                        lum_col,
                        luminance_cd_m2_at_voltage(
                            photo,
                            luminance_coefficient,
                            voltage,
                            calibration_model,
                        ),
                    )
        _atomic_save_workbook(wb, workbook_path)
    finally:
        wb.close()


def _stability_rows_from_workbook(
    workbook_path: Path,
    luminance_coefficient: float,
    pixel_area_mm2: float,
    calibration_model: Optional[Dict[str, Any]] = None,
) -> List[Dict[str, Any]]:
    rows: List[Dict[str, Any]] = []
    wb = load_workbook(workbook_path, data_only=True, read_only=True)
    try:
        ws = wb["Data"]
        header_row, headers = _find_header_row(ws, "Photodiode current (uA)")
        for row in range(header_row + 1, ws.max_row + 1):
            point = as_float_or_none(ws.cell(row, headers.get("Point", 1)).value)
            if point is None:
                continue
            mode = str(ws.cell(row, headers.get("Control mode", 4)).value or "current")
            current_mA = as_float_or_none(ws.cell(row, headers["Current OLED / LED (mA)"]).value) or 0.0
            photo_uA = as_float_or_none(ws.cell(row, headers["Photodiode current (uA)"]).value) or 0.0
            current_target = ws.cell(row, headers.get("Current setpoint (mA)", 5)).value
            voltage_target = ws.cell(row, headers.get("Voltage setpoint (V)", 6)).value
            target = current_target if mode == "current" else voltage_target
            calibration_voltage = ws.cell(
                row,
                headers.get("Applied voltage (V)", headers.get("Voltage OLED / LED (V)", 8)),
            ).value
            rows.append(
                {
                    "point": int(point),
                    "date_time": ws.cell(row, headers.get("Date time", 2)).value,
                    "elapsed_s": ws.cell(row, headers.get("Time (s)", 3)).value,
                    "control_mode": mode,
                    "target_setpoint": target,
                    "target_unit": "mA" if mode == "current" else "V",
                    "control_revision": "",
                    "current_setpoint_mA": current_target,
                    "voltage_setpoint_V": voltage_target,
                    "voltage_set_V": ws.cell(row, headers.get("Applied voltage (V)", 7)).value,
                    "voltage_led_measured_V": ws.cell(row, headers.get("Voltage OLED / LED (V)", 8)).value,
                    "current_led_A": float(current_mA) / 1000.0,
                    "voltage_photodiode_measured_V": ws.cell(row, headers.get("Voltage photodiode (V)", 11)).value,
                    "current_photodiode_A": -float(photo_uA) / 1_000_000.0,
                    "current_led_mA": current_mA,
                    "current_density_mA_cm2": current_density_mA_cm2(current_mA, pixel_area_mm2),
                    "current_photodiode_uA": photo_uA,
                    "luminance_cd_m2": luminance_cd_m2_at_voltage(
                        photo_uA,
                        luminance_coefficient,
                        calibration_voltage,
                        calibration_model,
                    ),
                }
            )
    finally:
        wb.close()
    return rows


def _update_stability_workbook(
    workbook_path: Path,
    luminance_coefficient: float,
    geometric_coefficient: float,
    calibration_model: Optional[Dict[str, Any]] = None,
) -> None:
    wb = load_workbook(workbook_path)
    try:
        ws = wb["Data"]
        _set_meta_value(ws, "Geometric coefficient", geometric_coefficient)
        _set_meta_value(ws, "Luminance conversion (cd/m^2 per uA)", luminance_coefficient)
        header_row, headers = _find_header_row(ws, "Photodiode current (uA)")
        lum_col = headers["Luminance (cd/m^2)"]
        for row in range(header_row + 1, ws.max_row + 1):
            photo = ws.cell(row, headers["Photodiode current (uA)"]).value
            if photo not in (None, ""):
                voltage = ws.cell(
                    row,
                    headers.get("Applied voltage (V)", headers.get("Voltage OLED / LED (V)", 8)),
                ).value
                ws.cell(
                    row,
                    lum_col,
                    luminance_cd_m2_at_voltage(
                        photo,
                        luminance_coefficient,
                        voltage,
                        calibration_model,
                    ),
                )
        _atomic_save_workbook(wb, workbook_path)
    finally:
        wb.close()


def _spectrum_summary_rows_from_workbook(
    workbook_path: Path,
    luminance_coefficient: float,
    pixel_area_mm2: float,
    calibration_model: Optional[Dict[str, Any]] = None,
) -> List[Dict[str, Any]]:
    rows: List[Dict[str, Any]] = []
    wb = load_workbook(workbook_path, data_only=True, read_only=True)
    try:
        ws = wb["Сводка"]
        header_row, headers = _find_header_row(ws, "I photodiode (uA)")
        for row in range(header_row + 1, ws.max_row + 1):
            point = as_float_or_none(ws.cell(row, headers.get("Point", 1)).value)
            if point is None:
                continue
            current_mA = as_float_or_none(ws.cell(row, headers.get("I LED (mA)", 4)).value) or 0.0
            photo_uA = as_float_or_none(ws.cell(row, headers["I photodiode (uA)"]).value) or 0.0
            t_ms = as_float_or_none(ws.cell(row, headers.get("T_int saved spectrum (ms)", 9)).value)
            calibration_voltage = ws.cell(
                row,
                headers.get("V set (V)", headers.get("V LED measured (V)", 3)),
            ).value
            rows.append(
                {
                    "point": int(point),
                    "date_time": "",
                    "voltage_set_V": ws.cell(row, headers.get("V set (V)", 2)).value,
                    "voltage_led_measured_V": ws.cell(row, headers.get("V LED measured (V)", 3)).value,
                    "current_led_A": float(current_mA) / 1000.0,
                    "current_led_mA": current_mA,
                    "current_density_mA_cm2": current_density_mA_cm2(current_mA, pixel_area_mm2),
                    "voltage_photodiode_measured_V": ws.cell(
                        row,
                        headers.get("V photodiode measured (V)", 6),
                    ).value,
                    "current_photodiode_A": -float(photo_uA) / 1_000_000.0,
                    "current_photodiode_uA": photo_uA,
                    "luminance_cd_m2": luminance_cd_m2_at_voltage(
                        photo_uA,
                        luminance_coefficient,
                        calibration_voltage,
                        calibration_model,
                    ),
                    "integration_time_s": None if t_ms is None else float(t_ms) / 1000.0,
                    "status": ws.cell(row, headers.get("Status", 13)).value,
                    "peak_nm": ws.cell(row, headers.get("Peak (nm)", 10)).value,
                    "peak_intensity_processed_counts": ws.cell(
                        row,
                        headers.get("Max intensity processed (counts)", 11),
                    ).value,
                    "fwhm_nm": ws.cell(row, headers.get("FWHM (nm)", 12)).value,
                    "peaks_detected": ws.cell(row, headers.get("Peaks detected", 15)).value,
                    "peaks_nm": ws.cell(row, headers.get("Peaks nm", 16)).value,
                    "baseline_value_raw_counts": ws.cell(
                        row,
                        headers.get("Background mean raw counts", 17),
                    ).value,
                    "baseline_region_nm": ws.cell(
                        row,
                        headers.get("Background region nm", 18),
                    ).value,
                }
            )
    finally:
        wb.close()
    return rows


def _spectrum_rows_from_workbook(workbook_path: Path) -> List[Dict[str, Any]]:
    rows: List[Dict[str, Any]] = []
    wb = load_workbook(workbook_path, data_only=True, read_only=True)
    try:
        ws_raw = wb["Raw spectra"]
        ws_dark = wb["Dark corrected"] if "Dark corrected" in wb.sheetnames else None
        data_header_row, _headers = _find_header_row(ws_raw, "Wavelength (nm)")
        wavelengths = [
            as_float_or_none(ws_raw.cell(row, 1).value)
            for row in range(data_header_row + 1, ws_raw.max_row + 1)
        ]
        for column in range(2, ws_raw.max_column + 1):
            point = as_float_or_none(ws_raw.cell(2, column).value)
            voltage = as_float_or_none(ws_raw.cell(3, column).value)
            t_ms = as_float_or_none(ws_raw.cell(10, column).value)
            if point is None:
                continue
            for row, wavelength in enumerate(wavelengths, start=data_header_row + 1):
                raw_value = as_float_or_none(ws_raw.cell(row, column).value)
                if wavelength is None or raw_value is None:
                    continue
                dark_value = None
                if ws_dark is not None:
                    dark_corrected = as_float_or_none(ws_dark.cell(row, column).value)
                    if dark_corrected is not None:
                        difference = float(raw_value) - float(dark_corrected)
                        if abs(difference) > 1e-12:
                            dark_value = difference
                rows.append(
                    {
                        "point": int(point),
                        "voltage_set_V": voltage,
                        "integration_time_s": None if t_ms is None else float(t_ms) / 1000.0,
                        "wavelength_nm": wavelength,
                        "raw_counts": raw_value,
                        "dark_counts": dark_value,
                    }
                )
    finally:
        wb.close()
    return rows


def _update_spectrum_workbook(
    workbook_path: Path,
    luminance_coefficient: float,
    geometric_coefficient: float,
    calibration_model: Optional[Dict[str, Any]] = None,
) -> None:
    wb = load_workbook(workbook_path)
    try:
        ws = wb["Сводка"]
        _set_meta_value(ws, "Geometric coefficient", geometric_coefficient)
        _set_meta_value(ws, "Luminance conversion (cd/m^2 per uA)", luminance_coefficient)
        header_row, headers = _find_header_row(ws, "I photodiode (uA)")
        lum_col = headers["Luminance (cd/m^2)"]
        for row in range(header_row + 1, ws.max_row + 1):
            photo = ws.cell(row, headers["I photodiode (uA)"]).value
            if photo not in (None, ""):
                voltage = ws.cell(
                    row,
                    headers.get("V set (V)", headers.get("V LED measured (V)", 3)),
                ).value
                ws.cell(
                    row,
                    lum_col,
                    luminance_cd_m2_at_voltage(
                        photo,
                        luminance_coefficient,
                        voltage,
                        calibration_model,
                    ),
                )
        _atomic_save_workbook(wb, workbook_path)
    finally:
        wb.close()


def recalculate_series_luminance(
    series,
    app_settings: Dict[str, Any],
    log: Optional[Callable[[str], None]] = None,
) -> LuminanceRecalculationReport:
    """Replace existing workbooks with values derived from current coefficients."""

    report = LuminanceRecalculationReport()
    units = app_settings.get("measurement_units", {})
    geometry = series.geometric_coefficient(app_settings)
    pixel_area = float(units.get("pixel_area_mm2", 1.0) or 1.0)
    raw_folder_name = str(raw_data_settings(app_settings)["folder_name"])
    seen: set[Path] = set()
    measurements = list(series.journal.list_measurements())
    latest_ivl_by_pixel: Dict[str, Tuple[float, int, Path]] = {}
    for order, measurement in enumerate(measurements):
        if str(measurement.get("Type") or "").upper() != "IVL":
            continue
        pixel_id = str(measurement.get("Pixel ID") or "")
        candidate = resolve_series_file(series.series_folder, measurement.get("File"))
        if candidate is None:
            continue
        sort_key = (candidate.stat().st_mtime, order, candidate)
        previous = latest_ivl_by_pixel.get(pixel_id)
        if previous is None or sort_key[:2] > previous[:2]:
            latest_ivl_by_pixel[pixel_id] = sort_key

    for measurement in measurements:
        measurement_type = str(measurement.get("Type") or "").upper()
        if measurement_type not in {"IVL", "SPECTRUM", "STABILITY"}:
            continue
        pixel_id = str(measurement.get("Pixel ID") or "")
        workbook_path = resolve_series_file(series.series_folder, measurement.get("File"))
        if workbook_path is None or workbook_path in seen:
            report.skipped += 1
            continue
        seen.add(workbook_path)
        luminance_coefficient = series.luminance_coefficient_for_pixel(pixel_id, app_settings)
        model_resolver = getattr(series, "luminance_model_for_pixel", None)
        calibration_model = (
            model_resolver(pixel_id, app_settings)
            if callable(model_resolver)
            else None
        )
        try:
            if measurement_type == "IVL":
                raw_path, existed = _existing_or_default_raw_path(
                    workbook_path,
                    "_raw.csv",
                    raw_folder_name,
                )
                if existed:
                    _recalculate_existing_raw(
                        raw_path,
                        IVL_RAW_HEADERS,
                        luminance_coefficient,
                        pixel_area,
                        calibration_model,
                    )
                    report.raw_files_updated += 1
                else:
                    _write_raw_rows(raw_path, IVL_RAW_HEADERS, _ivl_rows_from_workbook(workbook_path))
                    report.raw_files_restored += 1
                _update_ivl_workbook(
                    workbook_path,
                    luminance_coefficient,
                    geometry,
                    calibration_model,
                )
                latest_ivl = latest_ivl_by_pixel.get(pixel_id)
                if latest_ivl is not None and workbook_path == latest_ivl[2]:
                    create_ivl_thumbnail_from_workbook(workbook_path)
                    report.thumbnails_created += 1
            elif measurement_type == "STABILITY":
                raw_path, existed = _existing_or_default_raw_path(
                    workbook_path,
                    "_raw.csv",
                    raw_folder_name,
                )
                if existed:
                    _recalculate_existing_raw(
                        raw_path,
                        STABILITY_RAW_HEADERS,
                        luminance_coefficient,
                        pixel_area,
                        calibration_model,
                    )
                    report.raw_files_updated += 1
                else:
                    rows = _stability_rows_from_workbook(
                        workbook_path,
                        luminance_coefficient,
                        pixel_area,
                        calibration_model,
                    )
                    _write_raw_rows(raw_path, STABILITY_RAW_HEADERS, rows)
                    report.raw_files_restored += 1
                _update_stability_workbook(
                    workbook_path,
                    luminance_coefficient,
                    geometry,
                    calibration_model,
                )
            else:
                summary_path, summary_existed = _existing_or_default_raw_path(
                    workbook_path,
                    "_summary_raw.csv",
                    raw_folder_name,
                )
                spectra_path, spectra_existed = _existing_or_default_raw_path(
                    workbook_path,
                    "_spectra_raw.csv",
                    raw_folder_name,
                )
                if summary_existed:
                    _recalculate_existing_raw(
                        summary_path,
                        SPECTRUM_SUMMARY_RAW_HEADERS,
                        luminance_coefficient,
                        pixel_area,
                        calibration_model,
                    )
                    report.raw_files_updated += 1
                else:
                    summary_rows = _spectrum_summary_rows_from_workbook(
                        workbook_path,
                        luminance_coefficient,
                        pixel_area,
                        calibration_model,
                    )
                    _write_raw_rows(summary_path, SPECTRUM_SUMMARY_RAW_HEADERS, summary_rows)
                    report.raw_files_restored += 1
                if spectra_existed:
                    report.raw_files_updated += 1
                else:
                    _write_raw_rows(
                        spectra_path,
                        SPECTRUM_SPECTRA_RAW_HEADERS,
                        _spectrum_rows_from_workbook(workbook_path),
                    )
                    report.raw_files_restored += 1
                _update_spectrum_workbook(
                    workbook_path,
                    luminance_coefficient,
                    geometry,
                    calibration_model,
                )
            report.workbooks_updated += 1
            if log is not None:
                log(f"Пересчитан {measurement_type}: {workbook_path.name}")
        except Exception as exc:
            report.errors += 1
            if log is not None:
                log(f"Ошибка пересчёта {workbook_path}: {exc}")
    return report

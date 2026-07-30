"""Spectrum sensitivity correction and quarter-level integral calibration."""

from __future__ import annotations

import csv
import math
from dataclasses import asdict, dataclass
from functools import lru_cache
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Sequence, Tuple

import numpy as np
from openpyxl import Workbook, load_workbook
from openpyxl.chart import Reference, ScatterChart, Series
from openpyxl.chart.axis import ChartLines
from openpyxl.styles import Font, PatternFill

from oled_app.utils import (
    as_float_or_none,
    autosize_columns,
    now_str,
    safe_filename,
    spectral_integral_at_voltage,
    style_header_row,
    timestamp_for_file,
)


DEFAULT_SENSITIVITY_CSV = Path(__file__).resolve().parent.parent / "data" / "spectral_sensitivity.csv"


@dataclass(frozen=True)
class SpectralIntegral:
    """Corrected spectral integrals on the common CIE/BPW34 grid."""

    shape_integral: float
    weighted_integral: float
    peak_intensity: float
    wavelength_min_nm: float
    wavelength_max_nm: float
    point_count: int


@dataclass(frozen=True)
class QuarterIntegralCalibration:
    """Normalized CIE/BPW34 integral assigned to one physical quarter."""

    integral_coefficient: float
    coefficient: float
    geometric_coefficient: float
    effective_coefficient: float
    points_used: int
    relative_std_percent: Optional[float]
    integral_min: float
    integral_max: float
    source_pixel: str
    source_file: str
    calculated_at: str
    method: str = "normalized_shape_integral_filtered_median"
    points_total: int = 0
    points_rejected: int = 0
    inlier_threshold_percent: float = 10.0
    activation_voltage_V: Optional[float] = None
    reference_voltage_V: Optional[float] = None
    slope_integral_per_V: Optional[float] = None
    intercept_integral: Optional[float] = None
    r_squared: Optional[float] = None
    equation: str = ""
    included_point_numbers: Tuple[int, ...] = ()

    def as_dict(self) -> Dict[str, Any]:
        return asdict(self)

    @classmethod
    def from_dict(
        cls,
        payload: Dict[str, Any],
        *,
        integral_coefficient: Optional[float] = None,
        geometric_coefficient: Optional[float] = None,
        activation_voltage_V: Optional[float] = None,
    ) -> "QuarterIntegralCalibration":
        coefficient = float(payload.get("coefficient"))
        method = str(payload.get("method") or "normalized_shape_integral_median")
        configured_integral = float(
            integral_coefficient
            if integral_coefficient is not None
            else payload.get("integral_coefficient", 1.0)
        )
        geometry = float(
            geometric_coefficient
            if geometric_coefficient is not None
            else payload.get("geometric_coefficient", 1.0)
        )
        return cls(
            integral_coefficient=configured_integral,
            coefficient=coefficient,
            geometric_coefficient=geometry,
            effective_coefficient=coefficient * configured_integral * geometry,
            points_used=int(payload.get("points_used") or 0),
            relative_std_percent=as_float_or_none(payload.get("relative_std_percent")),
            integral_min=float(payload.get("integral_min", coefficient) or coefficient),
            integral_max=float(payload.get("integral_max", coefficient) or coefficient),
            source_pixel=str(payload.get("source_pixel") or ""),
            source_file=str(payload.get("source_file") or ""),
            calculated_at=str(payload.get("calculated_at") or ""),
            method=method,
            points_total=int(payload.get("points_total") or payload.get("points_used") or 0),
            points_rejected=int(payload.get("points_rejected") or 0),
            inlier_threshold_percent=float(payload.get("inlier_threshold_percent") or 10.0),
            activation_voltage_V=(
                as_float_or_none(
                    activation_voltage_V
                    if activation_voltage_V is not None
                    else payload.get("activation_voltage_V")
                )
                if method == "normalized_shape_integral_linear_voltage"
                else None
            ),
            reference_voltage_V=as_float_or_none(payload.get("reference_voltage_V")),
            slope_integral_per_V=as_float_or_none(payload.get("slope_integral_per_V")),
            intercept_integral=as_float_or_none(payload.get("intercept_integral")),
            r_squared=as_float_or_none(payload.get("r_squared")),
            equation=str(payload.get("equation") or ""),
            included_point_numbers=tuple(
                int(value)
                for value in (payload.get("included_point_numbers") or ())
            ),
        )

    def integral_at_voltage(self, voltage_V: Any = None) -> Optional[float]:
        value = spectral_integral_at_voltage(self.as_dict(), voltage_V)
        return None if value is None else float(value)


@lru_cache(maxsize=8)
def load_sensitivity_curve(path_text: str = "") -> Tuple[np.ndarray, np.ndarray, np.ndarray]:
    """Load and validate the machine-readable CIE/BPW34 sensitivity table."""

    path = Path(path_text) if path_text else DEFAULT_SENSITIVITY_CSV
    wavelengths: List[float] = []
    cie: List[float] = []
    bpw34: List[float] = []
    with path.open("r", encoding="utf-8", newline="") as handle:
        for row in csv.DictReader(handle):
            wavelength = as_float_or_none(row.get("wavelength_nm"))
            cie_value = as_float_or_none(row.get("cie_v_lambda"))
            bpw_value = as_float_or_none(row.get("bpw34_relative_response"))
            if wavelength is None or cie_value is None or bpw_value is None or bpw_value <= 0:
                continue
            wavelengths.append(float(wavelength))
            cie.append(float(cie_value))
            bpw34.append(float(bpw_value))
    if len(wavelengths) < 2:
        raise ValueError(f"В таблице чувствительности недостаточно точек: {path}")
    order = np.argsort(np.asarray(wavelengths, dtype=np.float64))
    wl = np.asarray(wavelengths, dtype=np.float64)[order]
    cie_values = np.asarray(cie, dtype=np.float64)[order]
    bpw_values = np.asarray(bpw34, dtype=np.float64)[order]
    return wl, cie_values, bpw_values


def _trapezoid(values: np.ndarray, wavelengths: np.ndarray) -> float:
    trapezoid = getattr(np, "trapezoid", None)
    if trapezoid is not None:
        return float(trapezoid(values, wavelengths))
    return float(np.trapz(values, wavelengths))


def calculate_spectral_integrals(
    wavelengths: Sequence[float],
    intensities_counts_per_s: Sequence[float],
    sensitivity_csv: Optional[Path] = None,
) -> SpectralIntegral:
    """Normalize a processed spectrum and integrate ``spectrum * CIE / BPW34``.

    ``shape_integral`` follows the user's manual check: the spectrum is first
    normalized to its maximum. ``weighted_integral`` is retained only as an
    absolute-amplitude diagnostic.
    """

    wl = np.asarray(wavelengths, dtype=np.float64)
    intensity = np.asarray(intensities_counts_per_s, dtype=np.float64)
    if wl.shape != intensity.shape or wl.ndim != 1:
        raise ValueError("Длины волн и интенсивности должны быть одномерными массивами одинаковой длины.")

    ref_wl, ref_cie, ref_bpw = load_sensitivity_curve(str(sensitivity_csv or ""))
    mask = (
        np.isfinite(wl)
        & np.isfinite(intensity)
        & (wl >= ref_wl[0])
        & (wl <= ref_wl[-1])
    )
    if int(np.count_nonzero(mask)) < 2:
        raise ValueError("Спектр не пересекается с диапазоном таблицы чувствительности.")

    wl = wl[mask]
    intensity = intensity[mask]
    order = np.argsort(wl)
    wl = wl[order]
    intensity = intensity[order]
    unique_wl, unique_indices = np.unique(wl, return_index=True)
    wl = unique_wl
    intensity = intensity[unique_indices]

    cie = np.interp(wl, ref_wl, ref_cie)
    bpw34 = np.interp(wl, ref_wl, ref_bpw)
    valid = np.isfinite(cie) & np.isfinite(bpw34) & (bpw34 > 0)
    wl = wl[valid]
    intensity = intensity[valid]
    correction = cie[valid] / bpw34[valid]
    if wl.size < 2:
        raise ValueError("После интерполяции чувствительности осталось недостаточно точек.")

    peak = float(np.nanmax(intensity))
    if not math.isfinite(peak) or peak <= 0:
        raise ValueError("Максимум обработанного спектра должен быть положительным.")
    normalized = intensity / peak
    return SpectralIntegral(
        shape_integral=_trapezoid(normalized * correction, wl),
        weighted_integral=_trapezoid(intensity * correction, wl),
        peak_intensity=peak,
        wavelength_min_nm=float(wl[0]),
        wavelength_max_nm=float(wl[-1]),
        point_count=int(wl.size),
    )


def spectral_luminance_cd_m2(
    photodiode_current_uA: Any,
    spectral_integral: Any,
    integral_coefficient: Any,
    geometric_coefficient: Any,
) -> Optional[float]:
    photodiode_current = as_float_or_none(photodiode_current_uA)
    integral = as_float_or_none(spectral_integral)
    coefficient = as_float_or_none(integral_coefficient)
    geometry = as_float_or_none(geometric_coefficient)
    if (
        photodiode_current is None
        or integral is None
        or coefficient is None
        or geometry is None
    ):
        return None
    return (
        float(photodiode_current)
        * float(integral)
        * float(coefficient)
        * float(geometry)
    )


def read_spectrum_integral_points(workbook_path: Path) -> List[Dict[str, Any]]:
    """Read per-voltage integral and photodiode pairs from a spectrum workbook."""

    wb = load_workbook(workbook_path, data_only=True, read_only=True)
    try:
        ws_sum = wb["Сводка"]
        summary_headers: Optional[Dict[str, int]] = None
        summary_by_point: Dict[int, Dict[str, Any]] = {}
        for row_number, values in enumerate(
            ws_sum.iter_rows(values_only=True),
            start=1,
        ):
            if summary_headers is None:
                candidates = {
                    str(value or "").strip(): index
                    for index, value in enumerate(values)
                }
                if "I photodiode (uA)" in candidates:
                    summary_headers = candidates
                    continue
                if row_number >= 60:
                    break
                continue

            def summary_value(header: str, default_index: int) -> Any:
                index = summary_headers.get(header, default_index)
                return values[index] if index < len(values) else None

            point = as_float_or_none(summary_value("Point", 0))
            if point is None:
                continue
            summary_by_point[int(point)] = {
                "point": int(point),
                "voltage_V": as_float_or_none(
                    summary_value("V set (V)", 1)
                ),
                "photodiode_uA": as_float_or_none(
                    summary_value("I photodiode (uA)", 6)
                ),
                "photodiode_luminance_cd_m2": as_float_or_none(
                    summary_value("Luminance (cd/m^2)", 7)
                ),
                "status": str(summary_value("Status", 12) or ""),
            }
        if summary_headers is None:
            raise ValueError(
                "В листе Сводка не найден столбец 'I photodiode (uA)'."
            )

        if "Processed counts per s" not in wb.sheetnames:
            raise ValueError("В книге нет листа 'Processed counts per s'.")
        ws_processed = wb["Processed counts per s"]
        point_by_column: Dict[int, int] = {}
        spectra_by_point: Dict[int, List[Tuple[float, float]]] = {}
        data_started = False
        for row_number, values in enumerate(
            ws_processed.iter_rows(values_only=True),
            start=1,
        ):
            first_value = str(values[0] or "").strip() if values else ""
            if not data_started:
                if first_value == "Point":
                    for column_index, value in enumerate(values[1:], start=1):
                        point_value = as_float_or_none(value)
                        if point_value is None:
                            continue
                        point = int(point_value)
                        point_by_column[column_index] = point
                        spectra_by_point.setdefault(point, [])
                if first_value == "Wavelength (nm)":
                    if not point_by_column:
                        raise ValueError(
                            "В листе 'Processed counts per s' не найдена строка Point."
                        )
                    data_started = True
                    continue
                if row_number >= 60:
                    break
                continue

            wavelength = as_float_or_none(values[0] if values else None)
            if wavelength is None:
                continue
            for column_index, point in point_by_column.items():
                intensity = as_float_or_none(
                    values[column_index]
                    if column_index < len(values)
                    else None
                )
                if intensity is not None:
                    spectra_by_point[point].append(
                        (float(wavelength), float(intensity))
                    )
        if not data_started:
            raise ValueError(
                "В листе 'Processed counts per s' не найден столбец "
                "'Wavelength (nm)'."
            )

        rows = []
        for point in point_by_column.values():
            pairs = spectra_by_point.get(point, [])
            if len(pairs) < 2:
                continue
            spectral = calculate_spectral_integrals(
                [pair[0] for pair in pairs],
                [pair[1] for pair in pairs],
            )
            row_info = dict(summary_by_point.get(point, {"point": point}))
            row_info.update(
                {
                    "shape_integral": spectral.shape_integral,
                    "weighted_integral": spectral.weighted_integral,
                    "integral_luminance_cd_m2": None,
                }
            )
            rows.append(row_info)
        return rows
    finally:
        wb.close()


def calibrate_quarter_spectral_integral(
    points: Iterable[Dict[str, Any]],
    geometric_coefficient: float,
    source_pixel: str,
    source_file: str,
    integral_coefficient: float = 1.0,
    activation_voltage_V: Optional[float] = None,
) -> QuarterIntegralCalibration:
    """Choose a filtered median or a linear integral-versus-voltage model."""

    geometry = float(geometric_coefficient)
    if geometry <= 0 or not math.isfinite(geometry):
        raise ValueError("Геометрический коэффициент должен быть положительным.")
    configured_integral = float(integral_coefficient)
    if configured_integral <= 0 or not math.isfinite(configured_integral):
        raise ValueError("Интегральный коэффициент должен быть положительным.")

    requested_activation = as_float_or_none(activation_voltage_V)
    valid_points: List[Tuple[int, float, float, float]] = []
    rejected_statuses = {"FAILED", "SATURATED", "NEEDS_REVIEW", "STOPPED", "NO_PEAK"}
    for point in points:
        status = str(point.get("status") or "").strip().upper()
        shape_integral = as_float_or_none(point.get("shape_integral"))
        photodiode_current = as_float_or_none(point.get("photodiode_uA"))
        voltage = as_float_or_none(point.get("voltage_V"))
        if (
            status in rejected_statuses
            or shape_integral is None
            or photodiode_current is None
            or voltage is None
            or shape_integral <= 0
            or photodiode_current <= 0
        ):
            continue
        point_number = int(as_float_or_none(point.get("point")) or len(valid_points) + 1)
        valid_points.append(
            (
                point_number,
                float(voltage),
                float(shape_integral),
                float(photodiode_current),
            )
        )

    if len(valid_points) < 2:
        raise ValueError("Для калибровки нужны минимум две пригодные точки спектра с разной светимостью.")
    if math.isclose(
        min(point[3] for point in valid_points),
        max(point[3] for point in valid_points),
    ):
        raise ValueError("Для калибровки нужны точки с различным током фотодетектора.")
    voltages = np.asarray([point[1] for point in valid_points], dtype=np.float64)
    if math.isclose(float(np.min(voltages)), float(np.max(voltages))):
        raise ValueError("Для калибровки нужны спектры при разных напряжениях.")

    integral_values = np.asarray([point[2] for point in valid_points], dtype=np.float64)
    median_integral = float(np.median(integral_values))
    stable_mask = np.abs(integral_values - median_integral) <= abs(median_integral) * 0.10
    stable_majority = int(np.count_nonzero(stable_mask)) > len(integral_values) / 2
    reference_voltage = float(np.median(voltages))
    activation_voltage: Optional[float] = None
    slope: Optional[float] = None
    intercept: Optional[float] = None
    r_squared: Optional[float] = None

    if stable_majority:
        selected_mask = stable_mask
        selected_values = integral_values[selected_mask]
        coefficient = float(np.median(selected_values))
        method = "normalized_shape_integral_filtered_median"
        equation = f"integral = {coefficient:.12g}"
    else:
        low_values = integral_values[voltages < reference_voltage]
        high_values = integral_values[voltages > reference_voltage]
        tolerance = max(abs(median_integral) * 1e-9, 1e-12)
        ascending = (
            low_values.size > 0
            and high_values.size > 0
            and bool(np.all(low_values <= median_integral + tolerance))
            and bool(np.all(high_values >= median_integral - tolerance))
        )
        descending = (
            low_values.size > 0
            and high_values.size > 0
            and bool(np.all(low_values >= median_integral - tolerance))
            and bool(np.all(high_values <= median_integral + tolerance))
        )
        if ascending or descending:
            linear_mask = (
                np.ones(len(voltages), dtype=bool)
                if requested_activation is None
                else voltages >= float(requested_activation)
            )
            if int(np.count_nonzero(linear_mask)) < 2:
                raise ValueError(
                    "Для линейной калибровки нужны минимум две пригодные точки "
                    "не ниже напряжения открытия."
                )
            linear_voltages = voltages[linear_mask]
            linear_integrals = integral_values[linear_mask]
            linear_currents = np.asarray(
                [point[3] for point in valid_points],
                dtype=np.float64,
            )[linear_mask]
            if math.isclose(
                float(np.min(linear_voltages)),
                float(np.max(linear_voltages)),
            ):
                raise ValueError(
                    "Для линейной калибровки нужны точки при разных напряжениях "
                    "не ниже напряжения открытия."
                )
            if math.isclose(
                float(np.min(linear_currents)),
                float(np.max(linear_currents)),
            ):
                raise ValueError(
                    "Для линейной калибровки нужны точки с различным током "
                    "фотодетектора не ниже напряжения открытия."
                )
            reference_voltage = float(np.median(linear_voltages))
            activation_voltage = (
                float(requested_activation)
                if requested_activation is not None
                else float(np.min(linear_voltages))
            )
            slope, intercept = (
                float(value)
                for value in np.polyfit(linear_voltages, linear_integrals, 1)
            )
            predicted = slope * linear_voltages + intercept
            residual_sum = float(np.sum((linear_integrals - predicted) ** 2))
            total_sum = float(
                np.sum((linear_integrals - np.mean(linear_integrals)) ** 2)
            )
            r_squared = (
                1.0
                if math.isclose(total_sum, 0.0)
                else 1.0 - residual_sum / total_sum
            )
            coefficient = float(slope * reference_voltage + intercept)
            selected_mask = linear_mask
            selected_values = linear_integrals
            method = "normalized_shape_integral_linear_voltage"
            equation = f"integral(V) = {slope:.12g} * V + {intercept:.12g}"
        else:
            absolute_deviation = np.abs(integral_values - median_integral)
            mad = float(np.median(absolute_deviation))
            selected_mask = (
                np.ones(len(integral_values), dtype=bool)
                if math.isclose(mad, 0.0)
                else absolute_deviation <= 3.5 * mad
            )
            selected_values = integral_values[selected_mask]
            if selected_values.size < 2:
                selected_mask = np.ones(len(integral_values), dtype=bool)
                selected_values = integral_values
            coefficient = float(np.median(selected_values))
            method = "normalized_shape_integral_robust_median"
            equation = f"integral = {coefficient:.12g}"

    if not math.isfinite(coefficient) or coefficient <= 0:
        raise ValueError("Рассчитанный спектральный интеграл некорректен.")
    effective = configured_integral * coefficient * geometry
    relative_std_percent = (
        None
        if coefficient == 0
        else float(np.std(selected_values) / coefficient * 100.0)
    )
    included_numbers = tuple(
        point[0]
        for point, included in zip(valid_points, selected_mask)
        if bool(included)
    )
    return QuarterIntegralCalibration(
        integral_coefficient=configured_integral,
        coefficient=coefficient,
        geometric_coefficient=geometry,
        effective_coefficient=effective,
        points_used=int(np.count_nonzero(selected_mask)),
        relative_std_percent=relative_std_percent,
        integral_min=float(np.min(selected_values)),
        integral_max=float(np.max(selected_values)),
        source_pixel=str(source_pixel),
        source_file=str(source_file),
        calculated_at=now_str(),
        method=method,
        points_total=len(valid_points),
        points_rejected=len(valid_points) - int(np.count_nonzero(selected_mask)),
        activation_voltage_V=activation_voltage,
        reference_voltage_V=reference_voltage,
        slope_integral_per_V=slope,
        intercept_integral=intercept,
        r_squared=r_squared,
        equation=equation,
        included_point_numbers=included_numbers,
    )


def fit_quarter_integral_coefficient(
    points: Iterable[Dict[str, Any]],
    geometric_coefficient: float,
    source_pixel: str,
    source_file: str,
    integral_coefficient: float = 1.0,
    activation_voltage_V: Optional[float] = None,
) -> QuarterIntegralCalibration:
    """Compatibility wrapper for the normalized-integral calibration."""

    return calibrate_quarter_spectral_integral(
        points,
        geometric_coefficient,
        source_pixel,
        source_file,
        integral_coefficient,
        activation_voltage_V,
    )


def spectral_recalculation_output_path(
    source_workbook: Path,
    pixel_id: str,
    output_dir: Optional[Path] = None,
) -> Path:
    source_workbook = Path(source_workbook)
    base_stem = f"SPECTRAL_RECALC_{safe_filename(pixel_id)}_{timestamp_for_file()}"
    target_dir = Path(output_dir) if output_dir is not None else source_workbook.parent
    candidate = target_dir / f"{base_stem}.xlsx"
    suffix = 2
    while candidate.exists():
        candidate = candidate.with_name(f"{base_stem}_{suffix}.xlsx")
        suffix += 1
    return candidate


def create_spectral_recalculation_workbook(
    output_path: Path,
    source_workbook: Path,
    pixel_id: str,
    quarter_number: int,
    points: Iterable[Dict[str, Any]],
    calibration: QuarterIntegralCalibration,
    sensitivity_csv: Optional[Path] = None,
    rgb_photodiode_coefficient: Optional[float] = None,
) -> Path:
    """Save on-demand CIE/BPW34 recalculation without modifying the source."""

    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    points = list(points)
    wb = Workbook()
    ws = wb.active
    ws.title = "Результаты"
    ws["A1"] = "Спектральный пересчёт CIE / BPW34"
    ws["A1"].font = Font(bold=True, size=14)
    effective_photo = as_float_or_none(rgb_photodiode_coefficient)
    base_color_coefficient = (
        None
        if effective_photo is None or calibration.geometric_coefficient == 0
        else effective_photo / calibration.geometric_coefficient
    )
    metadata = [
        ("Source spectrum", str(Path(source_workbook).resolve())),
        ("Pixel", pixel_id),
        ("Quarter", int(quarter_number)),
        ("Recalculated", now_str()),
        ("Sensitivity CSV", str(Path(sensitivity_csv or DEFAULT_SENSITIVITY_CSV).resolve())),
        ("RGB coefficient before geometry", base_color_coefficient),
        ("Previous photodiode coefficient (RGB * geometry)", effective_photo),
        ("Quarter spectral integral", calibration.coefficient),
        ("Integral coefficient (settings)", calibration.integral_coefficient),
        (
            "Coefficient replacing RGB",
            calibration.coefficient * calibration.integral_coefficient,
        ),
        ("Geometric coefficient", calibration.geometric_coefficient),
        ("Effective coefficient with geometry", calibration.effective_coefficient),
        ("Calibration source pixel", calibration.source_pixel),
        ("Calibration source file", calibration.source_file),
        ("Calibration date", calibration.calculated_at),
        ("Calibration method", calibration.method),
        ("Integral equation", calibration.equation),
        ("Linear fit R^2", calibration.r_squared),
        ("Linear model activation voltage (V)", calibration.activation_voltage_V),
        ("Reference voltage (V)", calibration.reference_voltage_V),
        ("Points used", calibration.points_used),
        ("Points total", calibration.points_total),
        ("Points rejected", calibration.points_rejected),
        ("Median inlier threshold (%)", calibration.inlier_threshold_percent),
        ("Integral relative std (%)", calibration.relative_std_percent),
        ("Integral minimum", calibration.integral_min),
        ("Integral maximum", calibration.integral_max),
        (
            "Formula",
            "L = I_photodiode_uA * integral(V) * integral_coefficient * geometric_coefficient",
        ),
    ]
    for row, (label, value) in enumerate(metadata, start=3):
        ws.cell(row, 1, label).font = Font(bold=True)
        ws.cell(row, 2, value)

    headers = [
        "Point",
        "V set (V)",
        "I photodiode (uA)",
        "Luminance by previous RGB coefficient (cd/m^2)",
        "Shape integral (CIE/BPW34)",
        "Weighted integral (counts/s*nm)",
        "Point coefficient replacing RGB",
        "Quarter spectral integral at V",
        "Quarter coefficient replacing RGB",
        "Luminance by quarter calibration (cd/m^2)",
        "Difference from previous RGB luminance (%)",
        "Point integral deviation from model (%)",
        "Used in calibration",
        "Linear calibration active",
        "Status",
    ]
    header_row = 3 + len(metadata) + 2
    for column, header in enumerate(headers, start=1):
        ws.cell(header_row, column, header)
    style_header_row(ws, header_row, 1, len(headers))

    for row, point in enumerate(points, start=header_row + 1):
        photodiode_current = as_float_or_none(point.get("photodiode_uA"))
        shape_integral = as_float_or_none(point.get("shape_integral"))
        photo_luminance = as_float_or_none(
            point.get("rgb_luminance_cd_m2", point.get("photodiode_luminance_cd_m2"))
        )
        model_integral = calibration.integral_at_voltage(point.get("voltage_V"))
        calibration_active = model_integral is not None
        if calibration.method == "normalized_shape_integral_linear_voltage":
            linear_activation_state = "YES" if calibration_active else "NO"
        else:
            linear_activation_state = "N/A"
        integral_luminance = (
            spectral_luminance_cd_m2(
                photodiode_current,
                model_integral,
                calibration.integral_coefficient,
                calibration.geometric_coefficient,
            )
            if calibration_active
            else photo_luminance
        )
        luminance_difference = None
        if (
            photo_luminance is not None
            and photo_luminance != 0
            and integral_luminance is not None
        ):
            luminance_difference = (
                (integral_luminance - photo_luminance) / photo_luminance * 100.0
            )
        integral_deviation = None
        if (
            shape_integral is not None
            and model_integral is not None
            and model_integral != 0
        ):
            integral_deviation = (
                (shape_integral - model_integral)
                / model_integral
                * 100.0
            )
        point_number = int(as_float_or_none(point.get("point")) or 0)
        values = [
            point.get("point"),
            point.get("voltage_V"),
            photodiode_current,
            photo_luminance,
            shape_integral,
            point.get("weighted_integral"),
            (
                None
                if shape_integral is None
                else shape_integral * calibration.integral_coefficient
            ),
            model_integral,
            (
                None
                if model_integral is None
                else model_integral * calibration.integral_coefficient
            ),
            integral_luminance,
            luminance_difference,
            integral_deviation,
            "YES" if point_number in calibration.included_point_numbers else "NO",
            linear_activation_state,
            point.get("status"),
        ]
        for column, value in enumerate(values, start=1):
            ws.cell(row, column, value)

    last_row = header_row + len(points)
    if last_row > header_row + 1:
        chart = ScatterChart()
        chart.title = "Светимость от тока фотодетектора"
        chart.x_axis.title = "I photodiode (uA)"
        chart.y_axis.title = "Luminance (cd/m^2)"
        chart.x_axis.majorGridlines = ChartLines()
        x_values = Reference(ws, min_col=3, min_row=header_row + 1, max_row=last_row)
        photo_values = Reference(
            ws,
            min_col=4,
            min_row=header_row + 1,
            max_row=last_row,
        )
        integral_values = Reference(
            ws,
            min_col=10,
            min_row=header_row + 1,
            max_row=last_row,
        )
        chart.series.append(
            Series(
                photo_values,
                x_values,
                title="Previous RGB luminance",
            )
        )
        chart.series.append(
            Series(
                integral_values,
                x_values,
                title="Quarter calibration",
            )
        )
        chart.height = 10
        chart.width = 18
        ws.add_chart(chart, "P3")

        integral_chart = ScatterChart()
        integral_chart.title = "Спектральный интеграл от напряжения"
        integral_chart.x_axis.title = "Voltage (V)"
        integral_chart.y_axis.title = "Integral (CIE/BPW34)"
        voltage_values = Reference(
            ws,
            min_col=2,
            min_row=header_row + 1,
            max_row=last_row,
        )
        measured_integrals = Reference(
            ws,
            min_col=5,
            min_row=header_row + 1,
            max_row=last_row,
        )
        model_integrals = Reference(
            ws,
            min_col=8,
            min_row=header_row + 1,
            max_row=last_row,
        )
        integral_chart.series.append(
            Series(measured_integrals, voltage_values, title="Measured integral")
        )
        integral_chart.series.append(
            Series(model_integrals, voltage_values, title="Calibration model")
        )
        integral_chart.height = 10
        integral_chart.width = 18
        ws.add_chart(integral_chart, "P23")

    ws.freeze_panes = f"A{header_row + 1}"
    ws.auto_filter.ref = f"A{header_row}:O{max(last_row, header_row)}"
    for row in range(header_row + 1, last_row + 1):
        ws.cell(row, 11).number_format = "0.00"
        ws.cell(row, 12).number_format = "0.00"
    ws["A1"].fill = PatternFill("solid", fgColor="D9EAF7")
    autosize_columns(ws, max_width=52)

    temp_path = output_path.with_name(f".{output_path.stem}.tmp.xlsx")
    wb.save(temp_path)
    wb.close()
    temp_path.replace(output_path)
    return output_path

"""General-purpose helpers shared across the OLED application."""

from __future__ import annotations

import math
import re
from datetime import date, datetime
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


def now_str() -> str:
    return datetime.now().strftime("%d-%m-%Y %H:%M:%S")


def timestamp_for_file() -> str:
    return datetime.now().strftime("%d-%m-%Y_%Hh%Mm%Ss")


def today_iso() -> str:
    return date.today().isoformat()


def safe_filename(text: str, fallback: str = "item") -> str:
    text = (text or "").strip()
    text = re.sub(r"[^0-9A-Za-zА-Яа-яЁё._\-]+", "_", text)
    text = re.sub(r"_+", "_", text).strip("_.-")
    return text or fallback


def as_float_or_none(value) -> Optional[float]:
    if value in (None, "", "—"):
        return None
    try:
        return float(value)
    except Exception:
        return None


def current_density_mA_cm2(current_mA: Any, pixel_area_mm2: Any) -> Optional[float]:
    current = as_float_or_none(current_mA)
    area = as_float_or_none(pixel_area_mm2)
    if current is None or area is None or area <= 0:
        return None
    return float(current) / (float(area) / 100.0)


def luminance_cd_m2(photo_uA: Any, conversion_cd_m2_per_uA: Any) -> Optional[float]:
    photo = as_float_or_none(photo_uA)
    coeff = as_float_or_none(conversion_cd_m2_per_uA)
    if photo is None or coeff is None:
        return None
    return float(photo) * float(coeff)


SPECTRAL_CALIBRATION_METHODS = {
    "normalized_shape_integral_median",
    "normalized_shape_integral_filtered_median",
    "normalized_shape_integral_robust_median",
    "normalized_shape_integral_linear_voltage",
}


def spectral_integral_at_voltage(
    calibration: Any,
    voltage_V: Any = None,
) -> Optional[float]:
    """Resolve the quarter spectral integral, including a voltage fit when present."""

    if not isinstance(calibration, dict):
        return None
    method = str(calibration.get("method") or "")
    if method not in SPECTRAL_CALIBRATION_METHODS:
        return None

    fallback = as_float_or_none(calibration.get("coefficient"))
    if method != "normalized_shape_integral_linear_voltage":
        return fallback if fallback is not None and fallback > 0 else None

    voltage = as_float_or_none(voltage_V)
    if voltage is None:
        voltage = as_float_or_none(calibration.get("reference_voltage_V"))
    activation_voltage = as_float_or_none(
        calibration.get("activation_voltage_V")
    )
    if (
        voltage is not None
        and activation_voltage is not None
        and voltage < activation_voltage
    ):
        return None
    slope = as_float_or_none(calibration.get("slope_integral_per_V"))
    intercept = as_float_or_none(calibration.get("intercept_integral"))
    if voltage is None or slope is None or intercept is None:
        return fallback if fallback is not None and fallback > 0 else None
    predicted = float(slope) * float(voltage) + float(intercept)
    if not math.isfinite(predicted) or predicted <= 0:
        return fallback if fallback is not None and fallback > 0 else None
    return predicted


def luminance_coefficient_at_voltage(
    default_coefficient: Any,
    voltage_V: Any = None,
    calibration: Any = None,
) -> Optional[float]:
    """Return the RGB or calibrated spectral coefficient for one voltage."""

    default = as_float_or_none(default_coefficient)
    integral = spectral_integral_at_voltage(calibration, voltage_V)
    if integral is None:
        return default
    integral_coefficient = as_float_or_none(calibration.get("integral_coefficient"))
    geometry = as_float_or_none(calibration.get("geometric_coefficient"))
    if integral_coefficient is None or geometry is None:
        return default
    coefficient = float(integral) * float(integral_coefficient) * float(geometry)
    return coefficient if math.isfinite(coefficient) and coefficient > 0 else default


def luminance_cd_m2_at_voltage(
    photo_uA: Any,
    default_coefficient: Any,
    voltage_V: Any = None,
    calibration: Any = None,
) -> Optional[float]:
    coefficient = luminance_coefficient_at_voltage(
        default_coefficient,
        voltage_V,
        calibration,
    )
    return luminance_cd_m2(photo_uA, coefficient)


def relative_to_or_abs(path: Path, base: Path) -> str:
    try:
        return str(path.resolve().relative_to(base.resolve()))
    except Exception:
        return str(path.resolve())


def resolve_series_file(series_folder: Path, file_value: Any) -> Optional[Path]:
    if not file_value:
        return None
    path = Path(str(file_value))
    if not path.is_absolute():
        path = Path(series_folder) / path
    return path if path.exists() else None


def read_spectrum_metrics_from_workbook(path: Optional[Path]) -> Dict[str, Any]:
    if not path or not path.exists():
        return {}
    try:
        wb = load_workbook(path, data_only=True, read_only=True)
        ws = wb.active
        header_row = 14
        headers = {}
        for candidate in range(1, min(ws.max_row, 40) + 1):
            row_headers = {str(ws.cell(candidate, c).value or ""): c for c in range(1, ws.max_column + 1)}
            if "Peaks detected" in row_headers or "Max intensity processed (counts)" in row_headers or "Max intensity processed (counts/s)" in row_headers or "Max intensity (counts)" in row_headers:
                header_row = candidate
                headers = row_headers
                break
        peaks_col = headers.get("Peaks detected")
        peaks_nm_col = headers.get("Peaks nm")
        max_int_col = headers.get("Max intensity processed (counts)") or headers.get("Max intensity processed (counts/s)") or headers.get("Max intensity (counts)")
        best = {"peak_count": "", "peaks_nm": "", "max_intensity": ""}
        best_intensity = -1.0
        for row in range(header_row + 1, ws.max_row + 1):
            intensity = as_float_or_none(ws.cell(row, max_int_col).value) if max_int_col else None
            if intensity is None:
                continue
            if intensity >= best_intensity:
                best_intensity = float(intensity)
                best = {
                    "peak_count": ws.cell(row, peaks_col).value if peaks_col else "",
                    "peaks_nm": ws.cell(row, peaks_nm_col).value if peaks_nm_col else "",
                    "max_intensity": round(float(intensity), 1),
                }
        wb.close()
        return best
    except Exception:
        return {}


def light_border() -> Border:
    side = Side(style="thin", color="D9D9D9")
    return Border(left=side, right=side, top=side, bottom=side)


def style_header_row(ws, row: int, min_col: int, max_col: int):
    fill = PatternFill("solid", fgColor="D9E1F2")
    font = Font(bold=True)
    border = light_border()
    for col in range(min_col, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border


def autosize_columns(ws, min_width: int = 10, max_width: int = 42):
    for col_cells in ws.columns:
        max_len = 0
        letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            value = cell.value
            if value is not None:
                max_len = max(max_len, len(str(value)))
        ws.column_dimensions[letter].width = max(min_width, min(max_width, max_len + 2))


def parse_float(text: str, field_name: str) -> float:
    try:
        return float(str(text).replace(",", "."))
    except Exception:
        raise ValueError(f"Поле '{field_name}' должно быть числом")


def parse_int(text: str, field_name: str) -> int:
    try:
        return int(float(str(text).replace(",", ".")))
    except Exception:
        raise ValueError(f"Поле '{field_name}' должно быть целым числом")


def build_report_voltage_grid(start: float, stop: float, step: float) -> List[float]:
    if step <= 0:
        raise ValueError("Шаг напряжения должен быть положительным")
    if stop < start:
        raise ValueError("Конец диапазона должен быть не меньше начала")
    values: List[float] = []
    current = float(start)
    while current <= stop + step / 2:
        values.append(round(current, 6))
        current += step
        if len(values) > 10000:
            raise ValueError("Слишком большая сетка напряжений для отчета")
    return values


def voltage_grid_missing(requested: Iterable[float], available: Iterable[float]) -> List[float]:
    available_set = {round(float(value), 6) for value in available}
    return [value for value in requested if round(float(value), 6) not in available_set]


def format_voltage(value: float) -> str:
    return f"{float(value):g}"

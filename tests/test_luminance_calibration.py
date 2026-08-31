from __future__ import annotations

import csv
import json
import os
import tempfile
import unittest
from pathlib import Path
from types import SimpleNamespace
from unittest.mock import patch

import numpy as np
from openpyxl import load_workbook
from openpyxl.worksheet._read_only import ReadOnlyWorksheet
from PIL import Image

from oled_app.gui.measurement_menu import (
    pixel_ids,
    refresh_ivl_thumbnails,
    refresh_pixel_table,
    show_ivl_hover_preview,
)
from oled_app.gui import spectral_calibration_window as spectral_calibration_gui
from oled_app.gui.spectral_calibration_window import (
    quarter_spectral_candidates,
    spectral_calibration_thresholds,
)
from oled_app.measurements.ivl import IVLParams
from oled_app.measurements.spectrum import SpectrumHelper, SpectrumParams
from oled_app.measurements.stability import StabilityParams
from oled_app.processing.ivl_preview import (
    _representative_cycles,
    create_ivl_thumbnail,
    ivl_thumbnail_needs_refresh,
    ivl_thumbnail_path,
)
from oled_app.processing.ivl_results import (
    confirmed_burned_cycle,
    final_ivl_status,
    save_ivl_workbook,
)
from oled_app.processing.luminance_recalculation import recalculate_series_luminance
from oled_app.processing.spectral_calibration import (
    calibrate_quarter_spectral_integral,
    calculate_spectral_integrals,
    create_spectral_recalculation_workbook,
    read_spectrum_integral_points,
    spectral_recalculation_output_path,
)
from oled_app.processing.spectrum_results import create_spectrum_workbook
from oled_app.processing.stability_results import create_stability_workbook
from oled_app.series.manager import SeriesManager
from oled_app.series.metadata import luminance_coefficient_for_color
from oled_app.series.paths import ensure_quarter_calibration_folder


class SpectralSensitivityTests(unittest.TestCase):
    def test_scope_calibration_is_saved_once_and_assigned_to_all_target_quarters(self):
        with tempfile.TemporaryDirectory() as folder:
            root = Path(folder)
            manager = object.__new__(SeriesManager)
            manager.series_folder = root
            manager.config_path = root / "series_config.json"
            manager.config = {
                "quarter_bases": {"1": "C", "2": "C", "3": "D", "4": "D"},
                "series_led_color": "red",
            }
            manager.journal = SimpleNamespace(config=manager.config)

            manager.save_scope_integral_calibration(
                (2, 1),
                {"method": "normalized_shape_integral_filtered_median", "coefficient": 3.49},
            )

            calibrations = manager.config["quarter_integral_calibrations"]
            self.assertEqual(calibrations["1"]["coefficient"], 3.49)
            self.assertEqual(calibrations["2"]["coefficient"], 3.49)
            self.assertEqual(calibrations["1"]["target_quarters"], [2, 1])
            self.assertEqual(
                calibrations["1"]["calibration_file"],
                calibrations["2"]["calibration_file"],
            )
            self.assertTrue((root / calibrations["1"]["calibration_file"]).is_file())

    def test_integral_calibration_files_live_in_quarter_folder(self):
        with tempfile.TemporaryDirectory() as folder:
            root = Path(folder)
            config = {
                "quarter_bases": {"1": "C"},
                "quarter_led_colors": {"1": "red"},
            }
            calibration_folder = ensure_quarter_calibration_folder(
                root,
                config,
                1,
            )
            source = root / "measurements" / "SPECTRUM_P1.xlsx"
            output = spectral_recalculation_output_path(
                source,
                "CR1_1_1",
                output_dir=calibration_folder,
            )
            manager = SeriesManager.__new__(SeriesManager)
            manager.series_folder = root
            manager.config = config
            manager.save_config = lambda: None

            manager.save_quarter_integral_calibration(
                1,
                {
                    "method": "normalized_shape_integral_median",
                    "coefficient": 3.0,
                },
            )

            json_path = root / "calibration" / "CR1" / "integral_calibration.json"
            self.assertEqual(output.parent, json_path.parent)
            self.assertTrue(json_path.exists())
            payload = json.loads(json_path.read_text(encoding="utf-8"))
            self.assertEqual(payload["coefficient"], 3.0)
            self.assertEqual(
                manager.config["quarter_integral_calibrations"]["1"][
                    "calibration_file"
                ],
                str(Path("calibration") / "CR1" / "integral_calibration.json"),
            )

    def test_spectrum_pixels_are_grouped_for_all_available_quarters(self):
        rows = [
            {
                "Pixel ID": "CB1_2_1",
                "Quarter number": 1,
                "Last spectrum file": "one.xlsx",
            },
            {
                "Pixel ID": "CB1_1_2",
                "Quarter number": 1,
                "Last spectrum file": "two.xlsx",
            },
            {
                "Pixel ID": "CB3_1_1",
                "Quarter number": 3,
                "Last spectrum file": "three.xlsx",
            },
            {
                "Pixel ID": "CB4_1_1",
                "Quarter number": 4,
                "Last spectrum file": "",
            },
        ]

        self.assertEqual(
            quarter_spectral_candidates(rows),
            {
                1: ["CB1_1_2", "CB1_2_1"],
                3: ["CB3_1_1"],
            },
        )

    def test_batch_spectral_recalculation_continues_after_one_error(self):
        rows = [
            {
                "Pixel ID": f"P{quarter}",
                "Quarter number": quarter,
                "Last spectrum file": f"p{quarter}.xlsx",
            }
            for quarter in range(1, 4)
        ]
        log_messages = []
        app = SimpleNamespace(
            series=SimpleNamespace(
                journal=SimpleNamespace(list_pixels=lambda: rows),
            ),
            log=log_messages.append,
        )
        completed = {
            "quarter": 1,
            "pixel": "P1",
            "coefficient": 3.49,
            "output": Path("SPECTRAL_RECALC_P1.xlsx"),
        }

        with (
            patch.object(
                spectral_calibration_gui,
                "ask_quarter_calibration_pixels",
                return_value=["P1", "P2", "P3"],
            ),
            patch.object(
                spectral_calibration_gui,
                "_calibrate_quarter_pixel",
                side_effect=[completed, ValueError("bad spectrum"), None],
            ) as process,
            patch.object(
                spectral_calibration_gui.messagebox,
                "showwarning",
            ) as warning,
        ):
            spectral_calibration_gui.calibrate_quarter_from_latest_spectrum(app)

        self.assertEqual(process.call_count, 3)
        warning.assert_called_once()
        self.assertIn("P2: bad spectrum", warning.call_args.args[1])
        self.assertIn("P3", warning.call_args.args[1])
        self.assertTrue(any("P2" in message for message in log_messages))

    def test_supplied_coupon_ratio_matches_manual_349_check(self):
        fixture = Path(__file__).resolve().parent / "fixtures" / "test_coupon_spectra.csv"
        wavelengths = []
        red = []
        green = []
        with fixture.open("r", encoding="utf-8", newline="") as handle:
            for row in csv.DictReader(handle):
                wavelengths.append(float(row["wavelength_nm"]))
                red.append(float(row["tcr_138_23"]))
                green.append(float(row["tcg_140_23"]))

        helper = SpectrumHelper(SpectrumParams(), lambda _message: None)
        wl = np.asarray(wavelengths, dtype=np.float64)
        red_values = np.asarray(red, dtype=np.float64)
        green_values = np.asarray(green, dtype=np.float64)
        _baseline, _region, red_background = helper.estimate_baseline(wl, red_values)
        _baseline, _region, green_background = helper.estimate_baseline(wl, green_values)

        red_integral = calculate_spectral_integrals(wl, red_values - red_background)
        green_integral = calculate_spectral_integrals(wl, green_values - green_background)
        ratio = green_integral.shape_integral / red_integral.shape_integral

        self.assertAlmostEqual(ratio, 3.4661641, places=5)
        self.assertAlmostEqual(ratio, 3.49, delta=0.05)

    def test_quarter_calibration_uses_median_normalized_integral(self):
        points = [
            {"voltage_V": 2.0, "shape_integral": 2.9, "photodiode_uA": 1.0, "status": "GOOD"},
            {"voltage_V": 3.0, "shape_integral": 3.0, "photodiode_uA": 2.0, "status": "GOOD"},
            {"voltage_V": 4.0, "shape_integral": 3.1, "photodiode_uA": 3.0, "status": "OK"},
        ]

        calibration = calibrate_quarter_spectral_integral(
            points,
            geometric_coefficient=2.0,
            source_pixel="CR1_1_1",
            source_file="spectrum.xlsx",
            integral_coefficient=1.5,
        )

        self.assertAlmostEqual(calibration.integral_coefficient, 1.5)
        self.assertAlmostEqual(calibration.coefficient, 3.0)
        self.assertAlmostEqual(calibration.effective_coefficient, 9.0)
        self.assertAlmostEqual(calibration.integral_min, 2.9)
        self.assertAlmostEqual(calibration.integral_max, 3.1)
        self.assertEqual(
            calibration.method,
            "normalized_shape_integral_filtered_median",
        )

    def test_quarter_calibration_removes_ten_percent_outlier(self):
        points = [
            {
                "point": index,
                "voltage_V": voltage,
                "shape_integral": integral,
                "photodiode_uA": float(index),
                "status": "GOOD",
            }
            for index, (voltage, integral) in enumerate(
                [(2.0, 3.0), (3.0, 3.05), (4.0, 2.95), (5.0, 8.0)],
                start=1,
            )
        ]

        calibration = calibrate_quarter_spectral_integral(
            points,
            geometric_coefficient=1.0,
            source_pixel="P1",
            source_file="spectrum.xlsx",
            activation_voltage_V=4.0,
        )

        self.assertAlmostEqual(calibration.coefficient, 3.0)
        self.assertEqual(calibration.points_used, 3)
        self.assertEqual(calibration.points_rejected, 1)
        self.assertEqual(calibration.included_point_numbers, (1, 2, 3))
        self.assertIsNone(calibration.activation_voltage_V)

    def test_quarter_calibration_fits_systematic_voltage_trend(self):
        points = [
            {
                "point": index,
                "voltage_V": voltage,
                "shape_integral": 0.8 * voltage + 0.5,
                "photodiode_uA": float(index),
                "status": "GOOD",
            }
            for index, voltage in enumerate((2.0, 3.0, 4.0, 5.0), start=1)
        ]

        calibration = calibrate_quarter_spectral_integral(
            points,
            geometric_coefficient=2.0,
            source_pixel="P1",
            source_file="spectrum.xlsx",
            integral_coefficient=3.0,
        )

        self.assertEqual(
            calibration.method,
            "normalized_shape_integral_linear_voltage",
        )
        self.assertAlmostEqual(calibration.slope_integral_per_V, 0.8)
        self.assertAlmostEqual(calibration.intercept_integral, 0.5)
        self.assertAlmostEqual(calibration.r_squared, 1.0)
        self.assertAlmostEqual(calibration.integral_at_voltage(4.0), 3.7)
        self.assertAlmostEqual(calibration.activation_voltage_V, 2.0)

    def test_median_tolerance_can_keep_a_wider_group_of_integrals(self):
        points = [
            {
                "point": index,
                "voltage_V": voltage,
                "shape_integral": integral,
                "photodiode_uA": float(index),
                "status": "GOOD",
            }
            for index, (voltage, integral) in enumerate(
                [(2.0, 1.0), (3.0, 1.2), (4.0, 1.4), (5.0, 1.6)],
                start=1,
            )
        ]

        calibration = calibrate_quarter_spectral_integral(
            points,
            geometric_coefficient=1.0,
            source_pixel="P1",
            source_file="spectrum.xlsx",
            median_tolerance_percent=30.0,
            linear_model_outlier_percent=50.0,
        )

        self.assertEqual(
            calibration.method,
            "normalized_shape_integral_filtered_median",
        )
        self.assertAlmostEqual(calibration.coefficient, 1.3)
        self.assertEqual(calibration.inlier_threshold_percent, 30.0)
        self.assertEqual(calibration.outlier_percent, 0.0)

    def test_linear_model_requires_configured_percentage_outside_tolerance(self):
        points = [
            {
                "point": index,
                "voltage_V": voltage,
                "shape_integral": integral,
                "photodiode_uA": float(index),
                "status": "GOOD",
            }
            for index, (voltage, integral) in enumerate(
                [(2.0, 1.0), (3.0, 1.2), (4.0, 1.4), (5.0, 1.6)],
                start=1,
            )
        ]

        median = calibrate_quarter_spectral_integral(
            points,
            geometric_coefficient=1.0,
            source_pixel="P1",
            source_file="spectrum.xlsx",
            median_tolerance_percent=10.0,
            linear_model_outlier_percent=75.0,
        )
        linear = calibrate_quarter_spectral_integral(
            points,
            geometric_coefficient=1.0,
            source_pixel="P1",
            source_file="spectrum.xlsx",
            median_tolerance_percent=10.0,
            linear_model_outlier_percent=50.0,
        )

        self.assertEqual(
            median.method,
            "normalized_shape_integral_filtered_median",
        )
        self.assertEqual(
            linear.method,
            "normalized_shape_integral_linear_voltage",
        )
        self.assertEqual(linear.outlier_percent, 50.0)
        self.assertEqual(linear.linear_model_outlier_threshold_percent, 50.0)

    def test_spectral_calibration_thresholds_are_validated(self):
        self.assertEqual(
            spectral_calibration_thresholds(
                {
                    "spectral_calibration": {
                        "median_tolerance_percent": 15,
                        "linear_model_outlier_percent": 60,
                    }
                }
            ),
            (15.0, 60.0),
        )
        with self.assertRaisesRegex(ValueError, "Допуск интеграла"):
            spectral_calibration_thresholds(
                {
                    "spectral_calibration": {
                        "median_tolerance_percent": 0,
                        "linear_model_outlier_percent": 60,
                    }
                }
            )

    def test_linear_calibration_ignores_points_before_opening_voltage(self):
        points = [
            {
                "point": index,
                "voltage_V": voltage,
                "shape_integral": voltage,
                "photodiode_uA": float(index),
                "status": "GOOD",
            }
            for index, voltage in enumerate((2.0, 3.0, 4.0, 5.0), start=1)
        ]

        calibration = calibrate_quarter_spectral_integral(
            points,
            geometric_coefficient=1.0,
            source_pixel="P1",
            source_file="spectrum.xlsx",
            activation_voltage_V=3.0,
        )

        self.assertEqual(
            calibration.method,
            "normalized_shape_integral_linear_voltage",
        )
        self.assertAlmostEqual(calibration.activation_voltage_V, 3.0)
        self.assertEqual(calibration.points_total, 4)
        self.assertEqual(calibration.points_used, 3)
        self.assertEqual(calibration.points_rejected, 1)
        self.assertAlmostEqual(calibration.integral_at_voltage(2.99), 3.0)
        self.assertAlmostEqual(calibration.integral_at_voltage(3.0), 3.0)
        self.assertAlmostEqual(calibration.integral_at_voltage(3.01), 3.01)

    def test_rgb_coefficient_is_multiplied_by_geometry(self):
        settings = {
            "measurement_units": {
                "luminance_red_cd_m2_per_uA": 7.5,
                "geometric_conversion_coefficient": 1.8,
            }
        }

        self.assertAlmostEqual(luminance_coefficient_for_color(settings, "red"), 13.5)

    def test_spectral_product_replaces_rgb_only_after_quarter_calibration(self):
        settings = {
            "measurement_units": {
                "luminance_red_cd_m2_per_uA": 5.0,
                "geometric_conversion_coefficient": 2.0,
                "integral_conversion_coefficient": 4.0,
            }
        }
        manager = SeriesManager.__new__(SeriesManager)
        manager.config = {
            "quarter_led_colors": {"1": "red"},
            "quarter_integral_calibrations": {},
        }
        manager.journal = SimpleNamespace(
            get_pixel=lambda _pixel: {"Quarter number": 1}
        )

        self.assertAlmostEqual(
            manager.luminance_coefficient_for_pixel("CR1_1_1", settings),
            10.0,
        )

        manager.config["quarter_integral_calibrations"]["1"] = {
            "method": "normalized_shape_integral_median",
            "coefficient": 3.0,
        }

        self.assertAlmostEqual(
            manager.luminance_coefficient_for_pixel("CR1_1_1", settings),
            24.0,
        )

        manager.config["quarter_integral_calibrations"]["1"] = {
            "method": "normalized_shape_integral_linear_voltage",
            "coefficient": 3.0,
            "reference_voltage_V": 3.0,
            "slope_integral_per_V": 1.0,
            "intercept_integral": 0.0,
        }
        manager.journal = SimpleNamespace(
            get_pixel=lambda _pixel: {
                "Quarter number": 1,
                "Opening voltage (V)": 3.0,
            }
        )
        self.assertAlmostEqual(
            manager.luminance_coefficient_for_pixel(
                "CR1_1_1",
                settings,
                voltage_V=2.99,
            ),
            24.0,
        )
        self.assertAlmostEqual(
            manager.luminance_coefficient_for_pixel(
                "CR1_1_1",
                settings,
                voltage_V=3.0,
            ),
            24.0,
        )
        self.assertAlmostEqual(
            manager.luminance_coefficient_for_pixel(
                "CR1_1_1",
                settings,
                voltage_V=4.0,
            ),
            32.0,
        )

    def test_on_demand_recalculation_saves_separate_workbook(self):
        with tempfile.TemporaryDirectory() as folder:
            root = Path(folder)
            source = root / "SPECTRUM_P1.xlsx"
            params = SpectrumParams(voltage_start=2.0, voltage_end=3.0, voltage_step=1.0)
            wb = create_spectrum_workbook(source, "P1", params, [2.0, 3.0])
            ws_sum = wb["Сводка"]
            ws_processed = wb["Processed counts per s"]
            for row, (point, voltage, photo, scale) in enumerate(
                [(1, 2.0, 2.0, 1.0), (2, 3.0, 4.0, 2.0)],
                start=22,
            ):
                ws_sum.cell(row, 1, point)
                ws_sum.cell(row, 2, voltage)
                ws_sum.cell(row, 7, photo)
                ws_sum.cell(row, 8, photo * 5.0)
                ws_sum.cell(row, 13, "GOOD")
                column = point + 1
                ws_processed.cell(2, column, point)
                ws_processed.cell(3, column, voltage)
                wavelengths = np.linspace(400.0, 700.0, 301)
                spectrum = scale * 1000.0 * np.exp(-0.5 * ((wavelengths - 530.0) / 35.0) ** 2)
                for data_row, (wavelength, intensity) in enumerate(
                    zip(wavelengths, spectrum),
                    start=21,
                ):
                    ws_processed.cell(data_row, 1, float(wavelength))
                    ws_processed.cell(data_row, column, float(intensity))
            wb.save(source)
            wb.close()
            source_before = source.read_bytes()

            with patch.object(
                ReadOnlyWorksheet,
                "cell",
                side_effect=AssertionError(
                    "read-only spectrum sheets must be streamed sequentially"
                ),
            ):
                points = read_spectrum_integral_points(source)
            for point in points:
                point["rgb_luminance_cd_m2"] = point["photodiode_uA"] * 5.0
            calibration = calibrate_quarter_spectral_integral(
                points,
                geometric_coefficient=2.0,
                source_pixel="P1",
                source_file=str(source),
            )
            output = root / "SPECTRAL_RECALC_P1.xlsx"
            create_spectral_recalculation_workbook(
                output,
                source,
                "P1",
                1,
                points,
                calibration,
                rgb_photodiode_coefficient=5.0,
            )

            self.assertTrue(output.exists())
            self.assertEqual(source.read_bytes(), source_before)
            result_wb = load_workbook(output, data_only=True)
            try:
                self.assertIn("Результаты", result_wb.sheetnames)
                ws_result = result_wb["Результаты"]
                self.assertEqual(ws_result["B4"].value, "P1")
                header_row = next(
                    row
                    for row in range(1, ws_result.max_row + 1)
                    if ws_result.cell(row, 1).value == "Point"
                )
                first_result_row = header_row + 1
                expected_luminance = (
                    ws_result.cell(first_result_row, 3).value
                    * calibration.coefficient
                    * calibration.integral_coefficient
                    * calibration.geometric_coefficient
                )
                self.assertAlmostEqual(
                    ws_result.cell(first_result_row, 10).value,
                    expected_luminance,
                )
            finally:
                result_wb.close()


class SpectrumPriorityTests(unittest.TestCase):
    def test_priority_pixels_are_listed_first(self):
        rows = [
            {"Pixel ID": "P1", "Spectrum priority": False},
            {"Pixel ID": "P2", "Spectrum priority": True},
            {"Pixel ID": "P3", "Spectrum priority": False},
        ]
        app = SimpleNamespace(
            series=SimpleNamespace(
                journal=SimpleNamespace(list_pixels=lambda: rows),
            )
        )

        self.assertEqual(pixel_ids(app), ["P2", "P1", "P3"])


class IvlStatusAndThumbnailTests(unittest.TestCase):
    def test_table_refresh_only_builds_thumbnails_when_explicitly_requested(self):
        class FakeTree:
            def get_children(self):
                return []

            def delete(self, _item):
                pass

            def insert(self, *_args, **_kwargs):
                pass

        app = SimpleNamespace(
            series=SimpleNamespace(
                series_folder=Path("series"),
                journal=SimpleNamespace(list_pixels=lambda: []),
            ),
            tree=FakeTree(),
        )
        with (
            patch("oled_app.gui.measurement_menu.render_status_holder_canvas"),
            patch("oled_app.gui.measurement_menu.refresh_ivl_history_tree"),
            patch("oled_app.gui.measurement_menu.refresh_ivl_thumbnails_async") as refresh_async,
        ):
            refresh_pixel_table(app)
            refresh_async.assert_not_called()

            refresh_pixel_table(app, refresh_thumbnails=True)
            refresh_async.assert_called_once_with(app)

    def test_missing_hover_thumbnail_is_scheduled_without_sync_workbook_read(self):
        with tempfile.TemporaryDirectory() as folder:
            root = Path(folder)
            workbook = root / "measurements" / "IVL_CR1_1_1.xlsx"
            workbook.parent.mkdir(parents=True)
            workbook.touch()
            app = SimpleNamespace(
                series=SimpleNamespace(
                    series_folder=root,
                    journal=SimpleNamespace(
                        get_pixel=lambda _pixel_id: {
                            "Last IVL file": str(workbook.relative_to(root)),
                        }
                    ),
                ),
            )
            event = SimpleNamespace(x_root=10, y_root=10)
            with (
                patch("oled_app.gui.measurement_menu.hide_ivl_hover_preview"),
                patch("oled_app.gui.measurement_menu.ivl_thumbnail_needs_refresh", return_value=True),
                patch("oled_app.gui.measurement_menu.refresh_ivl_thumbnails_async") as refresh_async,
                patch("oled_app.gui.measurement_menu.show_ivl_hover_message") as show_message,
                patch("oled_app.gui.measurement_menu.create_ivl_thumbnail_from_workbook") as create_sync,
            ):
                show_ivl_hover_preview(app, "CR1_1_1", event)

            refresh_async.assert_called_once_with(app, "CR1_1_1")
            show_message.assert_called_once()
            create_sync.assert_not_called()

    def test_refresh_action_checks_and_creates_latest_ivl_thumbnail(self):
        cycle = {
            "cycle": 1,
            "status": "WORKING",
            "status_desc": "ok",
            "current_limit_reached": False,
            "max_photo_uA": 1.0,
            "max_current_mA": 1.0,
            "opening_voltage": 2.0,
            "data": [
                {
                    "Point": 1,
                    "Voltage set (V)": 2.0,
                    "Voltage OLED / LED measured (V)": 2.0,
                    "Current OLED / LED (mA)": 1.0,
                    "Current density (mA/cm^2)": 100.0,
                    "Voltage photodiode measured (V)": -5.0,
                    "Photodiode current (uA)": 1.0,
                    "Luminance (cd/m^2)": 1.0,
                    "Measurement time (s)": 0.1,
                },
                {
                    "Point": 2,
                    "Voltage set (V)": 3.0,
                    "Voltage OLED / LED measured (V)": 3.0,
                    "Current OLED / LED (mA)": 2.0,
                    "Current density (mA/cm^2)": 200.0,
                    "Voltage photodiode measured (V)": -5.0,
                    "Photodiode current (uA)": 2.0,
                    "Luminance (cd/m^2)": 2.0,
                    "Measurement time (s)": 0.2,
                },
            ],
        }
        with tempfile.TemporaryDirectory() as folder:
            root = Path(folder)
            workbook = (
                root
                / "measurements"
                / "01_IVL"
                / "2026-07-30"
                / "CR1"
                / "CR1_1"
                / "CR1_1_1"
                / "IVL_CR1_1_1_20260730_120000.xlsx"
            )
            workbook.parent.mkdir(parents=True)
            save_ivl_workbook("CR1_1_1", workbook, IVLParams(), [cycle])
            app = SimpleNamespace(
                series=SimpleNamespace(
                    series_folder=root,
                    journal=SimpleNamespace(
                        list_pixels=lambda: [
                            {
                                "Pixel ID": "CR1_1_1",
                                "Last IVL file": str(workbook.relative_to(root)),
                            }
                        ]
                    ),
                ),
                log=lambda _message: None,
            )

            self.assertEqual(refresh_ivl_thumbnails(app), 1)
            self.assertTrue(
                root.joinpath(
                    "thumbnails",
                    "CR1_1_1_thumbnail.png",
                ).exists()
            )
            self.assertEqual(refresh_ivl_thumbnails(app), 0)

    def test_alive_confirmation_clears_preliminary_burned_status(self):
        cycles = [
            {"cycle": 1, "status": "BURNED", "data": [{"Point": 1}]},
            {"cycle": 2, "status": "WORKING", "data": [{"Point": 1}]},
        ]

        self.assertIsNone(confirmed_burned_cycle(cycles))
        self.assertEqual(final_ivl_status(cycles), "WORKING")
        self.assertEqual(_representative_cycles(cycles), [cycles[1]])

    def test_nonworking_confirmation_keeps_burned_curve_and_status(self):
        cycles = [
            {"cycle": 1, "status": "BURNED", "data": [{"Point": 1}]},
            {"cycle": 2, "status": "NONWORKING", "data": [{"Point": 1}]},
        ]

        self.assertEqual(confirmed_burned_cycle(cycles), 1)
        self.assertEqual(final_ivl_status(cycles), "BURNED")
        self.assertEqual(_representative_cycles(cycles), [cycles[0]])

    def test_thumbnail_path_is_fixed_per_pixel(self):
        first = Path("folder/IVL_CR1_2_1_01-01-2026_10h00m00s.xlsx")
        second = Path("folder/IVL_CR1_2_1_02-01-2026_11h00m00s.xlsx")

        self.assertEqual(ivl_thumbnail_path(first), ivl_thumbnail_path(second))
        self.assertEqual(
            ivl_thumbnail_path(first),
            Path("folder/thumbnails/CR1_2_1_thumbnail.png"),
        )

    def test_series_thumbnails_are_collected_in_one_root_folder(self):
        workbook = Path(
            "series/measurements/01_IVL/2026-07-30/CR1/CR1_2/CR1_2_1/"
            "IVL_CR1_2_1_20260730_120000.xlsx"
        )

        self.assertEqual(
            ivl_thumbnail_path(workbook),
            Path("series/thumbnails/CR1_2_1_thumbnail.png"),
        )

    def test_thumbnail_is_stale_when_new_ivl_workbook_is_newer(self):
        with tempfile.TemporaryDirectory() as folder:
            root = Path(folder)
            workbook = root / "IVL_P1_20260101_000000.xlsx"
            thumbnail = root / "thumbnails" / "P1_thumbnail.png"
            workbook.touch()
            create_ivl_thumbnail(thumbnail, [])
            newer = thumbnail.stat().st_mtime_ns + 1_000_000
            os.utime(workbook, ns=(newer, newer))

            self.assertTrue(
                ivl_thumbnail_needs_refresh(thumbnail, workbook)
            )

    def test_thumbnail_uses_live_colors_and_one_linear_scale(self):
        cycle = {
            "cycle": 1,
            "status": "WORKING",
            "data": [
                {
                    "Voltage OLED / LED measured (V)": 0.0,
                    "Current OLED / LED (mA)": 0.0,
                    "Photodiode current (uA)": 0.0,
                },
                {
                    "Voltage OLED / LED measured (V)": 1.0,
                    "Current OLED / LED (mA)": 1.0,
                    "Photodiode current (uA)": 2.0,
                },
            ],
        }
        with tempfile.TemporaryDirectory() as folder:
            thumbnail = Path(folder) / "preview.png"
            create_ivl_thumbnail(thumbnail, [cycle])

            with Image.open(thumbnail) as image:
                self.assertFalse(ivl_thumbnail_needs_refresh(thumbnail))
                rgb = image.convert("RGB")
                blue = (11, 97, 164)
                red = (196, 60, 48)
                right_edge = rgb.width - 18
                blue_end = [
                    y
                    for x in range(right_edge - 3, right_edge + 1)
                    for y in range(18, rgb.height - 20)
                    if rgb.getpixel((x, y)) == blue
                ]
                red_end = [
                    y
                    for x in range(right_edge - 3, right_edge + 1)
                    for y in range(18, rgb.height - 20)
                    if rgb.getpixel((x, y)) == red
                ]

            self.assertTrue(blue_end)
            self.assertTrue(red_end)
            self.assertGreater(min(blue_end), min(red_end) + 40)

    def test_legacy_thumbnail_is_refreshed_once(self):
        with tempfile.TemporaryDirectory() as folder:
            thumbnail = Path(folder) / "legacy.png"
            Image.new("RGB", (10, 10), "white").save(thumbnail)

            self.assertTrue(ivl_thumbnail_needs_refresh(thumbnail))


class ExistingSeriesRecalculationTests(unittest.TestCase):
    def test_ivl_raw_is_restored_and_workbook_is_replaced(self):
        cycle = {
            "cycle": 1,
            "status": "WORKING",
            "status_desc": "ok",
            "current_limit_reached": False,
            "max_photo_uA": 2.0,
            "max_current_mA": 1.0,
            "opening_voltage": 2.5,
            "data": [
                {
                    "Point": 1,
                    "Voltage set (V)": 2.5,
                    "Voltage OLED / LED measured (V)": 2.49,
                    "Current OLED / LED (mA)": 1.0,
                    "Current density (mA/cm^2)": 100.0,
                    "Voltage photodiode measured (V)": -5.0,
                    "Photodiode current (uA)": 2.0,
                    "Luminance (cd/m^2)": 2.0,
                    "Measurement time (s)": 0.1,
                }
            ],
        }
        with tempfile.TemporaryDirectory() as folder:
            root = Path(folder)
            workbook = root / "IVL_P1_20260101_000000.xlsx"
            save_ivl_workbook("P1", workbook, IVLParams(), [cycle])
            journal = SimpleNamespace(
                list_measurements=lambda: [
                    {"Type": "IVL", "Pixel ID": "P1", "File": str(workbook)}
                ]
            )
            series = SimpleNamespace(
                series_folder=root,
                journal=journal,
                luminance_coefficient_for_pixel=lambda _pixel, _settings: 6.0,
                geometric_coefficient=lambda _settings: 2.0,
            )
            settings = {
                "measurement_units": {"pixel_area_mm2": 1.0},
                "raw_data": {"folder_name": "raw_data"},
            }

            report = recalculate_series_luminance(series, settings)

            self.assertEqual(report.workbooks_updated, 1)
            self.assertEqual(report.raw_files_restored, 1)
            self.assertTrue(root.joinpath("raw_data", f"{workbook.stem}_raw.csv").exists())
            self.assertTrue(ivl_thumbnail_path(workbook).exists())
            wb = load_workbook(workbook, data_only=True)
            try:
                ws = wb["Cycle_1"]
                headers = {
                    str(ws.cell(4, column).value or ""): column
                    for column in range(1, ws.max_column + 1)
                }
                luminance = ws.cell(5, headers["Luminance (cd/m^2)"]).value
                self.assertAlmostEqual(luminance, 12.0)
            finally:
                wb.close()

    def test_spectrum_and_stability_raw_are_restored_from_workbooks(self):
        with tempfile.TemporaryDirectory() as folder:
            root = Path(folder)
            spectrum = root / "SPECTRUM_P1_20260101_000000.xlsx"
            spectrum_wb = create_spectrum_workbook(
                spectrum,
                "P1",
                SpectrumParams(voltage_start=2.0, voltage_end=2.0, voltage_step=1.0),
                [2.0],
            )
            ws_sum = spectrum_wb["Сводка"]
            summary_values = [
                1,
                2.0,
                1.99,
                0.5,
                50.0,
                -5.0,
                3.0,
                3.0,
                100.0,
                530.0,
                1000.0,
                40.0,
                "GOOD",
            ]
            for column, value in enumerate(summary_values, start=1):
                ws_sum.cell(22, column, value)
            for sheet_name in ("Raw spectra", "Dark corrected"):
                ws = spectrum_wb[sheet_name]
                ws.cell(2, 2, 1)
                ws.cell(3, 2, 2.0)
                ws.cell(10, 2, 100.0)
                for row, (wavelength, value) in enumerate(
                    [(500.0, 10.0), (510.0, 20.0), (520.0, 10.0)],
                    start=21,
                ):
                    ws.cell(row, 1, wavelength)
                    ws.cell(row, 2, value if sheet_name == "Raw spectra" else value - 1.0)
            spectrum_wb.save(spectrum)
            spectrum_wb.close()

            stability = root / "STABILITY_P1_20260101_000000.xlsx"
            stability_wb = create_stability_workbook(
                stability,
                "P1",
                StabilityParams(),
            )
            stability_wb["Data"].append(
                [
                    1,
                    "2026-01-01 00:00:00",
                    0.1,
                    "current",
                    1.0,
                    None,
                    2.0,
                    1.99,
                    0.5,
                    50.0,
                    -5.0,
                    3.0,
                    3.0,
                ]
            )
            stability_wb.save(stability)
            stability_wb.close()

            journal = SimpleNamespace(
                list_measurements=lambda: [
                    {"Type": "SPECTRUM", "Pixel ID": "P1", "File": str(spectrum)},
                    {"Type": "STABILITY", "Pixel ID": "P1", "File": str(stability)},
                ]
            )
            series = SimpleNamespace(
                series_folder=root,
                journal=journal,
                luminance_coefficient_for_pixel=lambda _pixel, _settings: 4.0,
                geometric_coefficient=lambda _settings: 2.0,
            )
            settings = {
                "measurement_units": {"pixel_area_mm2": 1.0},
                "raw_data": {"folder_name": "raw_data"},
            }

            report = recalculate_series_luminance(series, settings)

            self.assertEqual(report.errors, 0)
            self.assertEqual(report.workbooks_updated, 2)
            self.assertEqual(report.raw_files_restored, 3)
            raw_dir = root / "raw_data"
            self.assertTrue(raw_dir.joinpath(f"{spectrum.stem}_summary_raw.csv").exists())
            self.assertTrue(raw_dir.joinpath(f"{spectrum.stem}_spectra_raw.csv").exists())
            self.assertTrue(raw_dir.joinpath(f"{stability.stem}_raw.csv").exists())

            spectrum_result = load_workbook(spectrum, data_only=True)
            stability_result = load_workbook(stability, data_only=True)
            try:
                self.assertAlmostEqual(spectrum_result["Сводка"]["H22"].value, 12.0)
                self.assertAlmostEqual(stability_result["Data"]["M22"].value, 12.0)
            finally:
                spectrum_result.close()
                stability_result.close()

    def test_ivl_recalculation_uses_voltage_linear_integral(self):
        cycle = {
            "cycle": 1,
            "status": "WORKING",
            "status_desc": "ok",
            "current_limit_reached": False,
            "max_photo_uA": 1.0,
            "max_current_mA": 1.0,
            "opening_voltage": 2.0,
            "data": [
                {
                    "Point": point,
                    "Voltage set (V)": voltage,
                    "Voltage OLED / LED measured (V)": voltage - 0.05,
                    "Current OLED / LED (mA)": 1.0,
                    "Current density (mA/cm^2)": 100.0,
                    "Voltage photodiode measured (V)": -5.0,
                    "Photodiode current (uA)": 1.0,
                    "Luminance (cd/m^2)": 1.0,
                    "Measurement time (s)": 0.1 * point,
                }
                for point, voltage in ((1, 2.0), (2, 4.0))
            ],
        }
        with tempfile.TemporaryDirectory() as folder:
            root = Path(folder)
            workbook = root / "IVL_P1_20260101_000000.xlsx"
            save_ivl_workbook("P1", workbook, IVLParams(), [cycle])
            model = {
                "method": "normalized_shape_integral_linear_voltage",
                "coefficient": 3.0,
                "activation_voltage_V": 3.0,
                "reference_voltage_V": 3.0,
                "slope_integral_per_V": 1.0,
                "intercept_integral": 0.0,
                "integral_coefficient": 2.0,
                "geometric_coefficient": 3.0,
            }
            series = SimpleNamespace(
                series_folder=root,
                journal=SimpleNamespace(
                    list_measurements=lambda: [
                        {"Type": "IVL", "Pixel ID": "P1", "File": str(workbook)}
                    ]
                ),
                luminance_coefficient_for_pixel=lambda _pixel, _settings: 18.0,
                rgb_luminance_coefficient_for_pixel=lambda _pixel, _settings: 10.0,
                luminance_model_for_pixel=lambda _pixel, _settings: model,
                geometric_coefficient=lambda _settings: 3.0,
            )
            settings = {
                "measurement_units": {"pixel_area_mm2": 1.0},
                "raw_data": {"folder_name": "raw_data"},
            }

            report = recalculate_series_luminance(series, settings)

            self.assertEqual(report.errors, 0)
            result = load_workbook(workbook, data_only=True)
            try:
                ws = result["Cycle_1"]
                self.assertAlmostEqual(ws["H5"].value, 18.0)
                self.assertAlmostEqual(ws["H6"].value, 24.0)
            finally:
                result.close()


if __name__ == "__main__":
    unittest.main()

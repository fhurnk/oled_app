import csv
import tempfile
import time
import unittest
from pathlib import Path
from unittest.mock import patch

from openpyxl import load_workbook
import json
import urllib.request
import urllib.error
from oled_v2.server import LocalBackend
from oled_v2.config import CLIENT_HEADER, SESSION_HEADER
from oled_v2.ivl import IvlController


def wait_terminal(controller):
    deadline = time.monotonic() + 10
    while controller.snapshot()["active"] and time.monotonic() < deadline:
        time.sleep(0.01)
    state = controller.snapshot()
    if state["active"]:
        raise AssertionError("IVL did not terminate")
    return state


class IvlTests(unittest.TestCase):
    def setUp(self):
        self.temp = tempfile.TemporaryDirectory()
        self.controller = IvlController(Path(self.temp.name) / "Тест ВАЯХ")

    def tearDown(self):
        self.controller.shutdown()
        self.temp.cleanup()

    def test_reject_invalid_preflight_without_writing(self):
        for payload in ({"sweep_increment": 0}, {"sweep_end": -1},
                        {"sweep_end": 0}, {"sweep_end": float("nan")},
                        {"sweep_increment": 0.001}, {"current_limit_mA": True},
                        {"working_confirmation_points": 1.5}, {"hardware_mode": "real"}):
            with self.subTest(payload=payload), self.assertRaises(ValueError):
                self.controller.preflight(payload)
        self.assertFalse(self.controller.output_root.exists())

    def test_complete_uses_compatible_raw_and_workbook(self):
        self.controller.start({"sweep_end": 0.2, "sweep_increment": 0.1})
        state = wait_terminal(self.controller)
        self.assertEqual(state["status"], "completed", state)
        self.assertTrue(state["safe_shutdown_confirmed"])
        with Path(state["raw_file"]).open(encoding="utf-8") as stream:
            rows = list(csv.DictReader(stream))
        self.assertEqual(len(rows), len(state["points"]))
        self.assertGreater(len(rows), 1)
        wb = load_workbook(state["result"]["file"], read_only=True)
        try:
            self.assertIn("Summary", wb.sheetnames)
        finally:
            wb.close()
        state["points"].clear()
        self.assertTrue(self.controller.snapshot()["points"])

    def test_stop_keeps_raw_without_misrepresenting_partial_cycle(self):
        self.controller.start({"sweep_time_per_point": 0.05})
        deadline = time.monotonic() + 3
        while not self.controller.snapshot()["points"] and time.monotonic() < deadline:
            time.sleep(0.01)
        self.controller.stop()
        state = wait_terminal(self.controller)
        self.assertEqual(state["status"], "stopped")
        self.assertTrue(state["safe_shutdown_confirmed"])
        self.assertTrue(Path(state["raw_file"]).is_file())
        self.assertIsNone(state["result"])
        self.assertFalse(list(self.controller.output_root.rglob("*.xlsx")))

    def test_failure_still_shuts_down_device(self):
        with patch("oled_v2.ivl.run_ivl_cycle", side_effect=RuntimeError("measurement failed")):
            self.controller.start({})
            state = wait_terminal(self.controller)
        self.assertEqual(state["status"], "failed")
        self.assertEqual(state["error"], "measurement failed")
        self.assertTrue(state["safe_shutdown_confirmed"])

    def test_shutdown_stops_active_worker(self):
        self.controller.start({})
        self.controller.shutdown()
        self.assertFalse(self.controller.snapshot()["active"])
        self.assertEqual(self.controller.snapshot()["status"], "stopped")

    def test_api_auth_busy_and_snapshot_recovery(self):
        with LocalBackend(series_root=Path(self.temp.name)) as backend:
            controller = backend.server.config.app.state.ivl_controller
            controller.output_root = self.controller.output_root
            headers = {}
            def request(path, payload=None):
                req = urllib.request.Request(backend.session.origin + path, headers=headers,
                    data=json.dumps(payload).encode() if payload is not None else None)
                try:
                    with urllib.request.urlopen(req, timeout=5) as response:
                        return response.status, json.load(response)
                except urllib.error.HTTPError as exc:
                    return exc.code, json.load(exc)
            self.assertEqual(request("/api/ivl/state")[0], 401)
            headers.update({SESSION_HEADER: backend.session.token,
                            CLIENT_HEADER: "ivl-test-client-0001", "Content-Type": "application/json"})
            self.assertEqual(request("/api/ivl/preflight", {"sweep_increment": 0})[0], 422)
            code, started = request("/api/ivl/start", {"sweep_time_per_point": 0.5})
            self.assertEqual(code, 202)
            for path in ("/api/ivl/start", "/api/poc/start", "/api/poc/probe"):
                self.assertEqual(request(path, {})[0], 409)
            snapshot = request("/api/ivl/state")[1]
            self.assertEqual(snapshot["run_id"], started["run_id"])
            request("/api/ivl/stop", {})
            wait_terminal(controller)



if __name__ == "__main__":
    unittest.main()

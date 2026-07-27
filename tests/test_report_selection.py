import tempfile
import unittest
from datetime import datetime
from pathlib import Path
from types import SimpleNamespace
from unittest.mock import patch

from openpyxl import Workbook

from oled_app.gui.report_window import (
    collect_report_spectrum_candidates,
    report_output_name,
    selected_report_candidates,
)
from oled_app.gui.spectrum_window import spectrum_selection_visibility
from oled_app.reports.origin_report import (
    IvRecord,
    MeasurementPath,
    REPORT_MODE_FULL,
    REPORT_MODE_IVL,
    REPORT_MODE_SPECTRA,
    ReportData,
    SpectrumRecord,
    build_workbook,
    collect_iv_records,
    collect_report_data,
    collect_spectrum_records,
    create_origin_iv_book,
    parse_args,
    series_quarter_number,
)


class _Var:
    def __init__(self, value: str):
        self.value = value

    def get(self) -> str:
        return self.value


class _OriginSheet:
    def __init__(self):
        self.name = ""
        self.formulas = {}

    def from_list(self, *_args, **_kwargs):
        pass

    def set_formula(self, column: int, formula: str):
        self.formulas[column] = formula


class _OriginBook:
    def __init__(self):
        self.sheets = [_OriginSheet()]

    def __getitem__(self, index: int):
        return self.sheets[index]

    def add_sheet(self, name=""):
        sheet = _OriginSheet()
        sheet.name = name
        self.sheets.append(sheet)
        return sheet


class _OriginApp:
    def __init__(self):
        self.book = _OriginBook()

    def find_book(self, *_args):
        return self.book


def _write_spectrum(path: Path, voltage: float = 2.0) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "Processed counts per s"
    ws.cell(row=1, column=1, value="V set (V)")
    ws.cell(row=1, column=2, value=voltage)
    wb.save(path)


class ReportWindowSelectionTests(unittest.TestCase):
    def test_output_name_reflects_selected_report_mode(self):
        self.assertEqual(report_output_name("2026-07-20", "", report_mode=REPORT_MODE_IVL), "report_IVL_2026-07-20.opju")
        self.assertEqual(
            report_output_name("", "2026-07-21", report_mode=REPORT_MODE_SPECTRA),
            "report_Spctr_2026-07-21.opju",
        )
        self.assertEqual(
            report_output_name("2026-07-21", "2026-07-21", report_mode=REPORT_MODE_FULL),
            "report_2026-07-21.opju",
        )

    def test_spectrum_mode_hides_irrelevant_selector(self):
        self.assertEqual(spectrum_selection_visibility("single"), (True, False))
        self.assertEqual(spectrum_selection_visibility("substrate"), (True, True))

    def test_candidates_are_grouped_by_series_then_substrate(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            series_folder = Path(temp_dir)
            spectra_root = series_folder / "measurements" / "02_SPECTRA" / "2026-07-18" / "CG1"
            _write_spectrum(spectra_root / "CG1_1" / "CG1_1_1" / "SPECTRUM_CG1_1_1.xlsx")
            _write_spectrum(spectra_root / "CG1_2" / "CG1_2_1" / "SPECTRUM_CG1_2_1.xlsx")
            app = SimpleNamespace(series=SimpleNamespace(series_folder=series_folder))

            candidates = collect_report_spectrum_candidates(app, "2026-07-18")

            self.assertEqual(set(candidates), {"CG1"})
            self.assertEqual(set(candidates["CG1"]), {"CG1_1", "CG1_2"})
            self.assertEqual(set(candidates["CG1"]["CG1_2"]), {"CG1_2_1"})

    def test_excluded_quarter_is_hidden_from_spectrum_candidates(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            series_folder = Path(temp_dir)
            spectra_root = series_folder / "measurements" / "02_SPECTRA" / "2026-07-18"
            _write_spectrum(spectra_root / "CG1" / "CG1_1" / "CG1_1_1" / "SPECTRUM_CG1_1_1.xlsx")
            _write_spectrum(spectra_root / "CG2" / "CG2_1" / "CG2_1_1" / "SPECTRUM_CG2_1_1.xlsx")
            app = SimpleNamespace(series=SimpleNamespace(series_folder=series_folder))

            candidates = collect_report_spectrum_candidates(app, "2026-07-18", {2})

            self.assertEqual(set(candidates), {"CG1"})

    def test_selected_pixel_is_limited_to_selected_substrate(self):
        candidates = {
            "CG1": {
                "CG1_1": {"CG1_1_1": {"file": Path("one.xlsx"), "voltages": [2.0]}},
                "CG1_2": {"CG1_2_3": {"file": Path("two.xlsx"), "voltages": [2.1]}},
            }
        }

        selected = selected_report_candidates(
            candidates,
            {"CG1": _Var("CG1_2")},
            {"CG1": _Var("CG1_2_3")},
        )

        self.assertEqual(set(selected), {"CG1_2_3"})
        self.assertEqual(selected["CG1_2_3"]["series"], "CG1")
        self.assertEqual(selected["CG1_2_3"]["subseries"], "CG1_2")


class ReportBuilderSelectionTests(unittest.TestCase):
    def test_cli_accepts_each_report_mode(self):
        for mode in (REPORT_MODE_FULL, REPORT_MODE_IVL, REPORT_MODE_SPECTRA):
            with self.subTest(mode=mode):
                self.assertEqual(parse_args(["--report-mode", mode]).report_mode, mode)

    def test_cli_accepts_repeated_excluded_quarters(self):
        args = parse_args(["--exclude-quarter", "2", "--exclude-quarter", "4"])
        self.assertEqual(args.exclude_quarter, [2, 4])

    def test_series_quarter_number_uses_trailing_number(self):
        self.assertEqual(series_quarter_number("CLR3"), 3)
        self.assertEqual(series_quarter_number("quarter12"), 12)
        self.assertIsNone(series_quarter_number("UNKNOWN"))

    def test_excluded_measurement_root_is_not_required_in_strict_mode(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            measurements = Path(temp_dir)
            (measurements / "01_IVL_VAH").mkdir()
            ivl_args = parse_args(
                ["--measurements-dir", str(measurements), "--report-mode", REPORT_MODE_IVL, "--strict"]
            )
            self.assertEqual(collect_report_data(ivl_args).warnings, [])

        with tempfile.TemporaryDirectory() as temp_dir:
            measurements = Path(temp_dir)
            (measurements / "02_SPECTRA").mkdir()
            spectra_args = parse_args(
                ["--measurements-dir", str(measurements), "--report-mode", REPORT_MODE_SPECTRA, "--strict"]
            )
            self.assertEqual(collect_report_data(spectra_args).warnings, [])

    def test_workbook_contains_only_requested_sections(self):
        data = ReportData([], [], None, [])
        cases = {
            REPORT_MODE_FULL: ({"IVL_U_I_PD", "Spectra_by_voltage"}, set()),
            REPORT_MODE_IVL: ({"IVL_U_I_PD"}, {"Spectra_by_voltage", "Spectra_max_common"}),
            REPORT_MODE_SPECTRA: ({"Spectra_by_voltage"}, {"IVL_U_I_PD", "IVL_I_PD"}),
        }
        for mode, (included, excluded) in cases.items():
            with self.subTest(mode=mode):
                args = parse_args(["--report-mode", mode, "--format", "xlsx", "--no-charts"])
                workbook, _warnings = build_workbook(args, data)
                self.assertTrue(included.issubset(workbook.sheetnames))
                self.assertTrue(excluded.isdisjoint(workbook.sheetnames))

    def test_jl_columns_use_direct_origin_column_references(self):
        records = [
            IvRecord(
                Path("ivl.xlsx"),
                "2026-07-19",
                "CG1",
                "CG1_1",
                "CG1_1_1",
                "Cycle_1",
                "WORKING",
                [{"voltage": 2.0, "density": 1.5, "luminance": 100.0}],
            )
        ]

        with patch("oled_app.reports.origin_report.origin_set_folder"), patch(
            "oled_app.reports.origin_report.origin_move_page_to_folder"
        ):
            result = create_origin_iv_book(_OriginApp(), records)

        self.assertEqual(result["jl"].formulas, {0: "Sheet1!B", 1: "Sheet1!C"})

    def test_series_selection_keeps_one_substrate_and_pixel(self):
        timestamp = datetime(2026, 7, 18, 12, 0, 0)
        paths = {
            "CG1_1_1": MeasurementPath(Path("one.xlsx"), "2026-07-18", "CG1", "CG1_1", "CG1_1_1", timestamp),
            "CG1_2_3": MeasurementPath(Path("two.xlsx"), "2026-07-18", "CG1", "CG1_2", "CG1_2_3", timestamp),
        }

        def make_record(meta, _sheet_name, _warnings):
            return SpectrumRecord(
                meta.path,
                meta.date_dir,
                meta.series,
                meta.subseries,
                meta.pixel,
                [2.0],
                [500.0],
                [[100.0]],
            )

        warnings = []
        with patch("oled_app.reports.origin_report.latest_by_pixel", return_value=paths), patch(
            "oled_app.reports.origin_report.read_spectrum_record",
            side_effect=make_record,
        ):
            records = collect_spectrum_records(
                Path("spectra"),
                {},
                "Processed counts per s",
                True,
                warnings,
                "2026-07-18",
                {"CG1": "CG1_2_3"},
            )

        self.assertEqual([record.pixel for record in records], ["CG1_2_3"])
        self.assertEqual(records[0].subseries, "CG1_2")
        self.assertEqual(warnings, [])

    def test_excluded_quarter_is_skipped_before_reading_measurements(self):
        timestamp = datetime(2026, 7, 18, 12, 0, 0)
        paths = {
            "CG1_1_1": MeasurementPath(Path("one.xlsx"), "2026-07-18", "CG1", "CG1_1", "CG1_1_1", timestamp),
            "CG2_1_1": MeasurementPath(Path("two.xlsx"), "2026-07-18", "CG2", "CG2_1", "CG2_1_1", timestamp),
        }

        def make_spectrum_record(meta, _sheet_name, _warnings):
            return SpectrumRecord(
                meta.path,
                meta.date_dir,
                meta.series,
                meta.subseries,
                meta.pixel,
                [2.0],
                [500.0],
                [[100.0]],
            )

        def make_iv_record(meta, _warnings):
            return IvRecord(
                meta.path,
                meta.date_dir,
                meta.series,
                meta.subseries,
                meta.pixel,
                "Cycle_1",
                "WORKING",
                [],
            )

        with patch("oled_app.reports.origin_report.latest_by_pixel", return_value=paths), patch(
            "oled_app.reports.origin_report.read_spectrum_record",
            side_effect=make_spectrum_record,
        ), patch(
            "oled_app.reports.origin_report.read_iv_record",
            side_effect=make_iv_record,
        ):
            spectra = collect_spectrum_records(
                Path("spectra"),
                {},
                "Processed counts per s",
                False,
                [],
                excluded_quarters={2},
            )
            ivl = collect_iv_records(Path("ivl"), [], excluded_quarters={2})

        self.assertEqual([record.series for record in spectra], ["CG1"])
        self.assertEqual([record.series for record in ivl], ["CG1"])

    def test_cli_accepts_series_pixel_selection(self):
        args = parse_args(["--spectrum-series-pixel", "CG1=CG1_2_3"])
        self.assertEqual(args.spectrum_series_pixel, ["CG1=CG1_2_3"])


if __name__ == "__main__":
    unittest.main()
